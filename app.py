def build_admin_dashboard_data(user, branch_id=None):
    """Collect metrics, listings, and charts for the admin/HR dashboard using the current schema.
    
    Args:
        user: The current user object
        branch_id: Optional branch_id for filtering. If None, uses branch_scope for non-HR users.
                   For HR users, None means all branches, or a specific branch_id for filtering.
    """
    role = user.get('role') if user else None
    role_label = 'System Administrator' if role == 'admin' else 'HR Manager'
    # HR users can manage all branches - no scoping unless branch_id is explicitly provided
    if branch_id is None:
        branch_id = None if role == 'hr' else get_branch_scope(user)

    branch_info = {}
    if branch_id:
        branch_rows = fetch_rows(
            """
            SELECT branch_id, branch_name, address
            FROM branches
            WHERE branch_id = %s
            """,
            (branch_id,),
        )
        if branch_rows:
            branch_info = branch_rows[0]

    open_statuses = PUBLISHABLE_JOB_STATUSES
    pending_statuses = APPLICATION_REVIEW_STATUSES

    def scoped_query(base_query, params=None, alias='j', has_where=False):
        params = list(params or [])
        query = base_query
        if branch_id is not None:
            connector = 'WHERE' if not has_where else 'AND'
            query += f" {connector} {alias}.branch_id = %s"
            params.append(branch_id)
            has_where = True
        return query, tuple(params), has_where

    stats = {
        'open_jobs': 0,
        'total_applicants': 0,
        'total_applications': 0,
        'applied_applications': 0,
        'under_review_applications': 0,
        'interview_applications': 0,
        'hired_applications': 0,
        'rejected_applications': 0,
        'pending_applications': 0,
        'reviewed_applications': 0,
        'interviewed_applications': 0,
        'accepted_applications': 0,
        'pending_this_week': 0,
        'total_interviews': 0,
        'interviews_today': 0,
        'new_jobs_this_week': 0,
        'new_applications_today': 0,
        'new_applications_week': 0,
        'application_growth': 0,
        'success_rate': 0,
        'conversion_rate': 0,
        'mobile_applications': 0,
        'hr_accounts_count': 0,
        'active_branches': 0,
    }

    metrics = {'acceptance_rate': 0}
    chart_data = {
        'applications_over_time': {'labels': [], 'data': []},
        'status_distribution': {'labels': [], 'data': []},
        'top_jobs': {'labels': [], 'data': []},
        'branch_performance': {'labels': [], 'data': []},
    }

    # Open job count
    placeholders = ','.join(['%s'] * len(open_statuses))
    open_query, open_params, _ = scoped_query(
        f"""
        SELECT COUNT(*) AS count
        FROM jobs j
        WHERE j.status IN ({placeholders})
        """,
        open_statuses,
        alias='j',
        has_where=True,
    )
    stats['open_jobs'] = fetch_count(open_query, open_params)

    # Total applications
    total_apps_query, total_apps_params, _ = scoped_query(
        """
        SELECT COUNT(*) AS count
        FROM applications a
        JOIN jobs j ON j.job_id = a.job_id
        WHERE 1=1
        """,
        alias='j',
        has_where=True,
    )
    stats['total_applications'] = fetch_count(total_apps_query, total_apps_params)

    if branch_id:
        stats['total_applicants'] = fetch_count(
            """
            SELECT COUNT(DISTINCT a.applicant_id) AS count
            FROM applications a
            JOIN jobs j ON j.job_id = a.job_id
            WHERE j.branch_id = %s
            """,
            (branch_id,),
        )
    else:
        stats['total_applicants'] = fetch_count("SELECT COUNT(*) AS count FROM applicants")

    def applications_by_status(status_value):
        query = """
            SELECT COUNT(*) AS count
            FROM applications a
            JOIN jobs j ON j.job_id = a.job_id
            WHERE a.status = %s
        """
        params = [status_value]
        if branch_id:
            query += " AND j.branch_id = %s"
            params.append(branch_id)
        return fetch_count(query, tuple(params))

    # Use only the actual status values from the database enum: 'pending', 'scheduled', 'interviewed', 'hired', 'rejected'
    stats['pending_applications'] = applications_by_status('pending')
    stats['scheduled_applications'] = applications_by_status('scheduled')
    stats['interviewed_applications'] = applications_by_status('interviewed')
    stats['hired_applications'] = applications_by_status('hired')
    stats['rejected_applications'] = applications_by_status('rejected')
    # Legacy mappings for backward compatibility
    stats['reviewed_applications'] = applications_by_status('pending')  # Legacy: map reviewed to pending
    stats['accepted_applications'] = applications_by_status('hired')  # Legacy: map accepted to hired
    
    # Legacy status mappings for backward compatibility (these don't exist in database but kept for reference)
    stats['applied_applications'] = 0  # Maps to 'pending'
    stats['under_review_applications'] = 0  # Maps to 'reviewed'
    stats['interview_applications'] = 0  # Maps to 'interviewed'
    stats['hired_applications'] = stats['accepted_applications']  # Maps to 'accepted'

    pending_placeholders = ','.join(['%s'] * len(pending_statuses))
    pending_params = list(pending_statuses)
    pending_query = f"""
        SELECT COUNT(*) AS count
        FROM applications a
        JOIN jobs j ON j.job_id = a.job_id
        WHERE a.status IN ({pending_placeholders})
    """
    if branch_id:
        pending_query += " AND j.branch_id = %s"
        pending_params.append(branch_id)
    stats['pending_this_week'] = fetch_count(
        pending_query + " AND a.submitted_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)",
        tuple(pending_params),
    )

    def dated_application_count(condition):
        query = f"""
            SELECT COUNT(*) AS count
            FROM applications a
            JOIN jobs j ON j.job_id = a.job_id
            WHERE {condition}
        """
        params = []
        if branch_id:
            query += " AND j.branch_id = %s"
            params.append(branch_id)
        return fetch_count(query, tuple(params))

    stats['new_applications_today'] = dated_application_count("DATE(a.submitted_at) = CURDATE()")
    stats['new_applications_week'] = dated_application_count("a.submitted_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)")

    def interviews_count(condition):
        query = f"""
            SELECT COUNT(*) AS count
            FROM interviews i
            JOIN applications a ON a.application_id = i.application_id
            JOIN jobs j ON j.job_id = a.job_id
            WHERE {condition}
        """
        params = []
        if branch_id:
            query += " AND j.branch_id = %s"
            params.append(branch_id)
        return fetch_count(query, tuple(params))

    stats['total_interviews'] = interviews_count('1=1')
    stats['interviews_today'] = interviews_count('DATE(i.scheduled_date) = CURDATE()')

    # Ensure schema compatibility and get dynamic column expressions
    ensure_schema_compatibility()
    db = get_db()
    if db:
        cursor = db.cursor()
        try:
            _update_job_columns(cursor)
        finally:
            cursor.close()
    
    posted_at_expr = job_column_expr('posted_at', alternatives=['created_at'], default='j.created_at')
    new_jobs_query = f"SELECT COUNT(*) AS count FROM jobs j WHERE COALESCE({posted_at_expr}, j.created_at) >= DATE_SUB(NOW(), INTERVAL 7 DAY)"
    new_jobs_params = []
    if branch_id:
        new_jobs_query += " AND j.branch_id = %s"
        new_jobs_params.append(branch_id)
    stats['new_jobs_this_week'] = fetch_count(new_jobs_query, tuple(new_jobs_params))

    # HR accounts no longer have branch_id - they manage all branches
    hr_accounts_query = "SELECT COUNT(*) AS count FROM admins a JOIN users u ON u.user_id = a.user_id WHERE u.user_type = 'hr'"
    stats['hr_accounts_count'] = fetch_count(hr_accounts_query, None)

    stats['active_branches'] = fetch_count("SELECT COUNT(*) AS count FROM branches")

    current_week_pending = stats['pending_this_week']
    previous_week_pending = fetch_count(
        pending_query + " AND a.submitted_at >= DATE_SUB(NOW(), INTERVAL 14 DAY) AND a.submitted_at < DATE_SUB(NOW(), INTERVAL 7 DAY)",
        tuple(pending_params),
    )

    growth = 0
    if previous_week_pending:
        growth = ((current_week_pending - previous_week_pending) / previous_week_pending) * 100
    elif current_week_pending:
        growth = current_week_pending * 100
    stats['application_growth'] = round(growth, 1) if growth else 0

    total_apps = stats['total_applications'] or 0
    accepted_apps = stats['accepted_applications'] or 0
    acceptance_rate = round((accepted_apps / total_apps) * 100, 1) if total_apps else 0
    stats['success_rate'] = acceptance_rate
    stats['conversion_rate'] = acceptance_rate
    metrics['acceptance_rate'] = acceptance_rate

    trends = {
        'applications': {
            'current': current_week_pending,
            'previous': previous_week_pending,
            'change': stats['application_growth'],
            'direction': 'up' if stats['application_growth'] > 0 else 'down' if stats['application_growth'] < 0 else 'neutral',
        }
    }

    alerts = []
    if stats['pending_applications'] > 10:
        alerts.append({
            'type': 'warning',
            'icon': '⚠️',
            'message': f"{stats['pending_applications']} pending applications need review",
            'action': url_for('applications'),
        })
    if stats['interviews_today'] > 0:
        alerts.append({
            'type': 'info',
            'icon': '📅',
            'message': f"{stats['interviews_today']} interview(s) scheduled today",
            'action': url_for('admin_interviews' if role == 'admin' else 'hr_interviews'),
        })

    # Recent items ---------------------------------------------------------
    # Get dynamic column expressions
    job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
    posted_at_expr = job_column_expr('posted_at', alternatives=['created_at'], default='j.created_at')
    
    job_filter = "WHERE j.branch_id = %s" if branch_id else ""
    job_params = [branch_id] if branch_id else []
    recent_jobs = fetch_rows(
        f"""
        SELECT j.job_id,
               {job_title_expr} AS job_title,
               j.status,
               COALESCE({posted_at_expr}, j.created_at) AS posted_at,
               b.branch_name
        FROM jobs j
        LEFT JOIN branches b ON b.branch_id = j.branch_id
        {job_filter}
        ORDER BY COALESCE({posted_at_expr}, j.created_at) DESC
        LIMIT 5
        """,
        tuple(job_params),
    )
    formatted_recent_jobs = [
        {
            'job_id': row.get('job_id'),
            'title': row.get('job_title'),
            'status_key': (row.get('status') or '').lower(),
            'status_label': (row.get('status') or '').replace('_', ' ').title(),
            'status': (row.get('status') or '').lower(),
            'branch_name': row.get('branch_name'),
            'posted_at': format_human_datetime(row.get('posted_at') or row.get('created_at')),
        }
        for row in recent_jobs
    ]

    # Filter for pending applications automatically (but detect interviews to show proper badge)
    if branch_id:
        app_filter = "WHERE j.branch_id = %s AND a.status = 'pending'"
        app_params = [branch_id]
    else:
        app_filter = "WHERE a.status = 'pending'"
        app_params = []
    recent_applications = fetch_rows(
        f"""
        SELECT a.application_id,
               ap.full_name AS applicant_name,
               j.title AS job_title,
               a.status,
               a.submitted_at,
               (SELECT COUNT(*) FROM interviews i WHERE i.application_id = a.application_id) AS interview_count,
               (SELECT MAX(i.scheduled_date) FROM interviews i WHERE i.application_id = a.application_id) AS last_interview_date
        FROM applications a
        JOIN applicants ap ON ap.applicant_id = a.applicant_id
        JOIN jobs j ON j.job_id = a.job_id
        {app_filter}
        ORDER BY a.submitted_at DESC
        LIMIT 5
        """,
        tuple(app_params) if app_params else None,
    )
    formatted_recent_applications = []
    for row in recent_applications:
        raw_status = (row.get('status') or '').strip().lower()
        interview_count = row.get('interview_count', 0) or 0
        # Note: Status is already managed by interview scheduling/completion logic
        # 'scheduled' when interview is scheduled, 'interviewed' when interview is completed
        # So we don't auto-override status based on interview_count
        display_status = raw_status
        formatted_recent_applications.append({
            'application_id': row.get('application_id'),
            'applicant_name': row.get('applicant_name'),
            'position_applied': row.get('job_title'),
            'status_key': display_status,
            'status_label': display_status.replace('_', ' ').title(),
            'status': display_status,
            'submitted_at': format_human_datetime(row.get('submitted_at')),
            'has_interview': interview_count > 0,
            'last_interview_date': format_human_datetime(row.get('last_interview_date')) if row.get('last_interview_date') else None,
        })

    interview_filter = "AND j.branch_id = %s" if branch_id else ""
    interview_params = [branch_id] if branch_id else []
    upcoming_interviews = fetch_rows(
        f"""
        SELECT i.interview_id,
               ap.full_name AS applicant_name,
               j.title AS job_title,
               i.scheduled_date,
               COALESCE(i.interview_mode, 'in-person') AS interview_mode,
               i.location,
               a.status AS application_status
        FROM interviews i
        JOIN applications a ON a.application_id = i.application_id
        JOIN applicants ap ON ap.applicant_id = a.applicant_id
        JOIN jobs j ON j.job_id = a.job_id
        WHERE i.scheduled_date >= NOW() {interview_filter}
        ORDER BY i.scheduled_date ASC
        LIMIT 5
        """,
        tuple(interview_params),
    )
    formatted_upcoming_interviews = [
        {
            'interview_id': row.get('interview_id'),
            'applicant_name': row.get('applicant_name'),
            'position_applied': row.get('job_title'),
            'date_time': format_human_datetime(row.get('scheduled_date')),
            'type': (row.get('interview_mode') or 'in-person').replace('-', ' ').title(),
            'location': row.get('location'),
            'status_key': (row.get('application_status') or '').lower(),
            'status_label': (row.get('application_status') or '').replace('_', ' ').title(),
        }
        for row in upcoming_interviews
    ]

    recent_activity = []
    for item in formatted_recent_applications[:5]:
        recent_activity.append({
            'type': 'application',
            'description': f"{item['applicant_name']} applied for {item['position_applied']}",
            'timestamp': format_human_datetime(item['submitted_at']),
            'status': item['status_key'],
        })
    for item in formatted_upcoming_interviews[:5]:
        recent_activity.append({
            'type': 'interview',
            'description': f"Interview scheduled for {item['applicant_name']} ({item['position_applied']})",
            'timestamp': item['date_time'],
            'status': item['status_key'],
        })
    for item in formatted_recent_jobs[:3]:
        recent_activity.append({
            'type': 'job',
            'description': f"Job update: {item['title']}",
            'timestamp': item['posted_at'],
            'status': item['status_key'],
        })

    # Chart data - applications over last 30 days
    chart_params = []
    applications_over_time_query = """
        SELECT DATE(a.submitted_at) AS submitted_date,
               COUNT(*) AS total
        FROM applications a
        JOIN jobs j ON j.job_id = a.job_id
        WHERE a.submitted_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)
    """
    if branch_id:
        applications_over_time_query += " AND j.branch_id = %s"
        chart_params.append(branch_id)
    applications_over_time_query += " GROUP BY DATE(a.submitted_at) ORDER BY submitted_date ASC"
    applications_over_time = fetch_rows(applications_over_time_query, tuple(chart_params))
    chart_data['applications_over_time']['labels'] = [
        row['submitted_date'].strftime('%b %d') if row.get('submitted_date') else ''
        for row in applications_over_time
    ]
    chart_data['applications_over_time']['data'] = [row['total'] for row in applications_over_time]

    status_params = []
    status_distribution_query = """
        SELECT a.status, COUNT(*) AS total
        FROM applications a
        JOIN jobs j ON j.job_id = a.job_id
        WHERE 1=1
    """
    if branch_id:
        status_distribution_query += " AND j.branch_id = %s"
        status_params.append(branch_id)
    status_distribution_query += " GROUP BY a.status"
    status_distribution = fetch_rows(status_distribution_query, tuple(status_params))
    
    # Map statuses: normalize legacy statuses to canonical ones
    # Note: 'withdrawn' status has been removed - applications with withdrawn status should be manually reviewed
    status_map = {}
    status_normalization = {
        'pending': 'pending',
        'scheduled': 'scheduled',
        'interviewed': 'interviewed',
        'hired': 'hired',
        'rejected': 'rejected',
        'reviewed': 'pending',
        'shortlisted': 'pending',
        'accepted': 'hired',
        'applied': 'pending',
        'under_review': 'pending',
        'interview': 'interviewed'
    }
    
    for row in status_distribution:
        if row.get('status') and row.get('total', 0) > 0:  # Only include statuses with count > 0
            status = row['status'].lower()
            # Skip withdrawn status completely - don't include it in chart
            if status == 'withdrawn':
                continue
            # Normalize status to canonical one
            normalized_status = status_normalization.get(status, status)
            # Skip if normalized to withdrawn (shouldn't happen, but safety check)
            if normalized_status == 'withdrawn':
                continue
            # Combine counts for same normalized status
            if normalized_status in status_map:
                status_map[normalized_status] += row['total']
            else:
                status_map[normalized_status] = row['total']
    
    # Only include statuses with count > 0, in a consistent order
    # Define the order: scheduled, interviewed, hired, rejected (4 statuses only)
    # Note: withdrawn and pending are explicitly excluded
    status_order = ['scheduled', 'interviewed', 'hired', 'rejected']
    ordered_statuses = []
    for status in status_order:
        if status in status_map and status_map[status] > 0:
            ordered_statuses.append((status, status_map[status]))
    
    # Add any other statuses not in the standard order (but exclude withdrawn)
    for status, count in status_map.items():
        if status not in status_order and status != 'withdrawn' and count > 0:
            ordered_statuses.append((status, count))
    
    chart_data['status_distribution']['labels'] = [
        status.replace('_', ' ').title() for status, count in ordered_statuses
    ]
    chart_data['status_distribution']['data'] = [
        count for status, count in ordered_statuses
    ]

    top_job_params = []
    top_jobs_query = f"""
        SELECT {job_title_expr} AS job_title, COUNT(a.application_id) AS total
        FROM jobs j
        LEFT JOIN applications a ON a.job_id = j.job_id
        WHERE 1=1
    """
    if branch_id:
        top_jobs_query += " AND j.branch_id = %s"
        top_job_params.append(branch_id)
    top_jobs_query += f" GROUP BY j.job_id, {job_title_expr} ORDER BY total DESC, {job_title_expr} ASC LIMIT 5"
    top_jobs = fetch_rows(top_jobs_query, tuple(top_job_params))
    chart_data['top_jobs']['labels'] = [row['job_title'] for row in top_jobs]
    chart_data['top_jobs']['data'] = [row['total'] for row in top_jobs]

    if branch_id is None:
        branch_performance = fetch_rows(
            """
            SELECT b.branch_name, COUNT(a.application_id) AS total
            FROM branches b
            LEFT JOIN jobs j ON j.branch_id = b.branch_id
            LEFT JOIN applications a ON a.job_id = j.job_id
            GROUP BY b.branch_id, b.branch_name
            ORDER BY total DESC
            LIMIT 5
            """
        )
        chart_data['branch_performance']['labels'] = [row['branch_name'] for row in branch_performance]
        chart_data['branch_performance']['data'] = [row['total'] for row in branch_performance]

    # Active sessions snapshot
    active_sessions = fetch_rows(
        """
        SELECT s.session_id,
               s.login_time,
               u.email,
               u.user_type
        FROM auth_sessions s
        JOIN users u ON u.user_id = s.user_id
        WHERE s.is_active = 1
        ORDER BY s.login_time DESC
        LIMIT 5
        """
    )

    dashboard = {
        'user': {
            'full_name': user.get('name') if user else '',
            'email': user.get('email') if user else '',
            'user_type': role_label,
            'role': role,
        },
        'branch_info': branch_info,
        'stats': stats,
        'metrics': metrics,
        'chart_data': chart_data,
        'trends': trends,
        'system_alerts': alerts,
        'recent_jobs': formatted_recent_jobs,
        'recent_applications': formatted_recent_applications,
        'upcoming_interviews': formatted_upcoming_interviews,
        'recent_activity': recent_activity,
        'notifications': [],
        'notifications_count': 0,
        'active_sessions': [
            {
                'session_id': row.get('session_id'),
                'login_time': format_human_datetime(row.get('login_time')),
                'email': row.get('email'),
                'user_type': row.get('user_type'),
            }
            for row in active_sessions
        ],
        'branches': [] if branch_id else fetch_branches(),
        'hr_accounts': fetch_hr_accounts() if role == 'admin' else (
            fetch_rows(
                """
                SELECT a.admin_id, a.full_name, u.email
                FROM admins a
                JOIN users u ON u.user_id = a.user_id
                WHERE u.user_type = 'hr'
                ORDER BY a.full_name ASC
                """,
                None,
            )
        ),
    }

    return dashboard

import os
import mimetypes
import traceback

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, g, send_file, send_from_directory
from flask_wtf.csrf import CSRFProtect, CSRFError, generate_csrf
from functools import wraps
from datetime import datetime, date, timedelta, timezone
from uuid import uuid4
from decimal import Decimal, InvalidOperation
from threading import Lock

from config import Config
from utils.database import get_db, close_db, execute_query
from utils.auth import (
    hash_password,
    check_password,
    login_user,
    logout_user,
    get_current_user,
    is_logged_in,
)
from utils.helpers import save_uploaded_file
from utils.mailer import send_email
from utils.rate_limit import rate_limit

def _track_failed_login():
    """Helper function to track failed login attempts for rate limiting."""
    from utils.rate_limit import _rate_limit_store, _rate_limit_lock
    from datetime import datetime
    from flask import request
    
    identifier = request.remote_addr or 'unknown'
    key = f"login:{identifier}"
    with _rate_limit_lock:
        if key not in _rate_limit_store:
            _rate_limit_store[key] = []
        _rate_limit_store[key].append(datetime.now())


# ============================================================================
# AUTOMATION SYSTEM - Automatic Notifications and Emails
# ============================================================================

def auto_notify_and_email(cursor, application_id, message, email_subject=None, email_body=None, recipient_email=None, recipient_name=None):
    """
    Automatically create notification and send email for any system event.
    This ensures all events are automatically communicated to users.
    Prevents duplicate notifications by checking if the same message already exists.
    Also prevents JSON responses from being saved as notifications.
    """
    try:
        # Prevent JSON responses from being saved as notifications
        message_str = str(message).strip()
        if message_str.startswith('{') and ('"success"' in message_str or '"message"' in message_str or '"error"' in message_str):
            print(f'⚠️ Blocked JSON response from being saved as notification in auto_notify_and_email: {message_str[:100]}')
            return
        
        ensure_schema_compatibility()
        notification_columns = set()
        try:
            cursor.execute('SHOW COLUMNS FROM notifications')
            notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        except Exception as col_error:
            print(f"⚠️ Error checking notification columns: {col_error}")
        
        # Check if notification already exists to prevent duplicates
        cursor.execute(
            '''
            SELECT notification_id FROM notifications
            WHERE application_id = %s AND message = %s
            LIMIT 1
            ''',
            (application_id, message),
        )
        existing_notification = cursor.fetchone()
        
        if existing_notification:
            print(f'⚠️ Duplicate notification detected for application {application_id}. Skipping notification creation and email to prevent duplicates.')
            # Don't send email if notification already exists (prevents duplicate emails)
            return
        
        # Automatically create notification in system (only if not duplicate)
        if 'sent_at' in notification_columns:
            cursor.execute(
                '''
                INSERT INTO notifications (application_id, message, sent_at, is_read)
                VALUES (%s, %s, NOW(), 0)
                ''',
                (application_id, message),
            )
        else:
            cursor.execute(
                '''
                INSERT INTO notifications (application_id, message, is_read)
                VALUES (%s, %s, 0)
                ''',
                (application_id, message),
            )
        print(f'✅ Notification created in system for application {application_id}: {message[:50]}...')
        
        # Email notifications are always enabled (system settings removed)
        email_enabled = True
        # Automatically send email if enabled and recipient info provided
        # Only send email if notification was created (not duplicate)
        if email_enabled and recipient_email and email_subject and email_body:
            try:
                # Validate email parameters before sending
                if recipient_email and email_subject and email_body:
                    send_email(recipient_email, email_subject, email_body)
                    print(f'✅ Email sent to {recipient_email} for application {application_id}')
            except Exception as email_error:
                print(f"⚠️ Auto-email error (non-blocking): {email_error}")
                # Continue even if email fails - notification is created in system
        else:
            if not email_enabled:
                print(f'⚠️ Email not sent (disabled) for application {application_id}')
            elif not recipient_email:
                print(f'⚠️ Email not sent (missing recipient email) for application {application_id}')
            elif not email_subject or not email_body:
                print(f'⚠️ Email not sent (missing subject or body) for application {application_id}')
    except Exception as notify_error:
        import traceback
        print(f"⚠️ Auto-notification error (non-blocking): {notify_error}")
        print(f"Traceback: {traceback.format_exc()}")
        # Continue even if notification fails


def auto_update_application_status(cursor, application_id, new_status, reason=''):
    """
    Automatically update application status and notify applicant via system notification and email.
    """
    try:
        # Ensure schema compatibility to get correct job column names
        ensure_schema_compatibility()
        db_temp = get_db()
        if db_temp:
            cursor_temp = db_temp.cursor()
            try:
                _update_job_columns(cursor_temp)
            finally:
                cursor_temp.close()
        
        # Get job title expression - ensure it uses 'j' alias for jobs table
        job_title_expr = job_column_expr('job_title', alias='j', alternatives=['title'], default="'Untitled Job'")
        
        # Get applicant info before updating - ensure we get the correct job for THIS specific application
        cursor.execute(
            f'''
            SELECT ap.applicant_id, ap.email, ap.full_name, a.status AS old_status, 
                   a.job_id,
                   COALESCE({job_title_expr}, 'Untitled Job') AS job_title
            FROM applicants ap
            JOIN applications a ON ap.applicant_id = a.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            WHERE a.application_id = %s
            LIMIT 1
            ''',
            (application_id,)
        )
        applicant_info = cursor.fetchone()
        
        if applicant_info:
            old_status = applicant_info.get('old_status')
            
            # Update status in database
            cursor.execute(
                'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s',
                (new_status, application_id),
            )
            print(f'✅ Status updated: Application {application_id} -> {new_status}')
            
            # Auto-notify and email applicant
            status_display = new_status.replace('_', ' ').title()
            job_title = applicant_info.get('job_title') or 'Your Application'
            applicant_name = applicant_info.get('full_name') or 'Applicant'
            applicant_email = applicant_info.get('email')
            
            # Special handling for "hired" status
            if new_status.lower() == 'hired':
                message = f'Congratulations! You have been hired for the position: "{job_title}". Welcome to the team!'
                email_subject = f'Congratulations! You\'ve Been Hired - {job_title}'
                email_body = f"""Dear {applicant_name},

Congratulations! We are pleased to inform you that you have been selected for the position of {job_title}.

We are excited to welcome you to our team and look forward to working with you.

Please expect a call within 24hrs.

{f"Additional Notes: {reason}" if reason else ""}

Please log in to your account to view more details and next steps.

J&T Express Recruitment Team
                """.strip()
            elif new_status.lower() == 'rejected':
                message = f'Your application status for "{job_title}" has been updated to: {status_display}'
                if reason:
                    message += f' - {reason}'
                email_subject = f'Application Status Update - {job_title}'
                email_body = f"""Dear {applicant_name},

Thank you for your interest in the position "{job_title}".

After careful consideration, we regret to inform you that we have decided to move forward with other candidates at this time.

{f"Reason: {reason}" if reason else ""}

We appreciate your time and interest in our company. We encourage you to apply for future positions that match your qualifications.

Best regards,
J&T Express Recruitment Team
                """.strip()
            else:
                message = f'Your application status for "{job_title}" has been updated to: {status_display}'
                if reason:
                    message += f' - {reason}'
                email_subject = f'Application Status Update - {job_title}'
                email_body = f"""Dear {applicant_name},

Your application status for the job position "{job_title}" has been updated.

New Status: {status_display}
{f"Reason: {reason}" if reason else ""}

Please log in to your account to view more details.

Best regards,
J&T Express Recruitment Team
                """.strip()
            
            # Create notification and send email to applicant
            if applicant_email:
                auto_notify_and_email(
                    cursor, application_id, message,
                    email_subject, email_body,
                    applicant_email,
                    applicant_name
                )
                print(f'✅ Notification created and email sent to applicant {applicant_info.get("applicant_id")} ({applicant_email}) for application {application_id}')
            else:
                # Still create notification even if no email
                auto_notify_and_email(
                    cursor, application_id, message,
                    None, None,
                    None, None
                )
                print(f'✅ Notification created for application {application_id} (no email - applicant email missing)')
            
            return True
        else:
            print(f'⚠️ No applicant info found for application {application_id}')
    except Exception as auto_error:
        import traceback
        print(f"⚠️ Auto-status update error: {auto_error}")
        print(f"Traceback: {traceback.format_exc()}")
    return False


def auto_handle_job_status(cursor, job_id, new_status, old_status=None):
    """
    Automatically handle job status changes (posted_at, notifications, etc.)
    """
    try:
        # Automatically set posted_at when status becomes 'active' or 'open'
        # Always update to current server time (NOW()) to ensure accurate posting time
        if new_status in ('active', 'open'):
            cursor.execute(
                'UPDATE jobs SET posted_at = NOW() WHERE job_id = %s',
                (job_id,)
            )
        # Automatically clear posted_at when status becomes 'closed'
        elif new_status == 'closed':
            # Keep posted_at for historical record, just update status
            pass
        
        return True
    except Exception as auto_error:
        print(f"⚠️ Auto-job status error: {auto_error}")
    return False


VALID_JOB_STATUSES = ('active', 'closed')
PUBLISHABLE_JOB_STATUSES = ('open',)  # Only 'open' status is visible to applicants (database enum: 'open', 'closed')
VALID_EMPLOYMENT_TYPES = ('full_time', 'part_time', 'internship')
VALID_WORK_ARRANGEMENTS = ('onsite', 'remote', 'hybrid', 'field', 'flexible')
VALID_EXPERIENCE_LEVELS = ('entry', 'mid', 'senior', 'lead', 'manager')

# Database enum values: ENUM('pending', 'scheduled', 'interviewed', 'hired', 'rejected')
APPLICATION_STATUSES = ('pending', 'scheduled', 'interviewed', 'hired', 'rejected')
APPLICATION_PIPELINE_STATUSES = ('pending', 'scheduled', 'interviewed')
APPLICATION_REVIEW_STATUSES = ('pending',)
APPLICATION_SUCCESS_STATUSES = ('hired',)
APPLICATION_FAILED_STATUSES = ('rejected',)
APPLICATION_ACTIVE_STATUSES = ('pending', 'scheduled', 'interviewed')
APPLICATION_TERMINAL_STATUSES = ('hired', 'rejected')
APPLICATION_STATUS_LABELS = {
    'pending': 'Pending',
    'scheduled': 'Scheduled',
    'interviewed': 'Interviewed',
    'hired': 'Hired',
    'rejected': 'Rejected',
    # Legacy status mappings for backward compatibility
    'reviewed': 'Pending',
    'accepted': 'Hired',
    'applied': 'Pending',
    'under_review': 'Pending',
    'interview': 'Interviewed',
}
APPLICATION_STATUS_FLOW = ('pending', 'scheduled', 'interviewed', 'hired', 'rejected')


app = Flask(__name__)
app.config.from_object(Config)
csrf = CSRFProtect(app)
# Configure CSRF settings
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_TIME_LIMIT'] = None  # No time limit for CSRF tokens
# Disable redirect pages - use HTTP 302 redirects directly
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Configure secure session cookies
app.config['SESSION_COOKIE_SECURE'] = True  # Only send over HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Prevent JavaScript access
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF protection
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hour session timeout

_schema_lock = Lock()
_schema_checked = False
JOB_COLUMNS = set()


def immediate_redirect(location, code=302):
    """Create an immediate HTTP redirect without showing redirect page."""
    from flask import Response
    response = Response(status=code)
    response.headers['Location'] = location
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    # Ensure minimal response body for immediate redirect
    response.data = b''
    return response


def _update_job_columns(cursor):
    """Refresh the cached set of columns available on the jobs table."""
    global JOB_COLUMNS
    try:
        cursor.execute('SHOW COLUMNS FROM jobs')
        rows = cursor.fetchall() or []
        JOB_COLUMNS = {
            row.get('Field') if isinstance(row, dict) else row[0]
            for row in rows
        }
    except Exception as exc:
        print(f'⚠️ Failed to inspect jobs table columns: {exc}')
        JOB_COLUMNS = set()
    return JOB_COLUMNS


def job_column(preferred, *alternatives):
    """Return the present column name on jobs table, preferring the modern schema."""
    for candidate in (preferred,) + alternatives:
        if candidate in JOB_COLUMNS:
            return candidate
    return None


def job_column_expr(preferred, alias='j', default='NULL', alternatives=None):
    """Return a SQL expression pointing at an existing jobs column or a safe fallback."""
    alternatives = alternatives or []
    column_name = job_column(preferred, *alternatives)
    if column_name:
        return f'{alias}.{column_name}'
    return default


def job_column_name(preferred, alternatives=None, default=None):
    """Return the raw column name for updates/inserts, or a provided default."""
    alternatives = alternatives or []
    column_name = job_column(preferred, *alternatives)
    if column_name:
        return column_name
    return default


def ensure_schema_compatibility():
    """Best-effort guard to align dynamic queries with the current MySQL schema."""
    global _schema_checked
    if _schema_checked:
        return

    with _schema_lock:
        if _schema_checked:
            return

        db = get_db()
        if not db:
            return

        cursor = None
        updates_applied = False
        success = False

        def ensure_column(cur, table_name, column_name, column_definition, post_add=None):
            cur.execute(f"SHOW COLUMNS FROM {table_name} LIKE %s", (column_name,))
            if cur.fetchone():
                return False
            cur.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_definition}")
            if callable(post_add):
                try:
                    post_add()
                except Exception as copy_exc:
                    print(f'⚠️ Post-add hook for {table_name}.{column_name} failed: {copy_exc}')
            return True

        def ensure_table(cur, table_name, create_sql):
            """Ensure a table exists, create it if it doesn't."""
            try:
                cur.execute(f"SHOW TABLES LIKE '{table_name}'")
                if cur.fetchone():
                    return False
                cur.execute(create_sql)
                return True
            except Exception as e:
                print(f'⚠️ Failed to ensure table {table_name}: {e}')
                return False

        try:
            cursor = db.cursor()
            _update_job_columns(cursor)

            # Only ensure logout_time exists (last_activity is not in actual schema)
            updates_applied |= ensure_column(cursor, 'auth_sessions', 'logout_time', 'DATETIME NULL DEFAULT NULL')
            try:
                cursor.execute("ALTER TABLE auth_sessions MODIFY logout_time DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP")
            except Exception:
                pass
            try:
                cursor.execute("UPDATE auth_sessions SET logout_time = login_time WHERE logout_time IS NULL")
                if cursor.rowcount:
                    updates_applied = True
            except Exception:
                pass
            updates_applied |= ensure_column(cursor, 'interviews', 'interview_mode', "VARCHAR(50) NULL DEFAULT NULL")
            
            # Ensure interview status column exists
            # Ensure status column exists with all needed values including 'confirmed' and 'rescheduled'
            try:
                cursor.execute("SHOW COLUMNS FROM interviews LIKE 'status'")
                status_col = cursor.fetchone()
                if status_col:
                    # Check if 'confirmed' is in the enum
                    col_type = status_col.get('Type', '') if isinstance(status_col, dict) else str(status_col[1]) if len(status_col) > 1 else ''
                    if 'confirmed' not in col_type.lower():
                        # Modify enum to include 'confirmed' and 'rescheduled'
                        try:
                            cursor.execute("ALTER TABLE interviews MODIFY COLUMN status ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
                            updates_applied = True
                            print('✅ Added "confirmed" and "rescheduled" to interviews.status enum')
                        except Exception as enum_err:
                            print(f'⚠️ Could not modify status enum: {enum_err}')
                else:
                    # Column doesn't exist, create it
                    updates_applied |= ensure_column(cursor, 'interviews', 'status', "ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
            except Exception as status_check_err:
                print(f'⚠️ Error checking status column: {status_check_err}')
                # Fallback: try to ensure column exists
                updates_applied |= ensure_column(cursor, 'interviews', 'status', "ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
            
            # Ensure applications table status enum includes 'scheduled'
            try:
                cursor.execute("""
                    SELECT COLUMN_TYPE 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                    AND TABLE_NAME = 'applications' 
                    AND COLUMN_NAME = 'status'
                """)
                result = cursor.fetchone()
                if result:
                    current_enum = result[0].lower()
                    if 'scheduled' not in current_enum:
                        # Update enum to include 'scheduled'
                        cursor.execute("""
                            ALTER TABLE applications 
                            MODIFY COLUMN status ENUM('pending', 'scheduled', 'interviewed', 'hired', 'rejected', 'withdrawn') 
                            NOT NULL DEFAULT 'pending'
                        """)
                        updates_applied = True
                        print('✅ Updated applications.status enum to include "scheduled"')
            except Exception as enum_error:
                print(f'⚠️ Could not update applications.status enum: {enum_error}')
                # Continue - enum might already be correct or table might not exist yet
            
            # Ensure positions table exists
            positions_sql = """
                CREATE TABLE IF NOT EXISTS positions (
                    position_id INT AUTO_INCREMENT PRIMARY KEY,
                    title VARCHAR(200) NOT NULL,
                    department VARCHAR(200) NOT NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NULL DEFAULT NULL ON UPDATE CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """
            updates_applied |= ensure_table(cursor, 'positions', positions_sql)
            
            job_columns = _update_job_columns(cursor)

            updates_applied |= ensure_column(
                cursor,
                'notifications',
                'sent_at',
                'DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP'
            )
            updates_applied |= ensure_column(
                cursor,
                'notifications',
                'is_read',
                'TINYINT(1) NOT NULL DEFAULT 0'
            )
            # Ensure last login/logout columns exist
            updates_applied |= ensure_column(
                cursor,
                'users',
                'last_login',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'users',
                'last_logout',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'admins',
                'last_login',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'admins',
                'last_logout',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'last_login',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'last_logout',
                'DATETIME NULL DEFAULT NULL'
            )
            
            # Ensure applications.viewed_at column exists
            updates_applied |= ensure_column(
                cursor,
                'applications',
                'viewed_at',
                'DATETIME NULL'
            )
            
            # Ensure users.email_verified column exists for email verification
            updates_applied |= ensure_column(
                cursor,
                'users',
                'email_verified',
                'TINYINT(1) DEFAULT 0'
            )
            
            # Ensure applicants.verification_token column exists
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'verification_token',
                'VARCHAR(255) NULL DEFAULT NULL'
            )
            
            # Ensure applicants.verification_token_expires column exists for 60-second expiration
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'verification_token_expires',
                'DATETIME NULL DEFAULT NULL'
            )
            
            # Ensure applicants.last_profile_update column exists for tracking profile updates
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'last_profile_update',
                'DATETIME NULL DEFAULT NULL'
            )
            

            if updates_applied:
                try:
                    db.commit()
                except Exception:
                    pass

            success = True
        except Exception:
            pass
        finally:
            if cursor:
                cursor.close()

        if success:
            _schema_checked = True


# Register template filters
@app.template_filter('format_human_datetime')
def format_human_datetime_filter(value):
    """Produce a human-readable timestamp in 12-hour format (AM/PM)."""
    if isinstance(value, (datetime, date)):
        dt_value = value if isinstance(value, datetime) else datetime.combine(value, datetime.min.time())
        return dt_value.strftime('%b %d, %Y %I:%M %p')
    elif isinstance(value, str):
        try:
            # Try parsing common datetime formats
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d', '%Y-%m-%d %H:%M']:
                try:
                    dt_value = datetime.strptime(value, fmt)
                    return dt_value.strftime('%b %d, %Y %I:%M %p')
                except ValueError:
                    continue
        except:
            pass
    return value or ''


@app.teardown_appcontext
def teardown_database(exception=None):
    """Ensure database connections are closed after each request."""
    close_db(exception)


def fetch_notifications_for(scope=None, limit=5):
    """Fetch notifications with optional scoping.
    
    scope: dict with any of:
      - applicant_id: int → fetch applicant's notifications (application-related)
      - branch_id: int → fetch branch-scoped application notifications
      - system_only: bool → fetch system (application_id IS NULL)
    Returns (formatted_list, unread_count)
    """
    scope = scope or {}
    db = get_db()
    if not db:
        return ([], 0)
    cursor = db.cursor(dictionary=True)
    try:
        ensure_schema_compatibility()
        # Introspect columns
        cursor.execute('SHOW COLUMNS FROM notifications')
        notif_cols = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        sent_at_expr = 'COALESCE(n.sent_at, n.created_at, NOW())' if 'sent_at' in notif_cols else 'COALESCE(n.created_at, NOW())'
        is_read_expr = 'COALESCE(n.is_read, 0)' if 'is_read' in notif_cols else '0'
        has_application_fk = 'application_id' in notif_cols
        params = []
        joins = ''
        where_parts = []
        # Scope by system vs application related
        if scope.get('system_only'):
            where_parts.append('n.application_id IS NULL')
            # Ensure system-level notifications do not contain applicant-facing messages
            applicant_only_filters_sys = [
                "n.message NOT LIKE 'You applied for%'",
                "n.message NOT LIKE 'Congratulations! You have been hired%'",
                "n.message NOT LIKE 'Your application status%'",
                "n.message NOT LIKE 'Congratulations! You%'",
                "n.message NOT LIKE '%application status%'",
                "n.message NOT LIKE '%status has been updated%'",
            ]
            if where_parts:
                where_parts.append(' AND '.join(applicant_only_filters_sys))
        elif has_application_fk:
            # Application-related joins if needed
            joins = 'JOIN applications a ON n.application_id = a.application_id'
            if scope.get('applicant_id'):
                where_parts.append('a.applicant_id = %s')
                params.append(scope['applicant_id'])
            if scope.get('branch_id'):
                joins += ' JOIN jobs j ON a.job_id = j.job_id'
                where_parts.append('j.branch_id = %s')
                params.append(scope['branch_id'])
        # Build WHERE
        where_sql = ''
        if where_parts:
            where_sql = 'WHERE ' + ' AND '.join(where_parts)
        
            # If fetching for HR (branch_id scope but not applicant_id), exclude applicant-only notifications
            # CRITICAL: HR should NEVER see applicant-facing notifications - these are for applicants only
            if scope.get('branch_id') and not scope.get('applicant_id') and not scope.get('system_only'):
                applicant_only_filters = [
                    "n.message NOT LIKE 'You applied for%'",
                    "n.message NOT LIKE 'Congratulations! You have been hired%'",
                    "n.message NOT LIKE 'Your application status%'",
                    "n.message NOT LIKE 'Congratulations! You%'",  # Catch any variation of congratulations messages to applicants
                    "n.message NOT LIKE '%application status%'",  # Exclude all status update notifications
                    "n.message NOT LIKE '%status has been updated%'",  # Exclude status update messages
                ]
                if where_sql:
                    where_sql += ' AND ' + ' AND '.join(applicant_only_filters)
                else:
                    where_sql = 'WHERE ' + ' AND '.join(applicant_only_filters)
        
        # Fetch list
        cursor.execute(
            f'''
            SELECT n.notification_id,
                   n.message,
                   {sent_at_expr} AS sent_at,
                   {is_read_expr} AS is_read
            FROM notifications n
            {joins}
            {where_sql}
            ORDER BY {sent_at_expr} DESC
            LIMIT %s
            ''',
            tuple(params + [limit]),
        )
        rows = cursor.fetchall() or []
        formatted = []
        for r in rows:
            message = r.get('message', '')
            # Filter out JSON responses that might have been incorrectly stored
            if message and message.strip().startswith('{') and '"success"' in message:
                continue  # Skip JSON responses
            formatted.append({
                'title': 'Update' if not scope.get('system_only') else 'System',
                'message': message,
                'time': format_human_datetime(r.get('sent_at')),
                'is_read': r.get('is_read', False),
            })
        # Unread count
        unread_where = where_sql
        if unread_where:
            unread_where = unread_where + f' AND {is_read_expr} = 0'
        else:
            unread_where = f'WHERE {is_read_expr} = 0'
        cursor.execute(
            f'''
            SELECT COUNT(*) AS unread
            FROM notifications n
            {joins}
            {unread_where}
            ''',
            tuple(params),
        )
        unread_row = cursor.fetchone() or {}
        unread_count = unread_row.get('unread', 0) or 0
        return (formatted, unread_count)
    except Exception:
        return ([], 0)
    finally:
        try:
            cursor.close()
        except Exception:
            pass

def login_required(*roles):
    """Decorator enforcing authentication and optional role-based access control."""

    def decorator(view_func):
        @wraps(view_func)
        def wrapped_view(*args, **kwargs):
            if not is_logged_in():
                flash('Please login to access this page.', 'error')
                return immediate_redirect(url_for('login', _external=True))

            if roles and session.get('user_role') not in roles:
                flash('You do not have permission to access this page.', 'error')
                return immediate_redirect(url_for('index', _external=True))

            return view_func(*args, **kwargs)

        return wrapped_view

    return decorator


@app.before_request
def load_logged_in_user():
    """Attach the current user to the Flask global context for easy template access."""
    # Skip for logout, login, and resend_verification routes to prevent any redirect issues
    if request.endpoint in ('logout', 'login', 'resend_verification'):
        return
    
    ensure_schema_compatibility()
    g.current_user = get_current_user()

def determine_user_friendly_action(path, form_action, target_table):
    """Determine a user-friendly action name from path and form action."""
    path_lower = path.lower()
    form_action_lower = (form_action or '').lower()
    
    # Check form action first (most reliable)
    if form_action_lower in ['add', 'create', 'post']:
        return 'Add'
    elif form_action_lower in ['update', 'edit', 'modify']:
        return 'Update'
    elif form_action_lower == 'delete':
        return 'Delete'
    elif form_action_lower == 'schedule':
        return 'Schedule'
    elif form_action_lower == 'reschedule':
        return 'Reschedule'
    elif form_action_lower == 'cancel':
        return 'Cancel'
    elif form_action_lower == 'update_status':
        return 'Update Status'
    elif form_action_lower == 'bulk_update_status':
        return 'Bulk Update Status'
    
    # Check path patterns
    if '/add' in path_lower or '/create' in path_lower or '/post' in path_lower:
        return 'Add'
    elif '/update' in path_lower or '/edit' in path_lower:
        return 'Update'
    elif '/delete' in path_lower:
        return 'Delete'
    elif '/schedule' in path_lower:
        return 'Schedule'
    elif '/reschedule' in path_lower:
        return 'Reschedule'
    elif '/cancel' in path_lower:
        return 'Cancel'
    elif '/status' in path_lower:
        return 'Update Status'
    
    # Check target table for context
    if target_table == 'jobs' or 'job' in path_lower:
        if 'add' in path_lower or 'create' in path_lower or 'post' in path_lower:
            return 'Add Job'
        elif 'update' in path_lower or 'edit' in path_lower:
            return 'Update Job'
        elif 'delete' in path_lower:
            return 'Delete Job'
    elif target_table == 'applications' or 'application' in path_lower:
        if 'status' in path_lower or 'update_status' in form_action_lower:
            return 'Update Application Status'
        elif 'bulk' in path_lower or 'bulk_update' in form_action_lower:
            return 'Bulk Update Applications'
    elif target_table == 'interviews' or 'interview' in path_lower:
        if 'schedule' in path_lower or form_action_lower == 'schedule':
            return 'Schedule Interview'
        elif 'reschedule' in path_lower or form_action_lower == 'reschedule':
            return 'Reschedule Interview'
        elif 'cancel' in path_lower or form_action_lower == 'cancel':
            return 'Cancel Interview'
        elif 'update' in path_lower or 'edit' in path_lower:
            return 'Update Interview'
        elif 'delete' in path_lower:
            return 'Delete Interview'
    
    # Default fallback
    if form_action:
        return form_action.replace('_', ' ').title()
    elif path:
        path_parts = [p for p in path.split('/') if p and p not in ['admin', 'hr', 'applicant']]
        if path_parts:
            last_part = path_parts[-1]
            if last_part.isdigit() and len(path_parts) > 1:
                return path_parts[-2].replace('-', ' ').replace('_', ' ').title()
            else:
                return last_part.replace('-', ' ').replace('_', ' ').title()
    
    return 'Action'

def log_hr_activity(admin_id, action, target_table, target_id, details=None, skip_notification=False):
    """Log HR activity for admin monitoring."""
    try:
        db = get_db()
        if not db:
            return False
        cursor = db.cursor()
        try:
            # Ensure activity_logs table exists
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS activity_logs (
                    log_id INT AUTO_INCREMENT PRIMARY KEY,
                    admin_id INT,
                    action VARCHAR(255) NOT NULL,
                    target_table VARCHAR(255) NOT NULL,
                    target_id INT,
                    details TEXT,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_admin_id (admin_id),
                    INDEX idx_created_at (created_at),
                    INDEX idx_target (target_table, target_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            
            # Check if target_table column exists, add it if not
            try:
                cursor.execute("SHOW COLUMNS FROM activity_logs LIKE 'target_table'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_table VARCHAR(255) NOT NULL DEFAULT '' AFTER action")
                    print('✅ Added target_table column to activity_logs table')
            except Exception as col_err:
                # Column might already exist or table structure is different
                print(f'⚠️ Could not check/add target_table column: {col_err}')
            
            # Check if target_id column exists, add it if not
            try:
                cursor.execute("SHOW COLUMNS FROM activity_logs LIKE 'target_id'")
                if not cursor.fetchone():
                    # Determine position - after target_table if it exists, otherwise after action
                    try:
                        cursor.execute("SHOW COLUMNS FROM activity_logs LIKE 'target_table'")
                        if cursor.fetchone():
                            cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_id INT NULL AFTER target_table")
                        else:
                            cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_id INT NULL AFTER action")
                    except Exception as pos_err:
                        # Fallback: just add after action
                        try:
                            cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_id INT NULL AFTER action")
                        except Exception:
                            # Last resort: add at the end
                            cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_id INT NULL")
                    print('✅ Added target_id column to activity_logs table')
            except Exception as col_err:
                # Column might already exist or table structure is different
                print(f'⚠️ Could not check/add target_id column: {col_err}')
            
            # Insert activity log - use column check to handle missing columns gracefully
            try:
                cursor.execute(
                    '''
                    INSERT INTO activity_logs (admin_id, action, target_table, target_id, details)
                    VALUES (%s, %s, %s, %s, %s)
                    ''',
                    (admin_id, action, target_table, target_id, details)
                )
            except Exception as insert_err:
                # If columns don't exist, try with different combinations
                error_msg = str(insert_err).lower()
                if 'target_table' in error_msg or 'target_id' in error_msg or 'unknown column' in error_msg:
                    try:
                        # Try without target_table but with target_id
                        cursor.execute(
                            '''
                            INSERT INTO activity_logs (admin_id, action, target_id, details)
                            VALUES (%s, %s, %s, %s)
                            ''',
                            (admin_id, action, target_id, details)
                        )
                        print(f'⚠️ Logged activity without target_table column: {action}')
                    except Exception as fallback1_err:
                        # Try without both target_table and target_id
                        try:
                            cursor.execute(
                                '''
                                INSERT INTO activity_logs (admin_id, action, details)
                                VALUES (%s, %s, %s)
                                ''',
                                (admin_id, action, details)
                            )
                            print(f'⚠️ Logged activity without target_table and target_id columns: {action}')
                        except Exception as fallback2_err:
                            print(f'⚠️ Failed to log HR activity (all fallbacks failed): {fallback2_err}')
                else:
                    raise
            
            # Only create notification if not skipped (to avoid duplicates)
            # notify_admin_on_hr_actions already creates proper user-friendly notifications
            if not skip_notification:
                msg = f'HR Action: {action} on {target_table}'
                if target_id:
                    msg += f' (ID: {target_id})'
                if details:
                    msg += f' - {details}'
                
                # Prevent JSON responses from being saved as notifications
                msg_str = str(msg).strip()
                if msg_str.startswith('{') and ('"success"' in msg_str or '"message"' in msg_str or '"error"' in msg_str):
                    print(f'⚠️ Blocked JSON response from being saved as notification in log_hr_activity: {msg_str[:100]}')
                    return True
                
                cursor.execute(
                    '''
                    INSERT INTO notifications (application_id, message, sent_at, is_read)
                    VALUES (NULL, %s, NOW(), 0)
                    ''',
                    (msg,)
                )
                db.commit()
            return True
        except Exception as e:
            db.rollback()
            print(f'⚠️ Failed to log HR activity: {e}')
            return False
        finally:
            cursor.close()
    except Exception:
        return False


@app.before_request
def notify_admin_on_hr_actions():
    """Mirror all HR branch actions (POST) to Admin as system notifications and activity logs."""
    try:
        user = getattr(g, 'current_user', None)
        if not user or user.get('role') != 'hr':
            return
        if request.method != 'POST':
            return
        
        # Skip notification creation for delete notifications routes to prevent JSON responses from being saved
        if '/notifications/delete' in request.path or '/notifications/delete-all' in request.path:
            return
        
        hr_id = user.get('id')
        hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
        if not hr_id:
            return
            
        db = get_db()
        if not db:
            return
        
        cursor = db.cursor(dictionary=True)
        try:
            # Extract action details from request
            action = request.form.get('action', '')
            path = request.path
            branch_id = session.get('branch_id')
            
            # Determine target table and ID from path and form data
            target_table = 'unknown'
            target_id = None
            details = None  # Will be set based on action type
            admin_message = None
            
            # Parse common actions and create detailed admin notifications
            if 'applications' in path:
                target_table = 'applications'
                target_id = request.form.get('application_id', type=int)
                if action == 'update_status':
                    new_status = request.form.get('status', '')
                    details = f'Status changed to: {new_status}'
                    # Fetch applicant name and job title for better notification message
                    applicant_name = 'applicant'
                    job_title = 'position'
                    if target_id:
                        try:
                            cursor.execute('''
                                SELECT ap.full_name, COALESCE(j.title, 'N/A') AS job_title
                                FROM applications a
                                JOIN applicants ap ON a.applicant_id = ap.applicant_id
                                LEFT JOIN jobs j ON a.job_id = j.job_id
                                WHERE a.application_id = %s
                                LIMIT 1
                            ''', (target_id,))
                            app_info = cursor.fetchone()
                            if app_info:
                                applicant_name = app_info.get('full_name') or 'applicant'
                                job_title = app_info.get('job_title') or 'position'
                        except Exception as fetch_err:
                            print(f'⚠️ Error fetching applicant info for admin notification: {fetch_err}')
                    status_display = new_status.replace('_', ' ').title()
                    admin_message = f'HR {hr_name} updated application status to {status_display} for {applicant_name} ({job_title}).'
            elif 'applicants' in path:
                target_table = 'applicants'
                target_id = request.form.get('applicant_id', type=int)
                if action == 'bulk_update_status':
                    bulk_status = request.form.get('bulk_status', '')
                    details = f'Bulk status update: {bulk_status}'
                    admin_message = f'HR {hr_name} performed bulk status update to {bulk_status.replace("_", " ").title()}'
            elif 'job-postings' in path or 'job_postings' in path:
                target_table = 'jobs'
                target_id = request.form.get('job_id', type=int) or (path.split('/')[-1] if path.split('/')[-1].isdigit() else None)
                
                # Determine if this is an add or update based on path and action
                is_update = 'update' in path or action in ['update', 'edit']
                is_add = action in ['add', 'create'] and not is_update
                
                # Get job title if available for better notification
                job_title = None
                if target_id:
                    try:
                        cursor.execute('SELECT title FROM jobs WHERE job_id = %s LIMIT 1', (target_id,))
                        job_row = cursor.fetchone()
                        if job_row:
                            job_title = job_row.get('title') if isinstance(job_row, dict) else job_row[0]
                    except:
                        pass
                
                if is_add:
                    details = f'Job: {job_title}' if job_title else 'New job posting added'
                    if job_title:
                        admin_message = f'HR {hr_name} posted a new job: "{job_title}"'
                    else:
                        admin_message = f'HR {hr_name} posted a new job posting'
                elif is_update:
                    details = f'Job Updated: {job_title}' if job_title else 'Job posting updated'
                    if job_title:
                        admin_message = f'HR {hr_name} updated job posting: "{job_title}"'
                    else:
                        admin_message = f'HR {hr_name} updated a job posting'
                elif action == 'delete':
                    details = f'Job Deleted: {job_title}' if job_title else 'Job posting deleted'
                    if job_title:
                        admin_message = f'HR {hr_name} deleted job posting: "{job_title}"'
                    else:
                        admin_message = f'HR {hr_name} deleted a job posting'
                else:
                    # Fallback for other actions
                    details = f'Job Modified: {job_title}' if job_title else 'Job posting modified'
                    if job_title:
                        admin_message = f'HR {hr_name} modified job posting: "{job_title}"'
                    else:
                        admin_message = f'HR {hr_name} modified a job posting'
            elif 'interviews' in path:
                target_table = 'interviews'
                target_id = request.form.get('interview_id', type=int)
                if action in ['schedule', 'update', 'reschedule', 'cancel']:
                    details = f'Interview {action.title()}ed'
                    admin_message = f'HR {hr_name} {action}d interview'
                    if target_id:
                        admin_message += f' (Interview ID: {target_id})'
            elif 'communications' in path:
                target_table = 'communications'
                # HR communications route has been removed - no longer tracking
                admin_message = None
                details = None
            elif 'reports' in path:
                target_table = 'reports'
                details = 'Accessed reports'
                admin_message = f'HR {hr_name} accessed reports/analytics'
            
            # Determine user-friendly action name for logging early
            user_friendly_action = determine_user_friendly_action(path, action, target_table)

            # Default details if not set by any action handler — use friendly action as fallback
            if not details:
                details = f'Action: {user_friendly_action}' if user_friendly_action else 'System action'

            # Create admin notification if we have a message
            if admin_message:
                # Don't add branch ID to notification message - it's not user-friendly
                try:
                    create_admin_notification(cursor, admin_message)
                    db.commit()
                    print(f'✅ Admin notification created: {admin_message}')
                except Exception as notify_err:
                    print(f'⚠️ Error creating admin notification: {notify_err}')

            # Log the activity (for activity logs only, not notifications)
            # Don't create duplicate notification in log_hr_activity
            # Only append branch info when we have a string in details
            try:
                if branch_id and details:
                    details = f'{details} | Branch: {branch_id}'
            except Exception:
                # Safety: ensure details is a string
                details = str(details or user_friendly_action or 'System action')
                if branch_id:
                    details = f'{details} | Branch: {branch_id}'

            # Pass skip_notification=True to prevent duplicate notification creation
            log_hr_activity(hr_id, user_friendly_action, target_table, target_id, details, skip_notification=True)
        finally:
            cursor.close()
        
    except Exception as e:
        # Non-blocking; never break the request due to admin-notify mirror
        print(f'⚠️ Error in notify_admin_on_hr_actions: {e}')
        pass

@app.context_processor
def inject_user():
    """Provide current user information to all Jinja templates."""
    user = getattr(g, 'current_user', None)
    # Provide HR notifications (branch-scoped for specific branch, all notifications if managing all branches)
    notifications = []
    notif_count = 0
    branches = []
    if user and user.get('role') == 'hr':
        branch_id = user.get('branch_id')  # None means all branches
        if branch_id:
            # HR with specific branch: show branch-scoped notifications
            notifications, notif_count = fetch_notifications_for({'branch_id': branch_id}, limit=5)
        else:
            # HR managing all branches: show all notifications (no scope)
            notifications, notif_count = fetch_notifications_for(None, limit=5)
        # Fetch all branches for HR branch selector dropdown
        branches = fetch_branches()
    return {
        'current_user': user,
        'hr_notifications': notifications,
        'hr_notif_count': notif_count,
        'hr_branches': branches,
        'is_logged_in': is_logged_in,
    }

@app.context_processor
def inject_csrf_token():
    """Make CSRF token function available in all templates."""
    def csrf_token():
        """Generate CSRF token for forms."""
        return generate_csrf()
    return dict(csrf_token=csrf_token)

@app.context_processor
def inject_applicant_notifications():
    """Provide notification data to applicant templates."""
    if is_logged_in() and session.get('user_role') == 'applicant':
        applicant_id = session.get('user_id')
        if applicant_id:
            db = get_db()
            if db:
                cursor = db.cursor(dictionary=True)
                try:
                    ensure_schema_compatibility()
                    # Check notifications table columns
                    cursor.execute('SHOW COLUMNS FROM notifications')
                    notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                    
                    has_application_fk = 'application_id' in notification_columns
                    if not has_application_fk:
                        return {'unread_count': 0, 'recent_notifications': []}
                    
                    # Build sent_at expression
                    if 'sent_at' in notification_columns:
                        sent_at_expr = 'n.sent_at'
                    elif 'created_at' in notification_columns:
                        sent_at_expr = 'n.created_at'
                    else:
                        sent_at_expr = 'NOW()'
                    
                    # Build is_read expression
                    if 'is_read' in notification_columns:
                        is_read_expr = 'COALESCE(n.is_read, 0)'
                    else:
                        is_read_expr = '0'
                    
                    # Get dynamic job column expressions
                    _update_job_columns(cursor)
                    job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'System Notification'")
                    
                    select_fields = [
                        'n.notification_id',
                        'n.message',
                        f'{sent_at_expr} AS sent_at',
                        f'{is_read_expr} AS is_read',
                        'n.application_id',
                        'COALESCE(a.status, \'\') AS application_status',
                        f'COALESCE({job_title_expr}, \'System Notification\') AS job_title',
                    ]
                    query = f'''
                        SELECT DISTINCT {', '.join(select_fields)}
                        FROM notifications n
                        JOIN applications a ON n.application_id = a.application_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        WHERE a.applicant_id = %s
                        AND (
                            n.message LIKE 'You applied for%'
                            OR n.message LIKE 'Your application status%'
                            OR n.message LIKE 'Congratulations! You have been hired%'
                        )
                        ORDER BY {sent_at_expr} DESC
                        LIMIT 5
                    '''
                    cursor.execute(query, (applicant_id,))
                    notifications = cursor.fetchall() or []
                    
                    # Deduplicate by notification_id to prevent duplicates
                    seen_ids = set()
                    unique_notifications = []
                    for notif in notifications:
                        notif_id = notif.get('notification_id')
                        if notif_id and notif_id not in seen_ids:
                            seen_ids.add(notif_id)
                            unique_notifications.append(notif)
                    notifications = unique_notifications
                    
                    # Count unread
                    unread_count = len([n for n in notifications if not n.get('is_read')])
                    
                    # Format notifications same as communications page
                    formatted_notifications = []
                    for notif in notifications:
                        formatted_notifications.append({
                            'notification_id': notif.get('notification_id'),
                            'message': notif.get('message'),
                            'sent_at': format_human_datetime(notif.get('sent_at')),
                            'is_read': notif.get('is_read', False),
                            'application_id': notif.get('application_id'),
                            'job_title': notif.get('job_title'),
                            'application_status': notif.get('application_status'),
                        })
                    
                    return {'unread_count': unread_count, 'recent_notifications': formatted_notifications}
                except Exception as e:
                    print(f'⚠️ Error fetching applicant notifications: {e}')
                    return {'unread_count': 0, 'recent_notifications': []}
                finally:
                    if cursor:
                        cursor.close()
    return {'unread_count': 0, 'recent_notifications': []}

@app.context_processor
def inject_admin_notifications():
    """Provide admin notification data to all admin templates."""
    user = getattr(g, 'current_user', None)
    if user and user.get('role') in ('admin', 'hr'):
        try:
            # Fetch system-level notifications (application_id IS NULL)
            formatted, unread = fetch_notifications_for({'system_only': True}, limit=5)
            return {
                'admin_notifs': formatted,
                'admin_notif_count': unread,
                'admin_notif_display': '99+' if unread > 99 else unread
            }
        except Exception as e:
            print(f'⚠️ Error fetching admin notifications: {e}')
            return {
                'admin_notifs': [],
                'admin_notif_count': 0,
                'admin_notif_display': 0
            }
    return {
        'admin_notifs': [],
        'admin_notif_count': 0,
        'admin_notif_display': 0
    }

# ------------------------------


# ---------------------------------------------------------------------------
# HR Notifications endpoint (alias for HR communications/logs)
# Fix for templates expecting 'hr_notifications' endpoint.
# ---------------------------------------------------------------------------
@app.route('/hr/notifications')
@login_required('hr')
def hr_notifications():
    """View all HR notifications scoped to the HR user's branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('hr/notifications.html', notifications=[], unread_count=0)
    ensure_schema_compatibility()
    cursor = db.cursor(dictionary=True)
    try:
        # Clean up any JSON response notifications first (AGGRESSIVE CLEANUP)
        branch_id = session.get('branch_id')
        try:
            # First, delete JSON notifications that match the pattern
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    LEFT JOIN applications a ON n.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE (j.branch_id = %s OR n.application_id IS NULL)
                    AND (
                        (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                        OR n.message LIKE '%Notifications deleted%'
                        OR n.message LIKE '%All notifications deleted%'
                        OR n.message LIKE '%Notification deleted successfully%'
                    )
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE (
                        (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                        OR message = 'Notifications deleted.'
                        OR message LIKE '%Notifications deleted%'
                        OR message LIKE '%All notifications deleted%'
                        OR message LIKE '%Notification deleted successfully%'
                    )
                """)
            json_cleaned = cursor.rowcount
            if json_cleaned > 0:
                print(f'✅ Cleaned up {json_cleaned} JSON response notification(s) from HR notifications page')
                db.commit()
        except Exception as cleanup_err:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_err}')
            db.rollback()
        
        # Ensure notifications table and columns
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        sent_at_expr = 'COALESCE(n.sent_at, n.created_at, NOW())' if 'sent_at' in notification_columns else 'COALESCE(n.created_at, NOW())'
        is_read_expr = 'COALESCE(n.is_read, 0)' if 'is_read' in notification_columns else '0'
        # Scope to HR branch via jobs
        params = []
        where_sql = ''
        if branch_id:
            where_sql = 'WHERE j.branch_id = %s'
            params.append(branch_id)
        # HR focuses on application-related notifications
        # CRITICAL: Exclude ALL applicant-only notifications - these are for applicants only, NOT HR
        # HR should only see notifications about branch activities, not applicant-facing messages
        hr_where_sql = where_sql
        applicant_only_filters = [
            'n.message NOT LIKE \'You applied for%\'',
            'n.message NOT LIKE \'Congratulations! You have been hired%\'',
            'n.message NOT LIKE \'Your application status%\'',
            'n.message NOT LIKE \'Congratulations! You%\'',  # Catch any variation of congratulations messages to applicants
        ]
        if hr_where_sql:
            hr_where_sql += ' AND ' + ' AND '.join(applicant_only_filters)
        else:
            hr_where_sql = 'WHERE ' + ' AND '.join(applicant_only_filters)
        
        cursor.execute(
            f'''
            SELECT n.notification_id,
                   n.message,
                   {sent_at_expr} AS sent_at,
                   {is_read_expr} AS is_read,
                   a.application_id,
                   a.status AS application_status,
                   COALESCE(j.title, 'N/A') AS job_title,
                   COALESCE(ap.full_name, 'Unknown') AS applicant_name
            FROM notifications n
            JOIN applications a ON n.application_id = a.application_id
            JOIN jobs j ON a.job_id = j.job_id
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            {hr_where_sql}
            ORDER BY {sent_at_expr} DESC
            LIMIT 200
            ''',
            tuple(params)
        )
        rows = cursor.fetchall() or []
        # Filter out any JSON response notifications that might have slipped through
        filtered_rows = []
        for r in rows:
            message = r.get('message', '')
            # Skip JSON responses - check for exact JSON string and patterns
            if message:
                message_str = str(message).strip()
                # Skip if it's a JSON response
                if (message_str.startswith('{') and ('"success"' in message_str or '"message"' in message_str or '"error"' in message_str)):
                    continue
                # Skip exact JSON response string
                if message_str == '{"message":"Notifications deleted.","success":true}':
                    continue
                # Skip if message contains the JSON response pattern
                if '{"message":"Notifications deleted.' in message_str:
                    continue
            filtered_rows.append(r)
        
        unread_count = len([r for r in filtered_rows if not r.get('is_read')])
        notifications = [
            {
                'notification_id': r.get('notification_id'),
                'message': r.get('message', ''),
                'sent_at': format_human_datetime(r.get('sent_at')),
                'is_read': r.get('is_read', False),
                'application_id': r.get('application_id'),
                'job_title': r.get('job_title'),
                'applicant_name': r.get('applicant_name'),
                'application_status': r.get('application_status'),
            }
            for r in filtered_rows
        ]
        return render_template('hr/notifications.html', notifications=notifications, unread_count=unread_count)
    except Exception as exc:
        import traceback
        print('❌ HR notifications error:', exc)
        print(traceback.format_exc())
        flash('Unable to load HR notifications. Please try again later.', 'error')
        return render_template('hr/notifications.html', notifications=[], unread_count=0)
    finally:
        cursor.close()


@app.route('/hr/notifications/read-all', methods=['POST'])
@login_required('hr')
def mark_all_hr_notifications_read():
    """Mark all HR notifications as read, scoped to the HR user's branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_notifications'))
    
    cursor = db.cursor()
    try:
        ensure_schema_compatibility()
        branch_id = session.get('branch_id')
        
        # Check if is_read column exists
        cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
        has_is_read = cursor.fetchone() is not None
        
        if not has_is_read:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification read status not available.'}), 400
            flash('Notification read status not available.', 'error')
            return redirect(url_for('hr_notifications'))
        
        # Clean up any JSON responses that might have been stored as notifications
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE j.branch_id = %s
                    AND n.message LIKE '{%' 
                    AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%')
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE message LIKE '{%' 
                    AND (message LIKE '%"success"%' OR message LIKE '%"message"%')
                """)
        except Exception as cleanup_error:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_error}')
        
        # Mark all notifications as read, scoped to branch
        if branch_id:
            cursor.execute("""
                UPDATE notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                SET n.is_read = 1
                WHERE j.branch_id = %s AND n.is_read = 0
            """, (branch_id,))
        else:
            cursor.execute('UPDATE notifications SET is_read = 1 WHERE is_read = 0')
        
        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'All notifications marked as read'})
        flash('All notifications marked as read.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Mark all HR notifications read error: {exc}')
        print(f'Full traceback: {error_details}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to mark all notifications as read.'}), 500
        flash('Failed to mark all notifications as read.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('hr_notifications'))


@app.route('/hr/notifications/<int:notification_id>/read', methods=['POST'])
@login_required('hr')
def mark_hr_notification_read(notification_id):
    """Mark a notification as read, verifying it belongs to the HR user's branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_notifications'))
    
    cursor = db.cursor(dictionary=True)
    try:
        ensure_schema_compatibility()
        branch_id = session.get('branch_id')
        
        # Verify notification belongs to HR branch
        if branch_id:
            cursor.execute("""
                SELECT n.notification_id
                FROM notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                WHERE n.notification_id = %s AND j.branch_id = %s
                LIMIT 1
            """, (notification_id, branch_id))
        else:
            cursor.execute("""
                SELECT notification_id FROM notifications WHERE notification_id = %s LIMIT 1
            """, (notification_id,))
        
        notif_record = cursor.fetchone()
        if not notif_record:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification not found or access denied.'}), 404
            flash('Notification not found or access denied.', 'error')
            return redirect(url_for('hr_notifications'))
        
        # Check if this notification is a JSON response and delete it if so
        cursor.execute('SELECT message FROM notifications WHERE notification_id = %s', (notification_id,))
        notif_msg = cursor.fetchone()
        if notif_msg:
            message = notif_msg.get('message', '') if isinstance(notif_msg, dict) else (notif_msg[0] if notif_msg else '')
            if message and message.strip().startswith('{') and '"success"' in message:
                cursor.execute('DELETE FROM notifications WHERE notification_id = %s', (notification_id,))
                db.commit()
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                    return jsonify({'success': True, 'message': 'Invalid notification removed', 'notification_id': notification_id})
                flash('Invalid notification removed.', 'success')
                return redirect(url_for('hr_notifications'))
        
        # Check if is_read column exists
        cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
        has_is_read = cursor.fetchone() is not None
        
        if has_is_read:
            cursor.execute(
                'UPDATE notifications SET is_read = 1 WHERE notification_id = %s',
                (notification_id,)
            )
            db.commit()
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': True, 'message': 'Notification marked as read', 'notification_id': notification_id})
            flash('Notification marked as read.', 'success')
        else:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification read status not available.'}), 400
            flash('Notification read status not available.', 'error')
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Mark HR notification read error: {exc}')
        print(f'Full traceback: {error_details}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to mark notification as read.'}), 500
        flash('Failed to mark notification as read.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('hr_notifications'))


@app.route('/hr/notifications/<int:notification_id>/delete', methods=['POST'])
@login_required('hr')
def delete_hr_notification(notification_id):
    """Delete a notification, verifying it belongs to the HR user's branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_notifications'))
    
    cursor = db.cursor(dictionary=True)
    try:
        branch_id = session.get('branch_id')
        
        # Verify notification belongs to HR branch
        if branch_id:
            cursor.execute("""
                SELECT n.notification_id
                FROM notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                WHERE n.notification_id = %s AND j.branch_id = %s
                LIMIT 1
            """, (notification_id, branch_id))
        else:
            cursor.execute("""
                SELECT notification_id FROM notifications WHERE notification_id = %s LIMIT 1
            """, (notification_id,))
        
        notif_record = cursor.fetchone()
        if not notif_record:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification not found or access denied.'}), 404
            flash('Notification not found or access denied.', 'error')
            return redirect(url_for('hr_notifications'))
        
        # Clean up any JSON response notifications before deleting (AGGRESSIVE)
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    LEFT JOIN applications a ON n.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE (j.branch_id = %s OR n.application_id IS NULL)
                    AND (
                        (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                        OR n.message = 'Notifications deleted.'
                        OR n.message LIKE '%Notifications deleted%'
                        OR n.message LIKE '%All notifications deleted%'
                        OR n.message LIKE '%Notification deleted successfully%'
                    )
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE (
                        (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                        OR message = 'Notifications deleted.'
                        OR message = 'Notifications deleted.'
                        OR message LIKE '%Notifications deleted%'
                        OR message LIKE '%All notifications deleted%'
                        OR message LIKE '%Notification deleted successfully%'
                    )
                """)
            json_cleaned = cursor.rowcount
            if json_cleaned > 0:
                print(f'✅ Cleaned up {json_cleaned} JSON response notification(s) before delete')
                db.commit()
        except Exception as cleanup_err:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_err}')
            db.rollback()
        
        # Delete only the specific notification
        cursor.execute('DELETE FROM notifications WHERE notification_id = %s', (notification_id,))
        deleted_count = cursor.rowcount
        
        if deleted_count == 0:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification not found or already deleted.'}), 404
            flash('Notification not found or already deleted.', 'error')
            return redirect(url_for('hr_notifications'))
        
        db.commit()
        
        # Final cleanup AFTER commit to catch any JSON notifications created by other processes
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    LEFT JOIN applications a ON n.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE (j.branch_id = %s OR n.application_id IS NULL)
                    AND (
                        (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                        OR n.message = 'Notifications deleted.'
                        OR n.message LIKE '%Notifications deleted%'
                        OR n.message LIKE '%All notifications deleted%'
                        OR n.message LIKE '%Notification deleted successfully%'
                    )
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE (
                        (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                        OR message = 'Notifications deleted.'
                        OR message = 'Notifications deleted.'
                        OR message LIKE '%Notifications deleted%'
                        OR message LIKE '%All notifications deleted%'
                        OR message LIKE '%Notification deleted successfully%'
                    )
                """)
            final_cleaned = cursor.rowcount
            if final_cleaned > 0:
                print(f'✅ Final cleanup after delete: Removed {final_cleaned} JSON response notification(s)')
                db.commit()
        except Exception as final_cleanup_err:
            print(f'⚠️ Error in final cleanup: {final_cleanup_err}')
            db.rollback()
        
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'Notification deleted successfully', 'notification_id': notification_id})
        flash('Notification deleted successfully.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Delete HR notification error: {exc}')
        print(f'Full traceback: {error_details}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to delete notification.'}), 500
        flash('Failed to delete notification.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('hr_notifications'))


@app.route('/hr/notifications/delete-all', methods=['POST'])
@login_required('hr')
def delete_all_hr_notifications():
    """Delete all HR notifications, scoped to the HR user's branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_notifications'))
    
    cursor = db.cursor()
    try:
        branch_id = session.get('branch_id')
        
        # First, clean up any JSON response notifications
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE j.branch_id = %s
                    AND n.message LIKE '{%' 
                    AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%')
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE message LIKE '{%' 
                    AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%')
                """)
            json_cleaned = cursor.rowcount
            if json_cleaned > 0:
                print(f'✅ Cleaned up {json_cleaned} JSON response notification(s) before delete-all')
        except Exception as cleanup_err:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_err}')
        
        # Delete notifications scoped to branch (excluding JSON responses)
        if branch_id:
            cursor.execute("""
                DELETE n FROM notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                WHERE j.branch_id = %s
            """, (branch_id,))
        else:
            cursor.execute('DELETE FROM notifications')
        
        deleted_count = cursor.rowcount
        db.commit()
        
        # CRITICAL: Final cleanup BEFORE returning response to prevent JSON from being saved
        # This must run AFTER commit but BEFORE any response is returned
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    LEFT JOIN applications a ON n.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE (j.branch_id = %s OR n.application_id IS NULL)
                    AND (
                        (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                        OR n.message = 'Notifications deleted.'
                        OR n.message LIKE '%Notifications deleted%'
                        OR n.message LIKE '%All notifications deleted%'
                        OR n.message LIKE '%Notification deleted successfully%'
                        OR n.message = '{"message":"Notifications deleted.","success":true}'
                        OR n.message LIKE '%{"message":"Notifications deleted.%'
                    )
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE (
                        (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                        OR message = 'Notifications deleted.'
                        OR message LIKE '%Notifications deleted%'
                        OR message LIKE '%All notifications deleted%'
                        OR message LIKE '%Notification deleted successfully%'
                        OR message = '{"message":"Notifications deleted.","success":true}'
                        OR message LIKE '%{"message":"Notifications deleted.%'
                    )
                """)
            final_cleaned = cursor.rowcount
            if final_cleaned > 0:
                print(f'✅ Final cleanup: Removed {final_cleaned} JSON response notification(s) after delete-all')
                db.commit()
        except Exception as final_cleanup_err:
            print(f'⚠️ Error in final cleanup: {final_cleanup_err}')
            db.rollback()
        
        # Return response - JSON response should NEVER be saved as notification
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': f'All notifications deleted successfully ({deleted_count} notification(s) removed)'})
        flash(f'All notifications deleted successfully ({deleted_count} notification(s) removed).', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Delete all HR notifications error: {exc}')
        print(f'Full traceback: {error_details}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to delete all notifications.'}), 500
        flash('Failed to delete all notifications.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('hr_notifications'))


def generate_token():
    """Generate a random token suitable for email verification and password resets."""
    return uuid4().hex


def send_verification_email(email, token, applicant_name=None):
    """Send a verification email containing the confirmation link with HTML template."""
    from flask import render_template_string
    import os
    
    verification_url = url_for('verify_email', token=token, _external=True)
    # Generate absolute URL for logo image (needed for email clients)
    logo_url = url_for('static', filename='images/whitehat_logo.jpg', _external=True)
    subject = 'Verify your J&T Express applicant account'
    
    # Plain text version
    body = f"""
Hi{(' ' + applicant_name) if applicant_name else ''},

Thank you for registering with J&T Express.

Please confirm your email address by clicking the link below:
{verification_url}

If you did not create this account, you can safely ignore this email.

This verification link will expire after 60 seconds.

Regards,
J&T Express Recruitment Team
"""
    
    # HTML version
    html_template_path = os.path.join('templates', 'emails', 'verification_email.html')
    html_body = None
    
    try:
        if os.path.exists(html_template_path):
            with open(html_template_path, 'r', encoding='utf-8') as f:
                html_template = f.read()
            html_body = render_template_string(html_template, verification_url=verification_url, logo_url=logo_url)
    except Exception as e:
        print(f'⚠️ Could not load HTML email template: {e}')
    
    send_email(email, subject, body, html_body)


def send_password_reset_email(email, token):
    """Send password reset instructions to the applicant."""
    reset_url = url_for('reset_password', token=token, _external=True)
    subject = 'Reset your J&T Express password'
    body = f"""
    Hi,

    A password reset was requested for your account. To set a new password, please click the link below:
    {reset_url}

    This link will expire in 30 minutes. If you did not request a reset, you can ignore this email.

    Regards,
    J&T Express Recruitment Team
    """
    send_email(email, subject, body)


def get_valid_admin_id(admin_id):
    """Validate that admin_id exists in admins table. Returns admin_id if valid, None otherwise."""
    if not admin_id:
        return None
    
    db = get_db()
    if not db:
        return None
    
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            'SELECT admin_id FROM admins WHERE admin_id = %s LIMIT 1',
            (admin_id,)
        )
        result = cursor.fetchone()
        return result['admin_id'] if result else None
    except Exception:
        return None
    finally:
        cursor.close()


def log_profile_change(user_id, role, field, old_value, new_value):
    """Persist profile changes for audit history."""
    if old_value == new_value:
        return

    try:
        execute_query(
            """
            INSERT INTO profile_changes (user_id, role, field_changed, old_value, new_value)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (user_id, role, field, old_value or '', new_value or ''),
        )
    except Exception as exc:
        print(f'⚠️ Failed to log profile change: {exc}')


def ensure_default_accounts():
    """Create baseline System Administrator account if it does not exist. Only System Administrator can create HR accounts."""
    db = get_db()
    if not db:
        print('❌ Unable to verify default accounts because the database connection failed.')
        return False
            
    cursor = db.cursor(dictionary=True)

    try:
        # Only create System Administrator - no HR accounts
        account = {
                'full_name': 'System Administrator',
                'email': 'admin@whitehat88.com',
                'role': 'admin',
                'password': 'whitehat88@2025',
        }

        created_any = False

        # Check if System Administrator already exists
        cursor.execute(
            """
            SELECT u.user_id, u.user_type, a.admin_id
            FROM users u
            LEFT JOIN admins a ON a.user_id = u.user_id
            WHERE u.email = %s
            LIMIT 1
            """,
            (account['email'],),
        )

        exists = cursor.fetchone()

        # If exists and has admin record with correct settings, skip
        if exists and exists.get('admin_id'):
            # Verify it's set up correctly
            if exists.get('user_type') == 'super_admin' and exists.get('branch_id') is None:
                return True  # Already correctly configured
            else:
                # Fix existing account
                user_id = exists['user_id']
                # Update user_type to super_admin and ensure email is verified
                cursor.execute(
                    "UPDATE users SET user_type = 'super_admin', email_verified = 1 WHERE user_id = %s",
                    (user_id,)
                )
                # Update admin record
                cursor.execute(
                    "UPDATE admins SET full_name = 'System Administrator' WHERE admin_id = %s",
                    (exists['admin_id'],)
                )
                # Update password_hash
                password_hash = hash_password(account['password'])
                cursor.execute(
                    "UPDATE users SET password_hash = %s WHERE user_id = %s",
                    (password_hash, user_id)
                )
                db.commit()
                print(f"✅ Fixed System Administrator account: {account['email']}")
                return True

        user_type = 'super_admin'  # Always super_admin for System Administrator
            
        if exists and exists.get('user_id') and not exists.get('admin_id'):
            # User exists but no admin record, create admin record
            user_id = exists['user_id']
            user_email = account['email']
            # Update user_type to super_admin if needed and ensure email is verified
            cursor.execute(
                "UPDATE users SET user_type = 'super_admin', email_verified = 1 WHERE user_id = %s",
                (user_id,)
            )
        else:
            # Create new user account
            # Admin/HR accounts are automatically verified (no email verification required)
            password_hash = hash_password(account['password'])
            cursor.execute(
                """
                INSERT INTO users (email, password_hash, user_type, is_active, email_verified)
                VALUES (%s, %s, %s, %s, 1)
                """,
                (account['email'], password_hash, user_type, True),
            )
            user_id = cursor.lastrowid
            user_email = account['email']

        # Create or update admin record
        cursor.execute(
            """
            SELECT admin_id FROM admins WHERE user_id = %s LIMIT 1
            """,
            (user_id,),
        )
        admin_exists = cursor.fetchone()
        
        if not admin_exists:
            # Get password_hash from users table to copy to admins table
            cursor.execute(
                "SELECT password_hash FROM users WHERE user_id = %s LIMIT 1",
                (user_id,)
            )
            user_record = cursor.fetchone()
            password_hash = user_record.get('password_hash') if user_record else None
            
            # If password_hash is still None, create it
            if not password_hash:
                password_hash = hash_password(account['password'])
                cursor.execute(
                    "UPDATE users SET password_hash = %s WHERE user_id = %s",
                    (password_hash, user_id)
                )
            
            # Insert admin record with password_hash
            cursor.execute(
                """
                INSERT INTO admins (user_id, full_name, email, password_hash)
                VALUES (%s, %s, %s, %s)
                """,
                (user_id, account['full_name'], user_email, password_hash),
            )
            created_any = True
            print(
                f"✅ Created System Administrator account: {account['email']} / {account['password']}"
            )
        else:
            # Update existing admin record
            cursor.execute(
                "UPDATE admins SET full_name = 'System Administrator' WHERE admin_id = %s",
                (admin_exists['admin_id'],)
            )

        if created_any:
            db.commit()

        return True
        
    except Exception as exc:
        db.rollback()
        print(f'❌ Error creating default accounts: {exc}')
        return False
    finally:
        cursor.close()


def _first_value(row, default=0):
    """Return the first value from a dictionary result or a default."""
    if not row:
        return default

    for value in row.values():
        if value is None:
            return default
        if isinstance(value, Decimal):
            return float(value)
        return value

    return default


def fetch_count(query, params=None, default=0):
    """Execute a COUNT-style query and safely extract the resulting value."""
    row = execute_query(query, params, fetch_one=True)
    return _first_value(row, default)


def fetch_rows(query, params=None):
    """Execute a query returning multiple rows, ensuring a list result."""
    rows = execute_query(query, params, fetch_all=True)
    return rows or []


def fetch_branches():
    """Return all branches ordered alphabetically."""
    return fetch_rows('SELECT branch_id, branch_name, address FROM branches ORDER BY branch_name ASC')


def fetch_positions():
    """Return all job positions ordered alphabetically."""
    # Positions table has been removed, return empty list
    return []


def get_branch_scope(user):
    """Return the branch_id associated with an HR user, otherwise None.
    For HR users, always return None to allow management of all branches.
    HR can manage all branches without restrictions."""
    if user and user.get('role') == 'hr':
        # HR users can manage all branches - no scoping restrictions
        return None
    return None


def fetch_hr_accounts():
    """Fetch HR administrator accounts. All HR accounts manage all branches."""
    db = get_db()
    if not db:
        print('⚠️ No database connection in fetch_hr_accounts')
        return []
    
    cursor = db.cursor(dictionary=True)
    try:
        # First, check if there are any HR users in the users table
        cursor.execute("SELECT COUNT(*) as count FROM users WHERE user_type = 'hr'")
        hr_count = cursor.fetchone()
        print(f'🔍 Total HR users in users table: {hr_count.get("count", 0) if hr_count else 0}')
        
        # Check if there are any admins linked to HR users
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM admins a
            JOIN users u ON u.user_id = a.user_id
            WHERE u.user_type = 'hr'
        """)
        admin_hr_count = cursor.fetchone()
        print(f'🔍 Total HR admins (joined): {admin_hr_count.get("count", 0) if admin_hr_count else 0}')
        
        # Now fetch the actual HR accounts
        cursor.execute(
        """
        SELECT
            a.admin_id,
            a.full_name,
            u.email,
            NULL AS branch_id,
            'All Branches' AS branch_name,
            u.is_active,
            0 AS assigned_branch_count
        FROM admins a
        JOIN users u ON u.user_id = a.user_id
        WHERE u.user_type = 'hr'
        ORDER BY a.full_name ASC
        """
    )
        rows = cursor.fetchall()
        print(f'✅ fetch_hr_accounts: Found {len(rows) if rows else 0} HR accounts')
        if rows:
            print(f'🔍 Sample HR account: {rows[0]}')
        return rows or []
    except Exception as e:
        print(f'❌ Error fetching HR accounts: {e}')
        import traceback
        traceback.print_exc()
        return []
    finally:
        cursor.close()


def fetch_all_applications(user=None):
    """Fetch applications, automatically scoping to HR user's branch."""
    branch_id = get_branch_scope(user)

    if branch_id:
        # HR users: Only see applications for jobs in their branch
        query = """
            SELECT a.application_id,
                   ap.full_name AS applicant_name,
                   ap.email AS applicant_email,
                   j.title AS job_title,
                   COALESCE(b.branch_name,'Unassigned') AS branch_name,
                   a.status,
                   a.submitted_at
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE j.branch_id = %s
            ORDER BY a.submitted_at DESC
        """
        params = (branch_id,)
        print(f'🔍 fetch_all_applications: HR user - filtering by branch_id={branch_id}')
    else:
        # Admin users: See all applications from all branches
        query = """
            SELECT a.application_id,
                   ap.full_name AS applicant_name,
                   ap.email AS applicant_email,
                   j.title AS job_title,
                   COALESCE(b.branch_name,'Unassigned') AS branch_name,
                   a.status,
                   a.submitted_at
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            ORDER BY a.submitted_at DESC
        """
        params = None
        print(f'🔍 fetch_all_applications: Admin user - showing all applications from all branches')

    return fetch_rows(query, params)


def build_report_stats(user=None, date_filter='', date_params=None):
    """Build report statistics with optional date filtering."""
    branch_id = get_branch_scope(user)
    date_params = date_params or []

    # Status normalization: map legacy statuses to canonical ones
    pending_statuses = ['pending', 'reviewed', 'applied', 'under_review', 'shortlisted']
    interviewed_statuses = ['interviewed', 'interview']
    hired_statuses = ['hired', 'accepted']
    rejected_statuses = ['rejected']

    if branch_id:
        base = "FROM applications a JOIN jobs j ON a.job_id = j.job_id WHERE j.branch_id = %s"
        params = (branch_id,)
        date_clause = f" {date_filter}" if date_filter else ""
        all_params = list(params) + date_params

        # Build IN clauses for status groups
        pending_placeholders = ','.join(['%s'] * len(pending_statuses))
        interviewed_placeholders = ','.join(['%s'] * len(interviewed_statuses))
        hired_placeholders = ','.join(['%s'] * len(hired_statuses))
        rejected_placeholders = ','.join(['%s'] * len(rejected_statuses))

        stats = {
            'total_applications': fetch_count(f'SELECT COUNT(*) AS count {base}{date_clause}', tuple(all_params)),
            'pending': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'interviewed': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({interviewed_placeholders})", tuple(all_params + interviewed_statuses)),
            'hired': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({hired_placeholders})", tuple(all_params + hired_statuses)),
            'rejected': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({rejected_placeholders})", tuple(all_params + rejected_statuses)),
            'total_interviews': fetch_count(f"SELECT COUNT(DISTINCT i.interview_id) AS count FROM applications a JOIN jobs j ON a.job_id = j.job_id LEFT JOIN interviews i ON i.application_id = a.application_id WHERE j.branch_id = %s{date_clause.replace('a.submitted_at', 'i.scheduled_date') if date_filter else ''}", tuple(all_params)),
            # Legacy mappings for backward compatibility
            'applied': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'reviewed': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'accepted': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({hired_placeholders})", tuple(all_params + hired_statuses)),
        }
    else:
        # Build IN clauses for status groups
        pending_placeholders = ','.join(['%s'] * len(pending_statuses))
        interviewed_placeholders = ','.join(['%s'] * len(interviewed_statuses))
        hired_placeholders = ','.join(['%s'] * len(hired_statuses))
        rejected_placeholders = ','.join(['%s'] * len(rejected_statuses))

        if date_filter:
            # Remove leading 'AND ' only, not all occurrences
            date_filter_clean = date_filter.strip()
            if date_filter_clean.startswith('AND '):
                date_filter_clean = date_filter_clean[4:]  # Remove 'AND ' from start
            # Replace 'a.submitted_at' with 'submitted_at' since queries without branch_id don't use alias 'a'
            date_filter_clean = date_filter_clean.replace('a.submitted_at', 'submitted_at')
            date_clause = f" WHERE {date_filter_clean}"
            base_query = f'SELECT COUNT(*) AS count FROM applications{date_clause}'
            status_query_base = f'SELECT COUNT(*) AS count FROM applications{date_clause} AND status'
            all_params = date_params
            # For interview query, we need to use alias 'a', so create separate clause
            interview_date_filter = date_filter.strip()
            if interview_date_filter.startswith('AND '):
                interview_date_filter = interview_date_filter[4:]
            interview_date_clause = f" WHERE {interview_date_filter.replace('a.submitted_at', 'i.scheduled_date')}"
        else:
            base_query = 'SELECT COUNT(*) AS count FROM applications'
            status_query_base = 'SELECT COUNT(*) AS count FROM applications WHERE status'
            all_params = []
            date_clause = ''
            interview_date_clause = ''

        stats = {
            'total_applications': fetch_count(base_query, tuple(all_params) if all_params else None),
            'pending': fetch_count(f"{status_query_base} IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'interviewed': fetch_count(f"{status_query_base} IN ({interviewed_placeholders})", tuple(all_params + interviewed_statuses)),
            'hired': fetch_count(f"{status_query_base} IN ({hired_placeholders})", tuple(all_params + hired_statuses)),
            'rejected': fetch_count(f"{status_query_base} IN ({rejected_placeholders})", tuple(all_params + rejected_statuses)),
            'total_interviews': fetch_count(f"SELECT COUNT(DISTINCT i.interview_id) AS count FROM applications a LEFT JOIN interviews i ON i.application_id = a.application_id{interview_date_clause}", tuple(all_params) if all_params else None),
            # Legacy mappings for backward compatibility
            'applied': fetch_count(f"{status_query_base} IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'reviewed': fetch_count(f"{status_query_base} IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'under_review': fetch_count(f"{status_query_base} IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'interview': fetch_count(f"{status_query_base} IN ({interviewed_placeholders})", tuple(all_params + interviewed_statuses)),
            'accepted': fetch_count(f"{status_query_base} IN ({hired_placeholders})", tuple(all_params + hired_statuses)),
        }

    return stats


def to_iso(value):
    """Convert date/datetime objects to ISO strings for JavaScript consumption."""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat()
    return value


def format_human_datetime(value):
    """Produce a human-readable timestamp in 12-hour format (AM/PM)."""
    if isinstance(value, (datetime, date)):
        dt_value = value if isinstance(value, datetime) else datetime.combine(value, datetime.min.time())
        return dt_value.strftime('%b %d, %Y %I:%M %p')
    elif isinstance(value, str):
        try:
            # Try parsing common datetime formats
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d', '%Y-%m-%d %H:%M']:
                try:
                    dt_value = datetime.strptime(value, fmt)
                    return dt_value.strftime('%b %d, %Y %I:%M %p')
                except ValueError:
                    continue
        except:
            pass
    return value or ''


def create_admin_notification(cursor, message, application_id=None):
    """Insert a general notification entry for administrators/admin feed.
    Prevents duplicates by checking if notification with same message and application_id already exists.
    Also prevents JSON responses from being saved as notifications."""
    if not cursor or not message:
        return
    
    # Prevent JSON responses from being saved as notifications
    message_str = str(message).strip()
    if message_str.startswith('{') and ('"success"' in message_str or '"message"' in message_str or '"error"' in message_str):
        print(f'⚠️ Blocked JSON response from being saved as notification: {message_str[:100]}')
        return

    # Block or transform obvious applicant-facing messages from being saved to admin/HR feeds.
    # If message is applicant-facing but `application_id` is provided, attempt to rewrite it
    # into a third-person admin-friendly message (e.g., prefix with applicant name and job),
    # otherwise block and log.
    try:
        import re
        applicant_patterns = re.compile(r"^(\s)*(you|your)\b|\bcongratulations\b|you have been|you applied for|we are pleased to inform you",
                                        flags=re.IGNORECASE)
        if applicant_patterns.search(message_str):
            # Try to transform into admin-friendly message if we have an application_id to lookup
            if application_id is not None:
                try:
                    # Attempt to get applicant full name and job title for a better admin message
                    cursor.execute(
                        '''
                        SELECT a.application_id, ap.full_name AS applicant_name, COALESCE(j.title, '') AS job_title
                        FROM applications a
                        LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        WHERE a.application_id = %s
                        LIMIT 1
                        ''',
                        (application_id,)
                    )
                    app_row = cursor.fetchone()
                    applicant_name = None
                    job_title = None
                    if app_row:
                        applicant_name = app_row.get('applicant_name') if isinstance(app_row, dict) else (app_row[1] if len(app_row) > 1 else None)
                        job_title = app_row.get('job_title') if isinstance(app_row, dict) else (app_row[2] if len(app_row) > 2 else None)

                    # Remove leading applicant-directed fragments for a cleaner admin message
                    transformed = re.sub(r'^(congratulations[!]*\s*)', '', message_str, flags=re.IGNORECASE)
                    transformed = re.sub(r'^\s*(you|your)\b[:,]?\s*', '', transformed, flags=re.IGNORECASE)
                    transformed = transformed.strip().rstrip('.')

                    if applicant_name:
                        admin_message = f'Applicant {applicant_name}'
                        if job_title:
                            admin_message += f' ({job_title})'
                        admin_message += f': {transformed}'
                    else:
                        admin_message = transformed

                    print(f'ℹ️ Transformed applicant-facing message into admin message: {admin_message[:200]}')
                    # Replace original message with transformed admin_message for insertion below
                    message = admin_message
                    message_str = str(message).strip()
                except Exception as e:
                    print(f'⚠️ Could not rewrite applicant-facing notification to admin message: {e}')
                    # If we cannot safely rewrite, block to avoid leaking applicant-facing text
                    print(f'⚠️ Blocked applicant-facing notification from being saved to admin feed: {message_str[:200]}')
                    return
            else:
                print(f'⚠️ Blocked applicant-facing notification from being saved to admin feed (no application_id): {message_str[:200]}')
                return
    except Exception:
        # If regex check fails for any reason, continue — prefer not to block valid admin messages inadvertently
        pass
    
    try:
        cursor.execute("SHOW TABLES LIKE 'notifications'")
        if not cursor.fetchone():
            return
        
        cursor.execute('SHOW COLUMNS FROM notifications')
        columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in (cursor.fetchall() or []) if row}
        if 'message' not in columns:
            return
        
        # Check if notification already exists to prevent duplicates
        if application_id is not None and 'application_id' in columns:
            cursor.execute(
                '''
                SELECT notification_id FROM notifications
                WHERE application_id = %s AND message = %s
                LIMIT 1
                ''',
                (application_id, message)
            )
            existing_notification = cursor.fetchone()
            if existing_notification:
                # Notification already exists, skip creation
                return
        
        fields = []
        values = []
        params = []
        
        if application_id is not None and 'application_id' in columns:
            fields.append('application_id')
            values.append('%s')
            params.append(application_id)
        
        fields.append('message')
        values.append('%s')
        params.append(message)
        
        if 'sent_at' in columns:
            fields.append('sent_at')
            values.append('NOW()')
        if 'is_read' in columns:
            fields.append('is_read')
            values.append('0')
        
        sql = f"INSERT INTO notifications ({', '.join(fields)}) VALUES ({', '.join(values)})"
        cursor.execute(sql, tuple(params))
    except Exception as notify_err:
        print(f'⚠️ Notification insert error: {notify_err}')


def format_file_size(num_bytes):
    """Convert a byte value into a human-readable string."""
    if not isinstance(num_bytes, (int, float)) or num_bytes < 0:
        return 'Unknown'

    step = 1024.0
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    size = float(num_bytes)

    for unit in units:
        if size < step or unit == units[-1]:
            return f"{size:.1f} {unit}" if unit != 'B' else f"{int(size)} {unit}"
        size /= step


def parse_decimal_value(raw_value):
    """Convert an arbitrary numeric string into a Decimal or None."""
    if raw_value is None:
        return None

    cleaned = str(raw_value).strip().replace(',', '')
    if not cleaned:
        return None

    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def normalize_choice(value, valid_choices, default):
    """Return a sanitized enum choice using the provided defaults."""
    choice = (value or '').strip().lower()
    return choice if choice in valid_choices else default


def format_salary_range(currency_code, minimum, maximum):
    """Return a human-readable salary range string."""
    if minimum is None and maximum is None:
        return '—'

    def _format_amount(amount):
        if amount is None:
            return ''
        try:
            quantized = Decimal(amount)
        except (InvalidOperation, ValueError):
            return ''
        if quantized == quantized.to_integral():
            return f"{int(quantized):,}"
        return f"{quantized:,.2f}"

    currency = (currency_code or 'PHP').upper()
    min_str = _format_amount(minimum)
    max_str = _format_amount(maximum)

    if min_str and max_str:
        if min_str == max_str:
            return f"{currency} {min_str}"
        return f"{currency} {min_str} - {max_str}"
    if min_str:
        return f"{currency} {min_str}+"
    if max_str:
        return f"Up to {currency} {max_str}"
    return '—'


def get_application_status_label(value):
    """Return a user-friendly label for an application status."""
    status_key = (value or '').strip().lower()
    if not status_key:
        return 'Applied'
    return APPLICATION_STATUS_LABELS.get(status_key, status_key.replace('_', ' ').title())


def fetch_jobs_for_user(user):
    """Fetch job postings, automatically scoping to an HR user's branch."""
    params = []
    where_clause = ''

    branch_scope = get_branch_scope(user)
    if branch_scope is not None:
        where_clause = 'WHERE j.branch_id = %s'
        params.append(branch_scope)

    # Ensure schema compatibility and update job columns cache
    ensure_schema_compatibility()
    db = get_db()
    if db:
        cursor = db.cursor()
        try:
            _update_job_columns(cursor)
        finally:
            cursor.close()

    job_title_col = job_column('job_title', 'title')
    job_description_col = job_column('job_description', 'description')
    job_requirements_col = job_column('job_requirements', 'requirements')
    posted_at_col = job_column('posted_at', 'created_at') or 'created_at'
    created_at_col = job_column('created_at', 'posted_at') or 'created_at'

    job_title_expr = f'j.{job_title_col}' if job_title_col else "'Untitled Job'"
    job_description_expr = f'j.{job_description_col}' if job_description_col else 'NULL'
    job_requirements_expr = f'j.{job_requirements_col}' if job_requirements_col else 'NULL'
    
    # Build admin join for posted_by (actual schema)
    if 'posted_by' in JOB_COLUMNS:
        admin_join = 'LEFT JOIN admins a_posted ON j.posted_by = a_posted.admin_id'
    else:
        admin_join = ''

    query = f'''
        SELECT
            j.job_id,
            {job_title_expr} AS job_title,
            {job_description_expr} AS job_description,
            {job_requirements_expr} AS job_requirements,
            j.status,
            j.branch_id,
            j.created_at,
            j.posted_at,
            COALESCE(b.branch_name, 'Unassigned') AS branch_name,
            {job_title_expr} AS position_name,
            'General' AS department,
            COALESCE(a_posted.full_name, 'System') AS posted_by_name,
            (SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) AS application_count
        FROM jobs j
        LEFT JOIN branches b ON j.branch_id = b.branch_id
        {admin_join}
        {where_clause}
        ORDER BY COALESCE(j.posted_at, j.created_at) DESC
    '''

    return fetch_rows(query, tuple(params) if params else None)


def build_applicant_dashboard_data(applicant_id):
    """Collect metrics and listings tailored to a specific applicant."""
    dashboard = {
        'stats': {
            'total_applications': 0,
            'pending': 0,  # Maps from 'pending' status in database
            'interviewed': 0,  # Maps from 'interviewed' status in database
            'hired': 0,  # Maps from 'hired' status in database
            'rejected': 0,  # Maps from 'rejected' status in database
            # Legacy keys for backward compatibility
            'applied': 0,
            'under_review': 0,
            'interview': 0,
        },
        'applications': [],
        'upcoming_interviews': [],
        'notifications': [],
    }

    if not applicant_id:
        print(f'⚠️ build_applicant_dashboard_data: No applicant_id provided')
        return dashboard

    try:
        db = get_db()
        if not db:
            print(f'⚠️ build_applicant_dashboard_data: No database connection')
            return dashboard
        
        cursor = db.cursor(dictionary=True)
        job_title_expr = "'Untitled Job'"
        job_location_expr = "'Unassigned'"
        
        try:
            # Ensure schema compatibility and get dynamic job column expressions
            ensure_schema_compatibility()
            _update_job_columns(cursor)
            job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
            job_location_expr = job_column_expr('job_location', alternatives=['location'], default="'Unassigned'")

            try:
                dashboard['stats']['total_applications'] = fetch_count(
                    'SELECT COUNT(*) AS count FROM applications WHERE applicant_id = %s',
                    (applicant_id,),
                )
            except Exception as count_error:
                print(f'⚠️ Error fetching total_applications: {count_error}')
                dashboard['stats']['total_applications'] = 0

            try:
                status_rows = fetch_rows(
                    """
                    SELECT status, COUNT(*) AS count
                    FROM applications
                    WHERE applicant_id = %s
                    GROUP BY status
                    """,
                    (applicant_id,),
                )
                for row in (status_rows or []):
                    if row and row.get('status'):
                        status = row['status']
                        count = row.get('count', 0)
                        # Map database statuses to stats keys
                        if status == 'pending':
                            dashboard['stats']['pending'] = count
                            dashboard['stats']['under_review'] = count  # Legacy compatibility
                            dashboard['stats']['applied'] = count  # Legacy compatibility
                        elif status == 'interviewed':
                            dashboard['stats']['interviewed'] = count
                            dashboard['stats']['interview'] = count  # Legacy compatibility
                        elif status == 'hired':
                            dashboard['stats']['hired'] = count
                        elif status == 'rejected':
                            dashboard['stats']['rejected'] = count
                        # Also set directly if key exists (for backward compatibility)
                        if status in dashboard['stats']:
                            dashboard['stats'][status] = count
            except Exception as status_error:
                print(f'⚠️ Error fetching status rows: {status_error}')

            try:
                applications_rows = fetch_rows(
                    f"""
                    SELECT a.application_id,
                           {job_title_expr} AS job_title,
                           COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                           a.status,
                           a.submitted_at,
                           COALESCE({job_location_expr}, b.branch_name) AS job_location,
                           CASE 
                               WHEN EXISTS (
                                   SELECT 1 
                                   FROM interviews i 
                                   WHERE i.application_id = a.application_id 
                                   LIMIT 1
                               ) THEN 1 
                               ELSE 0 
                           END AS has_interview
                    FROM applications a
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    WHERE a.applicant_id = %s
                    ORDER BY a.submitted_at DESC
                    LIMIT 10
                    """,
                    (applicant_id,),
                )
                dashboard['applications'] = []
                for row in (applications_rows or []):
                    try:
                        submitted_at = row.get('submitted_at')
                        dashboard['applications'].append({
                            'application_id': row.get('application_id'),
                            'job_title': row.get('job_title') or 'Untitled Job',
                            'branch_name': row.get('branch_name') or 'Unassigned',
                            'company_name': row.get('branch_name') or 'Unassigned',
                            'job_location': row.get('job_location'),
                            'location': row.get('job_location'),
                            'status': row.get('status'),
                            'submitted_at': format_human_datetime(submitted_at) if submitted_at else '',
                            'applied_date': format_human_datetime(submitted_at) if submitted_at else '',
                            'updated_at': format_human_datetime(submitted_at) if submitted_at else '',
                            'has_interview': bool(row.get('has_interview')),
                        })
                    except Exception as row_error:
                        print(f'⚠️ Error processing application row: {row_error}')
                        continue
            except Exception as apps_error:
                print(f'⚠️ Error fetching applications: {apps_error}')
                dashboard['applications'] = []

            try:
                interviews_rows = fetch_rows(
                    f"""
                    SELECT i.scheduled_date,
                           COALESCE(i.interview_mode, 'in-person') AS interview_mode,
                           i.location,
                           {job_title_expr} AS job_title
                    FROM interviews i
                    JOIN applications a ON i.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.applicant_id = %s AND i.scheduled_date >= NOW()
                    ORDER BY i.scheduled_date ASC
                    LIMIT 5
                    """,
                    (applicant_id,),
                )
                dashboard['upcoming_interviews'] = []
                for row in (interviews_rows or []):
                    try:
                        scheduled_date = row.get('scheduled_date')
                        dashboard['upcoming_interviews'].append({
                            'job_title': row.get('job_title') or 'Interview',
                            'date_time': format_human_datetime(scheduled_date) if scheduled_date else '',
                            'type': (row.get('interview_mode') or 'in-person').replace('-', ' ').title(),
                            'location': row.get('location') or 'To be confirmed',
                        })
                    except Exception as row_error:
                        print(f'⚠️ Error processing interview row: {row_error}')
                        continue
            except Exception as interviews_error:
                print(f'⚠️ Error fetching interviews: {interviews_error}')
                dashboard['upcoming_interviews'] = []

            # Check if sent_at column exists in notifications table
            try:
                cursor.execute('SHOW COLUMNS FROM notifications')
                notification_columns_raw = cursor.fetchall()
                notification_columns = [col.get('Field') if isinstance(col, dict) else col[0] for col in (notification_columns_raw or [])]
                sent_at_expr = 'COALESCE(n.sent_at, n.created_at, NOW())' if 'sent_at' in notification_columns else 'COALESCE(n.created_at, NOW())'
            except Exception:
                sent_at_expr = 'COALESCE(n.created_at, NOW())'
            
            try:
                notification_rows = fetch_rows(
                    f"""
                    SELECT n.message, {sent_at_expr} AS sent_at
                    FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    WHERE a.applicant_id = %s
                    ORDER BY {sent_at_expr} DESC
                    LIMIT 5
                    """,
                    (applicant_id,),
                )
                dashboard['notifications'] = []
                for row in (notification_rows or []):
                    try:
                        sent_at = row.get('sent_at')
                        dashboard['notifications'].append({
                            'message': row.get('message') or '',
                            'time': format_human_datetime(sent_at) if sent_at else '',
                        })
                    except Exception as row_error:
                        print(f'⚠️ Error processing notification row: {row_error}')
                        continue
            except Exception as notif_error:
                print(f'⚠️ Error fetching notifications: {notif_error}')
                dashboard['notifications'] = []
        finally:
            cursor.close()

    except Exception as exc:
        print(f'❌ Failed to build applicant dashboard data: {exc}')
        import traceback
        traceback.print_exc()
        # Return empty dashboard on error to prevent 500 error

    return dashboard


def fetch_open_jobs(filters=None, applicant_id=None):
    """Retrieve open job postings with optional filters and smart matching.
    
    IMPORTANT: This function shows ALL jobs with status 'open' from ALL branches by default.
    Branch filtering only applies when branch_id is explicitly provided in filters.
    Applicants can see jobs from all branches regardless of who posted them (admin or HR).
    """
    # Ensure schema compatibility before querying
    ensure_schema_compatibility()
    db = get_db()
    has_position_name = False
    if db:
        cursor = db.cursor()
        try:
            _update_job_columns(cursor)
            # Check if position_name column exists
            cursor.execute('SHOW COLUMNS FROM jobs LIKE "position_name"')
            has_position_name = cursor.fetchone() is not None
        finally:
            cursor.close()
    
    # Use 'status' column directly - we know the schema uses 'status'
    # Filter by publishable statuses - only jobs with status 'open' are visible to applicants
    # Database enum supports: 'open', 'closed' - only 'open' is visible
    # CRITICAL: NO branch filtering is applied here - shows ALL branches by default
    # Only filter by branch if branch_id is explicitly provided in filters
    status_placeholders = ','.join(['%s'] * len(PUBLISHABLE_JOB_STATUSES))
    where_clauses = ['j.status IN ({})'.format(status_placeholders)]
    params = list(PUBLISHABLE_JOB_STATUSES)

    # Apply filters only if provided and not empty
    if filters and isinstance(filters, dict) and len(filters) > 0:
        if filters.get('keyword'):
            keyword = f"%{filters['keyword']}%"
            keyword_fields = [
                col for col in [
                    job_column('job_title', 'title'),
                    job_column('job_description', 'description'),
                    job_column('job_requirements', 'requirements'),
                ]
                if col
            ]
            # Build keyword search clauses for job fields
            keyword_clauses = []
            if keyword_fields:
                job_like_clauses = [f"j.{column} LIKE %s" for column in keyword_fields]
                keyword_clauses.extend(job_like_clauses)
                params.extend([keyword] * len(keyword_fields))
            
            # Add branch name to keyword search
            keyword_clauses.append("b.branch_name LIKE %s")
            params.append(keyword)
            
            if keyword_clauses:
                where_clauses.append(f"({' OR '.join(keyword_clauses)})")

        # Only filter by branch if explicitly provided - otherwise show all branches
        if filters.get('branch_id'):
            where_clauses.append('j.branch_id = %s')
            params.append(filters['branch_id'])

        # Position filter removed - positions table no longer exists
        # if filters.get('position_id'):
        #     where_clauses.append('j.position_id = %s')
        #     params.append(filters['position_id'])
    
    # Check if saved_only filter is set - need to handle this before building where_sql
    saved_only_filter = filters and filters.get('saved_only') == '1' if filters else False
    
    # Add saved_jobs join if applicant is logged in AND table exists
    saved_join = ''
    is_saved_select = '0 AS is_saved'  # Default to 0 if not logged in
    saved_jobs_table_exists = False
    if applicant_id:
        # Check if saved_jobs table exists before trying to join
        db_check = get_db()
        if db_check:
            cursor_check = db_check.cursor()
            try:
                cursor_check.execute("SHOW TABLES LIKE 'saved_jobs'")
                if cursor_check.fetchone():
                    # Table exists, use the join
                    saved_jobs_table_exists = True
                    # Filter by saved jobs only if saved_only filter is set
                    if saved_only_filter:
                        # Use INNER JOIN to only show saved jobs
                        saved_join = f'INNER JOIN saved_jobs sj ON sj.job_id = j.job_id AND sj.applicant_id = {applicant_id}'
                        is_saved_select = '1 AS is_saved'
                        where_clauses.append('sj.job_id IS NOT NULL')  # ✅ FIX: Use job_id instead of saved_job_id
                    else:
                        # Use LEFT JOIN to show all jobs with saved status
                        saved_join = f'LEFT JOIN saved_jobs sj ON sj.job_id = j.job_id AND sj.applicant_id = {applicant_id}'
                        is_saved_select = 'CASE WHEN sj.job_id IS NOT NULL THEN 1 ELSE 0 END AS is_saved'  # ✅ FIX: Use job_id instead of saved_job_id
                # If table doesn't exist, just use default 0 AS is_saved
            except Exception:
                # If check fails, just use default
                pass
            finally:
                cursor_check.close()

    where_sql = ' AND '.join(where_clauses)

    # Use direct column names since we know the schema: title, description, requirements, status, branch_id, posted_at
    # IMPORTANT: Show ALL jobs with status 'open' from ALL branches - no restrictions
    # This ensures applicants can see ALL jobs posted by HR or Admin with status 'open'
    # Build position_name expression conditionally based on whether column exists
    if has_position_name:
        position_name_expr = 'COALESCE(j.position_name, j.title)'
    else:
        position_name_expr = 'j.title'
    
    query = f"""
        SELECT
            j.job_id,
            j.title AS job_title,
            j.description AS job_description,
            j.requirements AS job_requirements,
            j.branch_id,
            j.status,
            COALESCE(j.posted_at, j.created_at) AS posted_at,
            COALESCE(b.branch_name, 'Unassigned') AS branch_name,
            {position_name_expr} AS position_title,
            'General' AS department,
            (SELECT COUNT(*) FROM applications a WHERE a.job_id = j.job_id) AS application_count,
            {is_saved_select}
        FROM jobs j
        LEFT JOIN branches b ON j.branch_id = b.branch_id
        {saved_join}
        WHERE {where_sql}
        ORDER BY COALESCE(j.posted_at, j.created_at) DESC
    """
    
    try:
        rows = fetch_rows(query, tuple(params))
    except Exception as query_error:
        print(f'❌ fetch_open_jobs query error: {query_error}')
        import traceback
        traceback.print_exc()
        rows = []
    

    jobs = []
    for row in rows:
        salary_display = format_salary_range(
            row.get('salary_currency'),
            row.get('salary_min'),
            row.get('salary_max'),
        )

        jobs.append(
            {
                'job_id': row.get('job_id'),
                'title': row.get('job_title'),
                'summary': (row.get('job_description') or '')[:200] if row.get('job_description') else '',
                'description': row.get('job_description'),
                'requirements': row.get('job_requirements'),
                'employment_type': row.get('employment_type'),
                'work_arrangement': row.get('work_arrangement'),
                'experience_level': row.get('experience_level'),
                'location': row.get('job_location'),
                'salary_currency': row.get('salary_currency'),
                'salary_min': row.get('salary_min'),
                'salary_max': row.get('salary_max'),
                'salary_display': salary_display,
                'branch_id': row.get('branch_id'),
                'branch_name': row.get('branch_name'),
                'position_id': row.get('position_id'),
                'position': row.get('position_title'),
                'position_title': row.get('position_title'),
                'position_name': row.get('position_title'),
                'department': row.get('department'),
                'application_deadline': row.get('application_deadline'),
                'status': row.get('status'),
                'posted_at': format_human_datetime(row.get('posted_at') or row.get('created_at')),
                'application_count': row.get('application_count', 0),
                'is_saved': bool(row.get('is_saved', 0)),
            }
        )

    # Smart matching: if applicant_id provided, calculate match scores
    if applicant_id:
        jobs = add_smart_matching(jobs, applicant_id)

    return jobs


def add_smart_matching(jobs, applicant_id):
    """Add smart matching scores based on applicant profile and application history."""
    db = get_db()
    if not db:
        return jobs
    
    cursor = db.cursor(dictionary=True)
    try:
        # Get applicant's application history
        cursor.execute(
            '''
            SELECT job_id, status
            FROM applications
            WHERE applicant_id = %s
            ''',
            (applicant_id,),
        )
        app_history = {row['job_id']: row['status'] for row in cursor.fetchall()}
        
        # Get applicant's resume keywords (simplified - could be enhanced with NLP)
        # Calculate match scores
        for job in jobs:
            score = 0
            
            if job['job_id'] not in app_history:
                score += 10
            else:
                if app_history[job['job_id']] in {'rejected', 'hired'}:
                    score -= 5
            
            job['match_score'] = max(0, min(100, score))
            job['already_applied'] = job['job_id'] in app_history
            job['application_status'] = app_history.get(job['job_id'])
    except Exception as exc:
        print(f'⚠️ Smart matching error: {exc}')
    finally:
        cursor.close()
    
    # Sort by match score (highest first)
    jobs.sort(key=lambda x: x.get('match_score', 0), reverse=True)
    return jobs




def fetch_applicants_summary(user=None):
    """Return applicants with aggregated application info, scoped for HR users."""
    branch_id = get_branch_scope(user)

    if branch_id:
        query = """
            SELECT ap.applicant_id,
                   ap.full_name,
                   ap.email,
                   ap.phone_number,
                   ap.created_at,
                   COUNT(a.application_id) AS total_applications,
                   SUM(CASE WHEN a.status = 'applied' THEN 1 ELSE 0 END) AS applied,
                   SUM(CASE WHEN a.status = 'under_review' THEN 1 ELSE 0 END) AS under_review,
                   SUM(CASE WHEN a.status = 'interview' THEN 1 ELSE 0 END) AS interview,
                   SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS hired,
                   SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected
            FROM applicants ap
            JOIN applications a ON a.applicant_id = ap.applicant_id
            JOIN jobs j ON a.job_id = j.job_id
            WHERE j.branch_id = %s
            GROUP BY ap.applicant_id, ap.full_name, ap.email, ap.phone_number, ap.created_at
            ORDER BY ap.created_at DESC
        """
        params = (branch_id,)
    else:
        query = """
            SELECT ap.applicant_id,
                   ap.full_name,
                   ap.email,
                   ap.phone_number,
                   ap.created_at,
                   COUNT(a.application_id) AS total_applications,
                   SUM(CASE WHEN a.status = 'applied' THEN 1 ELSE 0 END) AS applied,
                   SUM(CASE WHEN a.status = 'under_review' THEN 1 ELSE 0 END) AS under_review,
                   SUM(CASE WHEN a.status = 'interview' THEN 1 ELSE 0 END) AS interview,
                   SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS hired,
                   SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected
            FROM applicants ap
            LEFT JOIN applications a ON a.applicant_id = ap.applicant_id
            GROUP BY ap.applicant_id, ap.full_name, ap.email, ap.phone_number, ap.created_at
            ORDER BY ap.created_at DESC
        """
        params = None

    rows = fetch_rows(query, params)

    for r in rows:
        r['created_at'] = format_human_datetime(r.get('created_at'))
    return rows

@app.route('/')
def index():
    """Landing page."""
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle login for applicants, HR managers, and administrators."""
    # Apply rate limiting only to POST requests (actual login attempts), not GET requests (viewing page)
    if request.method == 'POST':
        from utils.rate_limit import _rate_limit_store, _rate_limit_lock
        from datetime import datetime
        
        identifier = request.remote_addr or 'unknown'
        key = f"login:{identifier}"
        now = datetime.now()
        max_requests = 10  # Allow 10 login attempts
        window_seconds = 900  # Within 15 minutes (was 5 attempts per 5 minutes)
        
        with _rate_limit_lock:
            # Clean old entries (older than window_seconds)
            if key in _rate_limit_store:
                _rate_limit_store[key] = [
                    timestamp for timestamp in _rate_limit_store[key]
                    if (now - timestamp).total_seconds() < window_seconds
                ]
            else:
                _rate_limit_store[key] = []
            
            # Check if limit exceeded
            if len(_rate_limit_store[key]) >= max_requests:
                flash('Too many login attempts. Please wait 15 minutes before trying again.', 'error')
                return immediate_redirect(url_for('login', _external=True))
    
    db = None
    cursor = None
    try:
        if request.method == 'POST':
            # Track this login attempt (only for failed attempts)
            from utils.rate_limit import _rate_limit_store, _rate_limit_lock
            from datetime import datetime
            
            email = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '').strip()
            user_type = request.form.get('user_type', 'applicant').strip().lower()
                
            if not email or not password:
                flash('Please fill in all required fields.', 'error')
                _track_failed_login()
                return render_template('login.html')
                
            db = get_db()
            if not db:
                flash('Database connection error. Please try again later.', 'error')
                return render_template('login.html')

            cursor = db.cursor(dictionary=True)

            try:
                type_map = {
                    'admin': 'super_admin',
                    'hr': 'hr',
                    'applicant': 'applicant'
                }
                target_type = type_map.get(user_type, 'applicant')

                if target_type in {'super_admin', 'hr'}:
                    cursor.execute(
                    """
                    SELECT
                        u.user_id,
                        u.password_hash,
                        u.user_type,
                        u.is_active,
                        a.admin_id,
                        a.full_name
                    FROM users u
                    LEFT JOIN admins a ON a.user_id = u.user_id
                    WHERE u.email = %s AND u.user_type = %s
                    LIMIT 1
                    """,
                    (email, target_type),
                )
                    account = cursor.fetchone()

                    if not account:
                        # Check if email exists with different user_type to provide helpful error
                        cursor.execute("""
                            SELECT user_type FROM users WHERE email = %s LIMIT 1
                        """, (email,))
                        existing_user = cursor.fetchone()
                        if existing_user:
                            actual_type = existing_user.get('user_type', '')
                            if actual_type == 'super_admin' and user_type != 'admin':
                                flash('Please select "Admin" as user type to login with this email.', 'error')
                            elif actual_type == 'hr' and user_type != 'hr':
                                flash('Please select "HR" as user type to login with this email.', 'error')
                            else:
                                flash('Invalid email or password.', 'error')
                        else:
                            flash('Invalid email or password.', 'error')
                        print(f'❌ Login failed: No account found for email={email}, user_type={target_type}')
                        _track_failed_login()
                        return render_template('login.html')

                    # Validate user_id - it should never be 0 or None
                    user_id = account.get('user_id')
                    if not user_id or user_id == 0 or not isinstance(user_id, int):
                        # Re-query to get the correct user_id directly from users table
                        print(f'⚠️ Invalid user_id={user_id} from query, re-querying users table...')
                        cursor.execute("""
                            SELECT user_id, email, password_hash, is_active, user_type
                            FROM users 
                            WHERE email = %s AND user_type = %s 
                            ORDER BY user_id DESC
                            LIMIT 1
                        """, (email, target_type))
                        user_record = cursor.fetchone()
                        if user_record:
                            found_user_id = user_record.get('user_id')
                            # Check if user_id is 0 - this is a data integrity issue
                            if found_user_id == 0:
                                print(f'⚠️ Found user_id=0 in database for email={email} - attempting to fix...')
                                # Try to find the correct user_id from admins table
                                cursor.execute("""
                                    SELECT a.user_id, a.admin_id
                                    FROM admins a
                                    WHERE a.email = %s
                                    LIMIT 1
                                """, (email,))
                                admin_record = cursor.fetchone()
                                if admin_record and admin_record.get('user_id') and admin_record.get('user_id') > 0:
                                    # Found valid user_id from admins table
                                    correct_user_id = admin_record.get('user_id')
                                    print(f'✅ Found valid user_id={correct_user_id} from admins table for email={email}')
                                    # Use the correct user_id from admins table for login
                                    # Note: We can't UPDATE user_id if it's a primary key, so we'll just use the correct one
                                    user_id = correct_user_id
                                    account['user_id'] = user_id
                                    # Also update password_hash and other fields from the user_record
                                    account['password_hash'] = user_record.get('password_hash') or account.get('password_hash')
                                    account['is_active'] = user_record.get('is_active', 1)
                                    account['user_type'] = user_record.get('user_type', target_type)
                                    print(f'✅ Using user_id={user_id} from admins table (users table has user_id=0 which is invalid)')
                                else:
                                    # No valid user_id in admins table either - try to find any valid user_id for this email
                                    cursor.execute("""
                                        SELECT user_id FROM users 
                                        WHERE email = %s AND user_id > 0 
                                        ORDER BY user_id DESC 
                                        LIMIT 1
                                    """, (email,))
                                    alt_user_record = cursor.fetchone()
                                    if alt_user_record and alt_user_record.get('user_id') and alt_user_record.get('user_id') > 0:
                                        user_id = alt_user_record.get('user_id')
                                        account['user_id'] = user_id
                                        print(f'✅ Found alternative user_id={user_id} for email={email}')
                                    else:
                                        print(f'❌ Cannot find valid user_id for email={email}')
                                        print(f'   User record: {user_record}')
                                        flash('Account configuration error. Invalid user ID in database. Please contact support.', 'error')
                                        _track_failed_login()
                                        return render_template('login.html')
                            elif found_user_id and isinstance(found_user_id, int) and found_user_id > 0:
                                user_id = found_user_id
                                # Update account dict with correct values from users table
                                account['user_id'] = user_id
                                account['password_hash'] = user_record.get('password_hash') or account.get('password_hash')
                                account['is_active'] = user_record.get('is_active', 1)
                                account['user_type'] = user_record.get('user_type', target_type)
                                print(f'✅ Corrected user_id={user_id} for email={email} from users table')
                            else:
                                print(f'❌ Invalid user_id={found_user_id} found in database for email={email}')
                                print(f'   User record: {user_record}')
                                flash('Account configuration error. Invalid user ID in database. Please contact support.', 'error')
                                _track_failed_login()
                                return render_template('login.html')
                        else:
                            # User doesn't exist in users table - this is a critical error
                            print(f'❌ Cannot find user record in users table for email={email}, user_type={target_type}')
                            print(f'   This means the user account does not exist in the users table.')
                            flash('Account not found. Please contact support to create your account.', 'error')
                            _track_failed_login()
                            return render_template('login.html')
                    else:
                        # user_id is valid, but let's verify it exists in the database
                        cursor.execute("""
                            SELECT user_id FROM users WHERE user_id = %s LIMIT 1
                        """, (user_id,))
                        verify_record = cursor.fetchone()
                        if not verify_record:
                            print(f'⚠️ Warning: user_id={user_id} from query does not exist in users table, re-querying...')
                            # Re-query by email
                            cursor.execute("""
                                SELECT user_id FROM users WHERE email = %s AND user_type = %s AND user_id > 0
                                ORDER BY user_id DESC
                                LIMIT 1
                            """, (email, target_type))
                            verify_record = cursor.fetchone()
                            if verify_record:
                                correct_user_id = verify_record.get('user_id')
                                if correct_user_id and correct_user_id > 0:
                                    user_id = correct_user_id
                                    account['user_id'] = user_id
                                    print(f'✅ Corrected user_id={user_id} for email={email}')
                                else:
                                    print(f'❌ Invalid user_id={correct_user_id} from verification query')
                                    flash('Account configuration error. Please contact support.', 'error')
                                    _track_failed_login()
                                    return render_template('login.html')
                            else:
                                print(f'❌ Cannot verify user_id={user_id} - user not found in database')
                                flash('Account configuration error. Please contact support.', 'error')
                                _track_failed_login()
                                return render_template('login.html')

                    if not account.get('admin_id'):
                        # Auto-fix: Create missing admin record for this user
                        print(f'⚠️ Auto-fixing: Creating missing admin record for email={email}, user_id={user_id}')
                        
                        try:
                            # Final validation: user_id must be a positive integer
                            if not user_id or not isinstance(user_id, int) or user_id <= 0:
                                raise Exception(f"Invalid user_id={user_id} (must be positive integer) before creating admin record")
                            
                            # Verify user exists - but use the user_id we already validated
                            # Don't re-query by email as it might return user_id=0 again
                            cursor.execute("""
                                SELECT user_id, email, password_hash, is_active, user_type 
                                FROM users 
                                WHERE user_id = %s
                                LIMIT 1
                            """, (user_id,))
                            user_info = cursor.fetchone()
                            if not user_info:
                                # User with this user_id doesn't exist in users table
                                # This is okay if we got user_id from admins table - we'll proceed anyway
                                print(f'⚠️ Warning: user_id={user_id} not found in users table, but proceeding with admin record creation')
                                # Get user info from the account dict or use defaults
                                user_email = email
                                password_hash = account.get('password_hash')
                                is_active = account.get('is_active', 1)
                            else:
                                # User exists, use the data from users table
                                user_email = user_info.get('email', email)
                                password_hash = user_info.get('password_hash') or account.get('password_hash')
                                is_active = user_info.get('is_active', 1)
                            
                            # Final validation before proceeding - user_id should already be validated
                            if not user_id or not isinstance(user_id, int) or user_id <= 0:
                                raise Exception(f"Invalid user_id={user_id} before creating admin record")
                            
                            print(f'✅ Proceeding with admin record creation for user_id={user_id}, email={email}')
                            
                            # Check which columns exist in admins table
                            cursor.execute('SHOW COLUMNS FROM admins')
                            admin_columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in cursor.fetchall()}
                            
                            # Build INSERT statement based on available columns
                            fields = ['user_id', 'full_name', 'email']
                            values = ['%s', '%s', '%s']
                            params = [user_id, account.get('full_name') or 'Admin User', user_email]
                            
                            if 'password_hash' in admin_columns and password_hash:
                                fields.append('password_hash')
                                values.append('%s')
                                params.append(password_hash)
                            
                            # branch_id column has been removed from admins table
                            # HR accounts manage all branches
                            
                            # Create admin record
                            sql = f"INSERT INTO admins ({', '.join(fields)}) VALUES ({', '.join(values)})"
                            cursor.execute(sql, tuple(params))
                            admin_id = cursor.lastrowid
                            
                            if not admin_id or admin_id == 0:
                                raise Exception(f"Failed to create admin record - admin_id={admin_id}")
                            
                            db.commit()
                            
                            # Update account dict with new admin_id and ensure user_id is correct
                            account['admin_id'] = admin_id
                            account['user_id'] = user_id  # Ensure user_id is set correctly
                            print(f'✅ Created admin record: admin_id={admin_id} for user_id={user_id}, email={user_email}')
                        except Exception as fix_error:
                            import traceback
                            print(f'❌ Failed to auto-fix admin record: {fix_error}')
                            print(f'Traceback: {traceback.format_exc()}')
                            db.rollback()
                            flash('Account configuration error. Admin/HR account not properly set up. Please contact support.', 'error')
                            _track_failed_login()
                            return render_template('login.html')

                    if not account.get('password_hash'):
                        flash('Account error. Please contact support.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not account.get('is_active'):
                        flash('This account has been deactivated. Contact your administrator.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not check_password(account['password_hash'], password):
                        print(f'❌ Login failed: Password check failed for email={email}, user_type={target_type}')
                        flash('Invalid email or password.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    # Final validation: Ensure user_id and admin_id are valid before proceeding
                    final_user_id = account.get('user_id')
                    final_admin_id = account.get('admin_id')
                    
                    if not final_user_id or final_user_id == 0:
                        print(f'❌ Login failed: Invalid user_id={final_user_id} for email={email}')
                        flash('Account error. User ID is invalid. Please contact support.', 'error')
                        return render_template('login.html')
                    
                    # If admin_id is still missing after auto-fix attempt, try to fetch it
                    if not final_admin_id or final_admin_id == 0:
                        print(f'⚠️ Admin record missing for email={email}, attempting to fetch or create...')
                        try:
                            # Try to get admin_id from admins table
                            cursor.execute("""
                                SELECT admin_id FROM admins WHERE user_id = %s LIMIT 1
                            """, (final_user_id,))
                            admin_record = cursor.fetchone()
                            if admin_record and admin_record.get('admin_id'):
                                final_admin_id = admin_record.get('admin_id')
                                account['admin_id'] = final_admin_id
                                print(f'✅ Found existing admin record: admin_id={final_admin_id}')
                            else:
                                # Admin record doesn't exist, this should have been created earlier
                                # But if it wasn't, we can't proceed without it
                                print(f'❌ Login failed: No admin record found for user_id={final_user_id}, email={email}')
                                flash('Account configuration error. Admin record is missing. Please contact support.', 'error')
                                return render_template('login.html')
                        except Exception as fetch_error:
                            print(f'❌ Error fetching admin record: {fetch_error}')
                            flash('Account configuration error. Please contact support.', 'error')
                            return render_template('login.html')

                    # Ensure final_admin_id is set after fetch attempt
                    if not final_admin_id or final_admin_id == 0:
                        final_admin_id = account.get('admin_id')
                    
                    if not final_admin_id or final_admin_id == 0:
                        print(f'❌ Login failed: Could not resolve admin_id for email={email}, user_id={final_user_id}')
                        flash('Account configuration error. Admin record is missing. Please contact support.', 'error')
                        return render_template('login.html')
                    
                    role = 'admin' if account['user_type'] == 'super_admin' else 'hr'
                    try:
                        print(f'✅ Attempting login for email={email}, role={role}, admin_id={final_admin_id}, user_id={final_user_id}')
                        login_user(
                            final_admin_id,
                            role,
                            email,
                            account.get('full_name', ''),
                            auth_user_id=final_user_id
                        )
                        print(f'✅ Login successful for email={email}, role={role}')

                        # HR accounts no longer have branch_id - they manage all branches
                        # Remove any existing branch_id from session
                        session.pop('branch_id', None)

                        flash('Welcome back!', 'success')
                        # Use immediate redirect to prevent redirect page
                        dashboard_name = 'hr_dashboard' if role == 'hr' else 'admin_dashboard'
                        return immediate_redirect(url_for(dashboard_name, _external=True))
                    except Exception as login_exc:
                        import traceback
                        print(f'❌ login_user error: {login_exc}')
                        print(f'Traceback: {traceback.format_exc()}')
                        flash('Error during login. Please try again.', 'error')
                        return render_template('login.html')

                elif target_type == 'applicant':
                    # Applicant authentication
                    cursor.execute(
                        """
                        SELECT
                            u.user_id,
                            u.password_hash,
                            u.user_type,
                            u.is_active,
                            u.email_verified,
                            ap.applicant_id,
                            ap.full_name,
                            ap.phone_number,
                            ap.verification_token
                        FROM users u
                        JOIN applicants ap ON ap.user_id = u.user_id
                        WHERE u.email = %s AND u.user_type = 'applicant'
                        LIMIT 1
                        """,
                        (email,),
                    )
                    applicant = cursor.fetchone()

                    if not applicant:
                        flash('Invalid email or password.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not applicant.get('password_hash'):
                        flash('Account error. Please contact support.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not check_password(applicant['password_hash'], password):
                        flash('Invalid email or password.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not applicant.get('is_active'):
                        flash('This account has been deactivated. Please contact support.', 'error')
                        _track_failed_login()
                        return render_template('login.html')
                    
                    # Check email verification
                    email_verified = applicant.get('email_verified', 0)
                    if not email_verified:
                        session['pending_verification_email'] = email
                        flash('Please verify your email address before logging in. Check your inbox for the verification link.', 'error')
                        return immediate_redirect(url_for('resend_verification', _external=True))

                    if not applicant.get('applicant_id'):
                        flash('Account error. Please contact support.', 'error')
                        return render_template('login.html')

                    if not applicant.get('user_id'):
                        flash('Account error. Please contact support.', 'error')
                        return render_template('login.html')

                    try:
                        login_user(
                            applicant['applicant_id'],
                            'applicant',
                            email,
                            applicant.get('full_name', ''),
                            auth_user_id=applicant['user_id']
                        )
                        session.pop('branch_id', None)

                        flash('Welcome back!', 'success')
                        # Use immediate redirect to prevent redirect page
                        return immediate_redirect(url_for('applicant_dashboard', _external=True))
                    except Exception as login_exc:
                        import traceback
                        print(f'❌ login_user error: {login_exc}')
                        print(f'Traceback: {traceback.format_exc()}')
                        flash('Error during login. Please try again.', 'error')
                        return render_template('login.html')

            except Exception as exc:
                import traceback
                error_details = traceback.format_exc()
                print(f'❌ Login error: {exc}')
                print(f'Full traceback: {error_details}')
                if db:
                    try:
                        db.rollback()
                    except Exception:
                        pass
                flash('An unexpected error occurred. Please try again.', 'error')
                return render_template('login.html')
            finally:
                if cursor:
                    try:
                        cursor.close()
                    except Exception:
                        pass
        else:
            # Handle GET request
            # Only redirect to dashboard if user is actually logged in (not just session data)
            if is_logged_in():
                # Double-check that session is valid by checking if user exists
                current_user = get_current_user()
                if current_user:
                    role = session.get('user_role')
                    # Redirect to appropriate dashboard based on role using immediate redirect
                    if role == 'applicant':
                        return immediate_redirect(url_for('applicant_dashboard', _external=True))
                    if role == 'hr':
                        return immediate_redirect(url_for('hr_dashboard', _external=True))
                    if role == 'admin':
                        return immediate_redirect(url_for('admin_dashboard', _external=True))
                else:
                    # Session data exists but user is invalid - clear it
                    session.clear()

            return render_template('login.html')
    except Exception as outer_exc:
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Login outer error: {outer_exc}')
        print(f'Full traceback: {error_details}')
        if db:
            try:
                db.rollback()
            except Exception:
                pass
        flash('An unexpected error occurred. Please try again.', 'error')
        return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Allow applicants to create accounts."""
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip().lower()
        phone_number = request.form.get('phone_number', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
            
        if not all([full_name, email, phone_number, password, confirm_password]):
            flash('Please complete all required fields.', 'error')
            return render_template('register.html')
            
        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('register.html')
            
        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('register.html')
            
        db = get_db()
        if not db:
            flash('Database connection error. Please try again later.', 'error')
            return render_template('register.html')
            
        cursor = db.cursor(dictionary=True)

        try:
            cursor.execute(
                'SELECT user_id FROM users WHERE email = %s LIMIT 1',
                (email,),
            )
            if cursor.fetchone():
                flash('Email address is already registered.', 'error')
                return render_template('register.html')

            password_hash = hash_password(password)
            verification_token = generate_token()
            
            # Set token expiration to 60 seconds from now
            from datetime import datetime, timedelta
            token_expires = datetime.now() + timedelta(seconds=60)
            
            # Insert user with email_verified = 0 (unverified)
            cursor.execute(
                """
                INSERT INTO users (email, password_hash, user_type, is_active, email_verified)
                VALUES (%s, %s, 'applicant', 1, 0)
                """,
                (email, password_hash),
            )
            user_id = cursor.lastrowid

            # Store password_hash in applicants table with verification token and expiration
            cursor.execute(
                """
                INSERT INTO applicants (user_id, full_name, email, phone_number, password_hash, verification_token, verification_token_expires)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (user_id, full_name, email, phone_number or None, password_hash, verification_token, token_expires),
            )
            applicant_id = cursor.lastrowid
            db.commit()
            
            # Send verification email
            try:
                send_verification_email(email, verification_token, applicant_name=full_name)
                flash('Registration successful! Please check your email to verify your account before logging in.', 'success')
            except Exception as email_error:
                print(f'⚠️ Failed to send verification email: {email_error}')
                flash('Registration successful! However, we could not send the verification email. Please contact support.', 'warning')
        except Exception as exc:
            db.rollback()
            print(f'❌ Registration error: {exc}')
            flash('An unexpected error occurred during registration.', 'error')
            return render_template('register.html')
        finally:
            cursor.close()

        return immediate_redirect(url_for('login', _external=True))
        
    return render_template('register.html')
    

@app.route('/verify-email/<token>')
def verify_email(token):
    """Verify email using token - checks for 60-second expiration."""
    db = get_db()
    if not db:
        flash('Database connection error. Please try again later.', 'error')
        return immediate_redirect(url_for('login', _external=True))

    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            '''
            SELECT a.applicant_id, a.email, a.user_id, a.verification_token_expires, u.email_verified
            FROM applicants a
            JOIN users u ON u.user_id = a.user_id
            WHERE a.verification_token = %s
            LIMIT 1
            ''',
            (token,),
        )
        applicant = cursor.fetchone()

        if not applicant:
            flash('The verification link is invalid or has expired.', 'error')
            return immediate_redirect(url_for('login', _external=True))
        
        # Check if already verified
        if applicant.get('email_verified'):
            flash('Your email is already verified. You can log in.', 'info')
            return immediate_redirect(url_for('login', _external=True))
        
        # Check token expiration (60 seconds)
        expires_at = applicant.get('verification_token_expires')
        if expires_at:
            from datetime import datetime
            if isinstance(expires_at, str):
                expires_at = datetime.strptime(expires_at, '%Y-%m-%d %H:%M:%S')
            if datetime.now() > expires_at:
                flash('The verification link has expired. Please request a new verification email.', 'error')
                return immediate_redirect(url_for('resend_verification', _external=True))

        # Mark email as verified in users table
        cursor.execute(
            '''
            UPDATE users
            SET email_verified = 1
            WHERE user_id = %s
            ''',
            (applicant['user_id'],),
        )
        
        # Clear verification token in applicants table
        cursor.execute(
            '''
            UPDATE applicants
            SET verification_token = NULL, verification_token_expires = NULL
            WHERE applicant_id = %s
            ''',
            (applicant['applicant_id'],),
        )
        db.commit()
        flash('Email verified successfully! You can now log in.', 'success')
    except Exception as exc:
        db.rollback()
        print(f'❌ Email verification error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to verify email at this time.', 'error')
    finally:
        cursor.close()

    return immediate_redirect(url_for('login', _external=True))


@app.route('/resend-verification', methods=['GET', 'POST'])
def resend_verification():
    """Resend verification email to user with 60-second token expiration."""
    preset_email = session.pop('pending_verification_email', None)

    if request.method == 'POST':
        email = (request.form.get('email') or preset_email or '').strip().lower()
        
        if not email:
            flash('Please enter your email address.', 'error')
            return render_template('resend_verification.html', preset_email=preset_email)

        db = get_db()
        if not db:
            flash('Database connection error. Please try again later.', 'error')
            return render_template('resend_verification.html', preset_email=preset_email)

        cursor = db.cursor(dictionary=True)
        try:
            # Check if user exists and is not verified
            cursor.execute(
                '''
                SELECT a.applicant_id, a.email, a.full_name, a.user_id, u.email_verified
                FROM applicants a
                JOIN users u ON u.user_id = a.user_id
                WHERE a.email = %s
                LIMIT 1
                ''',
                (email,),
            )
            applicant = cursor.fetchone()

            if not applicant:
                # Don't reveal if email exists for security
                flash('If this email is registered and not verified, a verification email will be sent.', 'info')
                return render_template('resend_verification.html', preset_email=preset_email)

            if applicant.get('email_verified'):
                flash('Your email is already verified. You can log in.', 'info')
                return immediate_redirect(url_for('login', _external=True))
            
            # Generate new token and set expiration (60 seconds)
            from datetime import datetime, timedelta
            verification_token = generate_token()
            token_expires = datetime.now() + timedelta(seconds=60)
            
            # Update verification token
            cursor.execute(
                '''
                UPDATE applicants
                SET verification_token = %s, verification_token_expires = %s
                WHERE applicant_id = %s
                ''',
                (verification_token, token_expires, applicant['applicant_id']),
            )
            db.commit()

            # Send verification email
            try:
                send_verification_email(email, verification_token, applicant_name=applicant.get('full_name'))
                flash('Verification email sent! Please check your inbox and click the link to verify your email. The link expires in 60 seconds.', 'success')
            except Exception as email_error:
                print(f'⚠️ Failed to send verification email: {email_error}')
                flash('Failed to send verification email. Please try again later or contact support.', 'error')
            
        except Exception as exc:
            db.rollback()
            print(f'❌ Resend verification error: {exc}')
            import traceback
            traceback.print_exc()
            flash('An error occurred. Please try again later.', 'error')
        finally:
            cursor.close()

        return immediate_redirect(url_for('login', _external=True))

    # Get preset email from session if available (from login redirect)
    return render_template('resend_verification.html', preset_email=preset_email)


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user_type = request.form.get('user_type', 'applicant').strip()  # applicant, admin, hr
        
        if not email:
            flash('Please enter your email address.', 'error')
            return redirect(url_for('forgot_password'))

        db = get_db()
        if not db:
            flash('Database connection error. Please try again later.', 'error')
            return redirect(url_for('forgot_password'))

        cursor = db.cursor(dictionary=True)
        try:
            user_found = False
            role = 'applicant'
            
            # Check applicants
            if user_type == 'applicant':
                cursor.execute(
                    '''
                    SELECT applicant_id, deleted_at
                    FROM applicants
                    WHERE email = %s
                    LIMIT 1
                    ''',
                    (email,),
                )
                user = cursor.fetchone()
                if user and not user.get('deleted_at'):
                    user_found = True
                    # Note: is_verified column doesn't exist in actual schema
                    if False:  # Skip verification check
                        flash('Please verify your email before requesting a password reset.', 'warning')
                        session['pending_verification_email'] = email
                        return redirect(url_for('resend_verification'))
                    role = 'applicant'
            
            # Check admins (both super_admin and hr)
            if not user_found:
                cursor.execute(
                    '''
                    SELECT a.admin_id, u.user_type, u.is_active
                    FROM admins a
                    JOIN users u ON u.user_id = a.user_id
                    WHERE a.email = %s AND u.is_active = 1
                    LIMIT 1
                    ''',
                    (email,),
                )
                admin = cursor.fetchone()
                if admin:
                    user_found = True
                    role = 'admin' if admin.get('user_type') == 'super_admin' else 'hr'

            if not user_found:
                flash('If an account exists for that email, you will receive a password reset link shortly.', 'info')
                return immediate_redirect(url_for('login', _external=True))

            token = generate_token()

            cursor.execute(
                'DELETE FROM password_resets WHERE user_email = %s',
                (email,),
            )
            cursor.execute(
                '''
                INSERT INTO password_resets (user_email, token, role, expired_at)
                VALUES (%s, %s, %s, DATE_ADD(NOW(), INTERVAL 30 MINUTE))
                ''',
                (email, token, role),
            )
            db.commit()

            send_password_reset_email(email, token)
            flash('If an account exists for that email, a password reset link has been sent.', 'info')
            return immediate_redirect(url_for('login', _external=True))
        except Exception as exc:
            db.rollback()
            print(f'❌ Forgot password error: {exc}')
            flash('Unable to process password reset right now.', 'error')
        finally:
            cursor.close()

    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    db = get_db()
    if not db:
        flash('Database connection error. Please try again later.', 'error')
        return redirect(url_for('forgot_password'))

    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            '''
            SELECT user_email, role, expired_at
            FROM password_resets
            WHERE token = %s
            ORDER BY created_at DESC
            LIMIT 1
            ''',
            (token,),
        )
        reset_record = cursor.fetchone()

        if not reset_record:
            flash('The password reset link is invalid or has expired.', 'error')
            return redirect(url_for('forgot_password'))

        if reset_record.get('expired_at') and reset_record['expired_at'] < datetime.now(timezone.utc):
            flash('The password reset link has expired. Please request a new one.', 'error')
            return redirect(url_for('forgot_password'))

        if request.method == 'POST':
            new_password = request.form.get('new_password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()

            if not new_password or not confirm_password:
                flash('Please enter and confirm your new password.', 'error')
                return redirect(url_for('reset_password', token=token))

            if len(new_password) < 6:
                flash('Password must be at least 6 characters long.', 'error')
                return redirect(url_for('reset_password', token=token))

            if new_password != confirm_password:
                flash('Passwords do not match.', 'error')
                return redirect(url_for('reset_password', token=token))

            role = reset_record.get('role', 'applicant')
            email = reset_record['user_email']
            
            if role == 'applicant':
                cursor.execute(
                    'SELECT applicant_id FROM applicants WHERE email = %s LIMIT 1',
                    (email,),
                )
                user = cursor.fetchone()
                if not user:
                    flash('Unable to locate the account for password reset.', 'error')
                    return redirect(url_for('forgot_password'))
                
                cursor.execute(
                    '''
                    UPDATE applicants
                    SET password_hash = %s
                    WHERE applicant_id = %s
                    ''',
                    (hash_password(new_password), user['applicant_id']),
                )
                log_profile_change(user['applicant_id'], 'applicant', 'password', '[updated]', '[updated]')
            else:  # admin or hr
                cursor.execute(
                    '''
                    SELECT a.admin_id, u.user_id
                    FROM admins a
                    JOIN users u ON u.user_id = a.user_id
                    WHERE a.email = %s
                    LIMIT 1
                    ''',
                    (email,),
                )
                admin = cursor.fetchone()
                if not admin:
                    flash('Unable to locate the account for password reset.', 'error')
                    return redirect(url_for('forgot_password'))
                
                cursor.execute(
                    'UPDATE users SET password_hash = %s WHERE user_id = %s',
                    (hash_password(new_password), admin['user_id']),
                )
            
            cursor.execute(
                'DELETE FROM password_resets WHERE user_email = %s',
                (email,),
            )
            db.commit()
            flash('Your password has been reset. You can now log in.', 'success')
            return immediate_redirect(url_for('login', _external=True))

        return render_template('reset_password.html', token=token)

    except Exception as exc:
        db.rollback()
        print(f'❌ Reset password error: {exc}')
        flash('Unable to reset password at this time.', 'error')
        return redirect(url_for('forgot_password'))
    finally:
        cursor.close()

@app.route('/applicant/dashboard')
@login_required('applicant')
def applicant_dashboard():
    """Applicant dashboard powered by live application data."""
    try:
        applicant_id = session.get('user_id')
        if not applicant_id:
            flash('Unable to identify your account. Please log in again.', 'error')
            return immediate_redirect(url_for('login', _external=True))
        
        dashboard = build_applicant_dashboard_data(applicant_id)
        dashboard = dashboard if isinstance(dashboard, dict) else {}
        raw_stats = dashboard.get('stats', {}) if isinstance(dashboard.get('stats'), dict) else {}
        raw_applications = dashboard.get('applications', []) if isinstance(dashboard.get('applications'), list) else []
        raw_interviews = dashboard.get('upcoming_interviews', []) if isinstance(dashboard.get('upcoming_interviews'), list) else []
        
        applicant_info = {
            'name': session.get('user_name') or session.get('user_email') or 'Applicant'
        }
        
        # Map database statuses to display stats
        # Database uses: 'pending', 'interviewed', 'hired', 'rejected'
        stats_view = {
            'total_applications': raw_stats.get('total_applications') or len(raw_applications),
            'pending': raw_stats.get('pending') or raw_stats.get('under_review') or raw_stats.get('reviewed') or 0,
            'interviewed': raw_stats.get('interviewed') or raw_stats.get('interview') or 0,
            'hired': raw_stats.get('hired') or raw_stats.get('accepted') or 0,
            'rejected': raw_stats.get('rejected') or 0,
        }
        
        recent_applications = []
        for app_entry in raw_applications[:5]:
            recent_applications.append({
                'id': app_entry.get('application_id'),
                'job_title': app_entry.get('job_title') or 'Job Removed',
                'company_name': app_entry.get('branch_name') or 'Unassigned Branch',
                'location': app_entry.get('job_location') or app_entry.get('branch_name'),
                'status': (app_entry.get('status') or '').lower(),
                'applied_date': app_entry.get('submitted_at') or 'Recently',
                'last_update': app_entry.get('updated_at') or '',
                'has_interview': bool(app_entry.get('has_interview')),
            })
        
        upcoming_interviews = []
        for interview in raw_interviews[:5]:
            scheduled_value = interview.get('date_time') or interview.get('scheduled_date') or 'To be determined'
            upcoming_interviews.append({
                'id': interview.get('interview_id'),
                'job_title': interview.get('job_title') or 'Interview',
                'company_name': interview.get('branch_name') or 'Hiring Team',
                'date': scheduled_value,
                'time': '',
                'time_remaining': '',
                'location': interview.get('location') or 'To be confirmed',
                'is_online': 'online' in ((interview.get('type') or interview.get('interview_mode') or '').lower()),
            })
        
        recent_activity = []
        for item in recent_applications:
            recent_activity.append({
                'type': 'application',
                'description': f"Applied to {item['job_title']}",
                'status': item['status'],
                'timestamp': item['applied_date'],
            })
        for interview in upcoming_interviews:
            recent_activity.append({
                'type': 'interview',
                'description': f"Interview for {interview['job_title']}",
                'status': 'interview',
                'timestamp': interview['date'],
            })
        
        # Fetch rejected applications for the modal
        rejected_applications = []
        try:
            db_rejected = get_db()
            if db_rejected:
                cursor = db_rejected.cursor(dictionary=True)
                try:
                    ensure_schema_compatibility()
                    _update_job_columns(cursor)
                    job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
                    
                    cursor.execute(
                        f'''
                        SELECT a.application_id,
                               a.status,
                               a.submitted_at,
                               {job_title_expr} AS job_title,
                               COALESCE(b.branch_name, 'Unassigned') AS branch_name
                        FROM applications a
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        LEFT JOIN branches b ON j.branch_id = b.branch_id
                        WHERE a.applicant_id = %s AND a.status = 'rejected'
                        ORDER BY a.submitted_at DESC
                        ''',
                        (applicant_id,),
                    )
                    rejected_rows = cursor.fetchall() or []
                    for row in rejected_rows:
                        rejected_applications.append({
                            'id': row.get('application_id'),
                            'job_title': row.get('job_title') or 'Job Removed',
                            'branch_name': row.get('branch_name') or 'Unassigned Branch',
                            'submitted_at': format_human_datetime(row.get('submitted_at')) if row.get('submitted_at') else 'Recently',
                        })
                finally:
                    cursor.close()
        except Exception as rejected_error:
            print(f'⚠️ Error fetching rejected applications: {rejected_error}')
        
        return render_template(
            'applicant/dashboard.html',
            applicant=applicant_info,
            stats=stats_view,
            recent_applications=recent_applications,
            recent_activity=recent_activity[:8],
            upcoming_interviews=upcoming_interviews,
            rejected_applications=rejected_applications,
        )
    except Exception as exc:
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant dashboard error: {exc}')
        print(f'Full traceback: {error_details}')
        flash('Unable to load dashboard. Please try again later.', 'error')
        # Return empty dashboard to prevent 500 error
        empty_dashboard = {
            'stats': {
                'total_applications': 0,
                'applied': 0,
                'under_review': 0,
                'interview': 0,
                'hired': 0,
                'rejected': 0,
            },
            'applications': [],
            'upcoming_interviews': [],
            'notifications': [],
        }
        return render_template('applicant/dashboard.html', dashboard=empty_dashboard)


@app.route('/applicant/applications')
@login_required('applicant')
def applicant_applications():
    """Applicant's application history with analytics."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/applications.html', applications=[], analytics={})
    
    cursor = db.cursor(dictionary=True)
    try:
        # Ensure schema compatibility
        ensure_schema_compatibility()
        
        # Get dynamic job column expressions
        _update_job_columns(cursor)
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
        
        # Get status filter from request
        status_filter = request.args.get('status', '').strip().lower()
        
        # Build WHERE clause with status filter
        where_clauses = ['a.applicant_id = %s']
        params = [applicant_id]
        
        if status_filter:
            # Map display statuses to database statuses
            status_map = {
                'pending': ['pending', 'reviewed', 'applied', 'under_review'],
                'scheduled': ['scheduled'],
                'interviewed': ['interviewed', 'interview'],
                'hired': ['hired', 'accepted'],
                'rejected': ['rejected', 'withdrawn'],  # Include 'withdrawn' as it's normalized to 'rejected' for display
            }
            db_statuses = status_map.get(status_filter, [status_filter])
            
            # Build IN clause for multiple status mappings
            if db_statuses:
                placeholders = ','.join(['%s'] * len(db_statuses))
                where_clauses.append(f'a.status IN ({placeholders})')
                params.extend(db_statuses)
                print(f"🔍 Applicant status filter: '{status_filter}' -> WHERE a.status IN {db_statuses}")
        
        where_sql = ' AND '.join(where_clauses)
        
        # Fetch all applications with job details, including viewed_at status
        # Normalize withdrawn to rejected in SQL query - remove withdrawn status completely
        # Use subqueries for interviews to prevent duplicates from multiple interviews per application
        cursor.execute(
            f'''
            SELECT DISTINCT a.application_id,
                   CASE 
                       WHEN a.status = 'withdrawn' THEN 'rejected'
                       ELSE a.status
                   END AS status,
                   a.submitted_at,
                   a.updated_at,
                   j.job_id,
                   {job_title_expr} AS job_title,
                   COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                   {job_title_expr} AS position_title,
                   (SELECT interview_id FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS interview_id,
                   (SELECT scheduled_date FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS scheduled_date,
                   COALESCE((SELECT interview_mode FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1), 'in-person') AS interview_mode,
                   (SELECT location FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS interview_location,
                   a.viewed_at
            FROM applications a
            JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_sql}
            ORDER BY a.submitted_at DESC
            ''',
            tuple(params),
        )
        applications = cursor.fetchall()
        
        # Deduplicate by application_id to prevent any remaining duplicates
        seen_app_ids = set()
        unique_applications = []
        for app in applications:
            app_id = app.get('application_id')
            if app_id and app_id not in seen_app_ids:
                seen_app_ids.add(app_id)
                unique_applications.append(app)
        applications = unique_applications
        
        # Calculate analytics - fetch all applications for accurate stats
        # Normalize withdrawn to rejected in SQL query
        cursor.execute(
            f'''
            SELECT CASE 
                       WHEN a.status = 'withdrawn' THEN 'rejected'
                       ELSE a.status
                   END AS status
            FROM applications a
            WHERE a.applicant_id = %s
            ''',
            (applicant_id,),
        )
        all_apps = cursor.fetchall()
        
        total = len(all_apps)
        status_counts = {
            'pending': 0,
            'reviewed': 0,
            'interviewed': 0,
            'hired': 0,
            'accepted': 0,
            'rejected': 0,
            'applied': 0,
            'under_review': 0,
            'interview': 0
        }
        
        for app in all_apps:
            status = (app.get('status') or 'pending').lower()
            # Status is already normalized to rejected in SQL query (withdrawn -> rejected)
            if status in status_counts:
                status_counts[status] += 1
        
        # Map for display (combine similar statuses)
        # Note: withdrawn is already normalized to rejected in SQL query
        display_counts = {
            'pending': status_counts.get('pending', 0) + status_counts.get('reviewed', 0) + status_counts.get('applied', 0) + status_counts.get('under_review', 0),
            'interviewed': status_counts.get('interviewed', 0) + status_counts.get('interview', 0),
            'hired': status_counts.get('hired', 0) + status_counts.get('accepted', 0),
            'rejected': status_counts.get('rejected', 0),  # Withdrawn already normalized to rejected in SQL
        }
        
        analytics = {
            'total_applications': total,
            'status_breakdown': display_counts,
            'response_rate': round(
                (
                    display_counts.get('interviewed', 0)
                    + display_counts.get('hired', 0)
                    + display_counts.get('rejected', 0)
                )
                / total
                * 100,
                1,
            ) if total > 0 else 0,
            'interview_count': len([a for a in applications if a.get('interview_id')]),
        }
        
        # Format applications
        formatted_apps = []
        for app in applications:
            # Normalize withdrawn to rejected - remove withdrawn status completely
            app_status = (app.get('status') or 'pending').lower()
            if app_status == 'withdrawn':
                app_status = 'rejected'
            
            formatted_apps.append({
                'application_id': app.get('application_id'),
                'job_id': app.get('job_id'),
                'job_title': app.get('job_title'),
                'branch_name': app.get('branch_name'),
                'position_title': app.get('position_title'),
                'status': app_status,  # Use normalized status (withdrawn -> rejected)
                'submitted_at': format_human_datetime(app.get('submitted_at')),
                'has_interview': app.get('interview_id') is not None,
                'interview_date': format_human_datetime(app.get('scheduled_date')) if app.get('scheduled_date') else None,
                'interview_mode': app.get('interview_mode'),
                'interview_location': app.get('interview_location'),
                'is_viewed': app.get('viewed_at') is not None,  # Check if application has been viewed by HR/Admin
            })
        
        return render_template('applicant/applications.html', applications=formatted_apps, analytics=analytics, status_filter=status_filter)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant applications error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Unable to load applications: {str(exc)}', 'error')
        return render_template('applicant/applications.html', applications=[], analytics={})
    finally:
        if cursor:
            cursor.close()


@app.route('/applicant/interviews', methods=['GET', 'POST'])
@login_required('applicant')
def applicant_interviews():
    """Applicant's interview schedule and management."""
    applicant_id = session.get('user_id')
    if not applicant_id:
        flash('Please log in to view your interviews.', 'error')
        return immediate_redirect(url_for('login', _external=True))
    
    # Validate applicant_id is an integer
    try:
        applicant_id = int(applicant_id)
    except (ValueError, TypeError):
        print(f'⚠️ Invalid applicant_id in session: {applicant_id}')
        flash('Unable to identify your account. Please log in again.', 'error')
        return immediate_redirect(url_for('login', _external=True))
    
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/interviews.html', interviews=[], upcoming=[], past=[])
    
    cursor = None
    try:
        # Ensure schema compatibility - wrap in try-except to prevent crashes
        try:
            ensure_schema_compatibility()
        except Exception as schema_error:
            print(f'⚠️ Schema compatibility check failed (non-critical): {schema_error}')
            # Continue anyway - schema might still be compatible
        
        cursor = db.cursor(dictionary=True)
        # Handle POST actions (confirm/cancel interview)
        if request.method == 'POST':
            action = request.form.get('action')
            interview_id = request.form.get('interview_id', type=int)
            
            if action == 'confirm_interview' and interview_id:
                # Verify interview belongs to this applicant and get current status
                cursor.execute(
                    '''
                    SELECT i.interview_id, i.status, i.scheduled_date
                    FROM interviews i
                    JOIN applications a ON i.application_id = a.application_id
                    WHERE i.interview_id = %s AND a.applicant_id = %s
                    LIMIT 1
                    ''',
                    (interview_id, applicant_id),
                )
                interview = cursor.fetchone()
                if not interview:
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Interview not found.'}), 404
                    flash('Interview not found.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                current_status = (interview.get('status') or 'scheduled').lower()
                
                # Check if already confirmed
                if current_status == 'confirmed':
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'This interview is already confirmed.'})
                    flash('This interview is already confirmed.', 'info')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if cancelled - cannot confirm a cancelled interview
                if current_status == 'cancelled':
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot confirm a cancelled interview. Please contact HR if you need to reschedule.'})
                    flash('Cannot confirm a cancelled interview. Please contact HR if you need to reschedule.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if already completed or no_show
                if current_status in ('completed', 'no_show'):
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot modify the status of a completed interview.'})
                    flash('Cannot modify the status of a completed interview.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if interview is in the past
                scheduled_date = interview.get('scheduled_date')
                if scheduled_date and isinstance(scheduled_date, datetime) and scheduled_date < datetime.now():
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot confirm past interviews.'})
                    flash('Cannot confirm past interviews.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Update interview status to confirmed
                # IMPORTANT: This action does TWO things:
                # 1. Updates the interview status in the database to 'confirmed'
                # 2. Notifies HR users about the confirmation
                # Both operations are committed together in a single transaction
                try:
                    # First, try to update status to 'confirmed'
                    try:
                        # First, ensure enum has 'confirmed' value
                        try:
                            cursor.execute("ALTER TABLE interviews MODIFY COLUMN status ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
                            print('✅ Ensured "confirmed" is in status enum')
                        except Exception as enum_check_err:
                            error_msg = str(enum_check_err).lower()
                            # If enum already has the value or table doesn't exist, that's okay
                            if 'duplicate' not in error_msg and 'already' not in error_msg:
                                print(f'⚠️ Could not modify enum (may already be correct): {enum_check_err}')
                        
                        # Now update the status - wrap in try-except to catch MySQL enum errors
                        try:
                            cursor.execute(
                                'UPDATE interviews SET status = %s WHERE interview_id = %s',
                                ('confirmed', interview_id),
                            )
                            rows_updated = cursor.rowcount
                            
                            if rows_updated == 0:
                                # Verify interview exists
                                cursor.execute('SELECT interview_id FROM interviews WHERE interview_id = %s', (interview_id,))
                                exists = cursor.fetchone()
                                if not exists:
                                    raise Exception(f'Interview {interview_id} does not exist.')
                                else:
                                    raise Exception(f'Update returned 0 rows but interview exists.')
                            
                            print(f'✅ Interview {interview_id} status updated to confirmed. Rows updated: {rows_updated}')
                        except Exception as mysql_err:
                            error_msg = str(mysql_err).lower()
                            # If it's an enum/invalid value error, the ALTER should have fixed it
                            # But if it still fails, add note as fallback
                            if 'enum' in error_msg or 'invalid' in error_msg or 'value' in error_msg:
                                print(f'⚠️ MySQL enum error (unexpected after ALTER): {mysql_err}')
                                # Try one more time after ensuring enum
                                try:
                                    cursor.execute("ALTER TABLE interviews MODIFY COLUMN status ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
                                    cursor.execute(
                                        'UPDATE interviews SET status = %s WHERE interview_id = %s',
                                        ('confirmed', interview_id),
                                    )
                                    rows_updated = cursor.rowcount
                                    if rows_updated > 0:
                                        print(f'✅ Interview {interview_id} status updated after retry. Rows updated: {rows_updated}')
                                    else:
                                        raise Exception('Retry also returned 0 rows')
                                except Exception as retry_err:
                                    print(f'⚠️ Retry also failed: {retry_err}')
                                    raise
                            else:
                                raise
                    except Exception as status_update_error:
                        # Final fallback: add a note
                        print(f'⚠️ Status update failed: {status_update_error}')
                        try:
                            cursor.execute(
                                'UPDATE interviews SET notes = CONCAT(COALESCE(notes, ""), "\n\n[Applicant Confirmed Attendance on ", NOW(), "]") WHERE interview_id = %s',
                                (interview_id,),
                            )
                            print(f'✅ Added confirmation note to interview {interview_id}')
                        except Exception as note_err:
                            print(f'❌ Failed to add confirmation note: {note_err}')
                            raise
                    # Get interview details for HR notification
                    cursor.execute(
                        '''
                        SELECT i.interview_id, i.scheduled_date, i.location, i.interview_mode,
                               a.application_id, a.applicant_id,
                               ap.full_name AS applicant_name, ap.email AS applicant_email,
                               j.job_id, COALESCE(j.title, 'Untitled Job') AS job_title,
                               b.branch_id, b.branch_name
                        FROM interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        LEFT JOIN branches b ON j.branch_id = b.branch_id
                        WHERE i.interview_id = %s
                        LIMIT 1
                        ''',
                        (interview_id,),
                    )
                    interview_details = cursor.fetchone()
                    
                    # Notify HR about the confirmation - ALWAYS notify HR when status is updated
                    if interview_details:
                        try:
                            branch_id = interview_details.get('branch_id')
                            applicant_name = interview_details.get('applicant_name') or 'Applicant'
                            job_title = interview_details.get('job_title') or 'Position'
                            scheduled_date = interview_details.get('scheduled_date')
                            scheduled_str = format_human_datetime(scheduled_date) if scheduled_date else 'TBD'
                            
                            notification_message = f'Applicant {applicant_name} has confirmed attendance for interview: {job_title} scheduled on {scheduled_str}.'
                            
                            # Check if notification already exists to prevent duplicates
                            cursor.execute(
                                '''
                                SELECT notification_id FROM notifications
                                WHERE application_id = %s AND message = %s
                                LIMIT 1
                                ''',
                                (interview_details.get('application_id'), notification_message),
                            )
                            existing_notification = cursor.fetchone()
                            
                            if not existing_notification:
                                # Get HR users to notify (for logging purposes)
                                # HR accounts manage all branches (branch_id column removed from admins table)
                                hr_users_to_notify = []
                                cursor.execute(
                                    '''
                                    SELECT a.admin_id, u.email, a.full_name
                                    FROM admins a
                                    JOIN users u ON u.user_id = a.user_id
                                    WHERE a.is_active = 1 AND u.user_type = 'hr' AND u.is_active = 1
                                    '''
                                )
                                all_hr = cursor.fetchall() or []
                                hr_users_to_notify.extend(all_hr)
                                
                                # Create notification (one notification for all HR users)
                                # Create notification regardless - HR users will see it when they check notifications
                                create_admin_notification(cursor, notification_message, interview_details.get('application_id'))
                                print(f'✅ HR notification created for interview confirmation: {notification_message}')
                                print(f'   - Status updated to: confirmed')
                                print(f'   - HR users to be notified: {len(hr_users_to_notify)}')
                            else:
                                print(f'⚠️ HR notification already exists for this confirmation - skipping duplicate notification')
                                print(f'   - Status updated to: confirmed')
                        except Exception as notify_err:
                            print(f'⚠️ Error creating HR notification for interview confirmation: {notify_err}')
                            import traceback
                            traceback.print_exc()
                            # Don't fail the whole operation if notification fails, but log it
                    
                    # Commit the transaction
                    db.commit()
                    print(f'✅ Transaction committed for interview {interview_id} confirmation')
                    
                    # Verify the update was successful
                    cursor.execute(
                        'SELECT status FROM interviews WHERE interview_id = %s LIMIT 1',
                        (interview_id,)
                    )
                    verify_status = cursor.fetchone()
                    actual_status = (verify_status.get('status') or '').lower() if verify_status else None
                    print(f'✅ Verified interview {interview_id} status after update: {actual_status}')
                    
                    if request.accept_mimetypes.accept_json:
                        return jsonify({
                            'success': True, 
                            'message': 'Interview attendance confirmed successfully.', 
                            'status': actual_status or 'confirmed'
                        })
                    flash('Interview attendance confirmed successfully.', 'success')
                except Exception as update_exc:
                    db.rollback()
                    print(f'⚠️ Error updating interview status: {update_exc}')
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Unable to confirm interview. Please contact HR.'}), 500
                    flash('Unable to confirm interview. Please contact HR.', 'error')
            
            elif action == 'cancel_interview' and interview_id:
                # Verify interview belongs to this applicant and get current status
                cursor.execute(
                    '''
                    SELECT i.interview_id, i.status, i.scheduled_date
                    FROM interviews i
                    JOIN applications a ON i.application_id = a.application_id
                    WHERE i.interview_id = %s AND a.applicant_id = %s
                    LIMIT 1
                    ''',
                    (interview_id, applicant_id),
                )
                interview = cursor.fetchone()
                if not interview:
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Interview not found.'}), 404
                    flash('Interview not found.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                current_status = (interview.get('status') or 'scheduled').lower()
                
                # Check if already cancelled
                if current_status == 'cancelled':
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'This interview is already cancelled.'})
                    flash('This interview is already cancelled.', 'info')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if already completed or no_show
                if current_status in ('completed', 'no_show'):
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot cancel a completed interview.'})
                    flash('Cannot cancel a completed interview.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if interview is in the past
                scheduled_date = interview.get('scheduled_date')
                if scheduled_date and isinstance(scheduled_date, datetime) and scheduled_date < datetime.now():
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot cancel past interviews.'})
                    flash('Cannot cancel past interviews.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Update interview status to cancelled
                # IMPORTANT: This action does TWO things:
                # 1. Updates the interview status in the database to 'cancelled'
                # 2. Notifies HR users about the cancellation
                # Both operations are committed together in a single transaction
                try:
                    # Try to update status to 'cancelled' (should be in enum already)
                    try:
                        cursor.execute(
                            'UPDATE interviews SET status = %s WHERE interview_id = %s',
                            ('cancelled', interview_id),
                        )
                        rows_updated = cursor.rowcount
                        
                        if rows_updated == 0:
                            # Check if interview exists
                            cursor.execute('SELECT interview_id FROM interviews WHERE interview_id = %s', (interview_id,))
                            exists = cursor.fetchone()
                            if not exists:
                                raise Exception(f'Interview {interview_id} does not exist.')
                            else:
                                raise Exception(f'Update failed but interview exists. Possible enum issue.')
                        
                        print(f'✅ Interview {interview_id} status updated to cancelled. Rows updated: {rows_updated}')
                    except Exception as status_update_error:
                        error_msg = str(status_update_error).lower()
                        print(f'⚠️ Status update error: {status_update_error}')
                        
                        # If it's an enum error, try to fix it
                        if 'enum' in error_msg or 'invalid' in error_msg or 'value' in error_msg:
                            try:
                                # Ensure 'cancelled' is in enum
                                cursor.execute("ALTER TABLE interviews MODIFY COLUMN status ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
                                # Try update again
                                cursor.execute(
                                    'UPDATE interviews SET status = %s WHERE interview_id = %s',
                                    ('cancelled', interview_id),
                                )
                                rows_updated = cursor.rowcount
                                if rows_updated > 0:
                                    print(f'✅ Interview {interview_id} status updated to cancelled after enum fix.')
                                else:
                                    raise Exception('Update still failed after enum fix')
                            except Exception as enum_fix_err:
                                print(f'⚠️ Could not fix enum: {enum_fix_err}')
                                # Fallback: add a note
                                cursor.execute(
                                    'UPDATE interviews SET notes = CONCAT(COALESCE(notes, ""), "\n\n[Applicant Cancelled on ", NOW(), "]") WHERE interview_id = %s',
                                    (interview_id,),
                                )
                                print(f'✅ Added cancellation note to interview {interview_id}')
                        else:
                            # Other error - try adding note
                            try:
                                cursor.execute(
                                    'UPDATE interviews SET notes = CONCAT(COALESCE(notes, ""), "\n\n[Applicant Cancelled on ", NOW(), "]") WHERE interview_id = %s',
                                    (interview_id,),
                                )
                                print(f'✅ Added cancellation note to interview {interview_id}')
                            except Exception as note_err:
                                print(f'❌ Failed to add cancellation note: {note_err}')
                                raise
                    # Get interview details for HR notification
                    cursor.execute(
                        '''
                        SELECT i.interview_id, i.scheduled_date, i.location, i.interview_mode,
                               a.application_id, a.applicant_id,
                               ap.full_name AS applicant_name, ap.email AS applicant_email,
                               j.job_id, COALESCE(j.title, 'Untitled Job') AS job_title,
                               b.branch_id, b.branch_name
                        FROM interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        LEFT JOIN branches b ON j.branch_id = b.branch_id
                        WHERE i.interview_id = %s
                        LIMIT 1
                        ''',
                        (interview_id,),
                    )
                    interview_details = cursor.fetchone()
                    
                    # Notify HR about the cancellation - ALWAYS notify HR when status is updated
                    if interview_details:
                        try:
                            branch_id = interview_details.get('branch_id')
                            applicant_name = interview_details.get('applicant_name') or 'Applicant'
                            job_title = interview_details.get('job_title') or 'Position'
                            scheduled_date = interview_details.get('scheduled_date')
                            scheduled_str = format_human_datetime(scheduled_date) if scheduled_date else 'TBD'
                            
                            notification_message = f'Applicant {applicant_name} has cancelled the interview: {job_title} scheduled on {scheduled_str}.'
                            
                            # Check if notification already exists to prevent duplicates
                            cursor.execute(
                                '''
                                SELECT notification_id FROM notifications
                                WHERE application_id = %s AND message = %s
                                LIMIT 1
                                ''',
                                (interview_details.get('application_id'), notification_message),
                            )
                            existing_notification = cursor.fetchone()
                            
                            if not existing_notification:
                                # Create notification for HR users in the same branch
                                hr_users_to_notify = []
                                
                                if branch_id:
                                    # HR accounts manage all branches (branch_id column removed from admins table)
                                    cursor.execute(
                                        '''
                                        SELECT a.admin_id, u.email, a.full_name
                                        FROM admins a
                                        JOIN users u ON u.user_id = a.user_id
                                        WHERE a.is_active = 1 AND u.user_type = 'hr' AND u.is_active = 1
                                        '''
                                    )
                                    all_hr = cursor.fetchall() or []
                                    hr_users_to_notify.extend(all_hr)
                                
                                # Create notification (one notification for all HR users)
                                # Create notification regardless - HR users will see it when they check notifications
                                create_admin_notification(cursor, notification_message, interview_details.get('application_id'))
                                print(f'✅ HR notification created for interview cancellation: {notification_message}')
                                print(f'   - Status updated to: cancelled')
                                print(f'   - HR users to be notified: {len(hr_users_to_notify)}')
                            else:
                                print(f'⚠️ HR notification already exists for this cancellation - skipping duplicate notification')
                                print(f'   - Status updated to: cancelled')
                        except Exception as notify_err:
                            print(f'⚠️ Error creating HR notification for interview cancellation: {notify_err}')
                            import traceback
                            traceback.print_exc()
                            # Don't fail the whole operation if notification fails, but log it
                    
                    # Commit the transaction
                    db.commit()
                    print(f'✅ Transaction committed for interview {interview_id} cancellation')
                    
                    # Verify the update was successful
                    cursor.execute(
                        'SELECT status FROM interviews WHERE interview_id = %s LIMIT 1',
                        (interview_id,)
                    )
                    verify_status = cursor.fetchone()
                    actual_status = (verify_status.get('status') or '').lower() if verify_status else None
                    print(f'✅ Verified interview {interview_id} status after update: {actual_status}')
                    
                    if request.accept_mimetypes.accept_json:
                        return jsonify({
                            'success': True, 
                            'message': 'Interview cancellation requested. HR will be notified.', 
                            'status': actual_status or 'cancelled'
                        })
                    flash('Interview cancellation requested. HR will be notified.', 'success')
                except Exception as update_exc:
                    db.rollback()
                    print(f'⚠️ Error updating interview status: {update_exc}')
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Unable to cancel interview. Please contact HR.'}), 500
                    flash('Unable to cancel interview. Please contact HR.', 'error')
            
            elif action == 'delete_all_interviews':
                # Delete all interviews for this applicant
                try:
                    # Get all interview IDs for this applicant before deletion (for notification)
                    cursor.execute(
                        '''
                        SELECT i.interview_id, i.application_id, i.scheduled_date,
                               j.job_id, COALESCE(j.title, 'Untitled Job') AS job_title,
                               b.branch_id, b.branch_name
                        FROM interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        LEFT JOIN branches b ON j.branch_id = b.branch_id
                        WHERE a.applicant_id = %s
                        ''',
                        (applicant_id,)
                    )
                    interviews_to_delete = cursor.fetchall()
                    deleted_count = len(interviews_to_delete)
                    
                    if deleted_count == 0:
                        if request.accept_mimetypes.accept_json:
                            return jsonify({'success': False, 'message': 'No interviews found to delete.'})
                        flash('No interviews found to delete.', 'info')
                        return immediate_redirect(url_for('applicant_interviews', _external=True))
                    
                    # Delete all interviews for this applicant
                    cursor.execute(
                        '''
                        DELETE i FROM interviews i
                        INNER JOIN applications a ON i.application_id = a.application_id
                        WHERE a.applicant_id = %s
                        ''',
                        (applicant_id,)
                    )
                    deleted_rows = cursor.rowcount
                    
                    # Notify HR about the deletion
                    if interviews_to_delete:
                        try:
                            # Get unique branch IDs
                            branch_ids = set()
                            for interview in interviews_to_delete:
                                branch_id = interview.get('branch_id')
                                if branch_id:
                                    branch_ids.add(branch_id)
                            
                            # Get applicant info
                            cursor.execute(
                                'SELECT full_name, email FROM applicants WHERE applicant_id = %s LIMIT 1',
                                (applicant_id,)
                            )
                            applicant_data = cursor.fetchone()
                            applicant_name = applicant_data.get('full_name') if applicant_data else 'Applicant'
                            
                            notification_message = f'Applicant {applicant_name} has deleted all {deleted_count} interview(s).'
                            
                            # Create notification for HR users
                            for branch_id in branch_ids:
                                if branch_id:
                                    # HR accounts manage all branches (branch_id column removed from admins table)
                                    cursor.execute(
                                        '''
                                        SELECT a.admin_id, u.email, a.full_name
                                        FROM admins a
                                        JOIN users u ON u.user_id = a.user_id
                                        WHERE a.is_active = 1 AND u.user_type = 'hr' AND u.is_active = 1
                                        '''
                                    )
                                    all_hr = cursor.fetchall() or []
                            
                            # Create notification (one notification for all HR users)
                            # Use the first application_id from deleted interviews if available
                            first_application_id = interviews_to_delete[0].get('application_id') if interviews_to_delete else None
                            create_admin_notification(cursor, notification_message, first_application_id)
                            print(f'✅ HR notification created for deleting all interviews: {notification_message}')
                        except Exception as notify_err:
                            print(f'⚠️ Error creating HR notification for deleting all interviews: {notify_err}')
                            # Don't fail the whole operation if notification fails
                    
                    db.commit()
                    print(f'✅ Deleted {deleted_rows} interview(s) for applicant {applicant_id}')
                    
                    if request.accept_mimetypes.accept_json:
                        return jsonify({
                            'success': True,
                            'message': f'All {deleted_rows} interview(s) deleted successfully. HR has been notified.',
                            'deleted_count': deleted_rows
                        })
                    flash(f'All {deleted_rows} interview(s) deleted successfully. HR has been notified.', 'success')
                except Exception as delete_exc:
                    db.rollback()
                    print(f'⚠️ Error deleting all interviews: {delete_exc}')
                    import traceback
                    traceback.print_exc()
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Unable to delete all interviews. Please try again.'}), 500
                    flash('Unable to delete all interviews. Please try again.', 'error')
            
            if request.method == 'POST':
                if request.accept_mimetypes.accept_json:
                    # Return JSON response for AJAX requests
                    return jsonify({'success': True, 'redirect': url_for('applicant_interviews')})
                return immediate_redirect(url_for('applicant_interviews', _external=True))
        
        # GET request - fetch interviews
        interviews = []
        try:
            # Check if interviews table exists
            table_exists = False
            try:
                cursor.execute("SHOW TABLES LIKE 'interviews'")
                if cursor.fetchone():
                    table_exists = True
            except Exception as table_check_error:
                print(f'⚠️ Error checking for interviews table: {table_check_error}')
                # Assume table exists and try to query anyway
                table_exists = True
            
            # Check if applications table exists
            applications_table_exists = False
            try:
                cursor.execute("SHOW TABLES LIKE 'applications'")
                if cursor.fetchone():
                    applications_table_exists = True
            except Exception as app_table_check_error:
                print(f'⚠️ Error checking for applications table: {app_table_check_error}')
                applications_table_exists = True
            
            if table_exists and applications_table_exists and cursor:
                # Use the simplest possible query first - no dynamic functions, no complex joins
                # This query should work regardless of schema variations
                query = '''
                    SELECT DISTINCT i.interview_id,
                           i.scheduled_date,
                           COALESCE(i.interview_mode, 'in-person') AS interview_mode,
                           i.location,
                           i.notes,
                           COALESCE(i.status, 'scheduled') AS interview_status,
                           a.application_id,
                           COALESCE(a.status, 'pending') AS application_status,
                           a.job_id,
                           'Untitled Job' AS job_title,
                           'Unassigned' AS branch_name
                    FROM interviews i
                    INNER JOIN applications a ON i.application_id = a.application_id
                    WHERE a.applicant_id = %s
                    ORDER BY i.scheduled_date DESC
                '''
                
                try:
                    cursor.execute(query, (applicant_id,))
                    interviews = cursor.fetchall() or []
                    
                    # Deduplicate by interview_id to prevent duplicates
                    seen_ids = set()
                    unique_interviews = []
                    for interview in interviews:
                        interview_id = interview.get('interview_id')
                        if interview_id and interview_id not in seen_ids:
                            seen_ids.add(interview_id)
                            unique_interviews.append(interview)
                    interviews = unique_interviews
                    
                    # Try to enrich with job title and branch name if possible
                    if interviews:
                        try:
                            # Get job IDs from interviews
                            job_ids = [int(i.get('job_id')) for i in interviews if i.get('job_id')]
                            if job_ids:
                                # Try to get job titles and branch names
                                jobs_query = '''
                                    SELECT j.job_id,
                                           COALESCE(j.title, 'Untitled Job') AS job_title,
                                           COALESCE(b.branch_name, 'Unassigned') AS branch_name
                                    FROM jobs j
                                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                                    WHERE j.job_id IN (%s)
                                ''' % ','.join(['%s'] * len(job_ids))
                                cursor.execute(jobs_query, tuple(job_ids))
                                jobs_data = {row['job_id']: row for row in cursor.fetchall()}
                                
                                # Update interviews with job data
                                for interview in interviews:
                                    job_id = interview.get('job_id')
                                    if job_id and job_id in jobs_data:
                                        job_data = jobs_data[job_id]
                                        interview['job_title'] = job_data.get('job_title', 'Untitled Job')
                                        interview['branch_name'] = job_data.get('branch_name', 'Unassigned')
                        except Exception as enrich_error:
                            print(f'⚠️ Error enriching interview data: {enrich_error}')
                            # Continue with default values - not critical
                            pass
                except Exception as execute_error:
                    print(f'⚠️ Error executing interview query: {execute_error}')
                    import traceback
                    traceback.print_exc()
                    # Last resort: absolute minimal query
                    try:
                        minimal_query = '''
                            SELECT DISTINCT i.interview_id,
                                   i.scheduled_date,
                                   i.interview_mode,
                                   i.location,
                                   i.notes,
                                   i.status AS interview_status,
                                   a.application_id,
                                   a.status AS application_status,
                                   a.job_id
                            FROM interviews i
                            INNER JOIN applications a ON i.application_id = a.application_id
                            WHERE a.applicant_id = %s
                            ORDER BY i.scheduled_date DESC
                        '''
                        cursor.execute(minimal_query, (applicant_id,))
                        interviews = cursor.fetchall() or []
                        
                        # Deduplicate by interview_id
                        seen_ids = set()
                        unique_interviews = []
                        for interview in interviews:
                            interview_id = interview.get('interview_id')
                            if interview_id and interview_id not in seen_ids:
                                seen_ids.add(interview_id)
                                unique_interviews.append(interview)
                        interviews = unique_interviews
                        
                        # Add default values for missing fields
                        for interview in interviews:
                            interview['job_title'] = 'Untitled Job'
                            interview['branch_name'] = 'Unassigned'
                            interview['application_status'] = interview.get('application_status') or 'pending'
                            interview['interview_status'] = interview.get('interview_status') or 'scheduled'
                            interview['interview_mode'] = interview.get('interview_mode') or 'in-person'
                    except Exception as minimal_error:
                        print(f'⚠️ Error executing minimal interview query: {minimal_error}')
                        import traceback
                        traceback.print_exc()
                        interviews = []
        except Exception as query_error:
            print(f'⚠️ Error fetching interviews: {query_error}')
            import traceback
            traceback.print_exc()
            interviews = []
        
        now = datetime.now()
        upcoming = []
        past = []
        
        for interview in interviews:
            try:
                scheduled = interview.get('scheduled_date')
                interview_status = interview.get('interview_status') or 'scheduled'
                
                if not scheduled:
                    # Skip interviews without a scheduled date
                    continue
                
                # Handle both datetime objects and date strings
                scheduled_dt = None
                if isinstance(scheduled, datetime):
                    scheduled_dt = scheduled
                elif isinstance(scheduled, date):
                    # date but not datetime
                    scheduled_dt = datetime.combine(scheduled, datetime.min.time())
                elif isinstance(scheduled, str):
                    try:
                        scheduled_dt = datetime.strptime(scheduled, '%Y-%m-%d %H:%M:%S')
                    except:
                        try:
                            scheduled_dt = datetime.strptime(scheduled, '%Y-%m-%d')
                        except:
                            try:
                                scheduled_dt = datetime.strptime(scheduled, '%Y-%m-%d %H:%M:%S.%f')
                            except:
                                print(f'⚠️ Could not parse scheduled_date: {scheduled}')
                                continue
                else:
                    print(f'⚠️ Unexpected scheduled_date type: {type(scheduled)}')
                    continue
                
                if not scheduled_dt:
                    continue
                
                # Format the date safely
                try:
                    formatted_date = format_human_datetime(scheduled_dt)
                except Exception as format_error:
                    print(f'⚠️ Error formatting date: {format_error}')
                    formatted_date = str(scheduled_dt)
                
                interview_data = {
                    'interview_id': interview.get('interview_id'),
                    'application_id': interview.get('application_id'),
                    'job_title': interview.get('job_title') or 'Unknown Job Position',
                    'branch_name': interview.get('branch_name') or 'Unassigned',
                    'scheduled_date': formatted_date,
                    'interview_mode': interview.get('interview_mode') or 'in-person',
                    'location': interview.get('location'),
                    'notes': interview.get('notes'),
                    'application_status': interview.get('application_status') or 'pending',
                    'interview_status': interview_status,
                }
                
                if scheduled_dt >= now:
                    upcoming.append(interview_data)
                else:
                    past.append(interview_data)
            except Exception as process_error:
                print(f'⚠️ Error processing interview {interview.get("interview_id")}: {process_error}')
                import traceback
                traceback.print_exc()
                continue
            
        # Ensure we always return valid data
        return render_template(
            'applicant/interviews.html', 
            interviews=upcoming + past, 
            upcoming=upcoming or [], 
            past=past or []
        )
    except Exception as exc:
        if db:
            try:
                if hasattr(db, 'is_connected'):
                    if db.is_connected():
                        db.rollback()
                else:
                    try:
                        db.rollback()
                    except:
                        pass
            except:
                pass
        print(f'❌ Applicant interviews error: {exc}')
        import traceback
        error_trace = traceback.format_exc()
        print(f'Full traceback: {error_trace}')
        
        # Ensure cursor is closed if it exists
        if cursor:
            try:
                cursor.close()
            except:
                pass
        
        # Try to flash the error, but don't fail if flashing fails
        try:
            flash('Unable to load interviews. Please try again later.', 'error')
        except:
            pass
        
        # Always return valid template data - return empty lists to show no interviews instead of error
        try:
            return render_template(
                'applicant/interviews.html', 
                interviews=[], 
                upcoming=[], 
                past=[]
            )
        except Exception as template_error:
            # If template rendering fails, return a simple text response
            print(f'❌ Error rendering template: {template_error}')
            from flask import Response
            return Response('Unable to load interviews. Please try again later.', status=200, mimetype='text/plain')
    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass


@app.route('/applicant/notifications')
@login_required('applicant')
def applicant_notifications():
    """Applicant notification center with preferences."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/communications.html', notifications=[], unread_count=0, preferences={})
    
    cursor = db.cursor(dictionary=True)
    try:
        # Fetch notifications
        # Ensure schema compatibility
        ensure_schema_compatibility()
        
        # Verify notifications table exists before continuing
        cursor.execute("SHOW TABLES LIKE 'notifications'")
        table_exists = cursor.fetchone()
        if not table_exists:
            return render_template(
                'applicant/communications.html',
                notifications=[],
                unread_count=0,
                preferences={
                    'email_enabled': True,
                    'email_frequency': 'immediate',
                    'categories': ['application_updates', 'interview_alerts', 'system_messages'],
                },
            )
        
        # Check notifications table columns
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        
        has_application_fk = 'application_id' in notification_columns
        if not has_application_fk:
            print('⚠️ Notifications table missing application_id column; skipping applicant notifications view.')
            preferences = {
                'email_enabled': True,
                'email_frequency': 'immediate',
                'categories': ['application_updates', 'interview_alerts', 'system_messages'],
            }
            return render_template(
                'applicant/communications.html',
                notifications=[],
                unread_count=0,
                preferences=preferences,
            )
        
        # Build sent_at expression
        if 'sent_at' in notification_columns:
            sent_at_expr = 'n.sent_at'
        elif 'created_at' in notification_columns:
            sent_at_expr = 'n.created_at'
        else:
            sent_at_expr = 'NOW()'
        
        # Build is_read expression
        if 'is_read' in notification_columns:
            is_read_expr = 'COALESCE(n.is_read, 0)'
        else:
            is_read_expr = '0'
        
        # Get dynamic job column expressions
        _update_job_columns(cursor)
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'System Notification'")
        
        select_fields = [
            'n.notification_id',
            'n.message',
            f'{sent_at_expr} AS sent_at',
            f'{is_read_expr} AS is_read',
            'n.application_id',
            'COALESCE(a.status, \'\') AS application_status',
            f'COALESCE({job_title_expr}, \'System Notification\') AS job_title',
        ]
        query = f'''
            SELECT DISTINCT {', '.join(select_fields)}
            FROM notifications n
            JOIN applications a ON n.application_id = a.application_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            WHERE a.applicant_id = %s
            AND (
                n.message LIKE 'You applied for%'
                OR n.message LIKE 'Your application status%'
                OR n.message LIKE 'Congratulations! You have been hired%'
            )
            ORDER BY {sent_at_expr} DESC
            LIMIT 50
        '''
        cursor.execute(query, (applicant_id,))
        notifications = cursor.fetchall()
        
        unread_count = len([n for n in notifications if not n.get('is_read')])
        
        # Fetch notification preferences (if table exists)
        preferences = {
            'email_enabled': True,
            'email_frequency': 'immediate',  # immediate, daily, weekly
            'categories': ['application_updates', 'interview_alerts', 'system_messages'],
        }
        
        formatted_notifications = []
        for notif in notifications:
            formatted_notifications.append({
                'notification_id': notif.get('notification_id'),
                'message': notif.get('message'),
                'sent_at': format_human_datetime(notif.get('sent_at')),
                'is_read': notif.get('is_read', False),
                'application_id': notif.get('application_id'),
                'job_title': notif.get('job_title'),
                'application_status': notif.get('application_status'),
            })
        
        return render_template(
            'applicant/communications.html',
            notifications=formatted_notifications,
            unread_count=unread_count,
            preferences=preferences,
        )
    except Exception as exc:
        if db:
            db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant notifications error: {exc}')
        print(f'Full traceback: {error_details}')
        flash('Unable to load notifications. Please try again later.', 'error')
        return render_template('applicant/communications.html', notifications=[], unread_count=0, preferences={})
    finally:
        if cursor:
            cursor.close()


@app.route('/applicant/notifications/read-all', methods=['POST'])
@login_required('applicant')
def mark_all_notifications_read():
    """Mark all notifications as read for the applicant."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        # Fallback to HTML flow
        if request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('applicant_notifications'))
    
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            '''
            UPDATE notifications n
            JOIN applications a ON n.application_id = a.application_id
            SET n.is_read = 1
            WHERE a.applicant_id = %s AND n.is_read = 0
            ''',
            (applicant_id,),
        )
        db.commit()
        # Content negotiation: JSON for AJAX, redirect with flash otherwise
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'All notifications marked as read'})
        flash('All notifications marked as read.', 'success')
        return redirect(url_for('applicant_notifications'))
    except Exception as exc:
        db.rollback()
        print(f'❌ Mark all notifications read error: {exc}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': str(exc)}), 500
        flash('Failed to mark notifications as read.', 'error')
        return redirect(url_for('applicant_notifications'))
    finally:
        cursor.close()


@app.route('/applicant/notifications/<int:notification_id>/read', methods=['POST'])
@login_required('applicant')
def mark_notification_read(notification_id):
    """Mark a notification as read."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        if request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('applicant_notifications'))
    
    cursor = db.cursor()
    try:
        # Verify ownership
        cursor.execute(
            '''
            UPDATE notifications n
            JOIN applications a ON n.application_id = a.application_id
            SET n.is_read = TRUE
            WHERE n.notification_id = %s AND a.applicant_id = %s
            ''',
            (notification_id, applicant_id),
        )
        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True})
        flash('Notification marked as read.', 'success')
        return redirect(url_for('applicant_notifications'))
    except Exception as exc:
        db.rollback()
        print(f'❌ Mark notification read error: {exc}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': str(exc)}), 500
        flash('Failed to mark notification as read.', 'error')
        return redirect(url_for('applicant_notifications'))
    finally:
        cursor.close()


@app.route('/applicant/notifications/preferences', methods=['POST'])
@login_required('applicant')
def update_notification_preferences():
    """Update notification preferences."""
    applicant_id = session.get('user_id')
    email_enabled = request.form.get('email_enabled') == 'on'
    email_frequency = request.form.get('email_frequency', 'immediate')
    
    # Store preferences (could be in a separate table)
    # For now, just flash success
    flash('Notification preferences updated successfully.', 'success')
    return redirect(url_for('applicant_notifications'))


@app.route('/applicant/notifications/delete-all', methods=['POST'])
@login_required('applicant')
def delete_all_applicant_notifications():
    """Delete all notifications for the current applicant."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicant_notifications'))
    
    applicant_id = session.get('user_id')
    cursor = db.cursor()
    try:
        cursor.execute("SHOW TABLES LIKE 'notifications'")
        if not cursor.fetchone():
            flash('No notifications to delete.', 'info')
            return redirect(url_for('applicant_notifications'))
        
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns = {row[0] if isinstance(row, (list, tuple)) else row.get('Field') for row in (cursor.fetchall() or []) if row}
        if 'application_id' not in notification_columns:
            flash('Unable to associate notifications to your account.', 'error')
            return redirect(url_for('applicant_notifications'))
        
        cursor.execute(
            '''
            DELETE n FROM notifications n
            JOIN applications a ON n.application_id = a.application_id
            WHERE a.applicant_id = %s
            ''',
            (applicant_id,),
        )
        db.commit()
        flash('All notifications deleted.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        print(f'❌ Delete all applicant notifications error: {exc}')
        print(traceback.format_exc())
        flash('Failed to delete notifications.', 'error')
    finally:
        cursor.close()
    return redirect(url_for('applicant_notifications'))


@app.route('/applicant/account')
@login_required('applicant')
def applicant_account():
    """Account settings and security overview for applicants."""
    applicant_id = session.get('user_id')
    auth_user_id = session.get('auth_user_id')  # Get user_id from users table
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/account.html', account={}, login_history=[])

    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            '''
            SELECT full_name,
                   email,
                   phone_number,
                   created_at,
                   last_login
            FROM applicants
            WHERE applicant_id = %s
            LIMIT 1
            ''',
            (applicant_id,),
        )
        account = cursor.fetchone() or {}

        for key in ['created_at', 'last_login']:
            if account.get(key):
                account[key] = format_human_datetime(account.get(key))

        # Ensure schema compatibility
        ensure_schema_compatibility()
        
        # Check auth_sessions table columns
        cursor.execute('SHOW COLUMNS FROM auth_sessions')
        session_columns_raw = cursor.fetchall() or []
        session_columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in session_columns_raw if row}
        
        # Build logout_time expression
        if 'last_activity' in session_columns and 'logout_time' in session_columns:
            logout_expr = 'COALESCE(last_activity, logout_time)'
        elif 'logout_time' in session_columns:
            logout_expr = 'logout_time'
        elif 'last_activity' in session_columns:
            logout_expr = 'last_activity'
        else:
            logout_expr = 'NULL'
        
        # Use auth_user_id (from users table) for auth_sessions query
        # If auth_user_id is not available, try to get it from applicants table
        if not auth_user_id:
            cursor.execute(
                'SELECT user_id FROM applicants WHERE applicant_id = %s LIMIT 1',
                (applicant_id,),
            )
            user_record = cursor.fetchone()
            if user_record:
                auth_user_id = user_record.get('user_id')
        
        login_history = []
        active_sessions = 0
        
        if auth_user_id:
            try:
                # Check which columns exist in auth_sessions
                has_ip_address = 'ip_address' in session_columns
                has_user_agent = 'user_agent' in session_columns
                
                # Build SELECT statement dynamically based on available columns
                select_fields = ['login_time', f'{logout_expr} AS logout_time', 'COALESCE(is_active, 1) AS is_active']
                if has_ip_address:
                    select_fields.append("COALESCE(ip_address, 'Unknown') AS ip_address")
                else:
                    select_fields.append("'Unknown' AS ip_address")
                
                if has_user_agent:
                    select_fields.append("COALESCE(user_agent, 'Unknown') AS user_agent")
                else:
                    select_fields.append("'Unknown' AS user_agent")
                
                cursor.execute(
                    f'''
                    SELECT {', '.join(select_fields)}
                    FROM auth_sessions
                    WHERE user_id = %s
                    ORDER BY login_time DESC
                    LIMIT 10
                    ''',
                    (auth_user_id,),
                )
                sessions = cursor.fetchall() or []
                
                for row in sessions:
                    is_active = bool(row.get('is_active', 1))
                    if is_active:
                        active_sessions += 1
                    logout_value = format_human_datetime(row.get('logout_time')) if row.get('logout_time') else None
                    if is_active:
                        logout_value = None
                    login_history.append(
                        {
                            'login_time': format_human_datetime(row.get('login_time')),
                            'logout_time': logout_value,
                            'ip_address': row.get('ip_address') or '—',
                            'user_agent': row.get('user_agent') or 'Unknown device',
                            'status': 'Active' if is_active else 'Signed out',
                            'is_active': is_active,
                        }
                    )
            except Exception as session_error:
                print(f'⚠️ Error fetching login history: {session_error}')
                import traceback
                traceback.print_exc()
                login_history = []
                active_sessions = 0
        
        # Check notifications table for is_read column
        try:
            cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
            has_is_read = cursor.fetchone() is not None
            
            if has_is_read:
                notif_query = '''
                    SELECT 
                        COUNT(*) AS total,
                        SUM(CASE WHEN COALESCE(n.is_read, 0) = 0 THEN 1 ELSE 0 END) AS unread
                    FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    WHERE a.applicant_id = %s
                '''
            else:
                notif_query = '''
                    SELECT 
                        COUNT(*) AS total,
                        0 AS unread
                    FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    WHERE a.applicant_id = %s
                '''
            cursor.execute(notif_query, (applicant_id,))
        except Exception:
            # If notifications table doesn't exist or query fails, set defaults
            pass
        communications_stats = cursor.fetchone() or {'total': 0, 'unread': 0}

        account_overview = {
            'full_name': account.get('full_name'),
            'email': account.get('email'),
            'phone_number': account.get('phone_number'),
            'joined_at': account.get('created_at'),
            'last_login': account.get('last_login'),
            'active_sessions': active_sessions,
            'login_history': login_history,
            'communications': communications_stats,
        }

        # Fetch notification preferences (default values if table doesn't exist)
        preferences = {
            'email_enabled': True,
            'email_notifications': True,
            'email_frequency': 'immediate',
            'categories': ['application_updates', 'interview_alerts', 'system_messages'],
        }

        return render_template('applicant/account.html', account=account_overview, preferences=preferences)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant account settings error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Unable to load account settings: {str(exc)}', 'error')
        preferences = {
            'email_enabled': True,
            'email_notifications': True,
            'email_frequency': 'immediate',
        }
        return render_template('applicant/account.html', account={}, login_history=[], preferences=preferences)
    finally:
        if cursor:
            cursor.close()


@app.route('/applicant/upload-resume', methods=['POST'])
@login_required('applicant')
def upload_resume_before_apply():
    """Upload resume before submitting application - allows applicants to upload resume first."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        return jsonify({'success': False, 'error': 'Database connection error.'}), 500
    
    cursor = db.cursor(dictionary=True)
    try:
        resume_file = request.files.get('resume_file')
        if not resume_file or not resume_file.filename:
            return jsonify({'success': False, 'error': 'No file provided.'}), 400
        
        file_info, error = save_uploaded_file(resume_file, applicant_id)
        if not file_info:
            return jsonify({'success': False, 'error': error or 'Unable to process the uploaded resume file.'}), 400
        
        cursor.execute(
            '''
            INSERT INTO resumes (applicant_id, file_name, file_path, file_size_bytes)
            VALUES (%s, %s, %s, %s)
            ''',
            (
                applicant_id,
                file_info['original_filename'],
                file_info.get('storage_path') or file_info.get('file_path', ''),
                file_info['file_size'],
            ),
        )
        resume_id = cursor.lastrowid
        db.commit()
        
        return jsonify({
            'success': True,
            'resume_id': resume_id,
            'file_name': file_info['original_filename'],
            'message': 'Resume uploaded successfully!'
        }), 200
    except Exception as exc:
        db.rollback()
        print(f'❌ Error uploading resume: {exc}')
        return jsonify({'success': False, 'error': str(exc)}), 500
    finally:
        cursor.close()


@app.route('/applicant/apply/<int:job_id>', methods=['GET', 'POST'])
@app.route('/applicant/apply', methods=['GET', 'POST'])
@login_required('applicant')
def apply_to_job(job_id=None):
    """Show application form (GET) or submit application (POST) for a job."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('jobs'))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Check if editing existing application
        edit_application_id = request.args.get('edit', type=int)
        existing_app = None
        if edit_application_id:
            # Verify application belongs to applicant and hasn't been viewed
            cursor.execute(
                'SELECT application_id, job_id, viewed_at FROM applications WHERE application_id = %s AND applicant_id = %s LIMIT 1',
                (edit_application_id, applicant_id),
            )
            existing_app = cursor.fetchone()
            if not existing_app:
                flash('Application not found.', 'error')
                return redirect(url_for('applicant_applications'))
            if existing_app.get('viewed_at'):
                flash('This application cannot be edited because it has already been viewed by HR/Admin.', 'error')
                return redirect(url_for('applicant_applications'))
            # Use the job_id from existing application if no job_id provided
            if not job_id:
                job_id = existing_app.get('job_id')
        
        # When editing, job cannot be changed - always use the existing job_id
        # (This prevents applicants from changing the job when editing their application)
        if edit_application_id and existing_app:
            # Force use of existing job_id - ignore any job_id from form
            job_id = existing_app.get('job_id')
        
        # Check if already applied (for new applications only)
        if not edit_application_id and job_id:
            cursor.execute(
                'SELECT application_id FROM applications WHERE applicant_id = %s AND job_id = %s LIMIT 1',
                (applicant_id, job_id),
            )
            if cursor.fetchone():
                flash('You have already applied for this position.', 'warning')
                return redirect(url_for('jobs'))
        
        # Get all available jobs for dropdown (when editing)
        ensure_schema_compatibility()
        _update_job_columns(cursor)
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
        job_description_expr = job_column_expr('job_description', alternatives=['description'])
        job_requirements_expr = job_column_expr('job_requirements', alternatives=['requirements'])
        
        # Fetch all open/active jobs for the job selector
        status_placeholders = ','.join(['%s'] * len(PUBLISHABLE_JOB_STATUSES))
        cursor.execute(
            f'''
            SELECT 
                j.job_id,
                {job_title_expr} AS job_title,
                {job_description_expr} AS job_description,
                {job_requirements_expr} AS job_requirements,
                COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                {job_title_expr} AS position_title
            FROM jobs j
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE j.status IN ({status_placeholders})
            ORDER BY j.posted_at DESC, j.job_id DESC
            ''',
            tuple(PUBLISHABLE_JOB_STATUSES),
        )
        all_jobs = cursor.fetchall()
        
        # Get specific job details if job_id is provided
        job = None
        if job_id:
            job = next((j for j in all_jobs if j.get('job_id') == job_id), None)
            if not job:
                flash('Job not found or no longer available.', 'error')
                return redirect(url_for('jobs'))
        elif edit_application_id and existing_app:
            # If editing and no job_id, use the existing job
            job_id = existing_app.get('job_id')
            job = next((j for j in all_jobs if j.get('job_id') == job_id), None)
        
        if not job and not edit_application_id:
            flash('Please select a job to apply for.', 'error')
            return redirect(url_for('jobs'))
        
        # GET: Show application form
        if request.method == 'GET':
            # Get applicant's existing resumes
            cursor.execute(
                '''
                SELECT resume_id, file_name, uploaded_at, file_size_bytes
                FROM resumes
                WHERE applicant_id = %s
                ORDER BY uploaded_at DESC
                ''',
                (applicant_id,),
            )
            resumes = cursor.fetchall()
            
            # Format resume data
            formatted_resumes = []
            for resume in resumes:
                formatted_resumes.append({
                    'resume_id': resume.get('resume_id'),
                    'file_name': resume.get('file_name'),
                    'uploaded_at': format_human_datetime(resume.get('uploaded_at')),
                    'file_size': resume.get('file_size_bytes', 0),
                })
            
            return render_template(
                'applicant/apply.html',
                job=job,
                jobs=all_jobs,  # Pass all jobs for dropdown
                resumes=formatted_resumes,
                edit_application_id=edit_application_id,
            )
        
        # POST: Submit application
        resume_id = None
        
        # Handle resume file upload if provided
        resume_file = request.files.get('resume_file')
        if resume_file and resume_file.filename:
            file_info, error = save_uploaded_file(resume_file, applicant_id)
            if file_info:
                cursor.execute(
                    '''
                    INSERT INTO resumes (applicant_id, file_name, file_path, file_size_bytes)
                    VALUES (%s, %s, %s, %s)
                    ''',
                    (
                        applicant_id,
                        file_info['original_filename'],
                        file_info.get('storage_path') or file_info.get('file_path', ''),
                        file_info['file_size'],
                    ),
                )
                resume_id = cursor.lastrowid
            else:
                flash(error or 'Unable to process the uploaded resume file.', 'error')
                return redirect(url_for('apply_to_job', job_id=job_id))
        
        # If no file uploaded, use resume_id from form or get latest resume
        if not resume_id:
            resume_id_input = request.form.get('resume_id', '').strip()
            if resume_id_input:
                # Verify resume belongs to applicant
                cursor.execute(
                    'SELECT resume_id FROM resumes WHERE resume_id = %s AND applicant_id = %s',
                    (resume_id_input, applicant_id),
                )
                if cursor.fetchone():
                    resume_id = int(resume_id_input)
                else:
                    flash('Invalid resume selection.', 'error')
                    return redirect(url_for('apply_to_job', job_id=job_id))
            else:
                # Get latest resume if no selection
                cursor.execute(
                    '''
                    SELECT resume_id FROM resumes 
                    WHERE applicant_id = %s 
                    ORDER BY uploaded_at DESC 
                    LIMIT 1
                    ''',
                    (applicant_id,),
                )
                resume_record = cursor.fetchone()
                resume_id = resume_record.get('resume_id') if resume_record else None
        
        # When editing, always use the existing job_id - job cannot be changed
        if edit_application_id and existing_app:
            # Force use of existing job_id - prevent changing job when editing
            job_id = existing_app.get('job_id')
        
        # Create or update application
        if edit_application_id:
            # Update existing application (only if not viewed) - job cannot be changed
            # job_id is already set to existing_app.job_id above, so it will remain unchanged
            cursor.execute(
                '''
                UPDATE applications 
                SET job_id = %s, resume_id = %s, submitted_at = NOW() 
                WHERE application_id = %s AND applicant_id = %s AND viewed_at IS NULL
                ''',
                (job_id, resume_id, edit_application_id, applicant_id),
            )
            if cursor.rowcount == 0:
                flash('Application cannot be updated because it has been viewed by HR/Admin.', 'error')
                return redirect(url_for('applicant_applications'))
            application_id = edit_application_id
        else:
            # Create new application - status must be 'pending' (database enum: 'pending', 'scheduled', 'interviewed', 'hired', 'rejected')
            cursor.execute(
                '''
                INSERT INTO applications (applicant_id, job_id, resume_id, status, submitted_at)
                VALUES (%s, %s, %s, 'pending', NOW())
                ''',
                (applicant_id, job_id, resume_id),
            )
            application_id = cursor.lastrowid
        print(f'✅ Application created - ID: {application_id}, Job ID: {job_id}, Applicant ID: {applicant_id}, Status: pending')
        
        # AUTOMATIC: Notify applicant about successful submission
        if application_id:
            # Get applicant and job info for email and HR notification
            cursor.execute(
                '''
                SELECT ap.email, ap.full_name, ap.applicant_id,
                       j.title AS job_title, j.branch_id,
                       b.branch_name
                FROM applicants ap
                JOIN jobs j ON j.job_id = %s
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                WHERE ap.applicant_id = %s
                LIMIT 1
                ''',
                (job_id, applicant_id)
            )
            info = cursor.fetchone()
            
            # Send email to applicant (but don't create notification - HR will see their own notification)
            job_title = info.get('job_title') or info.get('job_title_alt') or 'the position'
            applicant_email = info.get('email')
            applicant_name = info.get('full_name') or 'Applicant'
            
            if applicant_email:
                email_subject = f'Application Submitted - {job_title}'
                email_body = f"""Dear {applicant_name},

Thank you for applying to the position: {job_title}

Your application has been successfully submitted. Our team will review your application and contact you soon.

Best regards,
J&T Express Recruitment Team
            """.strip()
            
                try:
                    send_email(applicant_email, email_subject, email_body)
                    print(f'✅ Confirmation email sent to applicant {applicant_email}')
                except Exception as email_err:
                    print(f'⚠️ Error sending confirmation email: {email_err}')
            
            # AUTOMATIC: Notify HR about new application (system notification + email)
            # HR will see this notification when they view branch-scoped notifications
            try:
                branch_id = info.get('branch_id')
                branch_name = info.get('branch_name') or 'a branch'
                hr_notification_message = f'{applicant_name} applied for {job_title} at {branch_name}.'
                
                # Create HR notification linked to application (HR will see it through branch-scoped queries)
                # HR notifications are fetched by joining notifications with applications and jobs by branch_id
                notification_columns = set()
                try:
                    cursor.execute('SHOW COLUMNS FROM notifications')
                    notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                except Exception:
                    pass
                
                # Create notification linked to application for HR visibility
                # HR will see this through branch-scoped queries (joining through applications -> jobs -> branch_id)
                # Applicants won't see this notification because they only see notifications starting with "You applied"
                # HR notifications use third-person format "{applicant_name} applied" which is filtered out in applicant queries
                if 'application_id' in notification_columns and 'message' in notification_columns:
                    # Check if notification already exists to prevent duplicates
                    cursor.execute(
                        '''
                        SELECT notification_id FROM notifications
                        WHERE application_id = %s AND message = %s
                        LIMIT 1
                        ''',
                        (application_id, hr_notification_message)
                    )
                    existing_notification = cursor.fetchone()
                    
                    if not existing_notification:
                        # Only create notification if it doesn't already exist
                        if 'sent_at' in notification_columns:
                            cursor.execute(
                                '''
                                INSERT INTO notifications (application_id, message, sent_at, is_read)
                                VALUES (%s, %s, NOW(), 0)
                                ''',
                                (application_id, hr_notification_message)
                            )
                        else:
                            cursor.execute(
                                '''
                                INSERT INTO notifications (application_id, message, is_read)
                                VALUES (%s, %s, 0)
                                ''',
                                (application_id, hr_notification_message)
                            )
                        print(f'✅ HR system notification created for application {application_id}')
                    else:
                        print(f'⚠️ HR notification already exists for application {application_id} - skipping duplicate')
                else:
                    print(f'⚠️ Notifications table missing required columns')
                
                # Send email to HR users managing this branch
                try:
                    # HR accounts manage all branches (branch_id column removed from admins table)
                    hr_query = '''
                        SELECT DISTINCT u.email, a.full_name
                        FROM users u
                        JOIN admins a ON a.user_id = u.user_id
                        WHERE u.user_type = 'hr' AND u.is_active = 1 AND a.is_active = 1
                    '''
                    cursor.execute(hr_query)
                    hr_users = cursor.fetchall()
                    
                    if hr_users:
                        # Send email to each HR user
                        email_subject = f'New Application Received - {job_title}'
                        email_body = f"""Dear HR Team,

A new application has been submitted:

Applicant: {applicant_name}
Position: {job_title}
Branch: {branch_name}
Application ID: {application_id}

Please log in to the HR portal to review this application.

Best regards,
J&T Express Recruitment System
                        """.strip()
                        
                        for hr_user in hr_users:
                            hr_email = hr_user.get('email')
                            if hr_email:
                                try:
                                    send_email(hr_email, email_subject, email_body)
                                    print(f'✅ HR notification email sent to {hr_email} for application {application_id}')
                                except Exception as email_err:
                                    print(f'⚠️ Error sending HR notification email to {hr_email}: {email_err}')
                        print(f'✅ HR notification emails sent to {len(hr_users)} HR user(s) for application {application_id}')
                    else:
                        print(f'⚠️ No HR users found for branch {branch_id} - no emails sent')
                except Exception as hr_email_err:
                    print(f'⚠️ Error sending HR notification emails: {hr_email_err}')
                    import traceback
                    traceback.print_exc()
                    
            except Exception as hr_notify_err:
                print(f'⚠️ Error creating HR notification: {hr_notify_err}')
                import traceback
                traceback.print_exc()
            
            # AUTOMATIC: Notify Admin about new application (system notification)
            # NOTE: Admin notification is skipped since HR notification above already covers this
            # Both HR and Admin users see the same notification, so we don't need to create a duplicate
            # The HR notification created above is sufficient for both HR and Admin users
            # This prevents duplicate notifications with the same message
        
        db.commit()
        flash('Application submitted successfully! You have been automatically notified.', 'success')
        return redirect(url_for('applicant_applications'))
    except Exception as exc:
        db.rollback()
        print(f'❌ Apply to job error: {exc}')
        import traceback
        traceback.print_exc()
        
        # Check if it's a CSRF error
        error_str = str(exc).lower()
        if 'csrf' in error_str or 'token' in error_str:
            flash('Security token expired. Please refresh the page and try again.', 'error')
        else:
            flash('Unable to submit application. Please try again.', 'error')
        return redirect(url_for('apply_to_job', job_id=job_id))
    finally:
        cursor.close()


@app.route('/applicant/applications/<int:application_id>/delete', methods=['POST'])
@login_required('applicant')
def delete_application(application_id):
    """Permanently delete a job application and all related data."""
    applicant_id = session.get('user_id')
    if not applicant_id:
        flash('Please log in to delete applications.', 'error')
        return immediate_redirect(url_for('login', _external=True))
    
    # Validate applicant_id is an integer
    try:
        applicant_id = int(applicant_id)
    except (ValueError, TypeError):
        flash('Unable to identify your account. Please log in again.', 'error')
        return immediate_redirect(url_for('login', _external=True))
    
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return immediate_redirect(url_for('applicant_applications', _external=True))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Verify application belongs to applicant
        cursor.execute(
            '''
            SELECT application_id, status, viewed_at, job_id
            FROM applications
            WHERE application_id = %s AND applicant_id = %s
            LIMIT 1
            ''',
            (application_id, applicant_id),
        )
        application = cursor.fetchone()
        
        if not application:
            flash('Application not found or you do not have permission to delete it.', 'error')
            return immediate_redirect(url_for('applicant_applications', _external=True))
        
        # Check if application has been viewed by HR/Admin - warn but allow deletion
        if application.get('viewed_at'):
            # Still allow deletion but warn the user
            pass
        
        # Delete related records first to avoid foreign key constraints
        try:
            # Delete interviews related to this application
            cursor.execute('DELETE FROM interviews WHERE application_id = %s', (application_id,))
            print(f'✅ Deleted interviews for application {application_id}')
        except Exception as interview_error:
            print(f'⚠️ Error deleting interviews: {interview_error}')
            # Continue anyway
        
        try:
            # Delete notifications related to this application
            cursor.execute('DELETE FROM notifications WHERE application_id = %s', (application_id,))
            print(f'✅ Deleted notifications for application {application_id}')
        except Exception as notif_error:
            print(f'⚠️ Error deleting notifications: {notif_error}')
            # Continue anyway
        
        
        # Finally, delete the application itself
        cursor.execute('DELETE FROM applications WHERE application_id = %s AND applicant_id = %s', (application_id, applicant_id))
        
        deleted_count = cursor.rowcount
        if deleted_count > 0:
            db.commit()
            flash('Application permanently deleted successfully.', 'success')
            print(f'✅ Application {application_id} deleted by applicant {applicant_id}')
        else:
            db.rollback()
            flash('Failed to delete application. Please try again.', 'error')
        
        return immediate_redirect(url_for('applicant_applications', _external=True))
    except Exception as exc:
        db.rollback()
        print(f'❌ Delete application error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to delete application. Please try again later.', 'error')
        return immediate_redirect(url_for('applicant_applications', _external=True))
    finally:
        cursor.close()


@app.route('/applicant/jobs/<int:job_id>/save', methods=['POST'])
@login_required('applicant')
def save_job(job_id):   
    """Save a job posting."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('jobs'))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Check if saved_jobs table exists, if not create it
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS saved_jobs (
                saved_job_id INT AUTO_INCREMENT PRIMARY KEY,
                applicant_id INT NOT NULL,
                job_id INT NOT NULL,
                saved_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY unique_saved_job (applicant_id, job_id),
                FOREIGN KEY (applicant_id) REFERENCES applicants(applicant_id) ON DELETE CASCADE,
                FOREIGN KEY (job_id) REFERENCES jobs(job_id) ON DELETE CASCADE
            )
        """)
        
        # Check if already saved - use applicant_id and job_id instead of saved_job_id
        cursor.execute(
            'SELECT applicant_id FROM saved_jobs WHERE applicant_id = %s AND job_id = %s',
            (applicant_id, job_id),
        )
        if cursor.fetchone():
            flash('Job already saved.', 'info')
        else:
            cursor.execute(
                'INSERT INTO saved_jobs (applicant_id, job_id) VALUES (%s, %s)',
                (applicant_id, job_id),
            )
            db.commit()
            flash('Job saved successfully.', 'success')
        
        return redirect(url_for('saved_jobs', just_saved=job_id))
    except Exception as exc:
        db.rollback()
        print(f'❌ Save job error: {exc}')
        flash('Unable to save job. Please try again.', 'error')
        return redirect(url_for('jobs'))
    finally:
        cursor.close()


@app.route('/applicant/jobs/<int:job_id>/unsave', methods=['POST'])
@login_required('applicant')
def unsave_job(job_id):
    """Unsave a job posting."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('jobs'))
    
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            'DELETE FROM saved_jobs WHERE applicant_id = %s AND job_id = %s',
            (applicant_id, job_id),
        )
        db.commit()
        flash('Job removed from saved list.', 'success')
        return redirect(request.referrer or url_for('jobs'))
    except Exception as exc:
        db.rollback()
        print(f'❌ Unsave job error: {exc}')
        flash('Unable to remove saved job. Please try again.', 'error')
        return redirect(url_for('jobs'))
    finally:
        cursor.close()


@app.route('/applicant/jobs/saved')
@login_required('applicant')
def saved_jobs():
    """View saved job postings."""
    filters = {
        'keyword': request.args.get('keyword', '').strip(),
        'branch_id': request.args.get('branch_id', type=int),
        'position_id': request.args.get('position_id', type=int),
    }
    filters = {k: v for k, v in filters.items() if v}

    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/jobs.html', jobs=[], branches=[], positions=[], current_filters=filters, saved_mode=True)
    
    cursor = db.cursor(dictionary=True)
    try:
        # Check if saved_jobs table exists
        cursor.execute("SHOW TABLES LIKE 'saved_jobs'")
        if not cursor.fetchone():
            flash('No saved jobs found.', 'info')
            return render_template('applicant/jobs.html', jobs=[], branches=[], positions=[], current_filters=filters, saved_mode=True)
        
        # Check the schema of saved_jobs table
        cursor.execute("DESCRIBE saved_jobs")
        saved_columns = [col['Field'] for col in cursor.fetchall()]
        print(f"Saved jobs columns: {saved_columns}")  # Debug info
        
        # Ensure schema compatibility for jobs table
        ensure_schema_compatibility()
        
        # Fetch saved jobs - use dynamic column checking
        _update_job_columns(cursor)
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
        salary_currency_expr = job_column_expr('salary_currency', default="'PHP'")
        salary_min_expr = job_column_expr('salary_min', default='NULL')
        salary_max_expr = job_column_expr('salary_max', default='NULL')
        
        job_location_expr = job_column_expr('job_location', alternatives=['location'], default='NULL')
        
        job_description_expr = job_column_expr('job_description', alternatives=['description'], default='NULL')
        employment_type_expr = job_column_expr('employment_type', default='NULL')
        work_arrangement_expr = job_column_expr('work_arrangement', default='NULL')
        experience_level_expr = job_column_expr('experience_level', default='NULL')
        application_deadline_expr = job_column_expr('application_deadline', default='NULL')
        status_expr = job_column_expr('status', default="'open'")
        branch_id_expr = job_column_expr('branch_id', default='NULL')
        position_id_expr = job_column_expr('position_id', default='NULL')
        
        # Query for saved jobs - removed all references to saved_job_id
        cursor.execute(
            f'''
            SELECT j.job_id, {job_title_expr} AS job_title, {job_description_expr} AS job_description, 
                   {employment_type_expr} AS employment_type, 
                   {work_arrangement_expr} AS work_arrangement,
                   {experience_level_expr} AS experience_level, 
                   {job_location_expr} AS job_location, 
                   {salary_min_expr} AS salary_min, 
                   {salary_max_expr} AS salary_max, 
                   {salary_currency_expr} AS salary_currency,
                   {application_deadline_expr} AS application_deadline, 
                   {status_expr} AS status, 
                   {branch_id_expr} AS branch_id,
                   {position_id_expr} AS position_id,
                   COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                   {job_title_expr} AS position_name,
                   'General' AS department,
                   sj.saved_at
            FROM saved_jobs sj
            JOIN jobs j ON sj.job_id = j.job_id
            LEFT JOIN branches b ON {branch_id_expr} = b.branch_id
            WHERE sj.applicant_id = %s
            ORDER BY sj.saved_at DESC
            ''',
            (applicant_id,),
        )
        saved_jobs_list = cursor.fetchall()
        
        # Format jobs similar to fetch_open_jobs
        jobs = []
        for job in saved_jobs_list:
            jobs.append({
                'job_id': job.get('job_id'),
                'title': job.get('job_title'),
                'job_title': job.get('job_title'),
                'summary': (job.get('job_description') or '')[:200] if job.get('job_description') else '',
                'description': job.get('job_description'),
                'employment_type': job.get('employment_type'),
                'work_arrangement': job.get('work_arrangement'),
                'experience_level': job.get('experience_level'),
                'location': job.get('job_location'),
                'job_location': job.get('job_location'),
                'salary_min': job.get('salary_min'),
                'salary_max': job.get('salary_max'),
                'salary_currency': job.get('salary_currency'),
                'salary_display': format_salary_range(job.get('salary_currency'), job.get('salary_min'), job.get('salary_max')),
                'application_deadline': job.get('application_deadline'),
                'status': job.get('status'),
                'branch_name': job.get('branch_name'),
                'branch_id': job.get('branch_id'),
                'position_name': job.get('position_name'),
                'position_id': job.get('position_id'),
                'department': job.get('department'),
                'saved_at': job.get('saved_at'),
                'is_saved': True,
            })

        def matches_filters(job_data):
            if filters.get('keyword'):
                keyword = filters['keyword'].lower()
                haystacks = [
                    (job_data.get('title') or '').lower(),
                    (job_data.get('job_title') or '').lower(),
                    (job_data.get('summary') or '').lower(),
                    (job_data.get('description') or '').lower(),
                    (job_data.get('branch_name') or '').lower(),
                    (job_data.get('position_name') or '').lower(),
                ]
                if not any(keyword in text for text in haystacks):
                    return False
            if filters.get('branch_id') and str(filters['branch_id']) != str(job_data.get('branch_id') or ''):
                return False
            if filters.get('position_id') and str(filters['position_id']) != str(job_data.get('position_id') or ''):
                return False
            return job_data.get('status') in PUBLISHABLE_JOB_STATUSES

        jobs = [job for job in jobs if matches_filters(job)]
        
        branches = fetch_branches()
        positions = fetch_positions()
        
        return render_template(
            'applicant/jobs.html',
            jobs=jobs,
            branches=branches,
            positions=positions,
            current_filters=filters,
            saved_mode=True,
        )
    except Exception as exc:
        db.rollback()
        print(f'❌ Saved jobs error: {exc}')
        flash('Unable to load saved jobs.', 'error')
        return render_template('applicant/jobs.html', jobs=[], branches=[], positions=[], current_filters=filters, saved_mode=True)
    finally:
        cursor.close()


@app.route('/applicant/profile', methods=['GET', 'POST'])
@login_required('applicant')
def applicant_profile():
    """Applicant profile management with resume upload."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/profile.html', applicant=None, resumes=[], login_history=[], profile_history=[])

    cursor = db.cursor(dictionary=True)
    try:
        # Ensure schema compatibility
        ensure_schema_compatibility()
        
        cursor.execute(
            '''
            SELECT applicant_id, full_name, email, phone_number, password_hash,
                   last_login, created_at
            FROM applicants
            WHERE applicant_id = %s
            LIMIT 1
            ''',
            (applicant_id,),
        )
        applicant_record = cursor.fetchone()

        if not applicant_record:
            flash('Unable to load your profile. Please contact support.', 'error')
            return redirect(url_for('logout'))

        stored_password_hash = applicant_record.get('password_hash')

        if request.method == 'POST':
            action = request.form.get('action', 'update_profile')

            if action == 'update_profile':
                full_name = request.form.get('full_name', '').strip() or applicant_record.get('full_name')
                phone_number = request.form.get('phone_number', '').strip() or applicant_record.get('phone_number')
                email = request.form.get('email', '').strip().lower() or applicant_record.get('email')
                resume = request.files.get('resume')

                if not all([full_name, phone_number, email]):
                    flash('Full name, phone number, and email are required.', 'error')
                    return redirect(url_for('applicant_profile'))

                email_changed = email != applicant_record.get('email')
                verification_token = None
                token_expires = None

                if email_changed:
                    cursor.execute(
                        'SELECT applicant_id FROM applicants WHERE email = %s AND applicant_id <> %s LIMIT 1',
                        (email, applicant_id),
                    )
                    if cursor.fetchone():
                        flash('That email address is already in use. Please choose another.', 'error')
                        return redirect(url_for('applicant_profile'))
                    from datetime import datetime, timedelta
                    verification_token = generate_token()
                    token_expires = datetime.now() + timedelta(seconds=60)

                # Log changes before updating
                if full_name != applicant_record.get('full_name'):
                    log_profile_change(applicant_id, 'applicant', 'full_name', applicant_record.get('full_name'), full_name)
                if phone_number != applicant_record.get('phone_number'):
                    log_profile_change(applicant_id, 'applicant', 'phone_number', applicant_record.get('phone_number'), phone_number)
                if email_changed:
                    log_profile_change(applicant_id, 'applicant', 'email', applicant_record.get('email'), email)

                # Always update all fields to ensure changes are saved
                # Check if last_profile_update column exists before using it
                cursor.execute('SHOW COLUMNS FROM applicants LIKE "last_profile_update"')
                has_last_profile_update = cursor.fetchone() is not None
                
                if email_changed:
                    if has_last_profile_update:
                        cursor.execute(
                            '''
                            UPDATE applicants
                            SET full_name = %s,
                                phone_number = %s,
                                email = %s,
                                verification_token = %s,
                                verification_token_expires = %s,
                                last_profile_update = NOW()
                            WHERE applicant_id = %s
                            ''',
                            (full_name, phone_number, email, verification_token, token_expires, applicant_id),
                        )
                    else:
                        cursor.execute(
                            '''
                            UPDATE applicants
                            SET full_name = %s,
                                phone_number = %s,
                                email = %s,
                                verification_token = %s,
                                verification_token_expires = %s
                            WHERE applicant_id = %s
                            ''',
                            (full_name, phone_number, email, verification_token, token_expires, applicant_id),
                        )
                else:
                    if has_last_profile_update:
                        cursor.execute(
                            '''
                            UPDATE applicants
                            SET full_name = %s,
                                phone_number = %s,
                                last_profile_update = NOW()
                            WHERE applicant_id = %s
                            ''',
                            (full_name, phone_number, applicant_id),
                        )
                    else:
                        cursor.execute(
                            '''
                            UPDATE applicants
                            SET full_name = %s,
                                phone_number = %s
                            WHERE applicant_id = %s
                            ''',
                            (full_name, phone_number, applicant_id),
                        )
                
                # Verify the update was successful
                rows_affected = cursor.rowcount
                if rows_affected == 0:
                    flash('Failed to update profile. Please try again.', 'error')
                    db.rollback()
                    return redirect(url_for('applicant_profile'))

                resume_uploaded = False
                if resume and resume.filename:
                    file_info, error = save_uploaded_file(resume, applicant_id)
                    if file_info:
                        # Delete old resumes to ensure only one resume exists
                        cursor.execute(
                            'SELECT resume_id, file_path FROM resumes WHERE applicant_id = %s',
                            (applicant_id,)
                        )
                        old_resumes = cursor.fetchall()
                        for old_resume in old_resumes:
                            old_file_path = old_resume.get('file_path') if isinstance(old_resume, dict) else old_resume[1]
                            try:
                                import os
                                if old_file_path and os.path.exists(old_file_path):
                                    os.remove(old_file_path)
                            except Exception as del_err:
                                print(f'⚠️ Error deleting old resume file: {del_err}')
                        # Delete old resume records from database
                        cursor.execute(
                            'DELETE FROM resumes WHERE applicant_id = %s',
                            (applicant_id,)
                        )
                        # Insert new resume
                        cursor.execute(
                            '''
                            INSERT INTO resumes (applicant_id, file_name, file_path, file_size_bytes)
                            VALUES (%s, %s, %s, %s)
                            ''',
                            (
                                applicant_id,
                                file_info['original_filename'],
                                file_info.get('storage_path') or file_info.get('file_path', ''),
                                file_info['file_size'],
                            ),
                        )
                        log_profile_change(
                            applicant_id,
                            'applicant',
                            'resume',
                            'previous_upload',
                            file_info.get('storage_path') or file_info.get('file_path', ''),
                        )
                        resume_uploaded = True
                    else:
                        flash(error or 'Unable to process the uploaded file.', 'error')

                # Commit the transaction to ensure changes are saved
                try:
                    db.commit()
                    print(f'✅ Profile updated successfully for applicant {applicant_id}: name={full_name}, phone={phone_number}, email={email}')
                except Exception as commit_error:
                    db.rollback()
                    print(f'❌ Error committing profile update: {commit_error}')
                    flash('Failed to save profile changes. Please try again.', 'error')
                    return redirect(url_for('applicant_profile'))

                # Update session with new values immediately
                session['user_name'] = full_name
                session['user_email'] = email

                # AUTOMATIC: Notify applicant via system notification and email about profile update
                try:
                    # Get any application_id for this applicant (for notification linking)
                    cursor.execute(
                        'SELECT application_id FROM applications WHERE applicant_id = %s LIMIT 1',
                        (applicant_id,)
                    )
                    app_record = cursor.fetchone()
                    application_id = app_record.get('application_id') if app_record else None
                    
                    # Create notification message and email
                    notification_message = 'Your profile information has been updated successfully.'
                    email_subject = 'Profile Updated - J&T Express'
                    email_body = f"""Dear {full_name},

Your profile information has been successfully updated.

Updated Information:
- Name: {full_name}
- Email: {email}
- Phone: {phone_number}
{'• Resume: Uploaded' if resume_uploaded else ''}

If you did not make this change, please contact our support team immediately.

Best regards,
J&T Express Recruitment Team
                    """.strip()
                    
                    # Use auto_notify_and_email function if we have an application_id
                    if application_id:
                        auto_notify_and_email(
                            cursor, application_id, notification_message,
                            email_subject, email_body,
                            email, full_name
                        )
                    else:
                        # Just send email if no application_id
                        send_email(email, email_subject, email_body)
                    
                    print(f'✅ Profile update notification sent to applicant {email}')
                except Exception as notify_err:
                    print(f'⚠️ Error sending profile update notification: {notify_err}')

                if resume_uploaded:
                    flash('Resume uploaded successfully.', 'success')

                if email_changed and verification_token:
                    # Get applicant name for personalized email
                    cursor.execute(
                        '''
                        SELECT full_name FROM applicants WHERE applicant_id = %s LIMIT 1
                        ''',
                        (applicant_id,),
                    )
                    applicant_name_record = cursor.fetchone()
                    name = applicant_name_record.get('full_name') if applicant_name_record else None
                    send_verification_email(email, verification_token, applicant_name=name)
                    session['pending_verification_email'] = email
                    flash('Profile updated successfully. Please verify your new email address from your inbox.', 'success')
                else:
                    flash('Profile updated successfully. You have been notified via email.', 'success')

                # Reload the page to show updated information
                return redirect(url_for('applicant_profile'))

            if action == 'change_password':
                current_password = request.form.get('current_password', '').strip()
                new_password = request.form.get('new_password', '').strip()
                confirm_password = request.form.get('confirm_password', '').strip()

                if not all([current_password, new_password, confirm_password]):
                    flash('Please complete all password fields.', 'error')
                    return redirect(url_for('applicant_profile'))

                if not check_password(stored_password_hash, current_password):
                    flash('Current password is incorrect.', 'error')
                    return redirect(url_for('applicant_profile'))

                if len(new_password) < 6:
                    flash('New password must be at least 6 characters.', 'error')
                    return redirect(url_for('applicant_profile'))

                if new_password != confirm_password:
                    flash('New passwords do not match.', 'error')
                    return redirect(url_for('applicant_profile'))

                new_hash = hash_password(new_password)
                cursor.execute(
                    '''
                    UPDATE applicants
                    SET password_hash = %s
                    WHERE applicant_id = %s
                    ''',
                    (new_hash, applicant_id),
                )
                cursor.execute(
                    'DELETE FROM password_resets WHERE user_email = %s AND role = %s',
                    (applicant_record.get('email'), 'applicant'),
                )
                db.commit()
                log_profile_change(applicant_id, 'applicant', 'password', '[updated]', '[updated]')
                
                # AUTOMATIC: Notify applicant via system notification and email about password change
                try:
                    applicant_email = applicant_record.get('email')
                    applicant_name = applicant_record.get('full_name') or 'Applicant'
                    
                    # Get any application_id for this applicant (for notification linking)
                    cursor.execute(
                        'SELECT application_id FROM applications WHERE applicant_id = %s LIMIT 1',
                        (applicant_id,)
                    )
                    app_record = cursor.fetchone()
                    application_id = app_record.get('application_id') if app_record else None
                    
                    # Create notification message and email
                    notification_message = 'Your account password has been changed successfully.'
                    email_subject = 'Password Changed - J&T Express'
                    email_body = f"""Dear {applicant_name},

Your account password has been successfully changed.

If you did not make this change, please contact our support team immediately to secure your account.

Best regards,
J&T Express Recruitment Team
                    """.strip()
                    
                    # Use auto_notify_and_email function if we have an application_id
                    if application_id:
                        auto_notify_and_email(
                            cursor, application_id, notification_message,
                            email_subject, email_body,
                            applicant_email, applicant_name
                        )
                    else:
                        # Just send email if no application_id
                        send_email(applicant_email, email_subject, email_body)
                    
                    print(f'✅ Password change notification sent to applicant {applicant_email}')
                except Exception as notify_err:
                    print(f'⚠️ Error sending password change notification: {notify_err}')
                
                flash('Password updated successfully. You have been notified via email.', 'success')
                return redirect(url_for('applicant_profile'))

            if action == 'delete_account':
                try:
                    # Get user_id and applicant info from applicants table before deleting (needed to delete from users table and create notification)
                    cursor.execute(
                        'SELECT user_id, email, full_name FROM applicants WHERE applicant_id = %s LIMIT 1',
                        (applicant_id,),
                    )
                    applicant_user_record = cursor.fetchone()
                    user_id = applicant_user_record.get('user_id') if applicant_user_record else None
                    applicant_email = applicant_user_record.get('email') if applicant_user_record else None
                    applicant_name = applicant_user_record.get('full_name') if applicant_user_record else 'Unknown Applicant'
                    
                    if not user_id:
                        print(f'⚠️ WARNING: No user_id found for applicant {applicant_id} (email: {applicant_email})')
                        print(f'⚠️ This applicant may not be linked to a users table record')
                    else:
                        print(f'🔍 Found user_id={user_id} for applicant {applicant_id} (email: {applicant_email})')
                    
                    # Create admin notification before deleting applicant
                    try:
                        admin_notification_msg = f'Applicant "{applicant_name}" (Email: {applicant_email}) has deleted their account.'
                        create_admin_notification(cursor, admin_notification_msg)
                        print(f'✅ Admin notification created for applicant account deletion: {applicant_name}')
                    except Exception as notify_err:
                        print(f'⚠️ Error creating admin notification for applicant deletion: {notify_err}')
                    
                    # Delete related data first (in order to avoid foreign key constraints)
                    # 1. Delete notifications linked to applicant's applications
                    cursor.execute("SHOW TABLES LIKE 'notifications'")
                    if cursor.fetchone():
                        cursor.execute(
                            '''
                            DELETE FROM notifications
                            WHERE application_id IN (
                                SELECT application_id FROM applications WHERE applicant_id = %s
                            )
                            ''',
                            (applicant_id,),
                        )
                    
                    # 2. Delete password reset tokens
                    applicant_email = applicant_record.get('email') if applicant_record else None
                    if applicant_email:
                        cursor.execute(
                            'DELETE FROM password_resets WHERE user_email = %s AND role = %s',
                            (applicant_email, 'applicant'),
                        )
                    
                    # 3. Delete auth sessions (use user_id if available, otherwise use applicant_id)
                    cursor.execute("SHOW TABLES LIKE 'auth_sessions'")
                    if cursor.fetchone():
                        if user_id:
                            cursor.execute(
                                'DELETE FROM auth_sessions WHERE user_id = %s',
                                (user_id,),
                            )
                        else:
                            cursor.execute(
                                'DELETE FROM auth_sessions WHERE user_id = %s AND role = %s',
                                (applicant_id, 'applicant'),
                            )
                    
                    # 4. Delete profile changes history
                    cursor.execute("SHOW TABLES LIKE 'profile_changes'")
                    if cursor.fetchone():
                        if user_id:
                            cursor.execute(
                                'DELETE FROM profile_changes WHERE user_id = %s AND role = %s',
                                (user_id, 'applicant'),
                            )
                        else:
                            cursor.execute(
                                'DELETE FROM profile_changes WHERE user_id = %s AND role = %s',
                                (applicant_id, 'applicant'),
                            )
                    
                    # 5. Delete saved jobs
                    cursor.execute("SHOW TABLES LIKE 'saved_jobs'")
                    if cursor.fetchone():
                        cursor.execute(
                            'DELETE FROM saved_jobs WHERE applicant_id = %s',
                            (applicant_id,),
                        )
                    
                    # 6. Delete interviews related to applicant's applications
                    cursor.execute("SHOW TABLES LIKE 'interviews'")
                    if cursor.fetchone():
                        cursor.execute(
                            'DELETE FROM interviews WHERE application_id IN (SELECT application_id FROM applications WHERE applicant_id = %s)',
                            (applicant_id,),
                        )
                        print(f'✅ Deleted interviews for applicant {applicant_id}')
                    
                    # 7. Delete applications explicitly
                    cursor.execute(
                        'DELETE FROM applications WHERE applicant_id = %s',
                        (applicant_id,),
                    )
                    applications_deleted = cursor.rowcount
                    print(f'✅ Deleted {applications_deleted} application(s) for applicant {applicant_id}')
                    
                    # 8. Delete resumes explicitly
                    cursor.execute("SHOW TABLES LIKE 'resumes'")
                    if cursor.fetchone():
                        cursor.execute(
                            'DELETE FROM resumes WHERE applicant_id = %s',
                            (applicant_id,),
                        )
                        resumes_deleted = cursor.rowcount
                        print(f'✅ Deleted {resumes_deleted} resume(s) for applicant {applicant_id}')
                    
                    # 9. Delete the applicant record from database
                    cursor.execute(
                        'DELETE FROM applicants WHERE applicant_id = %s',
                        (applicant_id,),
                    )
                    applicant_deleted = cursor.rowcount > 0
                    print(f'✅ Deleted applicant record {applicant_id} from applicants table' if applicant_deleted else f'⚠️ No applicant record found with ID {applicant_id}')
                    
                    # 10. Finally, delete the user record from users table to remove from system users
                    # IMPORTANT: Delete from users table BEFORE committing to ensure it's removed from system users
                    if user_id:
                        try:
                            # First verify the user exists
                            cursor.execute(
                                'SELECT user_id, email, user_type FROM users WHERE user_id = %s AND user_type = %s',
                                (user_id, 'applicant'),
                            )
                            user_check = cursor.fetchone()
                            if user_check:
                                print(f'🔍 Found user record to delete: user_id={user_id}, email={user_check.get("email")}')
                                # Delete from users table
                                cursor.execute(
                                    'DELETE FROM users WHERE user_id = %s AND user_type = %s',
                                    (user_id, 'applicant'),
                                )
                                user_deleted = cursor.rowcount > 0
                                if user_deleted:
                                    print(f'✅ Successfully deleted user record {user_id} from users table')
                                else:
                                    print(f'⚠️ DELETE query executed but no rows affected for user_id {user_id}')
                            else:
                                print(f'⚠️ User record not found with user_id={user_id} and user_type=applicant')
                        except Exception as user_delete_error:
                            import traceback
                            print(f'❌ Error deleting user record {user_id}: {user_delete_error}')
                            print(f'❌ Traceback: {traceback.format_exc()}')
                            # Rollback and re-raise to prevent partial deletion
                            db.rollback()
                            raise
                    else:
                        print(f'⚠️ No user_id found for applicant {applicant_id} - cannot delete from users table')
                    
                    # Commit all deletions together
                    db.commit()
                    print(f'✅ Account deletion completed and committed for applicant {applicant_id}')
                    
                    # Close database cursor before clearing session
                    cursor.close()
                    
                    # Clear session completely before logout
                    logout_user()
                    
                    # Force clear ALL session data to prevent redirect loops
                    session.clear()
                    for key in list(session.keys()):
                        session.pop(key, None)
                    
                    # Return an HTML page with immediate JavaScript redirect to prevent redirect page
                    login_url = url_for('login', _external=True)
                    html_content = f'''<!DOCTYPE html>
<html>
<head>
    <meta http-equiv="refresh" content="0; url={login_url}">
    <script>window.location.replace("{login_url}");</script>
</head>
<body>
    <script>window.location.replace("{login_url}");</script>
</body>
</html>'''
                    from flask import Response
                    return Response(html_content, status=302, headers={'Location': login_url, 'Content-Type': 'text/html; charset=utf-8'})
                except Exception as delete_exc:
                    db.rollback()
                    import traceback
                    print(f'❌ Delete account error: {delete_exc}')
                    traceback.print_exc()
                    flash('Failed to delete account. Please contact support.', 'error')
                    return redirect(url_for('applicant_profile'))

            if action == 'terminate_sessions':
                current_session_id = session.get('auth_session_id')
                if current_session_id:
                    cursor.execute(
                        '''
                        UPDATE auth_sessions
                        SET logout_time = NOW(), is_active = 0
                        WHERE user_id = %s
                          AND role = 'applicant'
                          AND session_id <> %s
                        ''',
                        (applicant_id, current_session_id),
                    )
                else:
                    cursor.execute(
                        '''
                        UPDATE auth_sessions
                        SET logout_time = NOW(), is_active = 0
                        WHERE user_id = %s
                          AND role = 'applicant'
                        ''',
                        (applicant_id,),
                    )
                db.commit()
                flash('All other active sessions have been signed out.', 'success')
                return redirect(url_for('applicant_profile'))
            
            if action == 'delete_resume':
                resume_id = request.form.get('resume_id')
                if not resume_id:
                    flash('Invalid resume selection.', 'error')
                    return redirect(url_for('applicant_profile'))

                cursor.execute(
                    '''
                    SELECT file_path
                    FROM resumes
                    WHERE resume_id = %s AND applicant_id = %s
                    LIMIT 1
                    ''',
                    (resume_id, applicant_id),
                )
                resume_record = cursor.fetchone()

                if not resume_record:
                    flash('Resume not found or already removed.', 'warning')
                    return redirect(url_for('applicant_profile'))

                # Detach resume from any applications to avoid FK constraints
                try:
                    cursor.execute(
                        '''
                        UPDATE applications
                        SET resume_id = NULL
                        WHERE applicant_id = %s AND resume_id = %s
                        ''',
                        (applicant_id, resume_id),
                    )
                except Exception as fk_exc:
                    # Proceed even if this optional step fails; deletion might still succeed without FK
                    print(f'⚠️ Could not detach resume {resume_id} from applications: {fk_exc}')

                cursor.execute('DELETE FROM resumes WHERE resume_id = %s', (resume_id,))
                db.commit()

                file_path = resume_record.get('file_path')
                if file_path:
                    absolute_path = os.path.join(app.root_path, file_path)
                    try:
                        if os.path.exists(absolute_path):
                            os.remove(absolute_path)
                    except OSError as err:
                        print(f'⚠️ Failed to delete resume file {absolute_path}: {err}')

                log_profile_change(applicant_id, 'applicant', 'resume', file_path or '', 'deleted')
                flash('Resume removed successfully.', 'success')
                return redirect(url_for('applicant_profile'))

        # GET and post-processing context
        # Re-fetch applicant data to ensure we have the latest information (important after updates)
        # Check if last_profile_update column exists before selecting it
        cursor.execute('SHOW COLUMNS FROM applicants LIKE "last_profile_update"')
        has_last_profile_update = cursor.fetchone() is not None
        
        if has_last_profile_update:
            cursor.execute(
                '''
                SELECT applicant_id, full_name, email, phone_number, password_hash,
                       last_login, created_at, last_profile_update
                FROM applicants
                WHERE applicant_id = %s
                LIMIT 1
                ''',
                (applicant_id,),
            )
        else:
            cursor.execute(
                '''
                SELECT applicant_id, full_name, email, phone_number, password_hash,
                       last_login, created_at
                FROM applicants
                WHERE applicant_id = %s
                LIMIT 1
                ''',
                (applicant_id,),
            )
        fresh_applicant_record = cursor.fetchone() or applicant_record
        
        cursor.execute(
            '''
            SELECT resume_id,
                   file_name,
                   file_path,
                   file_size_bytes,
                   uploaded_at
            FROM resumes
            WHERE applicant_id = %s
            ORDER BY uploaded_at DESC
            ''',
            (applicant_id,),
        )
        resumes = cursor.fetchall() or []

        # Use fresh data for display
        applicant = {key: value for key, value in fresh_applicant_record.items() if key != 'password_hash'}
        for date_field in ['last_login', 'created_at']:
            if date_field in applicant:
                applicant[date_field] = format_human_datetime(applicant.get(date_field))

        # Check auth_sessions table columns
        cursor.execute('SHOW COLUMNS FROM auth_sessions')
        session_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        
        # Build logout_time expression
        if 'last_activity' in session_columns:
            logout_expr = 'COALESCE(last_activity, logout_time)'
        elif 'logout_time' in session_columns:
            logout_expr = 'logout_time'
        else:
            logout_expr = 'NULL'
        
        login_rows = fetch_rows(
            f'''
            SELECT session_id, login_time, {logout_expr} AS logout_time, COALESCE(is_active, 1) AS is_active
            FROM auth_sessions
            WHERE user_id = %s
            ORDER BY login_time DESC
            LIMIT 10
            ''',
            (applicant_id,),
        )
        active_session_id = session.get('auth_session_id')
        login_history = []
        for row in login_rows:
            is_active = bool(row.get('is_active', 1))
            logout_value = None if is_active else (format_human_datetime(row.get('logout_time')) if row.get('logout_time') else None)
            login_history.append({
                'session_id': row.get('session_id'),
                'login_time': format_human_datetime(row.get('login_time')),
                'logout_time': logout_value,
                'active': is_active,
                'is_current': row.get('session_id') == active_session_id,
            })

        profile_rows = fetch_rows(
            '''
            SELECT field_changed, old_value, new_value, changed_at
            FROM profile_changes
            WHERE user_id = %s AND role = 'applicant'
            ORDER BY changed_at DESC
            LIMIT 10
            ''',
            (applicant_id,),
        )
        profile_history = [
            {
                'field': row.get('field_changed'),
                'old': row.get('old_value'),
                'new': row.get('new_value'),
                'changed_at': format_human_datetime(row.get('changed_at')),
            }
            for row in profile_rows
        ]

        resumes_context = []
        for item in resumes:
            file_path = (item.get('file_path') or '').replace('\\', '/')
            file_name = item.get('file_name') or os.path.basename(file_path) or 'Resume'
            file_size = format_file_size(item.get('file_size_bytes') or 0)

            resumes_context.append(
                {
                    'resume_id': item.get('resume_id'),
                    'file_name': file_name,
                    'file_size': file_size,
                    'uploaded_at': format_human_datetime(item.get('uploaded_at')),
                    'download_url': url_for('download_resume', resume_id=item.get('resume_id')),
                    'view_url': url_for('preview_resume', resume_id=item.get('resume_id')),
                    'is_pdf': file_name.lower().endswith('.pdf'),
                    'file_path': file_path,
                }
            )

        return render_template(
            'applicant/profile.html',
            applicant=applicant,
            resumes=resumes_context,
            login_history=login_history,
            profile_history=profile_history,
            active_session_id=active_session_id,
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant profile error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Unable to load profile: {str(exc)}', 'error')
        return render_template('applicant/profile.html', applicant=None, resumes=[], login_history=[], profile_history=[])
    finally:
        if cursor:
            cursor.close()


@app.route('/applicant/resumes/<int:resume_id>/download')
@login_required('applicant')
def download_resume(resume_id):
    applicant_id = session.get('user_id')
    record = execute_query(
        '''
        SELECT file_path, file_name
        FROM resumes
        WHERE resume_id = %s AND applicant_id = %s
        LIMIT 1
        ''',
        (resume_id, applicant_id),
        fetch_one=True,
    )

    if not record or not record.get('file_path'):
        flash('Resume not found.', 'error')
        return redirect(url_for('applicant_profile'))

    file_path = record['file_path']
    absolute_path = os.path.join(app.root_path, file_path)

    if not os.path.exists(absolute_path):
        flash('Resume file is no longer available.', 'error')
        return redirect(url_for('applicant_profile'))

    download_name = record.get('file_name') or os.path.basename(file_path)
    mimetype_value = mimetypes.guess_type(download_name)[0] or 'application/octet-stream'
    return send_file(absolute_path, as_attachment=True, download_name=download_name, mimetype=mimetype_value)


@app.route('/applicant/resumes/<int:resume_id>/view')
@login_required('applicant')
def preview_resume(resume_id):
    applicant_id = session.get('user_id')
    record = execute_query(
        '''
        SELECT file_path, file_name
        FROM resumes
        WHERE resume_id = %s AND applicant_id = %s
        LIMIT 1
        ''',
        (resume_id, applicant_id),
        fetch_one=True,
    )

    if not record or not record.get('file_path'):
        flash('Resume not found.', 'error')
        return redirect(url_for('applicant_profile'))

    file_path = record['file_path']
    absolute_path = os.path.join(app.root_path, file_path)

    if not os.path.exists(absolute_path):
        flash('Resume file is no longer available.', 'error')
        return redirect(url_for('applicant_profile'))
    
    # Always send inline for viewing (opens in browser, not download)
    download_name = record.get('file_name') or os.path.basename(file_path)
    
    # Determine mimetype based on file extension
    file_ext = os.path.splitext(download_name.lower())[1]
    if file_ext == '.pdf':
        mimetype_value = 'application/pdf'
    elif file_ext in ['.doc', '.docx']:
        mimetype_value = 'application/msword' if file_ext == '.doc' else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    elif file_ext in ['.txt']:
        mimetype_value = 'text/plain'
    else:
        mimetype_value = mimetypes.guess_type(download_name)[0] or 'application/octet-stream'
    
    # For view route, always open inline in browser (not as attachment)
    # This allows PDFs to be displayed in browser viewer
    response = send_file(absolute_path, as_attachment=False, download_name=download_name, mimetype=mimetype_value)
    
    # Ensure inline display for PDFs
    if file_ext == '.pdf':
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename="' + download_name + '"'
        response.headers['Accept-Ranges'] = 'bytes'
    
    return response


@app.route('/admin/resumes/<int:resume_id>/download')
@login_required('admin', 'hr')
def admin_download_resume(resume_id):
    """Allow admin and HR to download applicant resumes."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicants'))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Verify resume exists and get file info
        cursor.execute(
            '''
            SELECT r.file_path, r.file_name, a.applicant_id
            FROM resumes r
            JOIN applicants a ON r.applicant_id = a.applicant_id
            WHERE r.resume_id = %s
            LIMIT 1
            ''',
            (resume_id,),
        )
        record = cursor.fetchone()
        
        if not record or not record.get('file_path'):
            flash('Resume not found.', 'error')
            return redirect(url_for('applicants'))
        
        # For HR users, verify the resume belongs to their branch
        if user.get('role') == 'hr':
            branch_id = get_branch_scope(user)
            if branch_id:
                cursor.execute(
                    '''
                    SELECT a.application_id
                    FROM applications a
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.applicant_id = %s AND a.resume_id = %s AND j.branch_id = %s
                    LIMIT 1
                    ''',
                    (record['applicant_id'], resume_id, branch_id),
                )
                if not cursor.fetchone():
                    flash('You can only access resumes from your branch.', 'error')
                    return redirect(url_for('applicants'))
        
        file_path = record['file_path']
        absolute_path = os.path.join(app.root_path, file_path)
        
        if not os.path.exists(absolute_path):
            flash('Resume file is no longer available.', 'error')
            return redirect(url_for('applicants'))
        
        download_name = record.get('file_name') or os.path.basename(file_path)
        mimetype_value = mimetypes.guess_type(download_name)[0] or 'application/octet-stream'
        return send_file(absolute_path, as_attachment=True, download_name=download_name, mimetype=mimetype_value)
    except Exception as exc:
        print(f'❌ Admin download resume error: {exc}')
        flash('Unable to download resume.', 'error')
        return redirect(url_for('applicants'))
    finally:
        cursor.close()


@app.route('/admin/resumes/<int:resume_id>/view')
@login_required('admin', 'hr')
def admin_view_resume(resume_id):
    """Allow admin and HR to view applicant resumes."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicants'))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Verify resume exists and get file info
        cursor.execute(
            '''
            SELECT r.file_path, r.file_name, a.applicant_id
            FROM resumes r
            JOIN applicants a ON r.applicant_id = a.applicant_id
            WHERE r.resume_id = %s
            LIMIT 1
            ''',
            (resume_id,),
        )
        record = cursor.fetchone()
        
        if not record or not record.get('file_path'):
            flash('Resume not found.', 'error')
            return redirect(url_for('applicants'))
        
        # For HR users, verify the resume belongs to their branch
        if user.get('role') == 'hr':
            branch_id = get_branch_scope(user)
            if branch_id:
                cursor.execute(
                    '''
                    SELECT a.application_id
                    FROM applications a
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.applicant_id = %s AND a.resume_id = %s AND j.branch_id = %s
                    LIMIT 1
                    ''',
                    (record['applicant_id'], resume_id, branch_id),
                )
                if not cursor.fetchone():
                    flash('You can only access resumes from your branch.', 'error')
                    return redirect(url_for('applicants'))
        
        file_path = record['file_path']
        absolute_path = os.path.join(app.root_path, file_path)
        
        if not os.path.exists(absolute_path):
            flash('Resume file is no longer available.', 'error')
            return redirect(url_for('applicants'))
        
        # Always send inline for viewing (opens in browser, not download)
        download_name = record.get('file_name') or os.path.basename(file_path)
        
        # Determine mimetype based on file extension
        file_ext = os.path.splitext(download_name.lower())[1]
        if file_ext == '.pdf':
            mimetype_value = 'application/pdf'
        elif file_ext in ['.doc', '.docx']:
            mimetype_value = 'application/msword' if file_ext == '.doc' else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        elif file_ext in ['.txt']:
            mimetype_value = 'text/plain'
        else:
            mimetype_value = mimetypes.guess_type(download_name)[0] or 'application/octet-stream'
        
        # For view route, always open inline in browser (not as attachment)
        # Read file directly and create Response with explicit headers
        # This gives us full control over Content-Disposition
        from flask import Response
        
        with open(absolute_path, 'rb') as f:
            file_data = f.read()
        
        # Create Response with file data
        response = Response(
            file_data,
            mimetype=mimetype_value,
            headers={
                'Content-Disposition': 'inline',  # CRITICAL: inline = display, not download
                'Content-Type': mimetype_value,
                'Content-Length': str(len(file_data)),
                'X-Content-Type-Options': 'nosniff'  # Prevent MIME sniffing
            }
        )
        
        # For PDFs, ensure proper headers
        if file_ext == '.pdf':
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Accept-Ranges'] = 'bytes'
        
        # Double-check: Remove any attachment-related headers
        # Some browsers check for 'attachment' keyword and force download
        if 'attachment' in str(response.headers.get('Content-Disposition', '')):
            response.headers['Content-Disposition'] = 'inline'
        
        return response
    except Exception as exc:
        print(f'❌ Admin view resume error: {exc}')
        flash('Unable to view resume.', 'error')
        return redirect(url_for('applicants'))
    finally:
        cursor.close()



@app.route('/admin/dashboard')
@login_required('admin', 'hr')
def admin_dashboard():
    dashboard_data = build_admin_dashboard_data(get_current_user())
    return render_template('admin/admin_dashboard.html', dashboard_data=dashboard_data)


@app.route('/hr/dashboard')
@login_required('hr')
def hr_dashboard():
    """HR Dashboard with multi-branch support."""
    try:
        user = get_current_user()
        if not user:
            flash('Please log in to access the dashboard.', 'error')
            return immediate_redirect(url_for('login', _external=True))
        
        # Get selected branch from query parameter for filtering (HR can still manage all)
        selected_branch_id = request.args.get('branch_id', type=int)
        
        # Build main dashboard data (shows all branches by default, filtered if branch_id provided)
        try:
            dashboard_data = build_admin_dashboard_data(user, branch_id=selected_branch_id)
        except Exception as build_error:
            print(f'❌ Error building dashboard data: {build_error}')
            import traceback
            traceback.print_exc()
            dashboard_data = {
                'user': {'full_name': user.get('name', 'HR User'), 'email': user.get('email', ''), 'role': 'hr'},
                'stats': {},
                'metrics': {},
                'branch_info': {},
                'notifications': [],
                'upcoming_interviews': [],
                'recent_applications': [],
                'recent_activity': [],
                'recent_jobs': [],
            }
        
        # Get branch information for selected branch (if filtering)
        branch_info = {}
        if selected_branch_id:
            try:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (selected_branch_id,))
                if branch_rows:
                    branch_info = branch_rows[0]
            except Exception as branch_error:
                print(f'⚠️ Error fetching branch info: {branch_error}')
        
        dashboard_data['branch_info'] = branch_info
        dashboard_data['selected_branch_id'] = selected_branch_id
        
        # Get all branches for HR - HR can manage all branches
        all_branches = []
        try:
            # HR manages all branches - get all branches
            all_branches = fetch_branches() or []
        except Exception as branch_fetch_error:
            print(f'⚠️ Error fetching branches: {branch_fetch_error}')
            all_branches = []
        
        # Get branch-specific statistics for each branch
        branch_stats = []
        for branch in all_branches:
            try:
                bid = branch.get('branch_id')
                if not bid:
                    continue
                    
                branch_open_jobs = fetch_count(
                    """
                    SELECT COUNT(*) AS count
                    FROM jobs j
                    WHERE j.branch_id = %s AND j.status IN ('open', 'published', 'active')
                    """,
                    (bid,)
                ) or 0
                branch_applicants = fetch_count(
                    """
                    SELECT COUNT(DISTINCT a.applicant_id) AS count
                    FROM applications a
                    JOIN jobs j ON j.job_id = a.job_id
                    WHERE j.branch_id = %s
                    """,
                    (bid,)
                ) or 0
                branch_interviews = fetch_count(
                    """
                    SELECT COUNT(*) AS count
                    FROM interviews i
                    JOIN applications a ON a.application_id = i.application_id
                    JOIN jobs j ON j.job_id = a.job_id
                    WHERE j.branch_id = %s AND DATE(i.scheduled_date) = CURDATE()
                    """,
                    (bid,)
                ) or 0
                branch_hires = fetch_count(
                    """
                    SELECT COUNT(*) AS count
                    FROM applications a
                    JOIN jobs j ON j.job_id = a.job_id
                    WHERE j.branch_id = %s AND a.status = 'hired'
                    """,
                    (bid,)
                ) or 0
                branch_pending = fetch_count(
                    """
                    SELECT COUNT(*) AS count
                    FROM applications a
                    JOIN jobs j ON j.job_id = a.job_id
                    WHERE j.branch_id = %s AND a.status = 'pending'
                    """,
                    (bid,)
                ) or 0
                
                branch_stats.append({
                    'branch_id': bid,
                    'branch_name': branch.get('branch_name', 'Unknown'),
                    'address': branch.get('address', ''),
                    'open_jobs': branch_open_jobs,
                    'applicants': branch_applicants,
                    'interviews_today': branch_interviews,
                    'hires': branch_hires,
                    'pending_applications': branch_pending,
                    'status': 'active',
                })
            except Exception as stat_error:
                print(f'⚠️ Error calculating stats for branch {branch.get("branch_id")}: {stat_error}')
                continue
        
        dashboard_data['branch_stats'] = branch_stats
        dashboard_data['all_branches'] = all_branches
        
        # Get recent job postings with branch info (filtered by selected branch if provided)
        try:
            if selected_branch_id:
                recent_jobs_with_branch = fetch_rows(
                    """
                    SELECT j.job_id,
                           COALESCE(j.title, 'Untitled Job') AS job_title,
                           j.status,
                           j.created_at AS posted_at,
                           b.branch_id,
                           b.branch_name,
                           (SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) AS application_count
                    FROM jobs j
                    LEFT JOIN branches b ON b.branch_id = j.branch_id
                    WHERE j.branch_id = %s
                    ORDER BY j.created_at DESC
                    LIMIT 10
                    """,
                    (selected_branch_id,)
                ) or []
            else:
                recent_jobs_with_branch = fetch_rows(
                    """
                    SELECT j.job_id,
                           COALESCE(j.title, 'Untitled Job') AS job_title,
                           j.status,
                           j.created_at AS posted_at,
                           b.branch_id,
                           b.branch_name,
                           (SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) AS application_count
                    FROM jobs j
                    LEFT JOIN branches b ON b.branch_id = j.branch_id
                    ORDER BY j.created_at DESC
                    LIMIT 10
                    """
                ) or []
            dashboard_data['recent_jobs_with_branch'] = recent_jobs_with_branch
        except Exception as jobs_error:
            print(f'⚠️ Error fetching recent jobs: {jobs_error}')
            dashboard_data['recent_jobs_with_branch'] = []
        
        # Get recent applicants with branch info (filtered by selected branch if provided)
        try:
            if selected_branch_id:
                recent_applicants_with_branch = fetch_rows(
                    """
                    SELECT ap.applicant_id,
                           ap.full_name AS applicant_name,
                           ap.email,
                           COALESCE(j.title, 'N/A') AS job_title,
                           b.branch_id,
                           b.branch_name,
                           a.status,
                           a.submitted_at
                    FROM applicants ap
                    JOIN applications a ON ap.applicant_id = a.applicant_id
                    JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    WHERE j.branch_id = %s
                    ORDER BY a.submitted_at DESC
                    LIMIT 10
                    """,
                    (selected_branch_id,)
                ) or []
            else:
                recent_applicants_with_branch = fetch_rows(
                    """
                    SELECT ap.applicant_id,
                           ap.full_name AS applicant_name,
                           ap.email,
                           COALESCE(j.title, 'N/A') AS job_title,
                           b.branch_id,
                           b.branch_name,
                           a.status,
                           a.submitted_at
                    FROM applicants ap
                    JOIN applications a ON ap.applicant_id = a.applicant_id
                    JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    ORDER BY a.submitted_at DESC
                    LIMIT 10
                    """
                ) or []
            dashboard_data['recent_applicants_with_branch'] = recent_applicants_with_branch
        except Exception as applicants_error:
            print(f'⚠️ Error fetching recent applicants: {applicants_error}')
            dashboard_data['recent_applicants_with_branch'] = []
        
        return render_template('hr/dashboard.html', dashboard_data=dashboard_data)
    except Exception as exc:
        print(f'❌ HR Dashboard error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to load dashboard. Please try again later.', 'error')
        # Return minimal dashboard data
        return render_template('hr/dashboard.html', dashboard_data={
            'user': {'full_name': 'HR User', 'email': '', 'role': 'hr'},
            'stats': {},
            'metrics': {},
            'branch_info': {},
            'notifications': [],
            'upcoming_interviews': [],
            'recent_applications': [],
            'recent_activity': [],
            'recent_jobs': [],
            'recent_jobs_with_branch': [],
            'recent_applicants_with_branch': [],
            'all_branches': [],
            'branch_stats': [],
            'selected_branch_id': None,
        })


# HR Communications route removed - feature disabled
# @app.route('/hr/communications', methods=['GET', 'POST'])
# @login_required('hr')
# def hr_communications():
#     """Branch-scoped communications center for HR."""
#     ... (route code removed)


# HR-Specific Routes (Branch-Scoped)
# Note: HR routes redirect to admin routes which already handle branch filtering
# The admin routes check user role and filter by branch_id for HR users
# HR templates will be updated to extend admin base template for consistency


# Admin Navigation Routes
@app.route('/api/admin/notifications')
@login_required('admin', 'hr')
def api_admin_notifications():
    """API endpoint to fetch admin notifications in JSON format."""
    user = get_current_user()
    if not user or user.get('role') not in ('admin', 'hr'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        formatted, unread = fetch_notifications_for({'system_only': True}, limit=5)
        return jsonify({
            'success': True,
            'notifications': formatted,
            'unread_count': unread,
            'unread_display': '99+' if unread > 99 else unread
        })
    except Exception as e:
        print(f'⚠️ Error fetching admin notifications API: {e}')
        return jsonify({
            'success': False,
            'error': str(e),
            'notifications': [],
            'unread_count': 0,
            'unread_display': 0
        }), 500

@app.route('/admin/notifications')
@login_required('admin', 'hr')
def admin_notifications():
    """System-wide notification management."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/notifications.html', notifications=[], unread_count=0)
    
    cursor = db.cursor(dictionary=True)
    try:
        ensure_schema_compatibility()
        
        # Check if is_read and sent_at columns exist
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns_raw = cursor.fetchall()
        notification_columns = [col.get('Field') if isinstance(col, dict) else col[0] for col in notification_columns_raw]
        
        # Build dynamic expressions
        sent_at_expr = 'COALESCE(n.sent_at, n.created_at, NOW())' if 'sent_at' in notification_columns else 'COALESCE(n.created_at, NOW())'
        is_read_expr = 'COALESCE(n.is_read, 0)' if 'is_read' in notification_columns else '0'
        
        # Show only system-level notifications for Admin page (exclude applicant/application-specific entries)
        params = []
        where_sql = 'WHERE n.application_id IS NULL'

        cursor.execute(
            f'''
            SELECT n.notification_id,
                   n.message,
                   {sent_at_expr} AS sent_at,
                   {is_read_expr} AS is_read,
                   a.application_id,
                   a.status AS application_status,
                   COALESCE(j.title, 'N/A') AS job_title,
                   COALESCE(ap.full_name, 'Unknown') AS applicant_name
            FROM notifications n
            LEFT JOIN applications a ON n.application_id = a.application_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
            {where_sql}
            ORDER BY {sent_at_expr} DESC
            LIMIT 200
            ''',
            tuple(params)
        )
        notifications = cursor.fetchall() or []
        
        unread_count = len([n for n in notifications if not n.get('is_read')])
        system_count = len(notifications)
        application_count = 0
        
        formatted_notifications = []
        for notif in notifications:
            formatted_notifications.append({
                'notification_id': notif.get('notification_id'),
                'message': notif.get('message'),
                'sent_at': format_human_datetime(notif.get('sent_at')),
                'is_read': notif.get('is_read', False),
                'application_id': notif.get('application_id'),
                'job_title': notif.get('job_title'),
                'applicant_name': notif.get('applicant_name'),
                'application_status': notif.get('application_status'),
                'type': 'system',
            })
        
        return render_template(
            'admin/notifications.html',
            notifications=formatted_notifications,
            unread_count=unread_count,
            system_count=system_count,
            application_count=application_count
        )
    except Exception as exc:
        if db:
            db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Admin notifications error: {exc}')
        print(f'Full traceback: {error_details}')
        flash('Unable to load notifications. Please try again later.', 'error')
        return render_template('admin/notifications.html', notifications=[], unread_count=0)
    finally:
        cursor.close()


@app.route('/admin/notifications/read-all', methods=['POST'])
@login_required('admin', 'hr')
def mark_all_admin_notifications_read():
    """Mark all notifications as read."""
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_notifications'))
    
    cursor = db.cursor()
    try:
        ensure_schema_compatibility()
        
        # Clean up any JSON responses that might have been stored as notifications
        try:
            cursor.execute("""
                DELETE FROM notifications 
                WHERE message LIKE '{%' 
                AND (message LIKE '%"success"%' OR message LIKE '%"message"%')
            """)
            deleted_count = cursor.rowcount
            if deleted_count > 0:
                print(f'✅ Cleaned up {deleted_count} JSON response notifications from database')
        except Exception as cleanup_error:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_error}')
        
        # Check if is_read column exists
        cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
        has_is_read = cursor.fetchone() is not None
        
        if has_is_read:
            cursor.execute('UPDATE notifications SET is_read = 1 WHERE is_read = 0')
            db.commit()
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': True, 'message': 'All notifications marked as read'})
            flash('All notifications marked as read.', 'success')
        else:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification read status not available.'}), 400
            flash('Notification read status not available.', 'error')
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Mark all notifications read error: {exc}')
        print(f'Full traceback: {error_details}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to mark all notifications as read.'}), 500
        flash('Failed to mark all notifications as read.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('admin_notifications'))


@app.route('/admin/notifications/<int:notification_id>/read', methods=['POST'])
@login_required('admin', 'hr')
def mark_admin_notification_read(notification_id):
    """Mark a notification as read."""
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_notifications'))
    
    cursor = db.cursor()
    try:
        ensure_schema_compatibility()
        
        # Check if this notification is a JSON response and delete it if so
        cursor.execute('SELECT message FROM notifications WHERE notification_id = %s', (notification_id,))
        notif_record = cursor.fetchone()
        if notif_record:
            message = notif_record.get('message', '') if isinstance(notif_record, dict) else (notif_record[0] if notif_record else '')
            if message and message.strip().startswith('{') and '"success"' in message:
                # Delete JSON response notifications instead of marking as read
                cursor.execute('DELETE FROM notifications WHERE notification_id = %s', (notification_id,))
                db.commit()
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                    return jsonify({'success': True, 'message': 'Invalid notification removed', 'notification_id': notification_id})
                flash('Invalid notification removed.', 'success')
                return redirect(url_for('admin_notifications'))
        
        # Check if is_read column exists
        cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
        has_is_read = cursor.fetchone() is not None
        
        if has_is_read:
            cursor.execute(
                'UPDATE notifications SET is_read = 1 WHERE notification_id = %s',
                (notification_id,)
            )
        else:
            # If column doesn't exist, just return success (column will be created by ensure_schema_compatibility)
            pass
        
        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'Notification marked as read', 'notification_id': notification_id})
        flash('Notification marked as read.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Mark notification read error: {exc}')
        print(f'Full traceback: {error_details}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to mark notification as read.'}), 500
        flash('Failed to mark notification as read.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('admin_notifications'))


@app.route('/admin/notifications/delete-all', methods=['POST'])
@login_required('admin', 'hr')
def delete_all_admin_notifications():
    """Delete notifications. Admin: all. HR with branch: only notifications for their branch. HR without branch: all."""
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_notifications'))
    
    cursor = db.cursor()
    try:
        ensure_schema_compatibility()
        user = get_current_user()
        role = user.get('role') if user else session.get('user_role')
        branch_id = get_branch_scope(user) if user else (session.get('branch_id') if role == 'hr' else None)
        
        # Verify notifications table and needed columns
        cursor.execute("SHOW TABLES LIKE 'notifications'")
        if not cursor.fetchone():
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': True, 'message': 'No notifications to delete'}), 200
            flash('No notifications to delete.', 'info')
            return redirect(url_for('admin_notifications'))
        
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns = {row[0] if isinstance(row, (list, tuple)) else row.get('Field') for row in (cursor.fetchall() or []) if row}
        has_application_fk = 'application_id' in notification_columns
        
        if role == 'admin' or (role == 'hr' and not branch_id):
            # Admin or HR managing all branches: delete all notifications
            cursor.execute('DELETE FROM notifications')
        else:
            # HR with specific branch: delete only branch-scoped notifications (requires application_id)
            if not has_application_fk:
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                    return jsonify({'success': False, 'error': 'Unable to scope notifications by branch.'}), 400
                flash('Unable to scope notifications by branch.', 'error')
                return redirect(url_for('admin_notifications'))
            cursor.execute(
                '''
                DELETE n FROM notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                WHERE j.branch_id = %s
                ''',
                (branch_id,),
            )
        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            # Final cleanup: Remove any JSON response notifications that might have been created
            try:
                if branch_id:
                    cursor.execute("""
                        DELETE n FROM notifications n
                        LEFT JOIN applications a ON n.application_id = a.application_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        WHERE (j.branch_id = %s OR n.application_id IS NULL)
                        AND (
                            (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                            OR n.message = 'Notifications deleted.'
                            OR n.message LIKE '%Notifications deleted%'
                            OR n.message LIKE '%All notifications deleted%'
                            OR n.message LIKE '%Notification deleted successfully%'
                        )
                    """, (branch_id,))
                else:
                    cursor.execute("""
                        DELETE FROM notifications 
                        WHERE (
                            (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                            OR message = 'Notifications deleted.'
                            OR message LIKE '%Notifications deleted%'
                            OR message LIKE '%All notifications deleted%'
                            OR message LIKE '%Notification deleted successfully%'
                        )
                    """)
                final_cleaned = cursor.rowcount
                if final_cleaned > 0:
                    print(f'✅ Final cleanup: Removed {final_cleaned} JSON response notification(s) after admin delete-all')
                    db.commit()
            except Exception as final_cleanup_err:
                print(f'⚠️ Error in final cleanup: {final_cleanup_err}')
                db.rollback()
            
            return jsonify({'success': True, 'message': 'Notifications deleted.'})
        flash('Notifications deleted.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        print(f'❌ Delete all admin notifications error: {exc}')
        print(traceback.format_exc())
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to delete notifications.'}), 500
        flash('Failed to delete notifications.', 'error')
    finally:
        cursor.close()
    return redirect(url_for('admin_notifications'))


@app.route('/admin/manage-branches', methods=['GET', 'POST'])
@login_required('admin')
def manage_branches():
    """Comprehensive branch management with CRUD operations."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/manage_branches.html', branches=[], total_branches=0)
    
    cursor = db.cursor(dictionary=True)
    
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            if action == 'add':
                branch_name = request.form.get('branch_name', '').strip()
                address = request.form.get('address', '').strip()
                operating_hours = request.form.get('operating_hours', '').strip() or None
                is_active = request.form.get('is_active')
                
                # Validation
                if not branch_name or not address:
                    error_msg = 'Branch name and address are required.'
                    flash(error_msg, 'error')
                    if is_ajax:
                        return jsonify({'success': False, 'error': error_msg}), 400
                else:
                    try:
                        # Check which columns exist in branches table
                        cursor.execute('SHOW COLUMNS FROM branches')
                        columns = {row.get('Field') for row in cursor.fetchall() if row}
                        
                        # Build INSERT query dynamically based on available columns
                        insert_fields = ['branch_name', 'address']
                        insert_values = [branch_name, address]
                        
                        # Add operating_hours if column exists
                        if 'operating_hours' in columns:
                            insert_fields.append('operating_hours')
                            insert_values.append(operating_hours)
                        
                        # Add is_active if column exists (dropdown sends '1' for Active, '0' for Inactive)
                        if 'is_active' in columns:
                            # Handle dropdown values: '1' (Active) or '0' (Inactive)
                            # Also handle checkbox legacy: 'on' (checked) or None/False (unchecked)
                            if is_active == '1' or is_active == 'on' or is_active is True:
                                is_active_bool = True
                            elif is_active == '0' or is_active is False:
                                is_active_bool = False
                            else:
                                # Default to Active if not explicitly set
                                is_active_bool = True
                            insert_fields.append('is_active')
                            insert_values.append(is_active_bool)
                        
                        # Build and execute INSERT query
                        placeholders = ', '.join(['%s'] * len(insert_fields))
                        fields = ', '.join(insert_fields)
                        
                        cursor.execute(
                            f'''
                            INSERT INTO branches ({fields})
                            VALUES ({placeholders})
                            ''',
                            tuple(insert_values),
                        )
                        db.commit()
                        flash('Branch added successfully.', 'success')
                        if is_ajax:
                            return jsonify({'success': True, 'redirect': url_for('manage_branches', _external=False)})
                    except Exception as add_error:
                        db.rollback()
                        print(f"❌ ERROR adding branch: {add_error}")
                        import traceback
                        traceback.print_exc()
                        error_msg = f'Error adding branch: {str(add_error)}'
                        flash(error_msg, 'error')
                        if is_ajax:
                            return jsonify({'success': False, 'error': error_msg}), 500
            
            elif action == 'update':
                branch_id = request.form.get('branch_id')
                branch_name = request.form.get('branch_name', '').strip()
                address = request.form.get('address', '').strip()
                # Get operating_hours from form (field name is 'operating_hours' with underscore)
                # Strip whitespace and convert empty/whitespace-only strings to None (same as add action)
                operating_hours = request.form.get('operating_hours', '').strip() or None
                is_active = request.form.get('is_active')
                
                if not branch_id or not branch_name or not address:
                    error_msg = 'Branch ID, name, and address are required.'
                    flash(error_msg, 'error')
                    if is_ajax:
                        return jsonify({'success': False, 'error': error_msg}), 400
                else:
                    # AUTOMATIC: Always update operating_hours (exists in schema)
                    # Check which columns exist in branches table for is_active
                    cursor.execute('SHOW COLUMNS FROM branches')
                    columns = {row.get('Field') for row in cursor.fetchall() if row}
                    
                    # Build UPDATE query with operating_hours always included
                    update_fields = [
                        'branch_name = %s',
                        'address = %s',
                        'operating_hours = %s',  # AUTOMATIC: Always included
                    ]
                    update_values = [branch_name, address, operating_hours]
                    
                    # Add is_active if column exists (dropdown sends '1' for Active, '0' for Inactive)
                    if 'is_active' in columns:
                        # Handle dropdown values: '1' (Active) or '0' (Inactive)
                        # Also handle checkbox legacy: 'on' (checked) or None/False (unchecked)
                        if is_active == '1' or is_active == 'on' or is_active is True:
                            is_active_bool = True
                        elif is_active == '0' or is_active is False:
                            is_active_bool = False
                        else:
                            # Default to Active if not explicitly set
                            is_active_bool = True
                        update_fields.append('is_active = %s')
                        update_values.append(is_active_bool)
                    
                    update_values.append(branch_id)
                    
                    # Execute UPDATE
                    try:
                        cursor.execute(
                            f'''
                            UPDATE branches
                            SET {', '.join(update_fields)}
                            WHERE branch_id = %s
                            ''',
                            tuple(update_values),
                        )
                        
                        # Verify the update was successful
                        rows_affected = cursor.rowcount
                        
                        if rows_affected > 0:
                            db.commit()
                            success_msg = f'Branch updated successfully. Operating hours: {operating_hours or "Not set"}'
                            flash(success_msg, 'success')
                            if is_ajax:
                                return jsonify({'success': True, 'redirect': url_for('manage_branches', _external=False)})
                        else:
                            db.rollback()
                            warning_msg = 'Branch not found or no changes made.'
                            flash(warning_msg, 'warning')
                            if is_ajax:
                                return jsonify({'success': False, 'error': warning_msg}), 404
                    except Exception as update_error:
                        db.rollback()
                        print(f"❌ ERROR updating branch: {update_error}")
                        import traceback
                        traceback.print_exc()
                        error_msg = f'Error updating branch: {update_error}'
                        flash(error_msg, 'error')
                        if is_ajax:
                            return jsonify({'success': False, 'error': error_msg}), 500
            
            # For non-AJAX requests, redirect normally
            return redirect(url_for('manage_branches', _external=False))
        
        # Get filters from query parameters
        keyword = request.args.get('keyword', '').strip()
        status_filter = request.args.get('status', '').strip()
        
        # Build WHERE clauses
        where_clauses = []
        params = []
        
        # Keyword search (case-insensitive, searches branch_name, address)
        if keyword:
            keyword_pattern = f"%{keyword}%"
            where_clauses.append('(LOWER(b.branch_name) LIKE LOWER(%s) OR LOWER(b.address) LIKE LOWER(%s))')
            params.extend([keyword_pattern, keyword_pattern])
        
        # Status filter
        if status_filter and status_filter != 'all':
            if status_filter == 'active':
                where_clauses.append('b.is_active = 1')
            elif status_filter == 'inactive':
                where_clauses.append('b.is_active = 0')
        
        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        
        # Fetch branches with comprehensive metrics
        rows = fetch_rows(
            f"""
            SELECT 
                b.branch_id AS id,
                b.branch_id,
                b.branch_name,
                b.address,
                b.operating_hours,
                b.is_active,
                (SELECT COUNT(*) FROM jobs j WHERE j.branch_id = b.branch_id AND j.status IN ('published', 'active', 'open')) AS active_jobs,
                (SELECT COUNT(*) FROM applications a JOIN jobs j ON a.job_id = j.job_id WHERE j.branch_id = b.branch_id) AS total_applications,
                (SELECT COUNT(*) FROM applications a JOIN jobs j ON a.job_id = j.job_id WHERE j.branch_id = b.branch_id AND a.status = 'hired') AS accepted_applications
            FROM branches b
            WHERE {where_sql}
            ORDER BY b.branch_name ASC
            """,
            tuple(params) if params else None
        )
        
        total_branches = len(rows)
        return render_template(
            'admin/manage_branches.html',
            branches=rows,
            total_branches=total_branches,
            current_filters={'keyword': keyword, 'status': status_filter},
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Branch management error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        return render_template('admin/manage_branches.html', branches=[], total_branches=0)
    finally:
        cursor.close()


@app.route('/admin/hr-accounts', methods=['GET', 'POST'])
@login_required('admin')
def hr_accounts():
    """Comprehensive HR account management with CRUD operations - Admin only."""
    user = get_current_user()
    # Explicitly deny HR users
    if user.get('role') == 'hr':
        flash('Access denied. HR account management is only available to administrators.', 'error')
        return redirect(url_for('hr_dashboard'))
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/hr_accounts_management.html', accounts=[], branches=[])
    
    cursor = db.cursor(dictionary=True)
    
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'add':
                full_name = request.form.get('full_name', '').strip()
                email = request.form.get('email', '').strip().lower()
                password = request.form.get('password', '').strip()
                # HR accounts manage all branches (branch_id column removed from admins table)
                is_active = request.form.get('is_active') == 'on'
                
                if not all([full_name, email, password]):
                    flash('Full name, email, and password are required.', 'error')
                elif len(password) < 6:
                    flash('Password must be at least 6 characters.', 'error')
                else:
                    
                    # Check if email already exists in users table
                    cursor.execute(
                        'SELECT user_id FROM users WHERE email = %s LIMIT 1',
                        (email,),
                    )
                    if cursor.fetchone():
                        flash('Email address is already registered.', 'error')
                    else:
                        # First, create user in users table
                        # Admin/HR accounts are automatically verified (no email verification required)
                        password_hash = hash_password(password)
                        cursor.execute(
                            '''
                            INSERT INTO users (email, password_hash, user_type, is_active, email_verified)
                            VALUES (%s, %s, 'hr', %s, 1)
                            ''',
                            (email, password_hash, is_active),
                        )
                        user_id = cursor.lastrowid
                        
                        # Get password_hash from users table to copy to admins table
                        cursor.execute(
                            'SELECT password_hash FROM users WHERE user_id = %s LIMIT 1',
                            (user_id,)
                        )
                        user_record = cursor.fetchone()
                        admin_password_hash = user_record.get('password_hash') if user_record else password_hash
                        
                        # Then, create admin record linked to the user with password_hash
                        # HR accounts manage all branches (branch_id column removed)
                        # role must be explicitly set to 'hr' for HR accounts
                        cursor.execute(
                            '''
                            INSERT INTO admins (user_id, full_name, email, password_hash, role, is_active)
                            VALUES (%s, %s, %s, %s, 'hr', %s)
                            ''',
                            (user_id, full_name, email, admin_password_hash, is_active),
                        )
                        admin_id = cursor.lastrowid
                        
                        # HR accounts manage all branches
                        
                        # Create admin notification for HR account creation
                        try:
                            admin_msg = f'New HR account created: {full_name} ({email}).'
                            create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating notification for HR account creation: {notify_err}')
                        
                        db.commit()
                        flash('HR account created successfully. Account can manage all branches.', 'success')
            
            elif action == 'update':
                admin_id = request.form.get('admin_id')
                full_name = request.form.get('full_name', '').strip()
                email = request.form.get('email', '').strip().lower()
                # HR accounts manage all branches (branch_id column removed from admins table)
                is_active = request.form.get('is_active') == 'on'
                
                if not admin_id or not full_name or not email:
                    flash('Admin ID, full name, and email are required.', 'error')
                else:
                    
                    # Get current user_id from admins table
                    cursor.execute(
                        'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                        (admin_id,),
                    )
                    admin_record = cursor.fetchone()
                    if not admin_record:
                        flash('HR account not found.', 'error')
                    else:
                        user_id = admin_record['user_id']
                        
                        # Check if email is already in use by another user
                        cursor.execute(
                            'SELECT user_id FROM users WHERE email = %s AND user_id <> %s LIMIT 1',
                            (email, user_id),
                        )
                        if cursor.fetchone():
                            flash('Email address is already in use.', 'error')
                        else:
                            # Update users table
                            cursor.execute(
                                '''
                                UPDATE users
                                SET email = %s, is_active = %s
                                WHERE user_id = %s
                                ''',
                                (email, is_active, user_id),
                            )
                            
                            # Update admins table
                            # HR accounts manage all branches (branch_id column removed)
                            # Ensure role remains 'hr' and is_active is synced
                            cursor.execute(
                                '''
                                UPDATE admins
                                SET full_name = %s,
                                    email = %s,
                                    role = 'hr',
                                    is_active = %s
                                WHERE admin_id = %s
                                ''',
                                (full_name, email, is_active, admin_id),
                            )
                            
                            # HR accounts manage all branches
                            
                            db.commit()
                            flash('HR account updated successfully. Account can manage all branches.', 'success')
            
            elif action == 'reset_password':
                admin_id = request.form.get('admin_id')
                new_password = request.form.get('new_password', '').strip()
                
                if not admin_id or not new_password or len(new_password) < 6:
                    flash('Admin ID and password (min 6 characters) are required.', 'error')
                else:
                    try:
                        # Get user_id from admins table
                        cursor.execute(
                            'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                            (admin_id,),
                        )
                        admin_record = cursor.fetchone()
                        if not admin_record:
                            flash('HR account not found.', 'error')
                        else:
                            user_id = admin_record['user_id']
                            # Hard update password in users table - permanently changes password in database
                            cursor.execute(
                                'UPDATE users SET password_hash = %s WHERE user_id = %s',
                                (hash_password(new_password), user_id),
                            )
                            db.commit()
                            flash('Password reset successfully in system and database.', 'success')
                    except Exception as exc:
                        db.rollback()
                        flash(f'Failed to reset password: {exc}', 'error')
            
            return redirect(url_for('hr_accounts'))
        
        accounts = fetch_hr_accounts()
        branches = fetch_branches()
        
        # Ensure accounts is a list
        if accounts is None:
            accounts = []
        
        print(f'🔍 HR Accounts Route: Fetched {len(accounts)} accounts')
        if accounts and len(accounts) > 0:
            print(f'🔍 First account: {accounts[0]}')
        else:
            print('⚠️ No accounts returned from fetch_hr_accounts()')
            # Try a direct query to see what's in the database
            try:
                cursor.execute("""
                    SELECT a.admin_id, a.full_name, u.email, u.user_type, u.is_active
                    FROM admins a
                    LEFT JOIN users u ON u.user_id = a.user_id
                    LIMIT 10
                """)
                all_admins = cursor.fetchall()
                print(f'🔍 All admins in database (first 10): {len(all_admins) if all_admins else 0}')
                if all_admins:
                    for admin in all_admins[:3]:
                        print(f'   - Admin ID: {admin.get("admin_id")}, Name: {admin.get("full_name")}, User Type: {admin.get("user_type")}')
                
                # Also check specifically for HR accounts
                cursor.execute("""
                    SELECT a.admin_id, a.full_name, u.email, u.user_type
                    FROM admins a
                    INNER JOIN users u ON u.user_id = a.user_id
                    WHERE u.user_type = 'hr'
                """)
                hr_admins = cursor.fetchall()
                print(f'🔍 HR admins found directly: {len(hr_admins) if hr_admins else 0}')
                if hr_admins:
                    for hr in hr_admins[:3]:
                        print(f'   - HR Admin ID: {hr.get("admin_id")}, Name: {hr.get("full_name")}, Email: {hr.get("email")}')
            except Exception as debug_error:
                print(f'⚠️ Debug query error: {debug_error}')
                import traceback
                traceback.print_exc()
        
        # Check auth_sessions table columns first
        cursor.execute('SHOW COLUMNS FROM auth_sessions')
        session_columns_raw = cursor.fetchall() or []
        session_columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in session_columns_raw if row}
        
        # Build logout_time expression
        if 'last_activity' in session_columns and 'logout_time' in session_columns:
            logout_expr = 'COALESCE(last_activity, logout_time)'
        elif 'logout_time' in session_columns:
            logout_expr = 'logout_time'
        elif 'last_activity' in session_columns:
            logout_expr = 'last_activity'
        else:
            logout_expr = 'NULL'
        
        # Get comprehensive login history for each account (Admin and HR)
        for account in accounts:
            # Get user_id from admins table for this account
            cursor.execute(
                'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                (account.get('admin_id'),),
            )
            admin_record = cursor.fetchone()
            if admin_record:
                user_id = admin_record['user_id']
                # Build SELECT statement dynamically based on available columns
                select_fields = [
                    'login_time',
                    f'{logout_expr} AS logout_time',
                    'COALESCE(is_active, 1) AS is_active'
                ]
                
                cursor.execute(
                    f'''
                    SELECT {', '.join(select_fields)}
                    FROM auth_sessions
                    WHERE user_id = %s
                    ORDER BY login_time DESC
                    LIMIT 20
                    ''',
                    (user_id,),
                )
                login_rows = cursor.fetchall() or []
                # Format login history
                account['login_history'] = []
                for row in login_rows:
                    is_active = bool(row.get('is_active', 1))
                    account['login_history'].append({
                        'login_time': format_human_datetime(row.get('login_time')) if row.get('login_time') else 'Never',
                        'logout_time': format_human_datetime(row.get('logout_time')) if row.get('logout_time') and not is_active else (None if is_active else 'Never'),
                        'is_active': is_active,
                    })
            else:
                account['login_history'] = []
        
        # Ensure accounts is always a list, even if empty
        if not accounts:
            accounts = []
        
        print(f'🔍 Rendering template with {len(accounts)} accounts')
        return render_template('admin/hr_accounts_management.html', accounts=accounts, branches=branches)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ HR accounts management error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        branches = fetch_branches() or []
        return render_template('admin/hr_accounts_management.html', accounts=[], branches=branches)
    finally:
        cursor.close()


@app.route('/admin/job-postings', methods=['GET', 'POST'])
@login_required('admin', 'hr')
def job_postings():
    """Enhanced job postings management with bulk operations and advanced filtering."""
    user = get_current_user()
    if not user:
        flash('Please login to access this page.', 'error')
        return immediate_redirect(url_for('login', _external=True))

    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        branches = fetch_branches() or []
        positions = fetch_positions() or []
        metadata = {
            'status_options': VALID_JOB_STATUSES,
            'employment_types': VALID_EMPLOYMENT_TYPES,
            'work_arrangements': VALID_WORK_ARRANGEMENTS,
            'experience_levels': VALID_EXPERIENCE_LEVELS,
        }
        template = 'hr/job_postings.html' if user.get('role') == 'hr' else 'admin/job_postings.html'
        return render_template(
            template,
            jobs=[],
            branches=branches,
            positions=positions,
            user=user or {},
            current_branch=None,
            current_filters={},
            branch_info=None,
            job_meta=metadata,
        )

    cursor = db.cursor(dictionary=True)

    try:
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        branch_scope = get_branch_scope(user)

        def extract_job_payload(form, admin_id=None):
            """Collect and sanitize job form fields."""
            errors = []
            job_title = (form.get('job_title') or form.get('title') or '').strip()
            job_description = (form.get('job_description') or form.get('description') or '').strip()
            job_requirements = (form.get('job_requirements') or form.get('requirements') or '').strip()
            # Removed fields: employment_type, work_arrangement, experience_level, job_location, salary_currency, salary_min, salary_max, application_deadline, position_name, position_id (not in actual schema)

            # Handle status - database uses 'open' but code can use 'active' (they're equivalent for visibility)
            # IMPORTANT: Default to 'open' so jobs are immediately visible to applicants from all branches
            status_input = form.get('status', '').strip().lower()
            # Map 'active' to 'open' for database compatibility, or use 'open' directly
            if status_input == 'active':
                status = 'open'  # Database uses 'open' as the active status
            elif status_input == 'open':
                status = 'open'
            elif status_input == 'closed':
                status = 'closed'  # Only explicitly closed jobs are hidden
            else:
                # Default to 'open' for any other value or empty - ensures jobs are visible by default
                status = 'open'
            
            # Ensure that when posting a new job, if status is not explicitly set, default to 'open' (visible)
            # This ensures jobs are immediately visible to applicants from ALL branches
            if not status_input or status_input == '':
                status = 'open'
            
            # HR users can select any branch - no scoping restrictions
            if branch_scope is not None:
                branch_id = branch_scope
            else:
                # HR can select any branch from the form
                branch_raw = (form.get('branch_id') or '').strip()
                branch_id = int(branch_raw) if branch_raw.isdigit() else None
                if branch_id is None:
                    # For HR users, branch is required but they can choose any branch
                    errors.append('Please select a branch for this job posting.')

            if not job_title:
                errors.append('Job title is required.')
            # job_summary removed - not in schema
            if not job_description:
                errors.append('Job description is required.')
            if not job_requirements:
                errors.append('Job requirements are required.')

            # Use actual schema columns: title, description, requirements, status, branch_id, posted_by, posted_at
            # posted_at will be set by MySQL NOW() function in the INSERT/UPDATE query for accurate server time
            # Set posted_at to None - MySQL NOW() will be used in SQL for accurate server timezone
            posted_at = None
            
            payload = {
                'title': job_title,  # Actual column name
                'description': job_description,  # Actual column name
                'requirements': job_requirements,  # Actual column name
                'status': status,
                'branch_id': branch_id,
                'posted_by': admin_id if admin_id else None,
                'posted_at': posted_at,
            }

            return payload, errors

        if request.method == 'POST':
            action = request.form.get('action')
            
            # Block admins from posting/editing/deleting jobs - they can only view
            if user.get('role') == 'admin' and action in ('add', 'edit', 'duplicate', 'bulk_update', 'bulk_delete'):
                flash('Admins can only view job postings. Job management is restricted to HR staff.', 'error')
                return redirect(url_for('job_postings'))
            
            actor_admin_id_raw = session.get('user_id')
            # Validate admin_id exists in admins table to satisfy foreign key constraint
            actor_admin_id = get_valid_admin_id(actor_admin_id_raw)

            if action == 'add':
                payload, errors = extract_job_payload(request.form, actor_admin_id)

                if errors:
                    for message in errors:
                        flash(message, 'error')
                    # Redirect back to form on validation errors
                    return redirect(url_for('job_postings'))
                else:
                    try:
                        # Build INSERT statement using actual schema columns
                        ensure_schema_compatibility()
                        _update_job_columns(cursor)
                        
                        # Use actual schema: title, description, requirements, status, branch_id, position_id, posted_by, posted_at
                        # Handle NULL values for optional foreign keys (position_id, posted_by can be NULL)
                        # Ensure title is not None or empty
                        job_title = payload['title'] or 'Untitled Job'
                        if not job_title.strip():
                            job_title = 'Untitled Job'
                        
                        # Check if position_name column exists, create if missing
                        cursor.execute('SHOW COLUMNS FROM jobs LIKE "position_name"')
                        has_position_name = cursor.fetchone() is not None
                        if not has_position_name:
                            try:
                                cursor.execute('ALTER TABLE jobs ADD COLUMN position_name VARCHAR(200) DEFAULT NULL AFTER title')
                                db.commit()
                                has_position_name = True
                                print('✅ Added position_name column to jobs table')
                            except Exception as alter_err:
                                print(f'⚠️ Could not add position_name column: {alter_err}')
                                db.rollback()
                        
                        if has_position_name:
                            # Get position_name from payload - can be None, empty string, or actual value
                            position_value = payload.get('position_name')
                            # Convert None to empty string, but keep empty string as is
                            if position_value is None:
                                position_value = ''
                            # Trim whitespace
                            position_value = str(position_value).strip() if position_value else ''
                            print(f'🔍 Inserting job with position_name: "{position_value}"')
                            # Use MySQL NOW() for accurate server time when status is active/open
                            if payload['status'] in ('active', 'open'):
                                cursor.execute(
                                    '''
                                    INSERT INTO jobs (title, position_name, description, requirements, status, branch_id, posted_by, posted_at)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
                                    ''',
                                    (
                                        job_title,
                                        position_value,
                                        payload['description'],
                                        payload['requirements'],
                                        payload['status'],
                                        payload['branch_id'],
                                        payload['posted_by'] if payload['posted_by'] else None,
                                    ),
                                )
                            else:
                                cursor.execute(
                                    '''
                                    INSERT INTO jobs (title, position_name, description, requirements, status, branch_id, posted_by, posted_at)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                                    ''',
                                    (
                                        job_title,
                                        position_value,
                                        payload['description'],
                                        payload['requirements'],
                                        payload['status'],
                                        payload['branch_id'],
                                        payload['posted_by'] if payload['posted_by'] else None,
                                        None,
                                    ),
                                )
                            job_id = cursor.lastrowid
                            # Verify position_name was saved
                            cursor.execute('SELECT position_name FROM jobs WHERE job_id = %s', (job_id,))
                            saved_job = cursor.fetchone()
                            if saved_job:
                                saved_position = saved_job.get('position_name') if isinstance(saved_job, dict) else (saved_job[0] if len(saved_job) > 0 else None)
                                print(f'✅ Verified - position_name saved to database: "{saved_position}"')
                        else:
                            # Use MySQL NOW() for accurate server time when status is active/open
                            if payload['status'] in ('active', 'open'):
                                cursor.execute(
                                    '''
                                    INSERT INTO jobs (title, description, requirements, status, branch_id, posted_by, posted_at)
                                    VALUES (%s, %s, %s, %s, %s, %s, NOW())
                                    ''',
                                    (
                                        job_title,
                                        payload['description'],
                                        payload['requirements'],
                                        payload['status'],
                                        payload['branch_id'],
                                        payload['posted_by'] if payload['posted_by'] else None,
                                    ),
                                )
                            else:
                                cursor.execute(
                                    '''
                                    INSERT INTO jobs (title, description, requirements, status, branch_id, posted_by, posted_at)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                                    ''',
                                    (
                                        job_title,
                                        payload['description'],
                                        payload['requirements'],
                                        payload['status'],
                                        payload['branch_id'],
                                        payload['posted_by'] if payload['posted_by'] else None,
                                        None,
                                    ),
                                )
                        job_id = cursor.lastrowid
                        
                        # AUTOMATIC: Handle job status (posted_at, etc.)
                        if job_id:
                            auto_handle_job_status(cursor, job_id, payload['status'])
                        
                        # Create admin notification for new job posting
                        try:
                            hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                            branch_name = 'Unknown Branch'
                            if payload.get('branch_id'):
                                cursor.execute('SELECT branch_name FROM branches WHERE branch_id = %s LIMIT 1', (payload['branch_id'],))
                                branch_row = cursor.fetchone()
                                if branch_row:
                                    branch_name = branch_row.get('branch_name', 'Unknown Branch')
                            admin_msg = f'HR {hr_name} posted a new job: "{payload.get("title", "Untitled")}" at {branch_name}.'
                            create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating notification for job posting: {notify_err}')
                        
                        db.commit()
                        # Verify the job was saved with correct status
                        cursor.execute('SELECT job_id, status, posted_at FROM jobs WHERE job_id = %s', (job_id,))
                        saved_job = cursor.fetchone()
                        print(f'✅ Job posted successfully - ID: {job_id}, Status in DB: {saved_job["status"] if saved_job else "N/A"}, Posted by: {payload.get("posted_by")}')
                        if payload['status'] in ('active', 'open'):
                            flash('Job posting created successfully and is now automatically visible to applicants.', 'success')
                        else:
                            flash('Job posting created successfully. Set status to "Active" to make it visible to applicants.', 'success')
                        # Redirect after successful insert
                        return redirect(url_for('job_postings'))
                    except Exception as db_error:
                        db.rollback()
                        import traceback
                        error_details = traceback.format_exc()
                        print(f'❌ Database insert error: {db_error}')
                        print(f'Full traceback: {error_details}')
                        print(f'Payload: {payload}')
                        # Check for specific database errors
                        error_msg = str(db_error)
                        if 'foreign key constraint' in error_msg.lower():
                            if 'branch_id' in error_msg.lower():
                                flash('Invalid branch selected. Please select a valid branch.', 'error')
                            elif 'position_id' in error_msg.lower():
                                flash('Invalid position selected. Please select a valid position.', 'error')
                            elif 'posted_by' in error_msg.lower() or 'admin' in error_msg.lower():
                                flash('Invalid user account. Please log out and log back in.', 'error')
                            else:
                                flash(f'Database constraint error: {error_msg}', 'error')
                        elif 'cannot be null' in error_msg.lower() or 'not null' in error_msg.lower():
                            flash('Missing required information. Please fill in all required fields.', 'error')
                        else:
                            flash(f'Failed to create job posting: {error_msg}. Please check the console for details.', 'error')
                        # Redirect back to form on database error
                        return redirect(url_for('job_postings'))

            elif action == 'bulk_update':
                job_ids = request.form.getlist('job_ids')
                bulk_status_input = normalize_choice(request.form.get('bulk_status'), VALID_JOB_STATUSES, None)
                
                # Map 'active' to 'open' for database compatibility (database enum: 'open', 'closed')
                if bulk_status_input == 'active':
                    bulk_status = 'open'
                elif bulk_status_input == 'closed':
                    bulk_status = 'closed'
                else:
                    bulk_status = bulk_status_input

                if job_ids and bulk_status:
                    placeholders = ','.join(['%s'] * len(job_ids))
                    # AUTOMATIC: Handle posted_at based on status - use 'open' for database
                    posted_at_value = None
                    if bulk_status == 'open':
                        # Use MySQL NOW() for accurate server time
                        cursor.execute('SELECT NOW() as current_time')
                        current_time = cursor.fetchone()[0] if cursor.fetchone() else None
                        posted_at_value = current_time
                    params = [
                        bulk_status,
                        posted_at_value,
                        *job_ids,
                    ]
                    branch_clause = ''
                    if branch_scope is not None:
                        branch_clause = ' AND branch_id = %s'
                        params.append(branch_scope)

                    cursor.execute(
                        f'''
                        UPDATE jobs
                        SET status = %s,
                            posted_at = %s
                        WHERE job_id IN ({placeholders}){branch_clause}
                        ''',
                        tuple(params),
                    )
                    
                    # AUTOMATIC: Handle status for each job
                    for job_id_val in job_ids:
                        try:
                            auto_handle_job_status(cursor, job_id_val, bulk_status)
                        except Exception:
                            pass  # Non-blocking
                    
                    # Create admin notification for bulk update
                    try:
                        hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                        admin_msg = f'HR {hr_name} bulk updated {len(job_ids)} job posting(s) to status: {bulk_status}.'
                        create_admin_notification(cursor, admin_msg)
                    except Exception as notify_err:
                        print(f'⚠️ Error creating notification for bulk update: {notify_err}')
                    
                    db.commit()
                    flash(f'{len(job_ids)} job posting(s) updated successfully. Status changes are automatically handled.', 'success')
                else:
                    flash('Select job postings and a valid status for bulk update.', 'warning')

            elif action == 'delete':
                job_id = request.form.get('job_id')
                if not job_id:
                    flash('Job ID is required.', 'error')
                else:
                    try:
                        # Verify job exists and belongs to user's branch (if HR)
                        cursor.execute(
                            'SELECT job_id, title, branch_id FROM jobs WHERE job_id = %s LIMIT 1',
                            (job_id,)
                        )
                        job = cursor.fetchone()
                        
                        if not job:
                            flash('Job posting not found.', 'error')
                        else:
                            # Check branch access for HR users
                            if branch_scope is not None:
                                if job.get('branch_id') != branch_scope:
                                    flash('You do not have permission to delete this job posting.', 'error')
                                    if is_ajax:
                                        return jsonify({'success': False, 'error': 'Permission denied.'}), 403
                                    return redirect(url_for('job_postings'))
                            
                            job_title = job.get('title') or 'Job Posting'
                            
                            # Hard delete from database - permanently removes job posting
                            cursor.execute('DELETE FROM jobs WHERE job_id = %s', (job_id,))
                            
                            # Create admin notification for job deletion
                            try:
                                hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                admin_msg = f'HR {hr_name} deleted job posting: "{job_title}"'
                                create_admin_notification(cursor, admin_msg)
                            except Exception as notify_err:
                                print(f'⚠️ Error creating notification for job deletion: {notify_err}')
                            
                            db.commit()
                            
                            if is_ajax:
                                return jsonify({'success': True, 'message': 'Job posting deleted successfully.'})
                            flash('Job posting deleted successfully from system and database.', 'success')
                    except Exception as exc:
                        db.rollback()
                        error_msg = f'Failed to delete job posting: {str(exc)}'
                        print(f'❌ Error deleting job: {exc}')
                        if is_ajax:
                            return jsonify({'success': False, 'error': error_msg}), 500
                        flash(error_msg, 'error')
            
            elif action == 'bulk_delete':
                job_ids = request.form.getlist('job_ids')
                if job_ids:
                    try:
                        placeholders = ','.join(['%s'] * len(job_ids))
                        params = [*job_ids]
                        branch_clause = ''
                        if branch_scope is not None:
                            branch_clause = ' AND branch_id = %s'
                            params.append(branch_scope)

                        # Hard delete from database - permanently removes job postings
                        cursor.execute(
                            f'DELETE FROM jobs WHERE job_id IN ({placeholders}){branch_clause}',
                            tuple(params),
                        )
                        
                        # Create admin notification for bulk delete
                        try:
                            hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                            admin_msg = f'HR {hr_name} deleted {len(job_ids)} job posting(s) from the system.'
                            create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating notification for bulk delete: {notify_err}')
                        
                        db.commit()
                        flash(f'{len(job_ids)} job posting(s) deleted successfully from system and database.', 'success')
                    except Exception as exc:
                        db.rollback()
                        flash(f'Failed to delete job postings: {exc}', 'error')
                else:
                    flash('Select at least one job to delete.', 'warning')

            elif action == 'edit':
                job_id_raw = request.form.get('job_id', '').strip()
                if not job_id_raw or not job_id_raw.isdigit():
                    flash('Invalid job ID.', 'error')
                    return redirect(url_for('job_postings'))
                
                job_id = int(job_id_raw)
                payload, errors = extract_job_payload(request.form, actor_admin_id)

                if errors:
                    for message in errors:
                        flash(message, 'error')
                    # Redirect back to form on validation errors
                    return redirect(url_for('job_postings'))
                else:
                    branch_clause = ''
                    # Automatic posted_at handling based on status
                    posted_at_value = payload.get('posted_at')
                    # If status changed to 'active', set posted_at; if changed to 'closed', keep existing posted_at
                    
                    # Map payload keys correctly
                    job_title = payload.get('title') or payload.get('job_title') or ''
                    job_description = payload.get('description') or payload.get('job_description') or ''
                    job_requirements = payload.get('requirements') or payload.get('job_requirements') or ''
                    
                    # Check if position_name column exists, create if missing
                    cursor.execute('SHOW COLUMNS FROM jobs LIKE "position_name"')
                    has_position_name = cursor.fetchone() is not None
                    if not has_position_name:
                        try:
                            cursor.execute('ALTER TABLE jobs ADD COLUMN position_name VARCHAR(200) DEFAULT NULL AFTER title')
                            db.commit()
                            has_position_name = True
                            print('✅ Added position_name column to jobs table')
                        except Exception as alter_err:
                            print(f'⚠️ Could not add position_name column: {alter_err}')
                            db.rollback()
                    
                    position_name = payload.get('position_name')
                    
                    if has_position_name:
                        params = [
                            job_title,
                            position_name,
                            job_description,
                            job_requirements,
                            payload.get('status', 'active'),
                            payload.get('branch_id'),
                            posted_at_value,
                            job_id,
                        ]
                        if branch_scope is not None:
                            branch_clause = ' AND branch_id = %s'
                            params.append(branch_scope)

                        cursor.execute(
                            f'''
                            UPDATE jobs
                            SET title = %s,
                                position_name = %s,
                                description = %s,
                                requirements = %s,
                                status = %s,
                                branch_id = %s,
                                posted_at = %s
                            WHERE job_id = %s{branch_clause}
                            ''',
                            tuple(params),
                        )
                    else:
                        params = [
                            job_title,
                            job_description,
                            job_requirements,
                            payload.get('status', 'active'),
                            payload.get('branch_id'),
                            posted_at_value,
                            job_id,
                        ]
                        if branch_scope is not None:
                            branch_clause = ' AND branch_id = %s'
                            params.append(branch_scope)

                        cursor.execute(
                            f'''
                            UPDATE jobs
                            SET title = %s,
                                description = %s,
                                requirements = %s,
                                status = %s,
                                branch_id = %s,
                                posted_at = %s
                            WHERE job_id = %s{branch_clause}
                            ''',
                            tuple(params),
                        )
                    # AUTOMATIC: Handle job status changes
                    auto_handle_job_status(cursor, job_id, payload.get('status', 'active'))
                    
                    # Create admin notification for job edit
                    try:
                        hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                        admin_msg = f'HR {hr_name} updated job posting: "{job_title}" (ID: {job_id}).'
                        create_admin_notification(cursor, admin_msg)
                    except Exception as notify_err:
                        print(f'⚠️ Error creating notification for job edit: {notify_err}')
                    
                    db.commit()
                    flash('Job posting updated successfully. Changes are now visible in the job postings list.', 'success')
                    # Redirect after successful update
                    return redirect(url_for('job_postings'))

            elif action == 'duplicate':
                job_id_raw = request.form.get('job_id', '').strip()
                if not job_id_raw or not job_id_raw.isdigit():
                    flash('Invalid job ID.', 'error')
                    return redirect(url_for('job_postings'))
                
                job_id = int(job_id_raw)
                branch_clause = ''
                params = [job_id]
                if branch_scope is not None:
                    branch_clause = ' AND branch_id = %s'
                    params.append(branch_scope)

                # Use actual schema columns
                cursor.execute(
                    f'''
                    SELECT title, description, requirements, status, branch_id, position_id, posted_by
                    FROM jobs
                    WHERE job_id = %s{branch_clause}
                    ''',
                    tuple(params),
                )
                original_job = cursor.fetchone()
                
                if original_job:
                    # Use actual schema columns for duplicate
                    cursor.execute(
                        '''
                        INSERT INTO jobs (title, description, requirements, status, branch_id, posted_by, posted_at)
                        VALUES (%s, %s, %s, %s, %s, %s, NOW())
                        ''',
                        (
                            f"{original_job.get('title') or original_job.get('job_title', 'Untitled')} (Copy)",
                            original_job.get('description') or original_job.get('job_description', ''),
                            original_job.get('requirements') or original_job.get('job_requirements', ''),
                            'open',  # Default to open for new job (active status)
                            original_job.get('branch_id'),
                            actor_admin_id,
                        ),
                    )
                    db.commit()
                    flash('Job posting duplicated successfully.', 'success')
                else:
                    flash('Job not found.', 'error')

            return redirect(url_for('job_postings'))

        # Apply filters - extract and validate all filter parameters
        filters = {}
        
        # Keyword filter (case-insensitive, trim whitespace)
        keyword_raw = (request.args.get('keyword', '') or request.args.get('search', '')).strip()
        if keyword_raw:
            filters['keyword'] = keyword_raw
        
        # Branch filter - HR users can filter by branch (they manage all branches)
        # Admin users can also filter by branch
        branch_id_raw = request.args.get('branch_id', '').strip()
        if branch_id_raw and branch_id_raw.isdigit():
            filters['branch_id'] = int(branch_id_raw)
        
        # Position filter
        position_id_raw = request.args.get('position_id', '').strip()
        if position_id_raw and position_id_raw.isdigit():
            filters['position_id'] = int(position_id_raw)
        
        # Status filter (validate against valid statuses)
        status_raw = request.args.get('status', '').strip().lower()
        if status_raw and status_raw in VALID_JOB_STATUSES:
            filters['status'] = status_raw
        
        # Employment type filter
        employment_type_raw = request.args.get('employment_type', '').strip().lower()
        if employment_type_raw and employment_type_raw in VALID_EMPLOYMENT_TYPES:
            filters['employment_type'] = employment_type_raw
        
        # Work arrangement filter
        work_arrangement_raw = request.args.get('work_arrangement', '').strip().lower()
        if work_arrangement_raw and work_arrangement_raw in VALID_WORK_ARRANGEMENTS:
            filters['work_arrangement'] = work_arrangement_raw
        
        # Experience level filter
        experience_level_raw = request.args.get('experience_level', '').strip().lower()
        if experience_level_raw and experience_level_raw in VALID_EXPERIENCE_LEVELS:
            filters['experience_level'] = experience_level_raw
        
        # Department filter
        department_raw = request.args.get('department', '').strip()
        if department_raw:
            filters['department'] = department_raw
        
        # Location filter
        location_raw = request.args.get('location', '').strip()
        if location_raw:
            filters['location'] = location_raw

        # Build WHERE clauses
        where_clauses = []
        params = []

        # Branch filter - apply branch_scope if set, otherwise use branch_id filter
        # HR users can filter by branch_id even though they manage all branches
        if branch_scope is not None:
            # If HR has a specific branch scope, use it (but this should be None for HR now)
            where_clauses.append('j.branch_id = %s')
            params.append(branch_scope)
        elif filters.get('branch_id'):
            # Apply branch filter if specified (for both HR and Admin)
            where_clauses.append('j.branch_id = %s')
            params.append(filters['branch_id'])

        # Keyword search (case-insensitive, searches title, description, and requirements)
        if filters.get('keyword'):
            keyword = f"%{filters['keyword']}%"
            # Use dynamic column checking for keyword fields
            job_title_col = job_column('job_title', 'title')
            job_desc_col = job_column('job_description', 'description')
            job_req_col = job_column('job_requirements', 'requirements')
            
            keyword_fields = [col for col in [job_title_col, job_desc_col, job_req_col] if col]
            if keyword_fields:
                # Use LOWER() for case-insensitive search
                like_clauses = [f"LOWER(j.{column}) LIKE LOWER(%s)" for column in keyword_fields]
                where_clauses.append(f"({' OR '.join(like_clauses)})")
                params.extend([keyword] * len(like_clauses))

        # Position filter removed - positions table no longer exists
        # if filters.get('position_id'):
        #     where_clauses.append('j.position_id = %s')
        #     params.append(filters['position_id'])

        # Status filter - map 'active' to show all active statuses ('open', 'published', 'active')
        if filters.get('status'):
            status_filter = filters['status']
            if status_filter == 'active':
                # Show all active statuses: 'open', 'published', 'active' (database may have different status values)
                where_clauses.append('j.status IN (%s, %s, %s)')
                params.extend(['open', 'published', 'active'])
            else:
                # For 'closed' or other statuses, use exact match
                where_clauses.append('j.status = %s')
                params.append(status_filter)

        # Department filter removed (positions table no longer exists)
        # if filters.get('department'):
        #     where_clauses.append('p.department = %s')
        #     params.append(filters['department'])

        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'

        # Ensure job columns are updated before building expressions
        _update_job_columns(cursor)
        
        # Check if position_name column exists in jobs table, create if missing
        cursor.execute('SHOW COLUMNS FROM jobs LIKE "position_name"')
        has_position_name_col = cursor.fetchone() is not None
        if not has_position_name_col:
            try:
                cursor.execute('ALTER TABLE jobs ADD COLUMN position_name VARCHAR(200) DEFAULT NULL AFTER title')
                db.commit()
                has_position_name_col = True
                print('✅ Added position_name column to jobs table')
            except Exception as alter_err:
                print(f'⚠️ Could not add position_name column: {alter_err}')
                db.rollback()
        
        # Define all column expressions for SELECT clause
        # Use COALESCE to handle NULL values and ensure we always get a title
        job_title_col = job_column('job_title', 'title')
        if job_title_col:
            job_title_expr = f"COALESCE(j.{job_title_col}, 'Untitled Job')"
        else:
            job_title_expr = "'Untitled Job'"
        job_description_expr = job_column_expr('job_description', alternatives=['description'])
        job_requirements_expr = job_column_expr('job_requirements', alternatives=['requirements'])
        
        # Build position_name expression - use j.position_name directly (positions table removed)
        # According to schema: jobs.position_name VARCHAR(200) DEFAULT NULL
        if has_position_name_col:
            # Use j.position_name directly (what user entered in the form)
            position_name_expr = f'COALESCE(j.position_name, "")'
        else:
            position_name_expr = '""'
        created_at_expr = job_column_expr('created_at')
        posted_at_expr = job_column_expr('posted_at', alternatives=['created_at'])
        position_id_expr = job_column_expr('position_id', default='NULL')
        branch_id_expr = job_column_expr('branch_id', default='NULL')
        status_expr = job_column_expr('status', default="'open'")  # Database uses 'open', not 'active'
        
        # Build admin join conditions dynamically for both created_by and posted_by
        if 'created_by_admin_id' in JOB_COLUMNS:
            created_by_join = 'LEFT JOIN admins a_created ON j.created_by_admin_id = a_created.admin_id'
        else:
            created_by_join = 'LEFT JOIN admins a_created ON NULL = a_created.admin_id'
        
        if 'posted_by' in JOB_COLUMNS:
            posted_by_join = 'LEFT JOIN admins a_posted ON j.posted_by = a_posted.admin_id'
        else:
            posted_by_join = 'LEFT JOIN admins a_posted ON NULL = a_posted.admin_id'

        # Sort order handling
        sort_order = request.args.get('sort', 'newest').strip().lower()
        order_by_clause = f'COALESCE({posted_at_expr}, {created_at_expr}) DESC'  # Default: newest first
        
        if sort_order == 'oldest':
            order_by_clause = f'COALESCE({posted_at_expr}, {created_at_expr}) ASC'
        elif sort_order == 'title_asc':
            order_by_clause = f'{job_title_expr} ASC'
        elif sort_order == 'title_desc':
            order_by_clause = f'{job_title_expr} DESC'
        elif sort_order == 'applications_desc':
            order_by_clause = '(SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) DESC'
        elif sort_order == 'applications_asc':
            order_by_clause = '(SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) ASC'
        
        # Store sort in filters for template
        if sort_order != 'newest':
            filters['sort'] = sort_order

        query = f'''
            SELECT
                j.job_id,
                COALESCE({job_title_expr}, 'Untitled Job') AS job_title,
                {job_description_expr} AS job_description,
                {job_requirements_expr} AS job_requirements,
                {status_expr} AS status,
                {branch_id_expr} AS branch_id,
                {position_id_expr} AS position_id,
                {created_at_expr} AS created_at,
                {posted_at_expr} AS posted_at,
                COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                {position_name_expr} AS position_name,
                'General' AS department,
                COALESCE(a_posted.full_name, 'System') AS posted_by_name,
                (SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) AS application_count
            FROM jobs j
            LEFT JOIN branches b ON {branch_id_expr} = b.branch_id
            {posted_by_join}
            WHERE {where_sql}
            ORDER BY {order_by_clause}
        '''

        cursor.execute(query, tuple(params) if params else None)
        jobs_raw = cursor.fetchall()
        
        # Debug: Check if position_name is in results
        if jobs_raw and len(jobs_raw) > 0:
            sample_job = jobs_raw[0]
            if isinstance(sample_job, dict):
                print(f'🔍 Sample job position_name from query: "{sample_job.get("position_name")}"')
            else:
                print(f'🔍 Sample job (non-dict): {sample_job}')
        
        # Ensure position_name exists in all job results
        for job in jobs_raw:
            if isinstance(job, dict):
                if 'position_name' not in job:
                    job['position_name'] = ''
                # Debug each job's position_name
                if job.get('job_id'):
                    print(f'🔍 Job ID {job.get("job_id")} position_name: "{job.get("position_name")}"')

        def build_option_list(values):
            return [{'value': value, 'label': value.replace('_', ' ').title()} for value in values]

        formatted_jobs = []
        for job in jobs_raw:
            salary_display = format_salary_range(job.get('salary_currency'), job.get('salary_min'), job.get('salary_max'))
            posted_ts = job.get('posted_at') or job.get('created_at')
            formatted_jobs.append(
                {
                    'job_id': job.get('job_id'),
                    'job_title': job.get('job_title'),
                    'title': job.get('job_title'),
                    'job_summary': (job.get('job_description') or '')[:200] if job.get('job_description') else '',
                    'job_description': job.get('job_description'),
                    'job_requirements': job.get('job_requirements'),
                    'employment_type': job.get('employment_type'),
                    'work_arrangement': job.get('work_arrangement'),
                    'experience_level': job.get('experience_level'),
                    'job_location': job.get('job_location'),
                    'salary_currency': job.get('salary_currency'),
                    'salary_min': job.get('salary_min'),
                    'salary_max': job.get('salary_max'),
                    'salary_display': salary_display,
                    'application_deadline': job.get('application_deadline'),
                    'application_deadline_display': format_human_datetime(job.get('application_deadline')),
                    'status': job.get('status'),
                    'branch_id': job.get('branch_id'),
                    'branch_name': job.get('branch_name'),
                    'position_id': job.get('position_id'),
                    'position_name': (job.get('position_name') or '').strip() if job.get('position_name') else '',  # Include position_name from database, trim whitespace
                    'position_title': job.get('position_name') or job.get('job_title') or job.get('title'),
                    'department': job.get('department'),
                    'created_by_name': job.get('created_by_name'),
                    'posted_by_name': job.get('posted_by_name'),
                    'application_count': job.get('application_count', 0),
                    'created_at': format_human_datetime(job.get('created_at')),
                    'updated_at': format_human_datetime(job.get('updated_at')),
                    'posted_at': format_human_datetime(posted_ts),
                }
            )

        branches = fetch_branches()
        positions = fetch_positions()

        current_branch = None
        if branch_scope is not None:
            current_branch = next((branch for branch in branches if branch.get('branch_id') == branch_scope), None)

        # Get unique departments and locations for filters
        # Positions table removed - no departments available
        departments = []

        # Location filter removed (not in actual schema)
        locations = []

        job_meta = {
            'status_options': build_option_list(VALID_JOB_STATUSES),
            'departments': departments,
            'locations': locations,
        }

        template = 'hr/job_postings.html' if user.get('role') == 'hr' else 'admin/job_postings.html'
        branch_info = current_branch

        # Normalize filters for template comparison (ensure IDs are integers)
        normalized_filters = {}
        for key, value in filters.items():
            if key in ['branch_id', 'position_id']:
                try:
                    normalized_filters[key] = int(value) if value is not None else None
                except (ValueError, TypeError):
                    normalized_filters[key] = None
            else:
                normalized_filters[key] = value

        return render_template(
            template,
            jobs=formatted_jobs,
            branches=branches,
            positions=positions,
            user=user,
            current_branch=current_branch,
            current_filters=normalized_filters,
            branch_info=branch_info,
            job_meta=job_meta,
        )
    except Exception as exc:
        if db:
            db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Job postings error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        try:
            branches = fetch_branches() or []
            positions = fetch_positions() or []
        except Exception:
            branches = []
            positions = []

        job_meta = {
            'status_options': [{'value': value, 'label': value.replace('_', ' ').title()} for value in VALID_JOB_STATUSES],
            'employment_types': [{'value': value, 'label': value.replace('_', ' ').title()} for value in VALID_EMPLOYMENT_TYPES],
            'work_arrangements': [{'value': value, 'label': value.replace('_', ' ').title()} for value in VALID_WORK_ARRANGEMENTS],
            'experience_levels': [{'value': value, 'label': value.replace('_', ' ').title()} for value in VALID_EXPERIENCE_LEVELS],
        }
        template = 'hr/job_postings.html' if (user or {}).get('role') == 'hr' else 'admin/job_postings.html'
        return render_template(
            template,
            jobs=[],
            branches=branches,
            positions=positions,
            user=user or {},
            current_branch=None,
            current_filters={},
            branch_info=None,
            job_meta=job_meta,
        )
    finally:
        cursor.close()


@app.route('/admin/job-postings/<int:job_id>/update', methods=['POST'])
@login_required('admin', 'hr')
def update_job_posting(job_id):
    user = get_current_user()
    
    # Block admins from updating jobs - they can only view
    if user.get('role') == 'admin':
        flash('Admins can only view job postings. Job updates are restricted to HR staff.', 'error')
        return redirect(url_for('job_postings'))
    
    branch_scope = get_branch_scope(user)
    actor_admin_id = session.get('user_id')

    payload, errors = None, []

    def extract_payload(form):
        local_payload, local_errors = {}, []

        job_title = (form.get('job_title') or form.get('title') or '').strip()
        job_summary = ''  # job_summary column doesn't exist in schema
        job_description = (form.get('job_description') or form.get('description') or '').strip()
        job_requirements = (form.get('job_requirements') or form.get('requirements') or '').strip()
        position_name = (form.get('position') or form.get('position_name') or '').strip()
        # Keep as empty string if empty (don't convert to None) - matches schema DEFAULT NULL but allows empty strings
        print(f'🔍 Extracted position_name from form (update): "{position_name}"')
        
        # Get position_id from form if position was selected from dropdown
        position_id_raw = form.get('position_id', '').strip()
        if position_id_raw:
            try:
                position_id = int(position_id_raw)
            except (ValueError, TypeError):
                position_id = None
        else:
            position_id = None

        employment_type = normalize_choice(
            form.get('employment_type'),
            VALID_EMPLOYMENT_TYPES,
            VALID_EMPLOYMENT_TYPES[0],
        )

        # Handle status - database uses 'open' but code can use 'active' (they're equivalent for visibility)
        status_input = form.get('status', '').strip()
        # Map 'active' to 'open' for database compatibility, or use 'open' directly
        if status_input == 'active':
            status = 'open'  # Database uses 'open' as the active status
        elif status_input == 'open':
            status = 'open'
        else:
            status = normalize_choice(status_input, VALID_JOB_STATUSES, 'open')
        
        # Ensure default to 'open' if not set
        if not status_input or status_input == '':
            status = 'open'

        if branch_scope is not None:
            branch_id = branch_scope
        else:
            branch_raw = (form.get('branch_id') or '').strip()
            branch_id = int(branch_raw) if branch_raw.isdigit() else None
            if branch_id is None:
                local_errors.append('Branch is required for job postings.')

        # Position field removed - always set to None
        position_id = None

        if not job_title:
            local_errors.append('Job title is required.')
        # job_summary removed - not in schema
        if not job_description:
            local_errors.append('Job description is required.')
        if not job_requirements:
            local_errors.append('Job requirements are required.')

        # Automatic status handling - posted_at automatically set based on status
        # Note: posted_at is handled in the UPDATE query, not here

        local_payload = {
            'job_title': job_title,
            'title': job_title,  # Also include as 'title' for compatibility
            'job_description': job_description,
            'description': job_description,  # Also include as 'description' for compatibility
            'job_requirements': job_requirements,
            'requirements': job_requirements,  # Also include as 'requirements' for compatibility
            'status': status,
            'branch_id': branch_id,
        }

        return local_payload, local_errors

    payload, errors = extract_payload(request.form)
    
    # Debug: Print form data
    import traceback
    print(f'🔍 Update job posting - Job ID: {job_id}')
    print(f'🔍 Form data received: {dict(request.form)}')
    print(f'🔍 Payload extracted: {payload}')
    print(f'🔍 Errors: {errors}')

    if errors:
        for message in errors:
            flash(message, 'error')
        return redirect(url_for('job_postings'))

    try:
        # Validate admin_id exists in admins table to satisfy foreign key constraint
        actor_admin_id = get_valid_admin_id(actor_admin_id)
        
        # Build UPDATE statement dynamically based on available columns
        ensure_schema_compatibility()
        db = get_db()
        if db:
            cursor_temp = db.cursor()
            try:
                _update_job_columns(cursor_temp)
            finally:
                cursor_temp.close()
        
        # Determine which columns to use
        use_updated_by = 'updated_by_admin_id' in JOB_COLUMNS
        
        # AUTOMATIC: Status handling - get status from payload (already mapped to 'open' if 'active')
        status = payload.get('status', 'open')  # Default to 'open' (database value)
        print(f'🔍 Update - Status from payload: {status}')
        
        # Use actual schema columns
        # Map payload keys to actual column values
        job_title = payload.get('job_title') or payload.get('title') or ''
        job_description = payload.get('job_description') or payload.get('description') or ''
        job_requirements = payload.get('job_requirements') or payload.get('requirements') or ''
        
        # Build SET clauses - update both title and job_title if both exist
        # IMPORTANT: Build in the correct order to match params
        set_clauses = []
        params = []
        
        # Update title column (primary) - add first
        if 'title' in JOB_COLUMNS:
            set_clauses.append('title = %s')
            params.append(job_title)
        
        # Core fields
        set_clauses.extend([
            'description = %s', 
            'requirements = %s',
            'status = %s'
        ])
        params.extend([
            job_description,
            job_requirements,
            status
        ])
        
        # Add position_name if column exists - check before building final query, create if missing
        position_name = payload.get('position_name')
        has_position_name = False
        if db:
            cursor_check = db.cursor()
            try:
                cursor_check.execute('SHOW COLUMNS FROM jobs LIKE "position_name"')
                has_position_name = cursor_check.fetchone() is not None
                if not has_position_name:
                    try:
                        cursor_check.execute('ALTER TABLE jobs ADD COLUMN position_name VARCHAR(200) DEFAULT NULL AFTER title')
                        db.commit()
                        has_position_name = True
                        print('✅ Added position_name column to jobs table')
                    except Exception as alter_err:
                        print(f'⚠️ Could not add position_name column: {alter_err}')
                        db.rollback()
                
                if has_position_name:
                    set_clauses.append('position_name = %s')
                    # Convert None to empty string, trim whitespace
                    position_value = str(position_name).strip() if position_name else ''
                    params.append(position_value)
                    print(f'🔍 position_name column exists, adding to update: "{position_value}"')
            except Exception as pos_err:
                print(f'⚠️ Error checking position_name column: {pos_err}')
                import traceback
                print(traceback.format_exc())
            finally:
                cursor_check.close()
        
        # Only update branch_id if user is admin (not HR) and branch_id is provided
        # HR users cannot change branch_id - it's set by their branch_scope
        if branch_scope is None and payload.get('branch_id'):
            set_clauses.append('branch_id = %s')
            params.append(payload.get('branch_id'))
        
        # Also update job_title if it exists (for compatibility)
        if 'job_title' in JOB_COLUMNS:
            set_clauses.append('job_title = %s')
            params.append(job_title)
        
        # AUTOMATIC: Handle posted_at based on status
        # When status is active/open, set posted_at to current server time (NOW())
        # This ensures the time matches exactly when the job was posted
        if status in ('active', 'open'):  # Both 'active' and 'open' are visible statuses
            set_clauses.append('posted_at = NOW()')
        # For 'closed', keep existing posted_at (historical record) - don't update it
        
        if use_updated_by:
            set_clauses.append('updated_by_admin_id = %s')
            params.append(actor_admin_id)
        
        # Only add updated_at if the column exists in the schema
        if 'updated_at' in JOB_COLUMNS:
            set_clauses.append('updated_at = NOW()')
        
        params.append(job_id)
        where_clause = 'job_id = %s'
        
        branch_clause = ''
        # For HR users, restrict updates to their branch only
        # For admin users, allow updates to any job
        if branch_scope is not None:
            branch_clause = ' AND branch_id = %s'
            params.append(branch_scope)
            print(f'🔍 Branch scope restriction applied: branch_id must be {branch_scope}')

        if not db:
            flash('Database connection error.', 'error')
            return redirect(url_for('job_postings'))
        
        cursor = db.cursor(dictionary=True)
        try:
            # First, check the current job's branch_id to verify it matches branch_scope
            if branch_scope is not None:
                cursor.execute('SELECT branch_id, status FROM jobs WHERE job_id = %s', (job_id,))
                job_check = cursor.fetchone()
                if job_check:
                    current_branch_id = job_check.get('branch_id') if isinstance(job_check, dict) else (job_check[0] if len(job_check) > 0 else None)
                    current_status = job_check.get('status') if isinstance(job_check, dict) else (job_check[1] if len(job_check) > 1 else None)
                    print(f'🔍 Current job branch_id: {current_branch_id}, Required branch_scope: {branch_scope}')
                    print(f'🔍 Current job status: {current_status}')
                    if current_branch_id != branch_scope:
                        flash(f'You do not have permission to update this job. It belongs to a different branch.', 'error')
                        return redirect(url_for('job_postings'))
            
            # Debug: Print update query details
            print(f'🔍 Update query - SET clauses: {set_clauses}')
            print(f'🔍 Update query - Params: {params}')
            print(f'🔍 Update query - WHERE: {where_clause}{branch_clause}')
            print(f'🔍 Update query - Job ID: {job_id}')
            print(f'🔍 Update query - Status value being saved: {status}')
            
            # Build the full SQL query for debugging
            full_query = f'''
                UPDATE jobs
                SET {', '.join(set_clauses)}
                WHERE {where_clause}{branch_clause}
            '''
            print(f'🔍 Full SQL Query: {full_query}')
            print(f'🔍 Query Parameters ({len(params)}): {params}')
            print(f'🔍 SET clauses ({len(set_clauses)}): {set_clauses}')
            
            # Count actual placeholders (excluding NULL and NOW())
            placeholder_count = sum(1 for clause in set_clauses if '= %s' in clause)
            placeholder_count += 1  # job_id in WHERE
            if branch_clause:
                placeholder_count += 1  # branch_id in WHERE
            
            if placeholder_count != len(params):
                error_msg = f'Parameter count mismatch: {placeholder_count} placeholders but {len(params)} parameters. SET clauses: {set_clauses}, Params: {params}'
                print(f'❌ {error_msg}')
                flash(f'Database error: Parameter mismatch. Please check the console for details.', 'error')
                db.rollback()
                return redirect(url_for('job_postings'))
            
            try:
                cursor.execute(
                    f'''
                    UPDATE jobs
                    SET {', '.join(set_clauses)}
                    WHERE {where_clause}{branch_clause}
                    ''',
                    tuple(params),
                )
            except Exception as sql_err:
                print(f'❌ SQL execution error: {sql_err}')
                print(f'🔍 SQL: UPDATE jobs SET {", ".join(set_clauses)} WHERE {where_clause}{branch_clause}')
                print(f'🔍 Params: {params}')
                raise
            
            print(f'🔍 Update query executed - Rows affected: {cursor.rowcount}')
            
            # Verify the update by checking the database
            if cursor.rowcount > 0:
                cursor.execute('SELECT status, position_name FROM jobs WHERE job_id = %s', (job_id,))
                updated_job = cursor.fetchone()
                if updated_job:
                    db_status = updated_job.get('status') if isinstance(updated_job, dict) else (updated_job[0] if len(updated_job) > 0 else None)
                    db_position = updated_job.get('position_name') if isinstance(updated_job, dict) else (updated_job[1] if len(updated_job) > 1 else None)
                    print(f'🔍 Verified - Job status in database after update: {db_status}')
                    print(f'🔍 Verified - Job position_name in database after update: "{db_position}"')
            
            if cursor.rowcount == 0:
                flash('Job posting not found or you do not have permission to update it.', 'error')
                print(f'⚠️ No rows updated for job_id: {job_id}, branch_scope: {branch_scope}')
                return redirect(url_for('job_postings'))
            else:
                # AUTOMATIC: Handle job status changes
                auto_handle_job_status(cursor, job_id, status)
                db.commit()
                print(f'✅ Job posting {job_id} updated successfully in database')
                flash('Job posting updated successfully. Changes are now visible in the job postings list.', 'success')
                # Redirect after successful update
                return redirect(url_for('job_postings'))
        except Exception as db_exc:
            db.rollback()
            import traceback
            error_details = traceback.format_exc()
            print(f'❌ Database update error: {db_exc}')
            print(f'Full traceback: {error_details}')
            flash(f'Failed to update job posting: {db_exc}. Please check the console for details.', 'error')
        finally:
            cursor.close()
            
    except Exception as exc:
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Update job posting error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Failed to update job posting: {exc}. Please check the console for details.', 'error')

    return redirect(url_for('job_postings'))


@app.route('/admin/applicants', methods=['GET', 'POST'])
@login_required('admin', 'hr')
def applicants():
    """Enhanced candidate tracking with filters, stats, and comprehensive data."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        template = 'hr/applicants.html' if user.get('role') == 'hr' else 'admin/applicants.html'
        return render_template(template, applications=[], branch_info=None, stats={}, filters={}, positions=[], jobs=[], branches=[], view_mode='list')
    
    cursor = db.cursor(dictionary=True)
    try:
        # Ensure schema compatibility before querying
        ensure_schema_compatibility()
        branch_id = get_branch_scope(user)
        
        # Handle POST actions (bulk updates, etc.)
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'delete':
                # Only admin can delete applicants from system users
                if user.get('role') != 'admin':
                    flash('Access denied. Only administrators can delete applicants.', 'error')
                    return redirect(url_for('applicants'))
                
                applicant_id = request.form.get('applicant_id')
                if applicant_id:
                    try:
                        # Get user_id from applicants table before deleting
                        cursor.execute(
                            'SELECT user_id, full_name, email FROM applicants WHERE applicant_id = %s LIMIT 1',
                            (applicant_id,),
                        )
                        applicant_record = cursor.fetchone()
                        if applicant_record:
                            user_id = applicant_record['user_id']
                            applicant_name = applicant_record.get('full_name', 'Unknown')
                            applicant_email = applicant_record.get('email', 'Unknown')
                            
                            print(f'🔍 Deleting applicant: applicant_id={applicant_id}, user_id={user_id}, name={applicant_name}, email={applicant_email}')
                            
                            # Create admin notification before deleting applicant
                            try:
                                admin_notification_msg = f'Admin deleted applicant account: "{applicant_name}" (Email: {applicant_email}).'
                                create_admin_notification(cursor, admin_notification_msg)
                                print(f'✅ Admin notification created for applicant account deletion: {applicant_name}')
                            except Exception as notify_err:
                                print(f'⚠️ Error creating admin notification for applicant deletion: {notify_err}')
                            
                            # Delete all related data (complete deletion from database and system)
                            # 1. Delete notifications linked to applicant's applications
                            cursor.execute('DELETE FROM notifications WHERE application_id IN (SELECT application_id FROM applications WHERE applicant_id = %s)', (applicant_id,))
                            
                            # 2. Delete interviews related to applicant's applications
                            cursor.execute("SHOW TABLES LIKE 'interviews'")
                            if cursor.fetchone():
                                cursor.execute('DELETE FROM interviews WHERE application_id IN (SELECT application_id FROM applications WHERE applicant_id = %s)', (applicant_id,))
                                print(f'✅ Deleted interviews for applicant {applicant_id}')
                            
                            # 3. Delete saved jobs
                            cursor.execute('DELETE FROM saved_jobs WHERE applicant_id = %s', (applicant_id,))
                            
                            # 4. Delete applications
                            cursor.execute('DELETE FROM applications WHERE applicant_id = %s', (applicant_id,))
                            applications_deleted = cursor.rowcount
                            print(f'✅ Deleted {applications_deleted} application(s) for applicant {applicant_id}')
                            
                            # 5. Delete resumes
                            cursor.execute('DELETE FROM resumes WHERE applicant_id = %s', (applicant_id,))
                            resumes_deleted = cursor.rowcount
                            print(f'✅ Deleted {resumes_deleted} resume(s) for applicant {applicant_id}')
                            
                            # 6. Delete password resets
                            if applicant_email:
                                cursor.execute('DELETE FROM password_resets WHERE user_email = %s', (applicant_email,))
                            
                            # 7. Delete auth sessions
                            cursor.execute("SHOW TABLES LIKE 'auth_sessions'")
                            if cursor.fetchone():
                                if user_id:
                                    cursor.execute('DELETE FROM auth_sessions WHERE user_id = %s', (user_id,))
                            
                            # 8. Delete profile changes history
                            cursor.execute("SHOW TABLES LIKE 'profile_changes'")
                            if cursor.fetchone():
                                if user_id:
                                    cursor.execute('DELETE FROM profile_changes WHERE user_id = %s AND role = %s', (user_id, 'applicant'))
                            
                            # 9. Delete applicant record
                            cursor.execute('DELETE FROM applicants WHERE applicant_id = %s', (applicant_id,))
                            applicant_deleted = cursor.rowcount > 0
                            print(f'✅ Deleted applicant record {applicant_id} from applicants table' if applicant_deleted else f'⚠️ No applicant record found with ID {applicant_id}')
                            
                            # 8. ALWAYS delete user record from users table to ensure removal from system users
                            if user_id:
                                # Try with user_type check first
                                cursor.execute('DELETE FROM users WHERE user_id = %s AND user_type = %s', (user_id, 'applicant'))
                                user_deleted = cursor.rowcount > 0
                                
                                # If no rows affected with user_type check, try without user_type (in case of data inconsistency)
                                if not user_deleted:
                                    print(f'⚠️ No user deleted with user_type check, trying without user_type check...')
                                    cursor.execute('DELETE FROM users WHERE user_id = %s', (user_id,))
                                    user_deleted = cursor.rowcount > 0
                                
                                if user_deleted:
                                    print(f'✅ Successfully deleted user record {user_id} from users table')
                                else:
                                    print(f'⚠️ DELETE query executed but no rows affected for user_id {user_id} - user may not exist in users table')
                            else:
                                print(f'⚠️ No user_id found for applicant {applicant_id} - cannot delete from users table')
                            
                            db.commit()
                            print(f'✅ Applicant deletion completed for applicant_id {applicant_id}')
                            flash(f'Applicant {applicant_name} deleted successfully from system and database.', 'success')
                        else:
                            flash('Applicant not found.', 'error')
                    except Exception as exc:
                        db.rollback()
                        import traceback
                        print(f'❌ Error deleting applicant: {exc}')
                        print(f'❌ Traceback: {traceback.format_exc()}')
                        flash(f'Failed to delete applicant: {exc}', 'error')
                
                return redirect(url_for('applicants'))
            
            elif action == 'bulk_update_status':
                # Restrict admin from changing status - only HR can change status
                if user.get('role') == 'admin':
                    flash('Access denied. Only HR users can change application status.', 'error')
                    return redirect(url_for('applicants'))
                
                application_ids = request.form.getlist('application_ids')
                new_status = request.form.get('bulk_status', '').strip()
                
                # Note: All statuses can now be manually changed via bulk update
                # HR has full control over application statuses
                
                # Map simplified statuses to database statuses
                # ALL statuses can now be manually changed by HR: pending, scheduled, interviewed, hired, rejected
                CANONICAL_STATUS_MAP = {
                    'pending': 'pending',
                    'scheduled': 'scheduled',
                    'interview': 'interviewed',
                    'interviewed': 'interviewed',
                    'hired': 'hired',
                    'rejected': 'rejected',
                    # Legacy mappings for backward compatibility
                    'applied': 'pending',
                    'under_review': 'pending',
                    'reviewed': 'pending',
                    'shortlisted': 'pending',
                    'accepted': 'hired',  # Map old 'accepted' to new 'hired'
                }
                new_status = CANONICAL_STATUS_MAP.get(new_status, new_status)
                
                if application_ids and new_status:
                    # AUTOMATIC: Update status and notify applicants for each application
                    updated = 0
                    for app_id in application_ids:
                        # HR can manage all branches - no branch verification needed
                        # Verify application exists
                        cursor.execute(
                            '''
                            SELECT a.application_id
                            FROM applications a
                            WHERE a.application_id = %s
                            ''',
                            (app_id,),
                        )
                        if not cursor.fetchone():
                            continue
                        if auto_update_application_status(cursor, app_id, new_status):
                            updated += 1
                    
                    # Create admin notification for bulk update
                    try:
                        if user.get('role') == 'hr':
                            hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                            status_display = new_status.replace('_', ' ').title()
                            admin_msg = f'HR {hr_name} bulk updated {updated} candidate(s) to {status_display}.'
                            create_admin_notification(cursor, admin_msg)
                    except Exception as notify_err:
                        print(f'⚠️ Error creating admin notification: {notify_err}')
                    
                    db.commit()
                    # Special message for hired status
                    if new_status.lower() == 'hired':
                        flash(f'Congratulations! {updated} candidate(s) have been marked as HIRED. All applicants have been automatically notified via email and notification.', 'success')
                    else:
                        flash(f'{updated} candidate(s) status updated across the system. Applicants have been automatically notified via email and notification.', 'success')
            
            elif action == 'add_to_talent_pool':
                application_ids = request.form.getlist('application_ids')
                # Talent pool is essentially candidates with status in ['interviewed', 'hired']
                # This is handled by filtering in the view
                flash(f'{len(application_ids)} candidate(s) added to talent pool.', 'success')
            
            return redirect(url_for('applicants'))
        
        # Get filters from request
        status_param = request.args.get('status', '').strip()
        branch_filter = request.args.get('branch_id', type=int)
        filters = {
            'status': status_param.lower() if status_param else '',
            'branch_id': branch_filter,
            'position_id': request.args.get('position_id', type=int),
            'job_id': request.args.get('job_id', type=int),
            'source': request.args.get('source', '').strip(),
            'date_from': request.args.get('date_from', '').strip(),
            'date_to': request.args.get('date_to', '').strip(),
            'search': request.args.get('search', '').strip(),
            'view_mode': request.args.get('view_mode', 'list').strip(),
        }
        # Keep status filter even if empty (empty means "All Status")
        # Keep branch_id if it's not None
        # Filter out other empty values but ALWAYS keep status (even if empty string)
        filters = {k: v for k, v in filters.items() if (v or k == 'status') and (v is not None or k != 'branch_id')}
        
        # Debug: Print received status filter
        print(f"🔍 Received status filter from request: '{status_param}' -> processed: '{filters.get('status', '')}'")
        print(f"🔍 Status filter type: {type(filters.get('status'))}, value: '{filters.get('status')}', empty check: {not filters.get('status')}")
        
        # Build WHERE clauses
        where_clauses = []
        params = []
        
        # HR users can manage all branches - no scoping restrictions
        # Admin users also see all applications from all branches
        # Filter by branch if explicitly requested via filter (not automatic scoping)
        if filters.get('branch_id'):
            where_clauses.append('j.branch_id = %s')
            params.append(filters['branch_id'])
        elif branch_id:
            # Legacy branch_id from user scope (should not apply for HR, but keep for backward compatibility)
            where_clauses.append('j.branch_id = %s')
            params.append(branch_id)
        
        # Status filter - only apply if status is provided and not empty
        # Empty status means "All Status" - show all applicants
        # HR can access all branches - no branch scoping restrictions
        status_filter_value = filters.get('status', '')
        db_status = None  # Initialize for debug logging
        if status_filter_value:
            # Ensure it's a string and strip whitespace
            status_filter_value = str(status_filter_value).strip().lower()
            
            # Database has 5 statuses: pending, scheduled, interviewed, hired, rejected
            # Map display statuses to database statuses
            status_map = {
                'pending': 'pending',
                'scheduled': 'scheduled',
                'interviewed': 'interviewed',
                'accepted': 'hired',  # Frontend sends 'accepted' but database uses 'hired'
                'hired': 'hired',
                'rejected': 'rejected',
            }
            db_status = status_map.get(status_filter_value, status_filter_value)
            
            # Only apply filter if it's a valid status
            # Use exact match for all statuses - when pending is selected, show ONLY pending
            if db_status in APPLICATION_STATUSES:
                where_clauses.append('a.status = %s')
                params.append(db_status)
                print(f"🔍 Status filter applied: '{status_filter_value}' -> WHERE a.status = '{db_status}' (exact match)")
                print(f"🔍 WHERE clause will be: a.status = '{db_status}'")
            else:
                print(f"⚠️ Unknown status filter value: '{status_filter_value}' (mapped to '{db_status}') - ignoring filter")
                db_status = None  # Reset if invalid
        else:
            print(f"🔍 No status filter - showing all applicants from all branches")
        
        # Position filter removed - positions table no longer exists
        # if filters.get('position_id'):
        #     where_clauses.append('j.position_id = %s')
        #     params.append(filters['position_id'])
        
        if filters.get('date_from'):
            where_clauses.append('DATE(a.submitted_at) >= %s')
            params.append(filters['date_from'])
        
        if filters.get('date_to'):
            where_clauses.append('DATE(a.submitted_at) <= %s')
            params.append(filters['date_to'])
        
        if filters.get('search'):
            search_term = f"%{filters['search']}%"
            where_clauses.append('(ap.full_name LIKE %s OR ap.email LIKE %s OR ap.phone_number LIKE %s)')
            params.extend([search_term, search_term, search_term])
        
        # Add job filter for "view applicants per job"
        if filters.get('job_id'):
            where_clauses.append('j.job_id = %s')
            params.append(filters['job_id'])
        
        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        
        # Debug logging
        print(f"🔍 Final WHERE clause: {where_sql}")
        print(f"🔍 Query params: {params}")
        print(f"🔍 Number of WHERE clauses: {len(where_clauses)}")
        
        # Get dynamic column expressions for position_name
        ensure_schema_compatibility()
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
        
        # Check if position_name column exists in jobs table
        cursor.execute('SHOW COLUMNS FROM jobs LIKE "position_name"')
        has_position_name_col = cursor.fetchone() is not None
        
        # Build position_name expression conditionally
        if has_position_name_col:
            position_name_expr = f'COALESCE(j.position_name, {job_title_expr})'
        else:
            position_name_expr = job_title_expr
        
        # Fetch applications with enhanced data including email verification
        # Fix: Check for ANY resume from applicant, not just the one linked to application
        query = f'''
            SELECT 
                a.application_id,
                a.status,
                a.submitted_at,
                ap.applicant_id,
                ap.full_name AS applicant_name,
                ap.email AS applicant_email,
                ap.phone_number AS applicant_phone,
                FALSE AS email_verified,
                ap.created_at AS applicant_created_at,
                j.job_id,
                j.title AS job_title,
                {position_name_expr} AS position_name,
                j.branch_id AS branch_id,
                COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                COALESCE(a.resume_id, (SELECT resume_id FROM resumes WHERE applicant_id = ap.applicant_id ORDER BY uploaded_at DESC LIMIT 1)) AS resume_id,
                r.file_name AS resume_file_name,
                r.file_path AS resume_path,
                (SELECT COUNT(*) FROM interviews i WHERE i.application_id = a.application_id) AS interview_count,
                (SELECT MAX(i.scheduled_date) FROM interviews i WHERE i.application_id = a.application_id) AS last_interview_date
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            LEFT JOIN resumes r ON COALESCE(a.resume_id, (SELECT resume_id FROM resumes WHERE applicant_id = ap.applicant_id ORDER BY uploaded_at DESC LIMIT 1)) = r.resume_id
            WHERE {where_sql}
            ORDER BY a.submitted_at DESC
            '''
        cursor.execute(query, tuple(params) if params else None)
        applications = cursor.fetchall() or []
        
        # Debug: Check status distribution in results
        if status_filter_value and applications:
            status_counts = {}
            for app in applications:
                status = app.get('status', 'unknown')
                status_counts[status] = status_counts.get(status, 0) + 1
            print(f"🔍 Results status distribution: {status_counts}")
            print(f"🔍 Expected status: '{db_status}', Found statuses: {list(status_counts.keys())}")
            if db_status not in status_counts or status_counts.get(db_status, 0) != len(applications):
                print(f"⚠️ WARNING: Filter mismatch! Expected all '{db_status}', but found: {status_counts}")
        
        # Calculate quick stats (without status filter for accurate counts)
        stats_where_clauses = []
        stats_params = []
        if branch_id:
            stats_where_clauses.append('j.branch_id = %s')
            stats_params.append(branch_id)
        # Position filter removed - positions table no longer exists
        # if filters.get('position_id'):
        #     stats_where_clauses.append('j.position_id = %s')
        #     stats_params.append(filters['position_id'])
        if filters.get('date_from'):
            stats_where_clauses.append('DATE(a.submitted_at) >= %s')
            stats_params.append(filters['date_from'])
        if filters.get('date_to'):
            stats_where_clauses.append('DATE(a.submitted_at) <= %s')
            stats_params.append(filters['date_to'])
        if filters.get('search'):
            search_term = f"%{filters['search']}%"
            stats_where_clauses.append('(ap.full_name LIKE %s OR ap.email LIKE %s OR ap.phone_number LIKE %s)')
            stats_params.extend([search_term, search_term, search_term])
        
        stats_where_sql = ' AND '.join(stats_where_clauses) if stats_where_clauses else '1=1'
        
        cursor.execute(
            f'''
            SELECT 
                COUNT(*) AS total_candidates,
                COALESCE(SUM(CASE WHEN DATE(a.submitted_at) = CURDATE() THEN 1 ELSE 0 END), 0) AS new_today,
                COALESCE(SUM(CASE WHEN a.status = 'reviewed' THEN 1 ELSE 0 END), 0) AS in_review,
                COALESCE(SUM(CASE WHEN a.status = 'interviewed' THEN 1 ELSE 0 END), 0) AS interview_stage
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            JOIN jobs j ON a.job_id = j.job_id
            WHERE {stats_where_sql}
            ''',
            tuple(stats_params) if stats_params else None,
        )
        stats_row = cursor.fetchone() or {}
        stats = {
            'total_candidates': stats_row.get('total_candidates', 0) or 0,
            'new_today': stats_row.get('new_today', 0) or 0,
            'in_review': stats_row.get('in_review', 0) or 0,
            'interview_stage': stats_row.get('interview_stage', 0) or 0,
        }
        
        # Get unique positions for filter
        # Positions table removed - return empty list
        positions = []
        
        # Get all branches for filter dropdown
        branches = fetch_branches()
        
        # Get unique jobs for filter dropdown
        job_params = []
        job_query = '''
            SELECT DISTINCT j.job_id, j.title AS job_title
            FROM jobs j
            JOIN applications a ON a.job_id = j.job_id
        '''
        if filters.get('branch_id'):
            job_query += ' WHERE j.branch_id = %s'
            job_params.append(filters['branch_id'])
        elif branch_id:
            job_query += ' WHERE j.branch_id = %s'
            job_params.append(branch_id)
        job_query += ' ORDER BY j.title'
        cursor.execute(job_query, tuple(job_params) if job_params else None)
        jobs = cursor.fetchall() or []
        
        # Format applications data
        formatted_applications = []
        for app in applications:
            try:
                interview_count = app.get('interview_count', 0) or 0
                status_value = (app.get('status') or 'pending').strip().lower()
                # Normalize withdrawn to rejected - remove withdrawn status completely
                if status_value == 'withdrawn':
                    status_value = 'rejected'
                # Note: Status is already managed by interview scheduling/completion logic
                # 'scheduled' when interview is scheduled, 'interviewed' when interview is completed
                # So we don't auto-override status based on interview_count
                formatted_applications.append({
                    'application_id': app.get('application_id'),
                    'applicant_id': app.get('applicant_id'),
                    'applicant_name': app.get('applicant_name') or '—',
                    'applicant_email': app.get('applicant_email') or '—',
                    'applicant_phone': app.get('applicant_phone') or '—',
                    'email_verified': app.get('email_verified', False),
                    'position_name': app.get('position_name') or 'Unassigned',
                    'position_id': app.get('position_id'),
                    'job_title': app.get('job_title') or '—',
                    'job_id': app.get('job_id'),
                    'branch_id': app.get('branch_id'),  # Add branch_id for filtering
                    'branch_name': app.get('branch_name') or 'Unassigned',
                    'status': status_value,  # Use normalized status (withdrawn -> rejected)
                    'submitted_at': format_human_datetime(app.get('submitted_at')) or '—',
                    'submitted_at_raw': app.get('submitted_at'),
                    'resume_id': app.get('resume_id'),
                    'resume_file_name': app.get('resume_file_name'),
                    'has_resume': app.get('resume_id') is not None,
                    'interview_count': app.get('interview_count', 0),
                    'last_interview_date': format_human_datetime(app.get('last_interview_date')) if app.get('last_interview_date') else None,
                    'applicant_created_at': format_human_datetime(app.get('applicant_created_at')),
                })
            except Exception as format_exc:
                print(f'⚠️ Error formatting application {app.get("application_id")}: {format_exc}')
                continue
        
        # Render HR template if user is HR, otherwise admin template
        template = 'hr/applicants.html' if user.get('role') == 'hr' else 'admin/applicants.html'
        branch_info = None
        if user.get('role') == 'hr':
            branch_id_session = session.get('branch_id')
            if branch_id_session:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id_session,))
                if branch_rows:
                    branch_info = branch_rows[0]
        
        return render_template(
            template,
            applications=formatted_applications or [],
            branch_info=branch_info,
            dashboard_data={'branch_info': branch_info},
            stats=stats,
            filters=filters,
            positions=positions,
            jobs=jobs,
            branches=branches,
            view_mode=filters.get('view_mode', 'list'),
        )
    except Exception as exc:
        db.rollback()
        print(f'❌ Applicants tracking error: {exc}')
        import traceback
        traceback.print_exc()
        flash('An error occurred while loading candidates.', 'error')
        template = 'hr/applicants.html' if user.get('role') == 'hr' else 'admin/applicants.html'
        return render_template(template, applications=[], branch_info=None, dashboard_data={}, stats={}, filters={}, positions=[], jobs=[], branches=[], view_mode='list')
    finally:
        cursor.close()


@app.route('/admin/applications', methods=['GET', 'POST'])
@login_required('admin', 'hr')
def applications():
    """Applications management - Disabled."""
    user = get_current_user()
    # Block all users - redirect to dashboard
    if user.get('role') == 'admin':
        flash('Applications feature has been disabled.', 'error')
        return redirect(url_for('admin_dashboard'))
    elif user.get('role') == 'hr':
        flash('Applications feature has been disabled.', 'error')
        return redirect(url_for('hr_dashboard'))
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/applications.html', applications=[], analytics={}, branches=[], jobs=[], current_filters={}, user=user or {})
    
    cursor = db.cursor(dictionary=True)
    
    try:
        allowed_statuses = set(APPLICATION_STATUSES)
        
        # Capture current filters from request (for preserving after POST)
        current_filters = {
            'keyword': request.args.get('keyword', '').strip(),
            'branch_id': request.args.get('branch_id', type=int),
            'job_id': request.args.get('job_id', type=int),
            'status': request.args.get('status', '').strip(),
        }

        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'update_status':
                # Restrict admin from changing status - only HR can change status
                if user.get('role') == 'admin':
                    flash('Access denied. Only HR users can change application status.', 'error')
                    params = {}
                    if current_filters.get('keyword'):
                        params['keyword'] = current_filters['keyword']
                    if current_filters.get('branch_id'):
                        params['branch_id'] = current_filters['branch_id']
                    if current_filters.get('job_id'):
                        params['job_id'] = current_filters['job_id']
                    if current_filters.get('status'):
                        params['status'] = current_filters['status']
                    return redirect(url_for('applications', **params))
                
                application_id = request.form.get('application_id')
                new_status_raw = request.form.get('status', '').strip()
                
                # Build redirect URL with preserved filters
                def build_redirect_url():
                    params = {}
                    if current_filters.get('keyword'):
                        params['keyword'] = current_filters['keyword']
                    if current_filters.get('branch_id'):
                        params['branch_id'] = current_filters['branch_id']
                    if current_filters.get('job_id'):
                        params['job_id'] = current_filters['job_id']
                    if current_filters.get('status'):
                        params['status'] = current_filters['status']
                    return url_for('applications', **params)
                
                # Validate application_id first
                if not application_id:
                    flash('Application ID is required.', 'error')
                    return redirect(build_redirect_url())
                
                # Validate status is provided
                if not new_status_raw:
                    error_msg = 'Status is required. Please select a status from the dropdown.'
                    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'error': error_msg}), 400
                    flash(error_msg, 'error')
                    return redirect(build_redirect_url())
                
                # Map simplified statuses to database statuses
                # ALL statuses can now be manually changed by HR: pending, scheduled, interviewed, hired, rejected
                CANONICAL_STATUS_MAP = {
                    'pending': 'pending',
                    'scheduled': 'scheduled',
                    'interview': 'interviewed',
                    'interviewed': 'interviewed',
                    'hired': 'hired',
                    'rejected': 'rejected',
                    # Legacy mappings for backward compatibility
                    'applied': 'pending',
                    'under_review': 'pending',
                    'reviewed': 'pending',
                    'shortlisted': 'pending',
                    'accepted': 'hired',  # Map old 'accepted' to new 'hired'
                }
                new_status = CANONICAL_STATUS_MAP.get(new_status_raw.lower(), new_status_raw.lower())
                
                # Validate mapped status is not empty
                if not new_status:
                    error_msg = f'Invalid status value: "{new_status_raw}". Please select a valid status.'
                    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'error': error_msg}), 400
                    flash(error_msg, 'error')
                    return redirect(build_redirect_url())
                
                if new_status not in allowed_statuses:
                    error_msg = f'Invalid status: {new_status}. Allowed statuses: {", ".join(allowed_statuses)}'
                    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'error': error_msg}), 400
                    flash(error_msg, 'error')
                    return redirect(build_redirect_url())
                
                if application_id and new_status in allowed_statuses:
                    # Verify application belongs to HR's branch (if HR is branch-scoped)
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT a.application_id
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE a.application_id = %s AND j.branch_id = %s
                            ''',
                            (application_id, branch_id),
                        )
                        if not cursor.fetchone():
                            error_msg = 'You can only update applications for your branch.'
                            # Check if this is an AJAX request
                            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                                return jsonify({'success': False, 'error': error_msg}), 403
                            flash(error_msg, 'error')
                            # Build redirect URL with preserved filters
                            params = {}
                            if current_filters.get('keyword'):
                                params['keyword'] = current_filters['keyword']
                            if current_filters.get('branch_id'):
                                params['branch_id'] = current_filters['branch_id']
                            if current_filters.get('job_id'):
                                params['job_id'] = current_filters['job_id']
                            if current_filters.get('status'):
                                params['status'] = current_filters['status']
                            return redirect(url_for('applications', **params))
                    # Get applicant email before updating status
                    # Use job_column_expr to handle both job_title and title columns
                    job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
                    cursor.execute(
                        f'''
                        SELECT ap.email, ap.full_name, a.status AS old_status, COALESCE({job_title_expr}, 'Untitled Job') AS job_title
                        FROM applicants ap
                        JOIN applications a ON ap.applicant_id = a.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        WHERE a.application_id = %s
                        LIMIT 1
                        ''',
                        (application_id,)
                    )
                    applicant_info = cursor.fetchone()
                    
                    if not applicant_info:
                        error_msg = 'Application not found or applicant information is missing.'
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            return jsonify({'success': False, 'error': error_msg}), 404
                        flash(error_msg, 'error')
                        return redirect(build_redirect_url())
                    
                    old_status = applicant_info.get('old_status') if applicant_info else None
                    
                    # AUTOMATIC: Update status and notify applicant
                    try:
                        update_success = auto_update_application_status(cursor, application_id, new_status)
                    except Exception as update_err:
                        db.rollback()
                        error_msg = f'Failed to update application status: {str(update_err)}'
                        print(f'❌ Error updating application status: {update_err}')
                        import traceback
                        traceback.print_exc()
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            return jsonify({'success': False, 'error': error_msg}), 500
                        flash(error_msg, 'error')
                        return redirect(build_redirect_url())
                    
                    if update_success:
                        # If HR performed the action, add Admin system notification
                        try:
                            if user.get('role') == 'hr':
                                hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                status_display = new_status.replace('_', ' ').title()
                                applicant_name = applicant_info.get('full_name') if applicant_info else 'applicant'
                                job_title = applicant_info.get('job_title') if applicant_info else 'position'
                                admin_msg = f'HR {hr_name} updated application status to {status_display} for {applicant_name} ({job_title}).'
                                create_admin_notification(cursor, admin_msg)
                                print(f'✅ Admin notification created: {admin_msg}')
                        except Exception as notify_err:
                            print(f'⚠️ Error creating admin notification: {notify_err}')
                        
                        # Commit the transaction
                        db.commit()
                        print(f'✅ Application {application_id} status updated to "{new_status}" - transaction committed')
                        
                        # Check if this is an AJAX request
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            # Return JSON response for AJAX requests
                            status_display = new_status.replace('_', ' ').title()
                            if new_status.lower() == 'hired':
                                message = f'Congratulations! {applicant_info.get("full_name", "Applicant")} has been marked as HIRED. They have been automatically notified via email and notification.'
                            else:
                                message = 'Application status updated successfully. Applicant has been automatically notified via email and notification.'
                            return jsonify({
                                'success': True,
                                'message': message,
                                'status': new_status,
                                'status_display': status_display,
                                'application_id': application_id
                            })
                        
                        # Special success message for hired status (for regular form submissions)
                        if new_status.lower() == 'hired':
                            flash(f'Congratulations! {applicant_info.get("full_name", "Applicant")} has been marked as HIRED. They have been automatically notified via email and notification.', 'success')
                        else:
                            flash('Application status updated successfully. Applicant has been automatically notified via email and notification.', 'success')
                    else:
                        # Fallback to manual update if auto function fails
                        cursor.execute(
                            'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s',
                            (new_status, application_id),
                        )
                        db.commit()  # Ensure commit happens even in fallback
                        
                        # Check if this is an AJAX request
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            status_display = new_status.replace('_', ' ').title()
                            return jsonify({
                                'success': True,
                                'message': f'Application status updated to {status_display}.',
                                'status': new_status,
                                'status_display': status_display,
                                'application_id': application_id,
                                'note': 'Status updated but notification may not have been sent.'
                            })
                        # Still try to create notification even if auto function failed
                        try:
                            cursor.execute(
                                '''
                                SELECT ap.applicant_id, ap.email, ap.full_name, j.title AS job_title
                                FROM applicants ap
                                JOIN applications a ON ap.applicant_id = a.applicant_id
                                LEFT JOIN jobs j ON a.job_id = j.job_id
                                WHERE a.application_id = %s
                                LIMIT 1
                                ''',
                                (application_id,)
                            )
                            applicant_info = cursor.fetchone()
                            if applicant_info:
                                status_display = new_status.replace('_', ' ').title()
                                job_title = applicant_info.get('job_title') or 'Your Application'
                                message = f'Your application status for "{job_title}" has been updated to: {status_display}'
                                
                                # Create notification - this notification goes to the APPLICANT (not HR)
                                # The notification is linked to the application, which is associated with the applicant
                                notification_columns = set()
                                try:
                                    cursor.execute('SHOW COLUMNS FROM notifications')
                                    notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                                except Exception:
                                    pass
                                
                                if 'sent_at' in notification_columns:
                                    cursor.execute(
                                        'INSERT INTO notifications (application_id, message, sent_at, is_read) VALUES (%s, %s, NOW(), 0)',
                                        (application_id, message)
                                    )
                                else:
                                    cursor.execute(
                                        'INSERT INTO notifications (application_id, message, is_read) VALUES (%s, %s, 0)',
                                        (application_id, message)
                                    )
                                
                                # Send email
                                try:
                                    email_subject = f'Application Status Update - {applicant_info.get("job_title") or "Your Application"}'
                                    email_body = f"""Dear {applicant_info.get('full_name') or 'Applicant'},

Your application status for the position "{applicant_info.get('job_title') or 'the position'}" has been updated.

New Status: {status_display}

Please log in to your account to view more details.

Best regards,
J&T Express Recruitment Team
                                    """.strip()
                                    send_email(applicant_info.get('email'), email_subject, email_body)
                                except Exception as email_err:
                                    print(f"⚠️ Email error (non-blocking): {email_err}")
                        except Exception as notify_err:
                            print(f"⚠️ Notification creation error (non-blocking): {notify_err}")
                        
                        # If HR performed the action, add Admin system notification
                        try:
                            if user.get('role') == 'hr':
                                hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                status_display = new_status.replace('_', ' ').title()
                                applicant_name = applicant_info.get('full_name') if applicant_info else 'applicant'
                                job_title = applicant_info.get('job_title') if applicant_info else 'position'
                                admin_msg = f'HR {hr_name} updated application status to {status_display} for {applicant_name} ({job_title}).'
                                create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating admin notification: {notify_err}')
                        db.commit()
                        
                        # Check if this is an AJAX request before redirecting
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            status_display = new_status.replace('_', ' ').title()
                            return jsonify({
                                'success': True,
                                'message': 'Application status updated successfully. Notification sent to applicant.',
                                'status': new_status,
                                'status_display': status_display,
                                'application_id': application_id
                            })
                        
                        flash('Application status updated successfully. Notification sent to applicant.', 'success')
                
                # Build redirect URL with preserved filters (only for non-AJAX requests)
                params = {}
                if current_filters.get('keyword'):
                    params['keyword'] = current_filters['keyword']
                if current_filters.get('branch_id'):
                    params['branch_id'] = current_filters['branch_id']
                if current_filters.get('job_id'):
                    params['job_id'] = current_filters['job_id']
                if current_filters.get('status'):
                    params['status'] = current_filters['status']
                return redirect(url_for('applications', **params))
            
            elif action == 'bulk_update_status':
                # Restrict admin from changing status - only HR can change status
                if user.get('role') == 'admin':
                    flash('Access denied. Only HR users can change application status.', 'error')
                    params = {}
                    if current_filters.get('keyword'):
                        params['keyword'] = current_filters['keyword']
                    if current_filters.get('branch_id'):
                        params['branch_id'] = current_filters['branch_id']
                    if current_filters.get('job_id'):
                        params['job_id'] = current_filters['job_id']
                    if current_filters.get('status'):
                        params['status'] = current_filters['status']
                    return redirect(url_for('applications', **params))
                
                application_ids = request.form.getlist('application_ids')
                bulk_status = request.form.get('bulk_status', '').strip()
                
                # Map simplified statuses to database statuses
                CANONICAL_STATUS_MAP = {
                    'pending': 'pending',
                    'interview': 'interviewed',
                    'interviewed': 'interviewed',
                    'hired': 'hired',
                    'rejected': 'rejected',
                    'applied': 'pending',
                    'under_review': 'pending',
                    'reviewed': 'pending',
                    'shortlisted': 'pending',
                    'accepted': 'hired',
                }
                bulk_status = CANONICAL_STATUS_MAP.get(bulk_status, bulk_status)
                
                if application_ids and bulk_status in allowed_statuses:
                    # Verify all applications belong to HR's branch
                    branch_id = get_branch_scope(user)
                    valid_ids = []
                    if branch_id:
                        placeholders = ','.join(['%s'] * len(application_ids))
                        cursor.execute(
                            f'''
                            SELECT a.application_id
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE a.application_id IN ({placeholders}) AND j.branch_id = %s
                            ''',
                            (*application_ids, branch_id),
                        )
                        valid_ids = [row['application_id'] for row in cursor.fetchall()]
                        if len(valid_ids) != len(application_ids):
                            flash('Some applications do not belong to your branch. Only valid applications were updated.', 'warning')
                    else:
                        valid_ids = application_ids
                    
                    if valid_ids:
                        # AUTOMATIC: Update status and notify applicants for each application
                        updated_count = 0
                        for app_id in valid_ids:
                            if auto_update_application_status(cursor, app_id, bulk_status):
                                updated_count += 1
                        
                        # Create admin notification for bulk update
                        try:
                            if user.get('role') == 'hr':
                                hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                status_display = bulk_status.replace('_', ' ').title()
                                admin_msg = f'HR {hr_name} bulk updated {updated_count} application(s) to {status_display}.'
                                create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating admin notification: {notify_err}')
                        
                        db.commit()
                        flash(f'{updated_count} application(s) updated successfully. Applicants have been automatically notified via email and notification.', 'success')
            
            # Build redirect URL with preserved filters for bulk update
            params = {}
            if current_filters.get('keyword'):
                params['keyword'] = current_filters['keyword']
            if current_filters.get('branch_id'):
                params['branch_id'] = current_filters['branch_id']
            if current_filters.get('job_id'):
                params['job_id'] = current_filters['job_id']
            if current_filters.get('status'):
                params['status'] = current_filters['status']
            return redirect(url_for('applications', **params))
        
        # Apply filters
        filters = {
            'keyword': request.args.get('keyword', '').strip(),
            'branch_id': request.args.get('branch_id', type=int),
            'job_id': request.args.get('job_id', type=int),
            'status': request.args.get('status', '').strip(),
            'date_from': request.args.get('date_from', '').strip(),
            'date_to': request.args.get('date_to', '').strip(),
        }
        # Keep keyword even if empty string to track search attempts
        # But only add to filters dict if it has a value
        keyword_value = filters.pop('keyword', '')
        filters = {k: v for k, v in filters.items() if v}
        if keyword_value:
            filters['keyword'] = keyword_value
        
        branch_id = get_branch_scope(user)
        where_clauses = []
        params = []
        
        if branch_id:
            where_clauses.append('j.branch_id = %s')
            params.append(branch_id)
        
        if filters.get('keyword'):
            keyword = filters['keyword'].strip()
            if keyword:
                keyword_pattern = f"%{keyword}%"
                # Search in applicant name, email, and job title - only for applicants who have applications
                # Use TRIM and LOWER() for case-insensitive search and to handle whitespace
                # Use COALESCE to handle NULL job titles
                where_clauses.append('(LOWER(TRIM(ap.full_name)) LIKE LOWER(%s) OR LOWER(TRIM(ap.email)) LIKE LOWER(%s) OR LOWER(TRIM(COALESCE(j.title, \'\'))) LIKE LOWER(%s))')
                params.extend([keyword_pattern, keyword_pattern, keyword_pattern])
        
        if filters.get('branch_id'):
            where_clauses.append('j.branch_id = %s')
            params.append(filters['branch_id'])
        
        if filters.get('job_id'):
            where_clauses.append('a.job_id = %s')
            params.append(filters['job_id'])
        
        if filters.get('status'):
            # Normalize status filter - map display statuses to database statuses
            status_filter = filters['status'].strip().lower()
            status_map = {
                'pending': ['pending', 'reviewed', 'applied', 'under_review', 'shortlisted'],
                'scheduled': ['scheduled'],
                'interviewed': ['interviewed', 'interview'],
                'hired': ['hired', 'accepted'],
                'rejected': ['rejected'],
            }
            db_statuses = status_map.get(status_filter, [status_filter])
            # Only apply filter if it's a valid status
            if db_statuses:
                placeholders = ','.join(['%s'] * len(db_statuses))
                where_clauses.append(f'a.status IN ({placeholders})')
                params.extend(db_statuses)
                print(f"🔍 Status filter applied: '{status_filter}' -> WHERE a.status IN {db_statuses}")
            else:
                print(f"⚠️ Invalid status filter: '{status_filter}' - not in status_map")
        
        if filters.get('date_from'):
            where_clauses.append('DATE(a.submitted_at) >= %s')
            params.append(filters['date_from'])
        
        if filters.get('date_to'):
            where_clauses.append('DATE(a.submitted_at) <= %s')
            params.append(filters['date_to'])
        
        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        
        # Get dynamic column expressions for job title
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
        
        # Check if position_name column exists in jobs table
        cursor.execute('SHOW COLUMNS FROM jobs LIKE "position_name"')
        has_position_name_col = cursor.fetchone() is not None
        
        # Build position_title expression conditionally
        if has_position_name_col:
            position_title_expr = f'COALESCE(j.position_name, {job_title_expr})'
        else:
            position_title_expr = job_title_expr
        
        # Fetch applications
        # IMPORTANT: 
        # 1. Query starts with applications table, ensuring only applicants with applications are returned
        # 2. INNER JOIN with applicants ensures applicant data exists
        # 3. HR users only see applications for jobs in their branch (enforced by branch_id filter above)
        # 4. Admin users see all applications from all branches
        # 5. Keyword search only searches within applicants who have submitted applications
        # Use simpler approach: get ANY interview for each application, and use EXISTS to check
        # Normalize withdrawn to rejected in SQL query - remove withdrawn status completely
        cursor.execute(
            f'''
            SELECT a.application_id,
                   CASE 
                       WHEN a.status = 'withdrawn' THEN 'rejected'
                       ELSE a.status
                   END AS status,
                   a.submitted_at,
                   a.resume_id,
                   ap.applicant_id,
                   ap.full_name AS applicant_name,
                   ap.email AS applicant_email,
                   ap.phone_number AS applicant_phone,
                   j.job_id,
                   {job_title_expr} AS job_title,
                   COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                   {position_title_expr} AS position_title,
                   (SELECT interview_id FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS interview_id,
                   (SELECT scheduled_date FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS scheduled_date,
                   (SELECT interview_mode FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS interview_mode,
                   (SELECT location FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS interview_location
            FROM applications a
            INNER JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_sql}
            ORDER BY a.submitted_at DESC
            ''',
            tuple(params) if params else None,
        )
        applications = cursor.fetchall()
        
        # Calculate analytics - use correct status values from database enum (pending, scheduled, interviewed, hired, rejected)
        # Normalize withdrawn to rejected in SQL query - remove withdrawn status completely
        # Use DISTINCT to prevent duplicates from LEFT JOIN with interviews
        cursor.execute(
            f'''
            SELECT 
                COUNT(DISTINCT a.application_id) AS total,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'pending' THEN a.application_id END) AS pending,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'scheduled' THEN a.application_id END) AS scheduled,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'interviewed' THEN a.application_id END) AS interviewed,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'hired' THEN a.application_id END) AS hired,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'rejected' THEN a.application_id END) AS rejected,
                COUNT(DISTINCT i.interview_id) AS interviews_scheduled,
                COUNT(DISTINCT CASE WHEN DATE(a.submitted_at) >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) THEN a.application_id END) AS this_month,
                COUNT(DISTINCT CASE WHEN DATE(a.submitted_at) >= DATE_SUB(CURDATE(), INTERVAL 7 DAY) THEN a.application_id END) AS this_week
            FROM applications a
            INNER JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN interviews i ON i.application_id = a.application_id
            WHERE {where_sql}
            ''',
            tuple(params) if params else None,
        )
        analytics = cursor.fetchone()
        
        # Note: We no longer auto-update status based on interview existence
        # Status flow: pending -> scheduled (when interview scheduled) -> interviewed (when interview completed) -> hired/rejected
        # Applications with 'scheduled' status should remain 'scheduled' until interview is marked as completed
        
        # Format applications
        formatted_apps = []
        status_filter_value = filters.get('status', '').strip().lower() if filters else ''
        for app in applications:
            # Get the original status from database - CRITICAL: preserve rejected/hired status
            app_status = app.get('status')
            if app_status:
                app_status = app_status.lower().strip()
            else:
                app_status = 'pending'
            
            # Normalize withdrawn to rejected - remove withdrawn status completely
            if app_status == 'withdrawn':
                app_status = 'rejected'
            
            application_id = app.get('application_id')
            
            # CRITICAL: Never override final statuses (hired, rejected) - preserve them exactly as stored
            # Status is already set to 'scheduled' when interview is scheduled, and 'interviewed' when interview is completed
            # So we don't need to auto-override based on has_interview anymore
            has_interview = app.get('interview_id') is not None
            
            # If status is rejected or hired, NEVER override it, regardless of interviews
            if app_status in ('rejected', 'hired'):
                # Preserve the original status - do not override
                pass
            # Note: Status is managed by interview scheduling/completion logic, so we preserve it as-is
            # For any other status, keep the original status
            
            formatted_apps.append({
                'application_id': app.get('application_id'),
                'applicant_id': app.get('applicant_id'),
                'applicant_name': app.get('applicant_name'),
                'applicant_email': app.get('applicant_email'),
                'applicant_phone': app.get('applicant_phone'),
                'job_id': app.get('job_id'),
                'job_title': app.get('job_title'),
                'branch_name': app.get('branch_name'),
                'position_title': app.get('position_title'),
                'status': app_status,  # Use normalized status (withdrawn -> rejected)
                'submitted_at': format_human_datetime(app.get('submitted_at')),
                'resume_id': app.get('resume_id'),
                'has_resume': app.get('resume_id') is not None,
                'has_interview': has_interview,
                'interview_date': format_human_datetime(app.get('scheduled_date')) if app.get('scheduled_date') else None,
                'interview_mode': app.get('interview_mode'),
                'interview_location': app.get('interview_location'),
            })
        
        branches = fetch_branches()
        jobs = fetch_jobs_for_user(user)
        
        # Check if JSON format is requested (for AJAX/modals)
        if request.args.get('format') == 'json' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            modal_type = request.args.get('modal', '').strip().lower()
            status_filter = request.args.get('status', '').strip()
            
            # Handle modal-specific requests
            if modal_type:
                if modal_type == 'total':
                    # Return all applications for total modal
                    return jsonify({
                        'applications': formatted_apps,
                        'analytics': analytics or {},
                        'total': len(formatted_apps)
                    })
                elif modal_type == 'pending':
                    # Return only pending applications
                    pending_apps = [app for app in formatted_apps if app.get('status') == 'pending']
                    return jsonify({
                        'applications': pending_apps,
                        'analytics': analytics or {},
                        'total': len(pending_apps)
                    })
                elif modal_type == 'hired':
                    # Return only hired applications
                    hired_apps = [app for app in formatted_apps if app.get('status') == 'hired']
                    return jsonify({
                        'applications': hired_apps,
                        'analytics': analytics or {},
                        'total': len(hired_apps)
                    })
                elif modal_type == 'rejected':
                    # Return only rejected applications
                    rejected_apps = [app for app in formatted_apps if app.get('status') == 'rejected']
                    return jsonify({
                        'applications': rejected_apps,
                        'analytics': analytics or {},
                        'total': len(rejected_apps)
                    })
                elif modal_type == 'interviewed':
                    # Return only interviewed applications
                    interviewed_apps = [app for app in formatted_apps if app.get('status') == 'interviewed']
                    return jsonify({
                        'applications': interviewed_apps,
                        'analytics': analytics or {},
                        'total': len(interviewed_apps)
                    })
                elif modal_type == 'scheduled':
                    # Return only scheduled applications
                    scheduled_apps = [app for app in formatted_apps if app.get('status') == 'scheduled']
                    return jsonify({
                        'applications': scheduled_apps,
                        'analytics': analytics or {},
                        'total': len(scheduled_apps)
                    })
            
            # Default: return filtered applications by status
            print(f"📊 JSON response for status='{status_filter}': {len(formatted_apps)} applications returned")
            if status_filter == 'rejected':
                rejected_apps = [app for app in formatted_apps if app.get('status') == 'rejected']
                print(f"📊 Rejected applications count: {len(rejected_apps)} out of {len(formatted_apps)} total")
                print(f"📊 Rejected application IDs: {[app.get('application_id') for app in rejected_apps]}")
            return jsonify({
                'applications': formatted_apps,
                'analytics': analytics or {},
                'total': len(formatted_apps)
            })
        
        # Render HR template if user is HR, otherwise admin template
        template = 'hr/applications.html' if user.get('role') == 'hr' else 'admin/applications.html'
        branch_info = None
        if user.get('role') == 'hr':
            branch_id = session.get('branch_id')
            if branch_id:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id,))
                if branch_rows:
                    branch_info = branch_rows[0]
        return render_template(
            template,
            applications=formatted_apps,
            analytics=analytics or {},
            branches=branches,
            jobs=jobs,
            current_filters=filters,
            user=user,
            branch_info=branch_info,
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applications management error: {exc}')
        print(f'Full traceback: {error_details}')
        
        # Check if this is an AJAX request
        if request.method == 'POST' and (request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json):
            return jsonify({
                'success': False,
                'error': f'An error occurred: {str(exc)}. Please check the console for details.'
            }), 500
        
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        template = 'hr/applications.html' if (user or {}).get('role') == 'hr' else 'admin/applications.html'
        return render_template(template, applications=[], analytics={}, branches=[], jobs=[], current_filters={}, user=user or {}, branch_info=None)
    finally:
        cursor.close()


@app.route('/admin/interviews/get-jobs', methods=['GET'])
@login_required('admin', 'hr')
def get_applicant_jobs():
    """Get jobs for a specific applicant."""
    user = get_current_user()
    applicant_id = request.args.get('applicant_id', type=int)
    
    if not applicant_id:
        return jsonify({'jobs': []})
    
    db = get_db()
    if not db:
        return jsonify({'jobs': []})
    
    cursor = db.cursor(dictionary=True)
    try:
        where_clause = 'a.applicant_id = %s'
        params = [applicant_id]
        branch_id = get_branch_scope(user)
        if branch_id:
            where_clause += ' AND j.branch_id = %s'
            params.append(branch_id)
        
        cursor.execute(
            f'''
            SELECT DISTINCT j.job_id, j.title, a.status AS application_status
            FROM applications a
            JOIN jobs j ON a.job_id = j.job_id
            WHERE {where_clause}
            ORDER BY j.title ASC
            ''',
            tuple(params),
        )
        jobs = cursor.fetchall()
        return jsonify({'jobs': jobs})
    except Exception as exc:
        print(f'❌ Get applicant jobs error: {exc}')
        return jsonify({'jobs': []})
    finally:
        cursor.close()


def _render_interviews():
    user = get_current_user()
    target_endpoint = 'admin_interviews' if user.get('role') == 'admin' else 'hr_interviews'
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        template = 'admin/interviews.html' if user.get('role') == 'admin' else 'hr/interviews.html'
        return render_template(template, user=user, interviews=[], upcoming=[], past=[], applications=[], applicants_for_schedule=[], current_filters={})
    
    cursor = db.cursor(dictionary=True)
    schedule_applications = []

    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            # Block scheduling for admins - they can only view
            if action == 'schedule' and user.get('role') == 'admin':
                flash('Admins can only view interviews. Scheduling is restricted to HR staff.', 'error')
                return redirect(url_for(target_endpoint))
            
            if action == 'schedule':
                application_id = request.form.get('application_id', '').strip()
                applicant_id = request.form.get('applicant_id', '').strip()
                job_id = request.form.get('job_id', '').strip()
                scheduled_date = request.form.get('scheduled_date', '').strip()
                scheduled_time = request.form.get('scheduled_time', '').strip()
                interview_mode = (request.form.get('interview_mode') or 'in_person').strip()
                location = (request.form.get('location') or '').strip()
                notes = (request.form.get('notes') or '').strip()
                
                print(f'🔍 Interview scheduling POST data: application_id={application_id}, applicant_id={applicant_id}, scheduled_date={scheduled_date}, scheduled_time={scheduled_time}')
                
                # Basic input validation
                if not scheduled_date or not scheduled_time:
                    flash('Interview date and time are required.', 'error')
                    return redirect(url_for(target_endpoint))
                elif not application_id and not applicant_id:
                    flash('Please select an applicant.', 'error')
                    return redirect(url_for(target_endpoint))
                else:
                    # Validate combined datetime is parseable and not in the past (grace: allow same minute)
                    try:
                        scheduled_dt = datetime.strptime(f"{scheduled_date} {scheduled_time}", "%Y-%m-%d %H:%M")
                        if scheduled_dt < datetime.now():
                            flash('Interview time must be in the future.', 'error')
                            return redirect(url_for(target_endpoint))
                    except Exception:
                        flash('Invalid date/time format.', 'error')
                        return redirect(url_for(target_endpoint))
                    # Normalize notes length
                    if len(notes) > 1000:
                        notes = notes[:1000]
                    if len(location) > 255:
                        location = location[:255]
                    # Attempt to resolve the target application via application_id first
                    # For admin users, don't filter by branch - allow scheduling for all branches
                    branch_id = None if user.get('role') == 'admin' else get_branch_scope(user)
                    application = None
                    
                    if application_id:
                        lookup_clause = 'a.application_id = %s'
                        lookup_params = [application_id]
                        if branch_id:
                            lookup_clause += ' AND j.branch_id = %s'
                            lookup_params.append(branch_id)
                        cursor.execute(
                            f'''
                            SELECT a.application_id, a.applicant_id, a.job_id
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE {lookup_clause}
                            LIMIT 1
                            ''',
                            tuple(lookup_params),
                        )
                        application = cursor.fetchone()
                    
                    # Fallback: attempt to locate by applicant_id only (auto-select most recent eligible application)
                    if not application and applicant_id:
                        # Find the most recent application for this applicant that's eligible for interview
                        # Priority: reviewed > pending > interviewed (using actual database enum values)
                        # Exclude rejected and withdrawn applications - cannot schedule interview for rejected/withdrawn
                        lookup_clause = 'a.applicant_id = %s AND a.status NOT IN (%s, %s)'
                        lookup_params = [applicant_id, 'rejected', 'withdrawn']
                        if branch_id:
                            lookup_clause += ' AND j.branch_id = %s'
                            lookup_params.append(branch_id)
                        cursor.execute(
                            f'''
                            SELECT a.application_id, a.applicant_id, a.job_id, a.status
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE {lookup_clause}
                            ORDER BY 
                                CASE a.status
                                    WHEN 'reviewed' THEN 1
                                    WHEN 'pending' THEN 2
                                    WHEN 'interviewed' THEN 3
                                    ELSE 5
                                END,
                                a.submitted_at DESC
                            LIMIT 1
                            ''',
                            tuple(lookup_params),
                        )
                        application = cursor.fetchone()
                    
                    # Fallback: attempt to locate by applicant + job if job_id was provided (for backward compatibility)
                    if not application and applicant_id and job_id:
                        lookup_clause = 'a.applicant_id = %s AND a.job_id = %s'
                        lookup_params = [applicant_id, job_id]
                        if branch_id:
                            lookup_clause += ' AND j.branch_id = %s'
                            lookup_params.append(branch_id)
                        cursor.execute(
                            f'''
                            SELECT a.application_id, a.applicant_id, a.job_id
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE {lookup_clause}
                            LIMIT 1
                            ''',
                            tuple(lookup_params),
                        )
                        application = cursor.fetchone()
                    
                    if not application:
                        flash('No eligible application found for this applicant. Please ensure the applicant has an application that is not rejected or withdrawn (status must be: pending, reviewed, or interviewed).', 'error')
                        return redirect(url_for(target_endpoint))
                    else:
                        resolved_application_id = application.get('application_id')
                        resolved_applicant_id = application.get('applicant_id')
                        resolved_job_id = application.get('job_id')
                        
                        # Check if application status is 'rejected' or 'withdrawn' - prevent scheduling
                        application_status = (application.get('status') or '').lower()
                        if application_status in ('rejected', 'withdrawn'):
                            status_label = 'rejected' if application_status == 'rejected' else 'withdrawn'
                            flash(f'Cannot schedule interview for a {status_label} application. Please select an applicant with an eligible application (pending, reviewed, or interviewed).', 'error')
                            return redirect(url_for(target_endpoint))
                        
                        # Double-check by querying the current status from database
                        cursor.execute(
                            'SELECT status FROM applications WHERE application_id = %s LIMIT 1',
                            (resolved_application_id,)
                        )
                        current_status_record = cursor.fetchone()
                        if current_status_record:
                            current_status = (current_status_record.get('status') or '').lower()
                            if current_status in ('rejected', 'withdrawn'):
                                status_label = 'rejected' if current_status == 'rejected' else 'withdrawn'
                                flash(f'Cannot schedule interview for a {status_label} application. The application status is {status_label}.', 'error')
                                return redirect(url_for(target_endpoint))
                        
                        scheduled_datetime = f"{scheduled_date} {scheduled_time}"
                        
                        # Get current admin_id for reference (not stored in interviews table per schema)
                        admin_id = user.get('id')
                        
                        # Force interview mode to 'in-person' only
                        interview_mode = 'in-person'
                        
                        try:
                            # Step 1: Insert interview into database
                            cursor.execute(
                                '''
                                INSERT INTO interviews (application_id, scheduled_date, interview_mode, location, notes, status)
                                VALUES (%s, %s, %s, %s, %s, 'scheduled')
                                ''',
                                (
                                    resolved_application_id,
                                    scheduled_datetime,
                                    interview_mode,
                                    location,
                                    notes,
                                ),
                            )
                            interview_id = cursor.lastrowid
                            print(f'✅ Interview saved to database - ID: {interview_id}, Application ID: {resolved_application_id}')
                            
                            # Step 2: AUTOMATIC: Update application status to 'scheduled' when interview is scheduled
                            # Update status to 'scheduled' unless it's already 'hired' or 'rejected' (final states)
                            cursor.execute(
                                'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s AND status NOT IN (%s, %s)',
                                ('scheduled', resolved_application_id, 'hired', 'rejected'),
                            )
                            rows_updated = cursor.rowcount
                            if rows_updated > 0:
                                print(f'✅ Application status updated to "scheduled" for application {resolved_application_id}')
                            else:
                                print(f'⚠️ Application status not updated (may already be hired/rejected) for application {resolved_application_id}')
                            
                            # Step 3: Get applicant email for automatic notification
                            cursor.execute(
                                '''
                                SELECT ap.applicant_id, ap.email, ap.full_name, COALESCE(j.title, 'Position') AS job_title
                                FROM applicants ap
                                JOIN applications a ON ap.applicant_id = a.applicant_id
                                LEFT JOIN jobs j ON a.job_id = j.job_id
                                WHERE a.application_id = %s
                                LIMIT 1
                                ''',
                                (resolved_application_id,)
                            )
                            applicant_info = cursor.fetchone()
                            
                            if not applicant_info:
                                print(f'⚠️ Warning: No applicant info found for application {resolved_application_id}')
                                flash('Interview scheduled, but could not find applicant information for notification.', 'warning')
                            else:
                                # Step 4: AUTOMATIC: Notify and email applicant
                                # Format notification message to match applicant notification query pattern
                                job_title_display = applicant_info.get('job_title') or 'the position'
                                notification_message = f'You applied for {job_title_display}. Interview scheduled for {scheduled_date} at {scheduled_time}. Mode: {interview_mode.replace("_", " ").title()}. Location: {location or "TBD"}.'
                                email_subject = f'Interview Scheduled - {applicant_info.get("job_title") or "Your Application"}'
                                email_body = f"""Dear {applicant_info.get('full_name') or 'Applicant'},

Your interview has been scheduled for the position: {applicant_info.get('job_title') or 'the position'}

Interview Details:
- Date: {scheduled_date}
- Time: {scheduled_time}
- Mode: {interview_mode.replace('_', ' ').title()}
- Location: {location or 'To be determined'}

Please arrive on time and bring any required documents.

Best regards,
J&T Express Recruitment Team
                                """.strip()
                                
                                # Check if we already sent an email for this interview (prevent duplicates)
                                # Check for any recent notification about interview scheduling for this application
                                cursor.execute(
                                    '''
                                    SELECT notification_id FROM notifications
                                    WHERE application_id = %s 
                                    AND message LIKE %s
                                    AND sent_at >= DATE_SUB(NOW(), INTERVAL 5 MINUTE)
                                    LIMIT 1
                                    ''',
                                    (resolved_application_id, f'%Interview scheduled for {scheduled_date}%'),
                                )
                                recent_notification = cursor.fetchone()
                                
                                if recent_notification:
                                    print(f'⚠️ Recent interview scheduling notification found for application {resolved_application_id}. Skipping duplicate notification and email.')
                                else:
                                    # Create notification and send email (only if no recent duplicate)
                                    auto_notify_and_email(
                                        cursor, resolved_application_id, notification_message,
                                        email_subject, email_body,
                                        applicant_info.get('email'),
                                        applicant_info.get('full_name')
                                    )
                                    print(f'✅ Notification and email sent to applicant {applicant_info.get("applicant_id")} ({applicant_info.get("email")})')
                                # Step 4b: If HR performed the action, create an Admin system notification
                                try:
                                    if user.get('role') == 'hr':
                                        hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                        applicant_name = applicant_info.get('full_name') if applicant_info else 'applicant'
                                        job_title = applicant_info.get('job_title') if applicant_info else 'position'
                                        admin_msg = f'HR {hr_name} scheduled interview for {applicant_name} ({job_title}) on {scheduled_date} {scheduled_time}.'
                                        create_admin_notification(cursor, admin_msg)
                                except Exception as notify_err:
                                    print(f'⚠️ Error creating admin notification: {notify_err}')
                            
                            # Step 5: Commit all changes to database
                            db.commit()
                            print(f'✅ All changes committed to database successfully')
                            flash('Interview scheduled successfully. Applicant has been automatically notified via email and notification.', 'success')
                            return redirect(url_for(target_endpoint))
                        except Exception as schedule_error:
                            db.rollback()
                            import traceback
                            print(f'❌ Error scheduling interview: {schedule_error}')
                            print(f'Traceback: {traceback.format_exc()}')
                            flash(f'Error scheduling interview: {str(schedule_error)}', 'error')
                            return redirect(url_for(target_endpoint))
            
            elif action == 'update' or action == 'reschedule':
                # Block updating/rescheduling for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Updates are restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id')
                scheduled_date = request.form.get('scheduled_date', '').strip()
                scheduled_time = request.form.get('scheduled_time', '').strip()
                interview_mode = request.form.get('interview_mode', 'in-person').strip()
                location = request.form.get('location', '').strip()
                notes = request.form.get('notes', '').strip()
                
                # Normalize interview_mode to match schema ENUM values
                normalized_mode = interview_mode.lower().replace('-', '_')
                if normalized_mode in ('in_person', 'in person', 'inperson'):
                    interview_mode = 'in-person'
                elif normalized_mode in ('online', 'remote', 'video', 'virtual', 'phone'):
                    interview_mode = 'remote'
                else:
                    interview_mode = 'in-person'  # Default
                
                if not interview_id or not scheduled_date or not scheduled_time:
                    flash('Interview ID, date, and time are required.', 'error')
                else:
                    # Validate datetime for reschedule
                    try:
                        scheduled_dt = datetime.strptime(f"{scheduled_date} {scheduled_time}", "%Y-%m-%d %H:%M")
                        if scheduled_dt < datetime.now():
                            flash('New interview time must be in the future.', 'error')
                            return redirect(url_for(target_endpoint))
                    except Exception:
                        flash('Invalid date/time format.', 'error')
                        return redirect(url_for(target_endpoint))
                    if len(notes) > 1000:
                        notes = notes[:1000]
                    if len(location) > 255:
                        location = location[:255]
                    scheduled_datetime = f"{scheduled_date} {scheduled_time}"
                    # Verify interview belongs to HR's branch
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id, i.application_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        interview_record = cursor.fetchone()
                        if not interview_record:
                            flash('You can only update interviews for your branch.', 'error')
                            return redirect(url_for(target_endpoint))
                        application_id = interview_record['application_id']
                    else:
                        cursor.execute('SELECT application_id FROM interviews WHERE interview_id = %s', (interview_id,))
                        interview_record = cursor.fetchone()
                        application_id = interview_record['application_id'] if interview_record else None
                    
                    cursor.execute(
                        '''
                        UPDATE interviews
                        SET scheduled_date = %s,
                            interview_mode = %s,
                            location = %s,
                            notes = %s
                        WHERE interview_id = %s
                        ''',
                        (scheduled_datetime, interview_mode, location, notes, interview_id),
                    )
                    
                    # Get applicant email for notification
                    if action == 'reschedule' and application_id:
                        cursor.execute(
                            '''
                            SELECT ap.email, ap.full_name, j.title
                            FROM applicants ap
                            JOIN applications a ON ap.applicant_id = a.applicant_id
                            LEFT JOIN jobs j ON a.job_id = j.job_id
                            WHERE a.application_id = %s
                            LIMIT 1
                            ''',
                            (application_id,)
                        )
                        applicant_info = cursor.fetchone()
                        
                        # Create notification for reschedule
                        try:
                            ensure_schema_compatibility()
                            notification_columns = set()
                            try:
                                cursor.execute('SHOW COLUMNS FROM notifications')
                                notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                            except Exception:
                                pass
                            
                            reschedule_message = f'Interview rescheduled to {scheduled_date} at {scheduled_time}. Please check your interview schedule.'
                            if 'sent_at' in notification_columns:
                                cursor.execute(
                                    '''
                                    INSERT INTO notifications (application_id, message, sent_at, is_read)
                                    VALUES (%s, %s, NOW(), 0)
                                    ''',
                                    (application_id, reschedule_message),
                                )
                            else:
                                cursor.execute(
                                    '''
                                    INSERT INTO notifications (application_id, message, is_read)
                                    VALUES (%s, %s, 0)
                                    ''',
                                    (application_id, reschedule_message),
                                )
                            
                            # Send email notification
                            if applicant_info and applicant_info.get('email'):
                                try:
                                    applicant_name = applicant_info.get('full_name') or 'Applicant'
                                    job_title = applicant_info.get('job_title') or 'the position'
                                    email_subject = f'Interview Rescheduled - {job_title}'
                                    email_body = f"""Dear {applicant_name},

Your interview for the position "{job_title}" has been rescheduled.

New Interview Details:
- Date: {scheduled_date}
- Time: {scheduled_time}
- Mode: {interview_mode.replace('_', ' ').title()}
- Location: {location or 'To be determined'}

Please update your calendar accordingly.

Best regards,
J&T Express Recruitment Team
                                    """.strip()
                                    send_email(applicant_info.get('email'), email_subject, email_body)
                                except Exception as email_error:
                                    print(f"Email notification error: {email_error}")
                            # Admin system notification when HR reschedules
                            try:
                                if user.get('role') == 'hr':
                                    hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                    applicant_name = applicant_info.get('full_name') if applicant_info else 'applicant'
                                    job_title = applicant_info.get('job_title') if applicant_info else 'position'
                                    admin_msg = f'HR {hr_name} rescheduled interview for {applicant_name} ({job_title}) to {scheduled_date} {scheduled_time}.'
                                    create_admin_notification(cursor, admin_msg)
                            except Exception:
                                pass
                        except Exception:
                            pass
                    
                    db.commit()
                    flash('Interview updated successfully. Applicant has been notified via email.', 'success')
            
            elif action == 'cancel':
                # Block canceling for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Canceling is restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id')
                if interview_id:
                    # Verify interview belongs to HR's branch
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id, i.application_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        interview_record = cursor.fetchone()
                        if not interview_record:
                            flash('You can only cancel interviews for your branch.', 'error')
                            return redirect(url_for(target_endpoint))
                        application_id = interview_record['application_id']
                    else:
                        cursor.execute('SELECT application_id FROM interviews WHERE interview_id = %s', (interview_id,))
                        interview_record = cursor.fetchone()
                        application_id = interview_record['application_id'] if interview_record else None
                    
                    # Update status to cancelled instead of deleting
                    cursor.execute(
                        'UPDATE interviews SET status = %s WHERE interview_id = %s',
                        ('cancelled', interview_id),
                    )
                    
                    # Get applicant email for notification
                    if application_id:
                        cursor.execute(
                            '''
                            SELECT ap.email, ap.full_name, j.title
                            FROM applicants ap
                            JOIN applications a ON ap.applicant_id = a.applicant_id
                            LEFT JOIN jobs j ON a.job_id = j.job_id
                            WHERE a.application_id = %s
                            LIMIT 1
                            ''',
                            (application_id,)
                        )
                        applicant_info = cursor.fetchone()
                        
                        # Create notification
                        try:
                            ensure_schema_compatibility()
                            notification_columns = set()
                            try:
                                cursor.execute('SHOW COLUMNS FROM notifications')
                                notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                            except Exception:
                                pass
                            
                            cancel_message = 'Your interview has been cancelled. Please contact HR for more information.'
                            if 'sent_at' in notification_columns:
                                cursor.execute(
                                    '''
                                    INSERT INTO notifications (application_id, message, sent_at, is_read)
                                    VALUES (%s, %s, NOW(), 0)
                                    ''',
                                    (application_id, cancel_message),
                                )
                            else:
                                cursor.execute(
                                    '''
                                    INSERT INTO notifications (application_id, message, is_read)
                                    VALUES (%s, %s, 0)
                                    ''',
                                    (application_id, cancel_message),
                                )
                            
                            # Send email notification
                            if applicant_info and applicant_info.get('email'):
                                try:
                                    applicant_name = applicant_info.get('full_name') or 'Applicant'
                                    job_title = applicant_info.get('job_title') or 'the position'
                                    email_subject = f'Interview Cancelled - {job_title}'
                                    email_body = f"""Dear {applicant_name},

We regret to inform you that your scheduled interview for the position "{job_title}" has been cancelled.

Please contact our HR department for more information or to reschedule.

Best regards,
J&T Express Recruitment Team
                                    """.strip()
                                    send_email(applicant_info.get('email'), email_subject, email_body)
                                except Exception as email_error:
                                    print(f"Email notification error: {email_error}")
                            
                            # Admin system notification when HR cancels
                            try:
                                if user.get('role') == 'hr':
                                    hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                    applicant_name = applicant_info.get('full_name') if applicant_info else 'applicant'
                                    job_title = applicant_info.get('job_title') if applicant_info else 'position'
                                    admin_msg = f'HR {hr_name} cancelled interview for {applicant_name} ({job_title}).'
                                    create_admin_notification(cursor, admin_msg)
                            except Exception as notify_err:
                                print(f'⚠️ Error creating admin notification: {notify_err}')
                        except Exception:
                            pass
                    
                    db.commit()
                    flash('Interview cancelled successfully. Applicant has been notified via email.', 'success')
            
            elif action == 'delete':
                # Block deleting for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Deleting is restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id') or request.args.get('interview_id')
                if not interview_id:
                    flash('Interview ID is required.', 'error')
                    return redirect(url_for(target_endpoint))
                
                try:
                    interview_id = int(interview_id)
                except (ValueError, TypeError):
                    flash('Invalid interview ID.', 'error')
                    return redirect(url_for(target_endpoint))
                
                # Verify interview exists and belongs to HR's branch (if scoped)
                branch_id = get_branch_scope(user)
                if branch_id:
                    cursor.execute(
                        '''
                        SELECT i.interview_id, i.application_id
                        FROM interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN jobs j ON a.job_id = j.job_id
                        WHERE i.interview_id = %s AND j.branch_id = %s
                        LIMIT 1
                        ''',
                        (interview_id, branch_id),
                    )
                    interview_record = cursor.fetchone()
                    if not interview_record:
                        flash('Interview not found or you do not have permission to delete it.', 'error')
                        return redirect(url_for(target_endpoint))
                else:
                    # HR managing all branches - verify interview exists
                    cursor.execute(
                        'SELECT interview_id, application_id FROM interviews WHERE interview_id = %s LIMIT 1',
                        (interview_id,)
                    )
                    interview_record = cursor.fetchone()
                    if not interview_record:
                        flash('Interview not found.', 'error')
                        return redirect(url_for(target_endpoint))
                
                # Delete the interview
                cursor.execute('DELETE FROM interviews WHERE interview_id = %s', (interview_id,))
                deleted_count = cursor.rowcount
                
                if deleted_count > 0:
                    db.commit()
                    print(f'✅ Interview {interview_id} deleted successfully')
                    if request.accept_mimetypes.accept_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({'success': True, 'message': 'Interview deleted successfully.'})
                    flash('Interview deleted successfully.', 'success')
                else:
                    if request.accept_mimetypes.accept_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({'success': False, 'error': 'Interview not found or already deleted.'}), 404
                    flash('Interview not found or already deleted.', 'error')
                
                return redirect(url_for(target_endpoint))
            
            elif action == 'mark_attendance':
                # Block marking attendance for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Marking attendance is restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id')
                attendance_status = request.form.get('attendance_status', 'completed').strip()
                
                if interview_id:
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        if not cursor.fetchone():
                            flash('You can only update interviews for your branch.', 'error')
                            return redirect(url_for(target_endpoint))
                    
                    # Get application_id for this interview
                    cursor.execute(
                        'SELECT application_id FROM interviews WHERE interview_id = %s',
                        (interview_id,)
                    )
                    interview_record = cursor.fetchone()
                    application_id = interview_record.get('application_id') if interview_record else None
                    
                    # Update interview status
                    cursor.execute(
                        'UPDATE interviews SET status = %s WHERE interview_id = %s',
                        (attendance_status, interview_id),
                    )
                    
                    # AUTOMATIC: If interview is marked as completed, update application status to 'interviewed'
                    if attendance_status == 'completed' and application_id:
                        # Only update if status is not already a final state (hired/rejected)
                        cursor.execute(
                            'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s AND status NOT IN (%s, %s)',
                            ('interviewed', application_id, 'hired', 'rejected'),
                        )
                        rows_updated = cursor.rowcount
                        if rows_updated > 0:
                            print(f'✅ Application status auto-updated to "interviewed" for application {application_id} (interview completed)')
                            
                            # Create admin notification for interview completion
                            try:
                                if user.get('role') == 'hr':
                                    hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                    cursor.execute(
                                        '''
                                        SELECT ap.full_name, j.title
                                        FROM applicants ap
                                        JOIN applications a ON ap.applicant_id = a.applicant_id
                                        LEFT JOIN jobs j ON a.job_id = j.job_id
                                        WHERE a.application_id = %s
                                        LIMIT 1
                                        ''',
                                        (application_id,)
                                    )
                                    app_info = cursor.fetchone()
                                    applicant_name = app_info.get('full_name') if app_info else 'applicant'
                                    job_title = app_info.get('job_title') if app_info else 'position'
                                    admin_msg = f'HR {hr_name} marked interview as completed for {applicant_name} ({job_title}).'
                                    create_admin_notification(cursor, admin_msg)
                            except Exception as notify_err:
                                print(f'⚠️ Error creating notification for interview completion: {notify_err}')
                    
                    db.commit()
                    flash(f'Interview marked as {attendance_status.replace("_", " ")}.', 'success')
            
            elif action == 'add_evaluation':
                interview_id = request.form.get('interview_id')
                remarks = request.form.get('remarks', '').strip()
                
                # Enhanced feedback form fields
                technical_rating = request.form.get('technical_rating', '0')
                communication_rating = request.form.get('communication_rating', '0')
                cultural_rating = request.form.get('cultural_rating', '0')
                experience_rating = request.form.get('experience_rating', '0')
                strengths = request.form.get('strengths', '').strip()
                weaknesses = request.form.get('weaknesses', '').strip()
                recommendation = request.form.get('recommendation', '').strip()
                next_steps = request.form.get('next_steps', '').strip()
                
                # Legacy fields for backward compatibility
                criteria = request.form.get('criteria', '').strip()
                score = request.form.get('score', '').strip()
                
                if not interview_id:
                    flash('Interview ID is required.', 'error')
                elif not remarks and not criteria:
                    flash('Feedback comments are required.', 'error')
                else:
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        if not cursor.fetchone():
                            flash('You can only evaluate interviews for your branch.', 'error')
                            return redirect(url_for(target_endpoint))
                    
                    evaluator_id = user.get('id') or session.get('user_id')
                    
                    # Build comprehensive evaluation notes
                    evaluation_notes = []
                    if technical_rating and technical_rating != '0':
                        evaluation_notes.append(f'Technical Skills: {technical_rating}/5')
                    if communication_rating and communication_rating != '0':
                        evaluation_notes.append(f'Communication: {communication_rating}/5')
                    if cultural_rating and cultural_rating != '0':
                        evaluation_notes.append(f'Cultural Fit: {cultural_rating}/5')
                    if experience_rating and experience_rating != '0':
                        evaluation_notes.append(f'Experience: {experience_rating}/5')
                    if strengths:
                        evaluation_notes.append(f'Strengths: {strengths}')
                    if weaknesses:
                        evaluation_notes.append(f'Weaknesses: {weaknesses}')
                    if recommendation:
                        evaluation_notes.append(f'Recommendation: {recommendation.replace("_", " ").title()}')
                    if remarks:
                        evaluation_notes.append(f'Comments: {remarks}')
                    if next_steps:
                        evaluation_notes.append(f'Next Steps: {next_steps}')
                    
                    # Legacy format for backward compatibility
                    if criteria and score:
                        evaluation_notes.append(f'Criteria: {criteria}, Score: {score}')
                    
                    evaluation_text = '\n\n'.join(evaluation_notes)
                    
                    # Update interview notes with comprehensive feedback
                    cursor.execute(
                        '''
                        UPDATE interviews 
                        SET notes = CONCAT(COALESCE(notes, ''), '\n\n--- Interview Feedback ---\n', %s)
                        WHERE interview_id = %s
                        ''',
                        (evaluation_text, interview_id),
                    )
                    
                    # Try to insert into evaluation table if it exists
                    try:
                        # Store each rating as separate evaluation record
                        ratings = [
                            ('technical_skills', technical_rating),
                            ('communication', communication_rating),
                            ('cultural_fit', cultural_rating),
                            ('experience', experience_rating),
                        ]
                        for criteria_name, rating in ratings:
                            if rating and rating != '0':
                                try:
                                    cursor.execute(
                                        '''
                                        INSERT INTO evaluations (interview_id, evaluator_id, criteria, score, remarks)
                                        VALUES (%s, %s, %s, %s, %s)
                                        ON DUPLICATE KEY UPDATE score = %s, remarks = %s
                                        ''',
                                        (interview_id, evaluator_id, criteria_name, float(rating), remarks, float(rating), remarks),
                                    )
                                except Exception:
                                    # Table might not exist or have different structure, continue
                                    pass
                    except Exception:
                        # Evaluation table doesn't exist or has different structure, that's okay
                        pass
                    
                    # Update interview status to completed if not already
                    cursor.execute(
                        'UPDATE interviews SET status = %s WHERE interview_id = %s AND status = %s',
                        ('completed', interview_id, 'scheduled'),
                    )
                    
                    db.commit()
                    flash('Interview feedback recorded successfully.', 'success')
            
            elif action == 'make_decision':
                # Block making decisions for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Making decisions is restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id')
                decision = request.form.get('decision', '').strip()  # 'hired' or 'rejected'
                decision_notes = request.form.get('decision_notes', '').strip()
                
                if not interview_id or not decision:
                    flash('Interview ID and decision are required.', 'error')
                else:
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id, i.application_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        interview_record = cursor.fetchone()
                        if not interview_record:
                            flash('You can only make decisions for your branch interviews.', 'error')
                            return redirect(url_for(target_endpoint))
                        application_id = interview_record['application_id']
                    else:
                        cursor.execute('SELECT application_id FROM interviews WHERE interview_id = %s', (interview_id,))
                        interview_record = cursor.fetchone()
                        application_id = interview_record['application_id'] if interview_record else None
                    
                    if application_id:
                        # AUTOMATIC: Update application status and notify applicant
                        decision_reason = f'Interview decision: {decision_notes}' if decision_notes else ''
                        if auto_update_application_status(cursor, application_id, decision, decision_reason):
                            # Update interview status to completed
                            cursor.execute(
                                'UPDATE interviews SET status = %s, notes = CONCAT(COALESCE(notes, ""), "\n\nDecision: ", %s, "\nNotes: ", %s) WHERE interview_id = %s',
                                ('completed', decision, decision_notes, interview_id),
                            )
                            db.commit()
                            flash(f'Decision recorded: {decision}. Application status automatically updated and applicant notified.', 'success')
                        else:
                            # Fallback
                            cursor.execute(
                                'UPDATE applications SET status = %s WHERE application_id = %s',
                                (decision, application_id),
                            )
                            cursor.execute(
                                'UPDATE interviews SET status = %s, notes = CONCAT(COALESCE(notes, ""), "\n\nDecision: ", %s, "\nNotes: ", %s) WHERE interview_id = %s',
                                ('completed', decision, decision_notes, interview_id),
                            )
                            db.commit()
                            flash(f'Decision recorded: {decision}. Application status updated.', 'success')
            
            return redirect(url_for(target_endpoint))
        
        # Get application_id from query parameter for pre-filling schedule form
        prefill_application_id = request.args.get('application_id', type=int)
        
        # Apply filters - preserve all filter values even if empty for proper filtering
        filters = {
            'status': request.args.get('status', '').strip(),  # scheduled, completed, cancelled, all
            'mode': request.args.get('mode', '').strip(),  # in-person, remote, video, phone
            'date_from': request.args.get('date_from', '').strip(),
            'date_to': request.args.get('date_to', '').strip(),
            'position_id': request.args.get('position_id', type=int),
            'job_id': request.args.get('job_id', type=int),
            'applicant_id': request.args.get('applicant_id', type=int),
            'keyword': request.args.get('keyword', '').strip(),
            'view_mode': request.args.get('view_mode', 'list').strip(),  # list, calendar
        }
        # Only remove filters that are None or empty strings - keep integer 0 if it exists
        filters = {k: v for k, v in filters.items() if v is not None and v != ''}
        
        # For admin users, don't filter by branch - show all branches
        branch_id = None if user.get('role') == 'admin' else get_branch_scope(user)
        where_clauses = []
        params = []
        
        if branch_id:
            where_clauses.append('j.branch_id = %s')
            params.append(branch_id)
        
        if filters.get('status') and filters['status'].strip():
            # Normalize status filter - handle variations like 'in_progress' vs 'in-progress'
            status_value = filters['status'].strip().lower()
            # Map common status variations to database values
            status_map = {
                'scheduled': 'scheduled',
                'completed': 'completed',
                'cancelled': 'cancelled',
                'canceled': 'cancelled',  # Handle typo
                'rescheduled': 'rescheduled',
                'in_progress': 'in_progress',
                'in-progress': 'in_progress',  # Handle hyphen variation
                'inprogress': 'in_progress',  # Handle no separator
                'no_show': 'no_show',
                'no-show': 'no_show',  # Handle hyphen variation
                'noshow': 'no_show',  # Handle no separator
            }
            normalized_status = status_map.get(status_value, status_value)
            where_clauses.append('i.status = %s')
            params.append(normalized_status)
        
        if filters.get('mode'):
            where_clauses.append('i.interview_mode = %s')
            params.append(filters['mode'])
        
        if filters.get('date_from'):
            where_clauses.append('DATE(i.scheduled_date) >= %s')
            params.append(filters['date_from'])
        
        if filters.get('date_to'):
            where_clauses.append('DATE(i.scheduled_date) <= %s')
            params.append(filters['date_to'])
        
        # Position filter removed - positions table no longer exists
        # if filters.get('position_id'):
        #     where_clauses.append('j.position_id = %s')
        #     params.append(filters['position_id'])
        
        if filters.get('job_id'):
            # Ensure job_id is a valid integer
            try:
                job_id_val = int(filters['job_id']) if filters['job_id'] else None
                if job_id_val and job_id_val > 0:
                    where_clauses.append('j.job_id = %s')
                    params.append(job_id_val)
            except (ValueError, TypeError):
                print(f'⚠️ Invalid job_id filter: {filters.get("job_id")}')
        
        if filters.get('applicant_id'):
            # Ensure applicant_id is a valid integer
            try:
                applicant_id_val = int(filters['applicant_id']) if filters['applicant_id'] else None
                if applicant_id_val and applicant_id_val > 0:
                    where_clauses.append('ap.applicant_id = %s')
                    params.append(applicant_id_val)
            except (ValueError, TypeError):
                print(f'⚠️ Invalid applicant_id filter: {filters.get("applicant_id")}')
        
        if filters.get('keyword') and filters['keyword'].strip():
            keyword = f"%{filters['keyword'].strip().lower()}%"
            where_clauses.append('('
                                 'LOWER(ap.full_name) LIKE %s OR '
                                 'LOWER(ap.email) LIKE %s OR '
                                 'LOWER(COALESCE(j.title, "")) LIKE %s OR '
                                 'LOWER(COALESCE(b.branch_name, "")) LIKE %s OR '
                                 'LOWER(COALESCE(i.location, "")) LIKE %s OR '
                                 'LOWER(COALESCE(i.notes, "")) LIKE %s'
                                 ')')
            params.extend([keyword, keyword, keyword, keyword, keyword, keyword])
        
        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        
        # Preload applications for scheduling dropdown
        # Updated status values to match database enum
        # Load ALL applications (not filtered by status) so any applicant can be scheduled for interview
        # For admin users, don't filter by branch - show all applications from all branches
        application_where = []
        application_params = []
        if branch_id:
            application_where.append('j.branch_id = %s')
            application_params.append(branch_id)
        # If prefill_application_id is provided, include it
        if prefill_application_id:
            application_where.append('(1=1 OR a.application_id = %s)')
            application_params.append(prefill_application_id)
        application_where_sql = ' AND '.join(application_where) if application_where else '1=1'
        # Exclude rejected applications from schedule dropdown - cannot schedule interview for rejected
        application_where_sql += ' AND a.status != %s'
        application_params_with_rejected = list(application_params) if application_params else []
        application_params_with_rejected.append('rejected')
        
        cursor.execute(
            f'''
            SELECT a.application_id,
                   a.applicant_id,
                   a.job_id,
                   ap.full_name AS applicant_name,
                   j.title AS job_title,
                   a.status
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            JOIN jobs j ON a.job_id = j.job_id
            WHERE {application_where_sql}
            ORDER BY ap.full_name ASC, j.title ASC
            ''',
            tuple(application_params_with_rejected),
        )
        schedule_applications = cursor.fetchall() or []

        # Fetch interviews with interviewer and status info
        # Ensure we only select interview_mode, never interview_type
        try:
            cursor.execute(
                f'''
            SELECT i.interview_id,
                   i.application_id,
                   i.scheduled_date,
                       COALESCE(i.interview_mode, 'in-person') AS interview_mode,
                   i.location,
                   i.notes,
                   i.status AS interview_status,
                   ap.applicant_id,
                   ap.full_name AS applicant_name,
                   ap.email AS applicant_email,
                   ap.phone_number AS applicant_phone,
                   j.job_id,
                   COALESCE(j.title, 'Untitled Job') AS job_title,
                   COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                   a.status AS application_status
            FROM interviews i
            JOIN applications a ON i.application_id = a.application_id
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_sql}
            ORDER BY i.scheduled_date DESC
            ''',
            tuple(params) if params else None,
        )
            interviews = cursor.fetchall()
        except Exception as query_error:
            error_msg = str(query_error)
            if 'interview_type' in error_msg.lower():
                # If error is about interview_type, try to identify and fix the issue
                print(f'⚠️ Interview query error (interview_type): {error_msg}')
                # Return empty list and log the error
                flash('Error loading interviews. Please contact support if this persists.', 'error')
                interviews = []
            else:
                # Re-raise if it's a different error
                raise
        
        
        # Calculate statistics - wrapped in try-except to handle any interview_type errors
        stats_where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        try:
            cursor.execute(
                f'''
                SELECT 
                    COUNT(*) AS total_scheduled,
                    SUM(CASE WHEN DATE(i.scheduled_date) >= DATE_SUB(NOW(), INTERVAL 7 DAY) AND i.status = 'completed' THEN 1 ELSE 0 END) AS completed_this_week
                FROM interviews i
                JOIN applications a ON i.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                JOIN applicants ap ON a.applicant_id = ap.applicant_id
                WHERE {stats_where_sql}
                ''',
                tuple(params) if params else None,
            )
            stats_row = cursor.fetchone() or {}
        except Exception as stats_error:
            error_msg = str(stats_error)
            if 'interview_type' in error_msg.lower():
                print(f'⚠️ Interview stats query error (interview_type): {error_msg}')
                stats_row = {}
            else:
                raise
        total_scheduled = stats_row.get('total_scheduled', 0) or 0
        completed_this_week = stats_row.get('completed_this_week', 0) or 0
        
        interview_stats = {
            'total_scheduled': total_scheduled,
            'completed_this_week': completed_this_week,
        }
        
        # Separate upcoming and past interviews
        now = datetime.now()
        upcoming = []
        past = []
        
        for interview in interviews:
            scheduled = interview.get('scheduled_date')
            if scheduled and isinstance(scheduled, datetime):
                interview_status = interview.get('interview_status', 'scheduled')
                application_status = interview.get('application_status')
                
                # AUTOMATIC: If interview is completed, ensure application status shows as 'interviewed'
                # This ensures past interviews always show the correct status
                if interview_status == 'completed' and application_status not in ('hired', 'rejected'):
                    application_status = 'interviewed'
                
                interview_data = {
                    'interview_id': interview.get('interview_id'),
                    'application_id': interview.get('application_id'),
                    'applicant_id': interview.get('applicant_id'),
                    'applicant_name': interview.get('applicant_name'),
                    'applicant_email': interview.get('applicant_email'),
                    'applicant_phone': interview.get('applicant_phone'),
                    'job_id': interview.get('job_id'),
                    'job_title': interview.get('job_title'),
                    'branch_name': interview.get('branch_name'),
                    'scheduled_date': format_human_datetime(scheduled),
                    'scheduled_datetime': scheduled,  # Keep original for editing
                    'scheduled_date_raw': scheduled.strftime('%Y-%m-%d') if isinstance(scheduled, datetime) else '',
                    'scheduled_time_raw': scheduled.strftime('%H:%M') if isinstance(scheduled, datetime) else '',
                    'interview_mode': interview.get('interview_mode'),
                    'location': interview.get('location'),
                    'notes': interview.get('notes'),
                    'application_status': application_status,  # Use potentially overridden status
                    'interview_status': interview_status,
                }
                
                # Only show scheduled interviews in upcoming, completed/cancelled in past
                if interview_status == 'scheduled' and scheduled >= now:
                    upcoming.append(interview_data)
                elif interview_status in ('completed', 'cancelled', 'no_show') or (scheduled < now and interview_status != 'scheduled'):
                    past.append(interview_data)
                elif scheduled < now:
                    past.append(interview_data)
        
        # Get applicants for scheduling (all applicants who have applied to jobs)
        # Only show applicants with non-rejected and non-withdrawn applications (cannot schedule interview for rejected/withdrawn)
        # For admin users, don't filter by branch - show all applicants from all branches
        where_clauses_schedule = []
        params_schedule = []
        # Exclude rejected and withdrawn applications - cannot schedule interview for rejected/withdrawn
        where_clauses_schedule.append('a.status NOT IN (%s, %s)')
        params_schedule.extend(['rejected', 'withdrawn'])
        if branch_id:
            where_clauses_schedule.append('j.branch_id = %s')
            params_schedule.append(branch_id)
        where_sql_schedule = ' AND '.join(where_clauses_schedule) if where_clauses_schedule else '1=1'
        # Query to get applicants who have at least one eligible application (not rejected, not withdrawn)
        # Only show applicants with eligible applications (pending, reviewed, interviewed, hired)
        cursor.execute(
            f'''
            SELECT DISTINCT 
                   ap.applicant_id,
                   ap.full_name AS applicant_name,
                   ap.email AS applicant_email,
                   ap.phone_number AS applicant_phone,
                   GROUP_CONCAT(DISTINCT CONCAT(COALESCE(j.title, 'Untitled'), ' (', a.status, ')') SEPARATOR ' | ') AS job_titles,
                   COUNT(DISTINCT a.application_id) AS application_count,
                   GROUP_CONCAT(DISTINCT a.application_id SEPARATOR ',') AS application_ids,
                   MAX(b.branch_name) AS branch_name
            FROM applicants ap
            JOIN applications a ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_sql_schedule}
            GROUP BY ap.applicant_id, ap.full_name, ap.email, ap.phone_number
            HAVING COUNT(DISTINCT a.application_id) > 0
            ORDER BY ap.full_name ASC
            LIMIT 200
            ''',
            tuple(params_schedule) if params_schedule else None,
        )
        applicants_for_schedule = cursor.fetchall()
        
        # Render HR template if user is HR, otherwise admin template
        template = 'hr/interviews.html' if user.get('role') == 'hr' else 'admin/interviews.html'
        branch_info = None
        if user.get('role') == 'hr':
            branch_id = session.get('branch_id')
            if branch_id:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id,))
                if branch_rows:
                    branch_info = branch_rows[0]
        # Get unique positions for filter
        # Positions table removed - return empty list
        positions = []
        
        # Get unique jobs for filter - only show jobs that have interviews
        job_where = []
        job_params = []
        job_query = '''
            SELECT DISTINCT j.job_id, COALESCE(j.title, 'Untitled Job') AS job_title
            FROM jobs j
            JOIN applications a ON a.job_id = j.job_id
            JOIN interviews i ON i.application_id = a.application_id
        '''
        if branch_id:
            job_where.append('j.branch_id = %s')
            job_params.append(branch_id)
        if job_where:
            job_query += ' WHERE ' + ' AND '.join(job_where)
        job_query += ' ORDER BY j.title ASC'
        cursor.execute(job_query, tuple(job_params) if job_params else None)
        jobs = cursor.fetchall() or []
        
        # Get unique applicants for filter - only show applicants who have interviews
        applicant_where = []
        applicant_params = []
        applicant_query = '''
            SELECT DISTINCT ap.applicant_id, ap.full_name AS applicant_name
            FROM applicants ap
            JOIN applications a ON a.applicant_id = ap.applicant_id
            JOIN interviews i ON i.application_id = a.application_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
        '''
        if branch_id:
            applicant_where.append('j.branch_id = %s')
            applicant_params.append(branch_id)
        if applicant_where:
            applicant_query += ' WHERE ' + ' AND '.join(applicant_where)
        applicant_query += ' ORDER BY ap.full_name ASC'
        cursor.execute(applicant_query, tuple(applicant_params) if applicant_params else None)
        applicants = cursor.fetchall() or []
        
        return render_template(
            template,
            interviews=upcoming + past,
            upcoming=upcoming,
            past=past,
            applications=schedule_applications,
            applicants_for_schedule=applicants_for_schedule or [],
            hr_interviewers=[],
            current_filters=filters,
            user=user,
            branch_info=branch_info,
            interview_stats=interview_stats,
            positions=positions,
            jobs=jobs,
            applicants=applicants,
            prefill_application_id=prefill_application_id,
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Interviews management error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        template = 'hr/interviews.html' if user.get('role') == 'hr' else 'admin/interviews.html'
        return render_template(template, user=user, interviews=[], upcoming=[], past=[], applications=[], applicants_for_schedule=[], current_filters={})
    finally:
        cursor.close()


@app.route('/admin/interviews', methods=['GET', 'POST'])
@login_required('admin')
def admin_interviews():
    return _render_interviews()

@app.route('/hr/interviews', methods=['GET', 'POST'])
@login_required('hr', 'admin')
def hr_interviews():
    user = get_current_user()
    if user.get('role') != 'hr':
        return redirect(url_for('admin_interviews'))
    return _render_interviews()


@app.route('/interviews/<int:interview_id>/status', methods=['POST'])
@login_required('admin', 'hr')
def update_interview_status(interview_id: int):
    """Update interview status (scheduled → completed/cancelled/no_show). HR scoped by branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))
    ensure_schema_compatibility()
    cursor = db.cursor(dictionary=True)
    try:
        # Safely get status from form or JSON
        new_status = ''
        if request.form:
            new_status = request.form.get('status', '').strip().lower()
        elif request.is_json and request.json:
            new_status = request.json.get('status', '').strip().lower()
        
        # Safely get notes from form or JSON
        notes = ''
        if request.form:
            notes = request.form.get('notes', '').strip()
        elif request.is_json and request.json:
            notes = request.json.get('notes', '').strip()
        
        # Validate status
        if not new_status or new_status not in ('completed', 'cancelled', 'no_show'):
            error_msg = 'Invalid or missing status value' if not new_status else 'Invalid status value'
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': error_msg}), 400
            flash(error_msg, 'error')
            return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))
        # HR can manage all branches - no scoping restrictions
        branch_id = None  # HR users can manage all branches
        scope_sql = ''
        params = [new_status, notes, interview_id]
        # First, get the application_id before updating
        cursor.execute(
            f'''
            SELECT i.application_id
            FROM interviews i
            JOIN applications a ON i.application_id = a.application_id
            JOIN jobs j ON a.job_id = j.job_id
            WHERE i.interview_id = %s{scope_sql}
            ''',
            (interview_id,)
        )
        interview_data = cursor.fetchone()
        application_id = interview_data.get('application_id') if interview_data else None
        
        # Verify and update interview status
        try:
            # Build update query - append to existing notes if notes provided
            if notes:
                cursor.execute(
                    f'''
                    UPDATE interviews i
                    JOIN applications a ON i.application_id = a.application_id
                    JOIN jobs j ON a.job_id = j.job_id
                    SET i.status = %s, i.notes = CONCAT(COALESCE(i.notes, ''), '\n', %s), i.updated_at = NOW()
                    WHERE i.interview_id = %s{scope_sql}
                    ''',
                    (new_status, notes, interview_id)
                )
            else:
                cursor.execute(
                    f'''
                    UPDATE interviews i
                    JOIN applications a ON i.application_id = a.application_id
                    JOIN jobs j ON a.job_id = j.job_id
                    SET i.status = %s, i.updated_at = NOW()
                    WHERE i.interview_id = %s{scope_sql}
                    ''',
                    (new_status, interview_id)
                )
            if cursor.rowcount == 0:
                db.rollback()
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                    return jsonify({'success': False, 'error': 'Interview not found or not in scope'}), 404
                flash('Interview not found or not in your branch.', 'error')
                return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))
        except Exception as update_err:
            db.rollback()
            print(f'❌ Error updating interview status in database: {update_err}')
            import traceback
            traceback.print_exc()
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': f'Database error: {str(update_err)}'}), 500
            flash(f'Database error: {str(update_err)}', 'error')
            return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))
        
        # AUTOMATIC: If interview is marked as completed, update application status from 'scheduled' to 'interviewed'
        if new_status == 'completed' and application_id:
            # Only update if status is 'scheduled' and not already a final state (hired/rejected)
            cursor.execute(
                'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s AND status = %s AND status NOT IN (%s, %s)',
                ('interviewed', application_id, 'scheduled', 'hired', 'rejected'),
            )
            rows_updated = cursor.rowcount
            if rows_updated > 0:
                print(f'✅ Application status auto-updated from "scheduled" to "interviewed" for application {application_id} (interview {interview_id} marked as completed)')
        
        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'Interview status updated', 'interview_id': interview_id, 'status': new_status})
        flash('Interview status updated.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        print('❌ Update interview status error:', exc)
        print(traceback.format_exc())
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to update interview status'}), 500
        flash('Failed to update interview status.', 'error')
    finally:
        cursor.close()
    return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))

def handle_report_export(cursor, export_format, section, export_type, period_summary, 
                        applicant_summary, applicant_summary_details, job_vacancy, job_vacancy_details,
                        hiring_outcome, hr_performance, job_title_expr, date_filter, date_params):
    """Handle report exports in PDF, Excel, or CSV format."""
    from io import StringIO, BytesIO
    import csv
    from flask import Response
    from datetime import datetime
    
    if export_format == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Recruitment Report'])
        if period_summary:
            writer.writerow(['Period', f"{period_summary.get('start_date', '')} to {period_summary.get('end_date', '')}"])
        writer.writerow(['Generated', period_summary.get('generated_at', datetime.now().strftime('%b %d, %Y %I:%M %p'))])
        writer.writerow([])
        
        if section == 'applicant_summary':
            writer.writerow(['Applicant Summary Report'])
            writer.writerow(['Metric', 'Value'])
            writer.writerow(['Total Applicants', applicant_summary.get('total_applicants', 0)])
            writer.writerow(['New Applicants', applicant_summary.get('new_applicants', 0)])
            writer.writerow(['Active Applications', applicant_summary.get('active_applications', 0)])
            writer.writerow(['With Resume', applicant_summary.get('with_resume', 0)])
            writer.writerow(['Verified Email', applicant_summary.get('verified_email', 0)])
            if export_type == 'detailed' and applicant_summary_details:
                writer.writerow([])
                writer.writerow(['Detailed Applicant Status Report'])
                writer.writerow([])
                
                # Group by status
                status_groups = {
                    'Hired': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'hired'],
                    'Rejected': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'rejected'],
                    'Pending': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'pending'],
                    'Interview Scheduled': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'scheduled'],
                    'Interviewed': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'interviewed']
                }
                
                for status_name, apps in status_groups.items():
                    if apps:
                        writer.writerow([f'{status_name} Applicants ({len(apps)})'])
                        writer.writerow(['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire'])
                        for app in apps:
                            writer.writerow([
                                app.get('applicant_name', ''),
                                app.get('job_title', ''),
                                app.get('branch_name', ''),
                                app.get('status_label', app.get('application_status', '')),
                                app.get('date_applied', ''),
                                app.get('interview_date', '—'),
                                app.get('time_to_hire', '—')
                            ])
                        writer.writerow([])
                
                # All applicants summary
                writer.writerow(['All Applicants Summary'])
                writer.writerow(['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire'])
                for app in applicant_summary_details:
                    writer.writerow([
                        app.get('applicant_name', ''),
                        app.get('job_title', ''),
                        app.get('branch_name', ''),
                        app.get('status_label', app.get('application_status', '')),
                        app.get('date_applied', ''),
                        app.get('interview_date', '—'),
                        app.get('time_to_hire', '—')
                    ])
        elif section == 'job_vacancy':
            writer.writerow(['Job Vacancy Report'])
            writer.writerow(['Metric', 'Value'])
            writer.writerow(['Total Jobs Posted', job_vacancy.get('total_jobs', 0)])
            writer.writerow(['Active Jobs', job_vacancy.get('active_jobs', 0)])
            writer.writerow(['Closed Jobs', job_vacancy.get('closed_jobs', 0)])
            writer.writerow(['Total Applications', job_vacancy.get('total_applications', 0)])
            writer.writerow(['Avg Applications/Job', round(job_vacancy.get('avg_applications_per_job', 0), 2)])
            if export_type == 'detailed' and job_vacancy_details:
                writer.writerow([])
                writer.writerow(['Detailed Job List with Applicants'])
                writer.writerow(['Job Title', 'Branch', 'Status', 'Total Applications', 'Posted Date'])
                for job in job_vacancy_details:
                    writer.writerow([
                        job.get('job_title', ''),
                        job.get('branch_name', ''),
                        job.get('status', ''),
                        job.get('application_count', 0),
                        job.get('posted_date', '')
                    ])
                    # Get applicants for this job from applicant_summary_details
                    job_applicants = [app for app in applicant_summary_details if app.get('job_title') == job.get('job_title')]
                    if job_applicants:
                        writer.writerow([])
                        writer.writerow(['Applicants for this Job:'])
                        writer.writerow(['Applicant Name', 'Email', 'Status', 'Date Applied', 'Active Application'])
                        for app in job_applicants:
                            is_active = app.get('application_status', '').lower() in ['pending', 'scheduled', 'interviewed']
                            writer.writerow([
                                app.get('applicant_name', ''),
                                app.get('email', ''),
                                app.get('status_label', app.get('application_status', '')),
                                app.get('date_applied', ''),
                                'Yes' if is_active else 'No'
                            ])
                        writer.writerow([])
        elif section == 'hiring_outcome':
            writer.writerow(['Hiring Outcome Report'])
            writer.writerow(['Metric', 'Value'])
            writer.writerow(['Total Hired', hiring_outcome.get('total_hired', 0)])
            writer.writerow(['Hire Rate', f"{hiring_outcome.get('hire_rate', 0)}%"])
            writer.writerow(['Avg Time to Hire', f"{hiring_outcome.get('avg_time_to_hire', 0)} days"])
            writer.writerow(['Interview to Hire Rate', f"{hiring_outcome.get('interview_to_hire_rate', 0)}%"])
            writer.writerow(['Rejection Rate', f"{hiring_outcome.get('rejection_rate', 0)}%"])
        elif section == 'hr_performance':
            writer.writerow(['HR Performance Report'])
            writer.writerow(['Metric', 'Value'])
            writer.writerow(['Interviews Scheduled', hr_performance.get('interviews_scheduled', 0)])
            writer.writerow(['Interviews Completed', hr_performance.get('interviews_completed', 0)])
            writer.writerow(['Applications Reviewed', hr_performance.get('applications_reviewed', 0)])
            writer.writerow(['Status Updates', hr_performance.get('status_updates', 0)])
            writer.writerow(['Avg Response Time', f"{hr_performance.get('avg_response_time', 0)} hours"])
        else:
            # Full report
            writer.writerow(['Applicant Summary'])
            writer.writerow(['Total Applicants', applicant_summary.get('total_applicants', 0)])
            writer.writerow(['New Applicants', applicant_summary.get('new_applicants', 0)])
            writer.writerow([])
            writer.writerow(['Job Vacancy'])
            writer.writerow(['Total Jobs', job_vacancy.get('total_jobs', 0)])
            writer.writerow(['Active Jobs', job_vacancy.get('active_jobs', 0)])
            writer.writerow([])
            writer.writerow(['Hiring Outcome'])
            writer.writerow(['Total Hired', hiring_outcome.get('total_hired', 0)])
            writer.writerow(['Hire Rate', f"{hiring_outcome.get('hire_rate', 0)}%"])
            writer.writerow([])
            writer.writerow(['HR Performance'])
            writer.writerow(['Interviews Scheduled', hr_performance.get('interviews_scheduled', 0)])
        
        output.seek(0)
        filename = f'report_{section or "full"}_{period_summary.get("start_date", "all")}.csv'.replace('/', '-').replace(' ', '_')
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recruitment Report"
            
            # Header
            ws['A1'] = 'Recruitment Report'
            ws['A1'].font = Font(bold=True, size=14)
            if period_summary:
                ws['A2'] = f"Period: {period_summary.get('start_date', '')} to {period_summary.get('end_date', '')}"
            ws['A3'] = f"Generated: {period_summary.get('generated_at', datetime.now().strftime('%b %d, %Y %I:%M %p'))}"
            
            row = 5
            if section == 'applicant_summary':
                ws[f'A{row}'] = 'Applicant Summary Report'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                ws[f'A{row}'] = 'Metric'
                ws[f'B{row}'] = 'Value'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                row += 1
                for key, label in [('total_applicants', 'Total Applicants'), ('new_applicants', 'New Applicants'), 
                                  ('active_applications', 'Active Applications'), ('with_resume', 'With Resume'), 
                                  ('verified_email', 'Verified Email')]:
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = applicant_summary.get(key, 0)
                    row += 1
                
                # Add detailed applicant list if requested
                if export_type == 'detailed' and applicant_summary_details:
                    row += 2
                    ws[f'A{row}'] = 'Detailed Applicant Status Report'
                    ws[f'A{row}'].font = Font(bold=True, size=12)
                    row += 2
                    
                    # Group by status
                    status_groups = {
                        'Hired': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'hired'],
                        'Rejected': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'rejected'],
                        'Pending': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'pending'],
                        'Interview Scheduled': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'scheduled'],
                        'Interviewed': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'interviewed']
                    }
                    
                    for status_name, apps in status_groups.items():
                        if apps:
                            ws[f'A{row}'] = f'{status_name} Applicants ({len(apps)})'
                            ws[f'A{row}'].font = Font(bold=True, size=11)
                            ws[f'A{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                            for col in range(1, 9):
                                ws.cell(row=row, column=col).font = Font(bold=True, color='FFFFFF')
                            row += 1
                            headers = ['Applicant Name', 'Email', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Date Updated', 'Active Application']
                            for col, header in enumerate(headers, start=1):
                                cell = ws.cell(row=row, column=col)
                                cell.value = header
                                cell.font = Font(bold=True)
                            row += 1
                            for app in apps:
                                ws[f'A{row}'] = app.get('applicant_name', '')
                                ws[f'B{row}'] = app.get('job_title', '')
                                ws[f'C{row}'] = app.get('branch_name', '')
                                ws[f'D{row}'] = app.get('status_label', app.get('application_status', ''))
                                ws[f'E{row}'] = app.get('date_applied', '')
                                ws[f'F{row}'] = app.get('interview_date', '—')
                                ws[f'G{row}'] = app.get('time_to_hire', '—')
                                row += 1
                            row += 1
                    
                    # All applicants summary sheet
                    row += 1
                    ws[f'A{row}'] = 'All Applicants Summary'
                    ws[f'A{row}'].font = Font(bold=True, size=12)
                    row += 1
                    headers = ['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire']
                    for col, header in enumerate(headers, start=1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = header
                        cell.font = Font(bold=True)
                    row += 1
                    for app in applicant_summary_details:
                        ws[f'A{row}'] = app.get('applicant_name', '')
                        ws[f'B{row}'] = app.get('job_title', '')
                        ws[f'C{row}'] = app.get('branch_name', '')
                        ws[f'D{row}'] = app.get('status_label', app.get('application_status', ''))
                        ws[f'E{row}'] = app.get('date_applied', '')
                        ws[f'F{row}'] = app.get('interview_date', '—')
                        ws[f'G{row}'] = app.get('time_to_hire', '—')
                        row += 1
            elif section == 'job_vacancy':
                ws[f'A{row}'] = 'Job Vacancy Report'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                ws[f'A{row}'] = 'Metric'
                ws[f'B{row}'] = 'Value'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                row += 1
                for key, label in [('total_jobs', 'Total Jobs Posted'), ('active_jobs', 'Active Jobs'), 
                                  ('closed_jobs', 'Closed Jobs'), ('total_applications', 'Total Applications'), 
                                  ('avg_applications_per_job', 'Avg Applications/Job')]:
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = job_vacancy.get(key, 0)
                    row += 1
                
                # Add detailed job list with applicants if requested
                if export_type == 'detailed' and job_vacancy_details:
                    row += 2
                    ws[f'A{row}'] = 'Detailed Job List with Applicants'
                    ws[f'A{row}'].font = Font(bold=True, size=12)
                    row += 1
                    headers = ['Job Title', 'Branch', 'Status', 'Total Applications', 'Posted Date']
                    for col, header in enumerate(headers, start=1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = header
                        cell.font = Font(bold=True)
                    row += 1
                    for job in job_vacancy_details:
                        ws[f'A{row}'] = job.get('job_title', '')
                        ws[f'B{row}'] = job.get('branch_name', '')
                        ws[f'C{row}'] = job.get('status', '')
                        ws[f'D{row}'] = job.get('application_count', 0)
                        ws[f'E{row}'] = job.get('posted_date', '')
                        row += 1
                        # Get applicants for this job
                        job_applicants = [app for app in applicant_summary_details if app.get('job_title') == job.get('job_title')]
                        if job_applicants:
                            ws[f'A{row}'] = f'Applicants for {job.get("job_title", "")}:'
                            ws[f'A{row}'].font = Font(bold=True, italic=True)
                            row += 1
                            app_headers = ['Applicant Name', 'Email', 'Status', 'Date Applied', 'Active Application']
                            for col, header in enumerate(app_headers, start=1):
                                cell = ws.cell(row=row, column=col)
                                cell.value = header
                                cell.font = Font(bold=True)
                            row += 1
                            for app in job_applicants:
                                is_active = app.get('application_status', '').lower() in ['pending', 'scheduled', 'interviewed']
                                ws[f'A{row}'] = app.get('applicant_name', '')
                                ws[f'B{row}'] = app.get('email', '')
                                ws[f'C{row}'] = app.get('status_label', app.get('application_status', ''))
                                ws[f'D{row}'] = app.get('date_applied', '')
                                ws[f'E{row}'] = 'Yes' if is_active else 'No'
                                row += 1
                            row += 1
            elif section == 'hiring_outcome':
                ws[f'A{row}'] = 'Hiring Outcome Report'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                ws[f'A{row}'] = 'Metric'
                ws[f'B{row}'] = 'Value'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                row += 1
                ws[f'A{row}'] = 'Total Hired'
                ws[f'B{row}'] = hiring_outcome.get('total_hired', 0)
                row += 1
                ws[f'A{row}'] = 'Hire Rate'
                ws[f'B{row}'] = f"{hiring_outcome.get('hire_rate', 0)}%"
                row += 1
                ws[f'A{row}'] = 'Avg Time to Hire'
                ws[f'B{row}'] = f"{hiring_outcome.get('avg_time_to_hire', 0)} days"
                row += 1
                ws[f'A{row}'] = 'Interview to Hire Rate'
                ws[f'B{row}'] = f"{hiring_outcome.get('interview_to_hire_rate', 0)}%"
                row += 1
                ws[f'A{row}'] = 'Rejection Rate'
                ws[f'B{row}'] = f"{hiring_outcome.get('rejection_rate', 0)}%"
            elif section == 'hr_performance':
                ws[f'A{row}'] = 'HR Performance Report'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                ws[f'A{row}'] = 'Metric'
                ws[f'B{row}'] = 'Value'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                row += 1
                for key, label in [('interviews_scheduled', 'Interviews Scheduled'), 
                                  ('interviews_completed', 'Interviews Completed'), 
                                  ('applications_reviewed', 'Applications Reviewed'), 
                                  ('status_updates', 'Status Updates'), 
                                  ('avg_response_time', 'Avg Response Time (hours)')]:
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = hr_performance.get(key, 0)
                    row += 1
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f'report_{section or "full"}_{period_summary.get("start_date", "all")}.xlsx'.replace('/', '-').replace(' ', '_')
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename={filename}'}
            )
        except ImportError:
            flash('Excel export requires openpyxl library. Please install it: pip install openpyxl', 'error')
            from flask import redirect, url_for, request
            user = get_current_user()
            redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
            return redirect(request.referrer or redirect_url)
    elif export_format == 'pdf':
        # Basic PDF export using reportlab or weasyprint
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title = Paragraph("Recruitment Report", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Period info
            if period_summary:
                period_text = f"Period: {period_summary.get('start_date', '')} to {period_summary.get('end_date', '')}"
                elements.append(Paragraph(period_text, styles['Normal']))
                elements.append(Spacer(1, 0.1*inch))
            
            # Report data
            if section == 'applicant_summary':
                data = [['Metric', 'Value']]
                data.append(['Total Applicants', str(applicant_summary.get('total_applicants', 0))])
                data.append(['New Applicants', str(applicant_summary.get('new_applicants', 0))])
                data.append(['Active Applications', str(applicant_summary.get('active_applications', 0))])
                data.append(['With Resume', str(applicant_summary.get('with_resume', 0))])
                data.append(['Verified Email', str(applicant_summary.get('verified_email', 0))])
                
                # Add status breakdown summary even for summary type
                if applicant_summary_details:
                    data.append([])
                    data.append(['Applicant Status Breakdown'])
                    # Group by status to get counts
                    status_counts = {
                        'Hired': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'hired']),
                        'Rejected': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'rejected']),
                        'Pending': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'pending']),
                        'Interview Scheduled': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'scheduled']),
                        'Interviewed': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'interviewed'])
                    }
                    data.append(['Status', 'Count'])
                    for status_name, count in status_counts.items():
                        if count > 0:
                            data.append([status_name, str(count)])
                
                # Add detailed applicant list if requested
                if export_type == 'detailed' and applicant_summary_details:
                    data.append([])
                    data.append(['Detailed Applicant Status Report'])
                    data.append([])
                    
                    # Group by status
                    status_groups = {
                        'Hired': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'hired'],
                        'Rejected': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'rejected'],
                        'Pending': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'pending'],
                        'Interview Scheduled': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'scheduled'],
                        'Interviewed': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'interviewed']
                    }
                    
                    for status_name, apps in status_groups.items():
                        if apps:
                            data.append([f'{status_name} Applicants ({len(apps)})'])
                            data.append(['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire'])
                            for app in apps:
                                data.append([
                                    app.get('applicant_name', ''),
                                    app.get('job_title', ''),
                                    app.get('branch_name', ''),
                                    app.get('status_label', app.get('application_status', '')),
                                    app.get('date_applied', ''),
                                    app.get('interview_date', '—'),
                                    app.get('time_to_hire', '—')
                                ])
                            data.append([])
                    
                    # All applicants summary
                    data.append(['All Applicants Summary'])
                    data.append(['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire'])
                    for app in applicant_summary_details:
                        data.append([
                            app.get('applicant_name', ''),
                            app.get('job_title', ''),
                            app.get('branch_name', ''),
                            app.get('status_label', app.get('application_status', '')),
                            app.get('date_applied', ''),
                            app.get('interview_date', '—'),
                            app.get('time_to_hire', '—')
                        ])
            elif section == 'job_vacancy':
                data = [['Metric', 'Value']]
                data.append(['Total Jobs Posted', str(job_vacancy.get('total_jobs', 0))])
                data.append(['Active Jobs', str(job_vacancy.get('active_jobs', 0))])
                data.append(['Closed Jobs', str(job_vacancy.get('closed_jobs', 0))])
                data.append(['Total Applications', str(job_vacancy.get('total_applications', 0))])
                data.append(['Avg Applications/Job', str(round(job_vacancy.get('avg_applications_per_job', 0), 2))])
                
                # Add detailed job list with applicants if requested
                if export_type == 'detailed' and job_vacancy_details:
                    data.append([])
                    data.append(['Detailed Job List with Applicants'])
                    data.append(['Job Title', 'Branch', 'Status', 'Total Applications', 'Posted Date'])
                    for job in job_vacancy_details:
                        data.append([
                            job.get('job_title', ''),
                            job.get('branch_name', ''),
                            job.get('status', ''),
                            str(job.get('application_count', 0)),
                            job.get('posted_date', '')
                        ])
                        # Get applicants for this job
                        job_applicants = [app for app in applicant_summary_details if app.get('job_title') == job.get('job_title')]
                        if job_applicants:
                            data.append([])
                            data.append([f'Applicants for {job.get("job_title", "")}:'])
                            data.append(['Applicant Name', 'Email', 'Status', 'Date Applied', 'Active Application'])
                            for app in job_applicants:
                                is_active = app.get('application_status', '').lower() in ['pending', 'scheduled', 'interviewed']
                                data.append([
                                    app.get('applicant_name', ''),
                                    app.get('email', ''),
                                    app.get('status_label', app.get('application_status', '')),
                                    app.get('date_applied', ''),
                                    'Yes' if is_active else 'No'
                                ])
                            data.append([])
            elif section == 'hiring_outcome':
                data = [['Metric', 'Value']]
                data.append(['Total Hired', str(hiring_outcome.get('total_hired', 0))])
                data.append(['Hire Rate', f"{hiring_outcome.get('hire_rate', 0)}%"])
                data.append(['Avg Time to Hire', f"{hiring_outcome.get('avg_time_to_hire', 0)} days"])
                data.append(['Interview to Hire Rate', f"{hiring_outcome.get('interview_to_hire_rate', 0)}%"])
                data.append(['Rejection Rate', f"{hiring_outcome.get('rejection_rate', 0)}%"])
            elif section == 'hr_performance':
                data = [['Metric', 'Value']]
                data.append(['Interviews Scheduled', str(hr_performance.get('interviews_scheduled', 0))])
                data.append(['Interviews Completed', str(hr_performance.get('interviews_completed', 0))])
                data.append(['Applications Reviewed', str(hr_performance.get('applications_reviewed', 0))])
                data.append(['Status Updates', str(hr_performance.get('status_updates', 0))])
                data.append(['Avg Response Time', f"{hr_performance.get('avg_response_time', 0)} hours"])
            else:
                # Full report
                data = [['Section', 'Metric', 'Value']]
                data.append(['Applicant Summary', 'Total Applicants', str(applicant_summary.get('total_applicants', 0))])
                data.append(['Applicant Summary', 'New Applicants', str(applicant_summary.get('new_applicants', 0))])
                data.append(['Job Vacancy', 'Total Jobs', str(job_vacancy.get('total_jobs', 0))])
                data.append(['Job Vacancy', 'Active Jobs', str(job_vacancy.get('active_jobs', 0))])
                data.append(['Hiring Outcome', 'Total Hired', str(hiring_outcome.get('total_hired', 0))])
                data.append(['Hiring Outcome', 'Hire Rate', f"{hiring_outcome.get('hire_rate', 0)}%"])
                data.append(['HR Performance', 'Interviews Scheduled', str(hr_performance.get('interviews_scheduled', 0))])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            
            doc.build(elements)
            buffer.seek(0)
            filename = f'report_{section or "full"}_{period_summary.get("start_date", "all") if period_summary else "all"}.pdf'.replace('/', '-').replace(' ', '_')
            return Response(
                buffer.getvalue(),
                mimetype='application/pdf',
                headers={'Content-Disposition': f'attachment; filename={filename}'}
            )
        except ImportError:
            flash('PDF export requires reportlab library. Please install it: pip install reportlab', 'error')
            from flask import redirect, url_for, request
            user = get_current_user()
            redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
            return redirect(request.referrer or redirect_url)
        except Exception as pdf_error:
            print(f'⚠️ Error generating PDF: {pdf_error}')
            flash(f'Error generating PDF: {str(pdf_error)}', 'error')
            from flask import redirect, url_for, request
            user = get_current_user()
            redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
            return redirect(request.referrer or redirect_url)
    else:
        flash('Invalid export format.', 'error')
        from flask import redirect, url_for, request
        user = get_current_user()
        redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
        return redirect(request.referrer or redirect_url)

@app.route('/api/reports/data', methods=['GET'])
@login_required('admin', 'hr')
def api_get_reports_data():
    """API endpoint to fetch real report data for client-side rendering.
    
    Query parameters:
    - period: 'week', 'month', 'year', 'all'
    - branch: branch_id filter
    - position: job_title filter
    - status: application_status filter
    - dateFrom: YYYY-MM-DD
    - dateTo: YYYY-MM-DD
    """
    from datetime import datetime, timedelta
    
    user = get_current_user()
    db = get_db()
    
    if not db:
        return jsonify({'error': 'Database connection error'}), 500
    
    try:
        ensure_schema_compatibility()
        cursor = db.cursor(dictionary=True)
        
        # Get filters from query params
        period = request.args.get('period', 'month')
        branch_filter = request.args.get('branch', '').strip()
        job_filter = request.args.get('position', '').strip()
        status_filter = request.args.get('status', '').strip()
        date_from = request.args.get('dateFrom', '').strip()
        date_to = request.args.get('dateTo', '').strip()
        
        # Build WHERE clause for filtering
        where_clauses = []
        params = []
        
        # Apply period-based date filter if no custom dates provided
        if not date_from and not date_to:
            today = datetime.now().date()
            if period == 'week':
                start_date = today - timedelta(days=today.weekday())
                where_clauses.append('DATE(a.submitted_at) >= %s')
                params.append(start_date)
            elif period == 'month':
                start_date = today.replace(day=1)
                where_clauses.append('DATE(a.submitted_at) >= %s')
                params.append(start_date)
            elif period == 'year':
                start_date = today.replace(month=1, day=1)
                where_clauses.append('DATE(a.submitted_at) >= %s')
                params.append(start_date)
            # For 'all', no date filter is applied
        
        # Branch filter
        if branch_filter and branch_filter.lower() != 'all':
            # Allow branch filter to be passed as either branch name or branch id
            try:
                branch_id_val = int(branch_filter)
            except Exception:
                branch_id_val = None

            if branch_id_val is not None:
                where_clauses.append('b.branch_id = %s')
                params.append(branch_id_val)
            else:
                where_clauses.append('b.branch_name = %s')
                params.append(branch_filter)

        # Job position filter - treat common 'all' variants as no filter
        jf_norm = job_filter.strip().lower() if job_filter else ''
        if jf_norm and jf_norm not in ('all', '', 'all positions', 'any'):
            # Accept either job id or job title
            try:
                job_id_val = int(job_filter)
            except Exception:
                job_id_val = None

            if job_id_val is not None:
                where_clauses.append('j.job_id = %s')
                params.append(job_id_val)
            else:
                # Match by title using dynamic job column detection to avoid referencing missing columns
                title_expr = job_column_expr('title', alternatives=['job_title', 'position_name'], default='j.title')
                where_clauses.append(f'({title_expr} = %s)')
                params.append(job_filter)
        
        # Status filter
        if status_filter and status_filter.lower() != 'all':
            where_clauses.append('a.status = %s')
            params.append(status_filter)
        
        # Date range filter (overrides period filter if provided)
        if date_from:
            try:
                date_obj = datetime.strptime(date_from, '%Y-%m-%d')
                where_clauses.append('DATE(a.submitted_at) >= %s')
                params.append(date_obj.date())
            except ValueError:
                pass
        
        if date_to:
            try:
                date_obj = datetime.strptime(date_to, '%Y-%m-%d')
                where_clauses.append('DATE(a.submitted_at) <= %s')
                params.append(date_obj.date())
            except ValueError:
                pass
        
        where_sql = 'WHERE ' + ' AND '.join(where_clauses) if where_clauses else ''
        
        # Fetch all applicants with their application data
        query = f'''
            SELECT 
                ap.full_name,
                COALESCE(j.title, 'Unknown') AS position,
                COALESCE(b.branch_name, 'Unassigned') AS branch,
                a.status,
                a.submitted_at,
                a.application_id,
                a.updated_at,
                (SELECT MAX(i.scheduled_date) FROM interviews i WHERE i.application_id = a.application_id) AS interview_date,
                (SELECT COUNT(*) FROM interviews i WHERE i.application_id = a.application_id AND i.status IN ('completed', 'confirmed')) AS interview_count,
                DATEDIFF(COALESCE(a.updated_at, (SELECT MAX(i.scheduled_date) FROM interviews i WHERE i.application_id = a.application_id)), a.submitted_at) AS time_to_hire
            FROM applications a
            LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            {where_sql}
            ORDER BY a.submitted_at DESC
        '''
        
        # Debug: print final query and params to help diagnose empty result issues
        try:
            print('--- reports API: executing query ---')
            print(query)
            print('params:', tuple(params))
        except Exception:
            pass

        cursor.execute(query, tuple(params))
        all_records = cursor.fetchall() or []
        print(f'--- reports API: fetched {len(all_records)} records')
        
        # Calculate KPIs
        total_applicants = len(all_records)
        interviews_scheduled = len([r for r in all_records if r.get('interview_date')])
        hired = len([r for r in all_records if r.get('status') == 'hired'])
        rejected = len([r for r in all_records if r.get('status') == 'rejected'])
        pending = len([r for r in all_records if r.get('status') == 'pending'])
        
        # Calculate average time to hire
        hired_records = [r for r in all_records if r.get('status') == 'hired' and r.get('time_to_hire')]
        avg_time_to_hire = 0
        if hired_records:
            avg_time_to_hire = sum([r.get('time_to_hire') or 0 for r in hired_records]) / len(hired_records)
            avg_time_to_hire = round(avg_time_to_hire, 1)
        
        # Calculate conversion rate
        conversion_rate = round((hired / total_applicants * 100), 1) if total_applicants > 0 else 0
        
        # Get all branches from database to ensure they all display in chart
        # If a specific branch is filtered, extract actual branch names from the fetched records
        if branch_filter and branch_filter.lower() != 'all':
            # Extract distinct branch names from the filtered records (handles branch ID vs name mismatch)
            all_branches = list(set([r.get('branch', 'Unknown') for r in all_records]))
            all_branches.sort()
        else:
            cursor.execute('SELECT DISTINCT branch_name FROM branches ORDER BY branch_name')
            all_branches_result = cursor.fetchall()
            all_branches = [b.get('branch_name') for b in (all_branches_result or [])]
        
        # Group by branch performance - initialize all branches with zero values
        branch_stats = {}
        for branch_name in all_branches:
            branch_stats[branch_name] = {
                'applicants': 0,
                'hired': 0,
                'conversion_rate': 0,
                'avg_time_to_hire': None,
                '_total_time_to_hire': 0,
                '_hired_count_for_time': 0
            }

        # Now populate with actual data from filtered records
        for record in all_records:
            branch = record.get('branch', 'Unknown')
            if branch not in branch_stats:
                branch_stats[branch] = {
                    'applicants': 0,
                    'hired': 0,
                    'conversion_rate': 0,
                    'avg_time_to_hire': None,
                    '_total_time_to_hire': 0,
                    '_hired_count_for_time': 0
                }
            branch_stats[branch]['applicants'] += 1
            if record.get('status') == 'hired':
                branch_stats[branch]['hired'] += 1
                # accumulate time_to_hire for average calculation
                t = record.get('time_to_hire')
                if t is not None:
                    try:
                        branch_stats[branch]['_total_time_to_hire'] += float(t)
                        branch_stats[branch]['_hired_count_for_time'] += 1
                    except Exception:
                        pass

        # Calculate conversion rates and average time to hire per branch
        for branch in branch_stats:
            total = branch_stats[branch]['applicants']
            hired_count = branch_stats[branch]['hired']
            branch_stats[branch]['conversion_rate'] = round((hired_count / total * 100), 1) if total > 0 else 0
            if branch_stats[branch]['_hired_count_for_time'] > 0:
                avg = branch_stats[branch]['_total_time_to_hire'] / branch_stats[branch]['_hired_count_for_time']
                branch_stats[branch]['avg_time_to_hire'] = round(avg, 1)
            else:
                branch_stats[branch]['avg_time_to_hire'] = None
            # cleanup internal keys
            branch_stats[branch].pop('_total_time_to_hire', None)
            branch_stats[branch].pop('_hired_count_for_time', None)
        
        # Group by job position
        job_stats = {}
        for record in all_records:
            job = record.get('position', 'Unknown')
            if job not in job_stats:
                job_stats[job] = 0
            job_stats[job] += 1
        
        # Group by status
        status_breakdown = {
            'pending': pending,
            'scheduled': len([r for r in all_records if r.get('status') == 'scheduled']),
            'interviewed': len([r for r in all_records if r.get('status') == 'interviewed']),
            'hired': hired,
            'rejected': rejected
        }
        
        # Group by date for trend chart - adapt based on period
        daily_stats = {}
        today = datetime.now()
        
        # Determine number of days to display based on period
        if period == 'week':
            days_to_show = 7
        elif period == 'month':
            days_to_show = 30
        elif period == 'year':
            days_to_show = 365
        else:  # 'all'
            days_to_show = 30  # Default to last 30 days for 'all'
        
        # Initialize daily stats for the period and keep full date objects for correct sorting
        date_map = {}
        # Use a readable month-day label (e.g. 'Dec 02') consistently for keys so counting matches
        for i in range(days_to_show):
            date_obj = (today - timedelta(days=i))
            date_key = date_obj.strftime('%b %d')
            daily_stats[date_key] = 0
            # store the exact date object for this label to avoid ambiguous parsing later
            date_map[date_key] = date_obj
        
        # Count hires (hired status) per day — use hire date (updated_at) when available
        for record in all_records:
            # For trend of hires, prefer the updated_at date when status == 'hired'
            date_candidate = None
            try:
                if record.get('status') == 'hired' and record.get('updated_at'):
                    date_candidate = record.get('updated_at')
                else:
                    date_candidate = record.get('submitted_at')
                if not date_candidate:
                    continue

                # Normalize to datetime object or parse string
                parsed = None
                if hasattr(date_candidate, 'strftime'):
                    parsed = date_candidate
                else:
                    s = str(date_candidate)
                    for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                        try:
                            parsed = datetime.strptime(s, fmt)
                            break
                        except Exception:
                            parsed = None
                if not parsed:
                    continue

                date_key = parsed.strftime('%b %d')
                # Only count hires (status == 'hired') for the hiring trend
                if record.get('status') == 'hired' and date_key in daily_stats:
                    daily_stats[date_key] += 1
            except Exception:
                continue
        
        # Sort dates from oldest to newest using the stored date objects (avoids year-less parsing)
        try:
            # date_map currently maps labels like '%m/%d' to date objects; convert keys to our display format
            converted_map = { (date_map[k].strftime('%b %d') if k in date_map else k): date_map.get(k, today) for k in date_map }
            trend_labels = sorted(list(daily_stats.keys()), key=lambda k: converted_map.get(k, today))
        except Exception:
            trend_labels = list(daily_stats.keys())
        trend_data = [daily_stats[label] for label in trend_labels]
        
        cursor.close()
        
        # Return formatted JSON response
        return jsonify({
            'success': True,
            'kpis': {
                'applicants': total_applicants,
                'interviews': interviews_scheduled,
                'hired': hired,
                'timeToHire': avg_time_to_hire,
                'conversionRate': conversion_rate,
                'rejected': rejected,
                'pending': pending
            },
            'branchPerformance': branch_stats,
            'jobStats': job_stats,
            'statusBreakdown': status_breakdown,
            'trendChart': {
                'labels': trend_labels,
                'data': trend_data
            },
            'applicantData': [
                {
                    'name': r.get('full_name'),
                    'job': r.get('position'),
                    'branch': r.get('branch'),
                    'status': r.get('status'),
                    'date': str(r.get('submitted_at')).split(' ')[0],
                    'interview': str(r.get('interview_date')).split(' ')[0] if r.get('interview_date') else '',
                    'time': int(r.get('time_to_hire')) if (r.get('status') == 'hired' and r.get('time_to_hire')) else None
                }
                for r in all_records
            ]
        })
        
    except Exception as e:
        import traceback
        print(f'⚠️ Error fetching report data: {e}')
        print(f'Traceback: {traceback.format_exc()}')
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            cursor.close()
        except:
            pass


@app.route('/api/jobs', methods=['GET'])
@login_required('admin', 'hr')
def api_get_jobs():
    """Return a list of job postings visible to the current user.

    Response JSON: { success: True, jobs: [{ job_id, title, branch_name }, ...] }
    """
    try:
        user = get_current_user()
        jobs = fetch_jobs_for_user(user) or []
        out = []
        seen = set()
        for j in jobs:
            title = (j.get('job_title') or j.get('position_name') or j.get('position_title') or '').strip()
            if not title:
                continue
            # avoid duplicates by title
            if title in seen:
                continue
            seen.add(title)
            out.append({
                'job_id': j.get('job_id'),
                'title': title,
                'branch_name': j.get('branch_name')
            })

        return jsonify({'success': True, 'jobs': out})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/branches', methods=['GET'])
@login_required('admin', 'hr')
def api_get_branches():
    """Return a list of branches for filters.

    Response JSON: { success: True, branches: [{ branch_id, branch_name, address }, ...] }
    """
    try:
        branches = fetch_branches() or []
        out = []
        for b in branches:
            out.append({
                'branch_id': b.get('branch_id'),
                'branch_name': b.get('branch_name')
            })
        return jsonify({'success': True, 'branches': out})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/reports-analytics', endpoint='admin_reports_analytics')
@app.route('/hr/reports-analytics', endpoint='hr_reports_analytics')
@login_required('admin', 'hr')
def reports_analytics():
    """Comprehensive analytics and reporting dashboard with period filtering and export."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        template = 'hr/reports_analytics.html' if user.get('role') == 'hr' else 'admin/reports.html'
        # Get period from request
        period = request.args.get('period', 'all')
        selected_week = request.args.get('week', '')
        selected_week_month = request.args.get('week_month', '')
        selected_week_number = request.args.get('week_number', '')
        selected_month = request.args.get('month', '')
        selected_year = request.args.get('year', '')
        return render_template(
            template, 
            stats={}, 
            trends={}, 
            branch_stats=[], 
            metrics={},
            branch_comparison=[],
            system_stats={},
            job_performance=[],
            status_breakdown=[],
            funnel_data={},
            source_effectiveness=[],
            branch_info=None,
            period=period, 
            period_summary={}, 
            available_years=[], 
            selected_week=selected_week, 
            selected_week_month=selected_week_month,
            selected_week_number=selected_week_number,
            selected_month=selected_month, 
            selected_year=selected_year, 
            applicant_summary={}, 
            job_vacancy={}, 
            hiring_outcome={}, 
            hr_performance={}, 
            applicant_summary_details=[], 
            job_vacancy_details=[]
        )
    
    # Ensure schema compatibility before proceeding
    ensure_schema_compatibility()
    
    cursor = db.cursor(dictionary=True)
    
    try:
        # Get period parameters
        period = request.args.get('period', 'all')
        selected_week = request.args.get('week', '')  # Legacy support
        selected_week_month = request.args.get('week_month', '')
        selected_week_number = request.args.get('week_number', '')
        selected_month = request.args.get('month', '')
        selected_year = request.args.get('year', '')
        export_format = request.args.get('export', '')
        export_section = request.args.get('section', '')
        export_type = request.args.get('type', 'summary')
        generate_section = request.args.get('generate_section', '') or request.form.get('generate_section', '')
        
        # Calculate date range based on period
        from datetime import datetime, timedelta
        from datetime import date as date_class
        period_summary = {}
        date_filter = ''
        date_params = []
        
        if period == 'week':
            # New week selector: month + week number (1-5)
            if selected_week_month and selected_week_number:
                try:
                    year, month = map(int, selected_week_month.split('-'))
                    week_num = int(selected_week_number)
                    
                    # Calculate the first day of the month
                    month_start = datetime(year, month, 1)
                    first_day_weekday = month_start.weekday()  # 0=Monday, 6=Sunday
                    
                    # Calculate start of week (Monday)
                    # Find the Monday of the week containing the 1st
                    days_to_monday = (first_day_weekday) % 7
                    first_monday = month_start - timedelta(days=days_to_monday)
                    
                    # Week 1 starts on the first Monday on or before the 1st
                    # Week 2 starts 7 days after that, etc.
                    days_to_add = (week_num - 1) * 7
                    start_date = first_monday + timedelta(days=days_to_add)
                    
                    # Ensure start_date is within the month (at least the 1st)
                    if start_date < month_start:
                        start_date = month_start
                    
                    # End date is 6 days after start (Sunday)
                    end_date = start_date + timedelta(days=6)
                    
                    # Make sure we don't go beyond the month
                    if month == 12:
                        month_end = datetime(year + 1, 1, 1) - timedelta(days=1)
                    else:
                        month_end = datetime(year, month + 1, 1) - timedelta(days=1)
                    
                    if end_date > month_end:
                        end_date = month_end
                    
                except (ValueError, IndexError) as e:
                    print(f'Error parsing week: {e}')
                    today = datetime.now()
                    start_date = today - timedelta(days=today.weekday())
                    end_date = start_date + timedelta(days=6)
            # Legacy support for old week format
            elif selected_week:
                try:
                    year, week = map(int, selected_week.split('-W'))
                    jan4 = date_class(year, 1, 4)
                    jan4_weekday = jan4.weekday()
                    days_since_jan4 = (week - 1) * 7 - jan4_weekday
                    week_start = jan4 + timedelta(days=days_since_jan4)
                    start_date = datetime.combine(week_start, datetime.min.time())
                    end_date = start_date + timedelta(days=6)
                except (ValueError, IndexError):
                    today = datetime.now()
                    start_date = today - timedelta(days=today.weekday())
                    end_date = start_date + timedelta(days=6)
            else:
                # Default to current week (Monday to Sunday)
                today = datetime.now()
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)
                # Set default selected_week_month and selected_week_number for template
                selected_week_month = f"{today.year}-{today.month:02d}"
                # Calculate which week of the month we're in
                first_day = datetime(today.year, today.month, 1)
                first_monday = first_day - timedelta(days=first_day.weekday())
                week_num = ((start_date - first_monday).days // 7) + 1
                if week_num < 1:
                    week_num = 1
                elif week_num > 5:
                    week_num = 5
                selected_week_number = str(week_num)
            
            date_filter = "AND DATE(a.submitted_at) BETWEEN %s AND %s"
            date_params = [start_date.date(), end_date.date()]
            period_summary = {
                'start_date': start_date.strftime('%b %d, %Y'),
                'end_date': end_date.strftime('%b %d, %Y'),
                'total_days': (end_date.date() - start_date.date()).days + 1,
                'generated_at': datetime.now().strftime('%b %d, %Y %I:%M %p')
            }
        elif period == 'month':
            if selected_month:
                try:
                    year, month = map(int, selected_month.split('-'))
                    start_date = datetime(year, month, 1)
                    if month == 12:
                        end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                    else:
                        end_date = datetime(year, month + 1, 1) - timedelta(days=1)
                except (ValueError, IndexError):
                    # Default to current month if parsing fails
                    today = datetime.now()
                    start_date = datetime(today.year, today.month, 1)
                    if today.month == 12:
                        end_date = datetime(today.year + 1, 1, 1) - timedelta(days=1)
                    else:
                        end_date = datetime(today.year, today.month + 1, 1) - timedelta(days=1)
                    # Update selected_month to current month for template
                    selected_month = f"{today.year}-{today.month:02d}"
            else:
                # Default to current month
                today = datetime.now()
                start_date = datetime(today.year, today.month, 1)
                if today.month == 12:
                    end_date = datetime(today.year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(today.year, today.month + 1, 1) - timedelta(days=1)
                # Set default selected_month for template
                selected_month = f"{today.year}-{today.month:02d}"
            
            date_filter = "AND DATE(a.submitted_at) BETWEEN %s AND %s"
            date_params = [start_date.date(), end_date.date()]
            period_summary = {
                'start_date': start_date.strftime('%b %d, %Y'),
                'end_date': end_date.strftime('%b %d, %Y'),
                'total_days': (end_date - start_date).days + 1,
                'generated_at': datetime.now().strftime('%b %d, %Y %I:%M %p')
            }
        elif period == 'year':
            if selected_year:
                try:
                    year = int(selected_year)
                except ValueError:
                    year = datetime.now().year
            else:
                # Default to current year
                year = datetime.now().year
                # Set default selected_year for template
                selected_year = str(year)
            
            start_date = datetime(year, 1, 1)
            end_date = datetime(year, 12, 31)
            date_filter = "AND YEAR(a.submitted_at) = %s"
            date_params = [year]
            period_summary = {
                'start_date': start_date.strftime('%b %d, %Y'),
                'end_date': end_date.strftime('%b %d, %Y'),
                'total_days': 365 if year % 4 != 0 else 366,
                'generated_at': datetime.now().strftime('%b %d, %Y %I:%M %p')
            }
        
        # Get available years for dropdown
        cursor.execute("SELECT DISTINCT YEAR(submitted_at) as year FROM applications WHERE submitted_at IS NOT NULL ORDER BY year DESC")
        available_years = [str(row['year']) for row in cursor.fetchall() if row.get('year')]
        
        # Get user's branch scope first (needed for filtering)
        branch_id = get_branch_scope(user)
        
        # Get branches for filter dropdown
        branches = fetch_branches()
        
        # Get unique job titles for filter dropdown
        branch_filter_id = request.args.get('branch_id', type=int)
        
        # Build job filter query - filter by branch if specified, otherwise show all jobs
        job_filter_where = ""
        job_filter_params = []
        if branch_filter_id:
            job_filter_where = "WHERE j.branch_id = %s"
            job_filter_params = [branch_filter_id]
        elif branch_id:
            # If user has branch scope, limit jobs to that branch
            job_filter_where = "WHERE j.branch_id = %s"
            job_filter_params = [branch_id]
        
        cursor.execute(f'''
            SELECT DISTINCT j.job_id, COALESCE(j.title, 'Untitled Job') AS job_title
            FROM jobs j
            {job_filter_where}
            ORDER BY j.title ASC
        ''', tuple(job_filter_params) if job_filter_params else None)
        jobs = cursor.fetchall() or []
        
        # Determine available job columns for backward compatibility
        cursor.execute('SHOW COLUMNS FROM jobs')
        job_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}

        def job_expr(candidates, fallback):
            for column in candidates:
                if column in job_columns:
                    return f'j.{column}'
            return fallback

        def job_coalesce(candidates, fallback):
            expressions = [f'j.{column}' for column in candidates if column in job_columns]
            if not expressions:
                return fallback
            if len(expressions) == 1:
                return expressions[0]
            return f"COALESCE({', '.join(expressions)})"

        job_title_expr = job_expr(['job_title', 'title'], "'Untitled Job'")
        job_posted_expr = job_coalesce(['posted_at', 'created_at'], 'j.created_at')
        time_to_apply_expr = f'CASE WHEN a.submitted_at IS NOT NULL AND {job_posted_expr} IS NOT NULL THEN DATEDIFF(a.submitted_at, {job_posted_expr}) ELSE NULL END'
        where_templates = []
        params_reports = []
        
        # Apply branch filter from request (takes precedence over user's branch scope)
        # branch_filter_id is already retrieved above for job filtering
        if branch_filter_id:
            where_templates.append('{alias}.branch_id = %s')
            params_reports.append(branch_filter_id)
        elif branch_id:
            # Use user's branch scope only if no filter is specified
            where_templates.append('{alias}.branch_id = %s')
            params_reports.append(branch_id)
        
        # Apply job filter from request
        job_filter_id = request.args.get('job_id', type=int)
        if job_filter_id:
            where_templates.append('{alias}.job_id = %s')
            params_reports.append(job_filter_id)
        
        # Apply HR User filter from request
        hr_filter_id = request.args.get('hr_user_id', type=int)
        hr_filter_condition = None
        if hr_filter_id:
            # HR filter will be applied separately to queries that need it
            hr_filter_condition = hr_filter_id
        
        # Apply Status filter from request
        status_filter = request.args.get('status', '').strip()
        status_filter_condition = None
        if status_filter and status_filter.lower() != 'all':
            status_filter_condition = status_filter.lower()

        def build_where(alias):
            if not where_templates:
                return ''
            clause = ' AND '.join(template.format(alias=alias) for template in where_templates)
            return f'WHERE {clause}'

        base_where = build_where('j')
        
        # Apply date filter to all queries
        date_filter_clause = date_filter if date_filter else ''
        date_params_all = date_params if date_params else []
        
        # Comprehensive statistics (with period filter)
        stats = build_report_stats(user, date_filter_clause, date_params_all)
        
        # Comprehensive recruitment metrics
        metrics = {}
        try:
            if base_where:
                # For branch-scoped users, total_branches should be 1 (their branch)
                # But we'll get it from branches table to be accurate
                branch_count_query = 'SELECT COUNT(*) AS count FROM branches WHERE branch_id = %s'
                branch_count_params = (branch_id,)
                total_branches_count = fetch_count(branch_count_query, branch_count_params) or 0
                
                metrics_query = f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT j.job_id) AS total_jobs,
                        {total_branches_count} AS total_branches,
                        COUNT(DISTINCT i.interview_id) AS total_interviews,
                        AVG({time_to_apply_expr}) AS avg_time_to_apply,
                        AVG(CASE WHEN a.submitted_at IS NOT NULL AND i.scheduled_date IS NOT NULL 
                            THEN DATEDIFF(i.scheduled_date, a.submitted_at) ELSE NULL END) AS avg_time_to_interview,
                        AVG(CASE WHEN a.status = 'hired' AND a.submitted_at IS NOT NULL AND a.updated_at IS NOT NULL 
                            THEN DATEDIFF(a.updated_at, a.submitted_at) ELSE NULL END) AS avg_time_to_hire,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS total_hires,
                        COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) AS applications_with_interviews,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) AS hired_with_interview,
                        CASE 
                            WHEN COUNT(*) > 0 THEN SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) / COUNT(*) * 100 
                            ELSE 0 
                        END AS acceptance_rate,
                        CASE 
                            WHEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) * 100
                            ELSE 0 
                        END AS interview_to_hire_rate,
                        CASE 
                            WHEN COUNT(*) > 0 
                            THEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(*) * 100
                            ELSE 0 
                        END AS application_to_interview_rate
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN interviews i ON i.application_id = a.application_id
                    {base_where}
                    '''
                cursor.execute(metrics_query, tuple(params_reports) if params_reports else None)
            else:
                # For admin (no branch scope), count all branches from branches table
                total_branches_count = fetch_count('SELECT COUNT(*) AS count FROM branches') or 0
                
                metrics_query = f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT j.job_id) AS total_jobs,
                        {total_branches_count} AS total_branches,
                        COUNT(DISTINCT i.interview_id) AS total_interviews,
                        AVG({time_to_apply_expr}) AS avg_time_to_apply,
                        AVG(CASE WHEN a.submitted_at IS NOT NULL AND i.scheduled_date IS NOT NULL 
                            THEN DATEDIFF(i.scheduled_date, a.submitted_at) ELSE NULL END) AS avg_time_to_interview,
                        AVG(CASE WHEN a.status = 'hired' AND a.submitted_at IS NOT NULL AND a.updated_at IS NOT NULL 
                            THEN DATEDIFF(a.updated_at, a.submitted_at) ELSE NULL END) AS avg_time_to_hire,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS total_hires,
                        COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) AS applications_with_interviews,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) AS hired_with_interview,
                        CASE 
                            WHEN COUNT(*) > 0 THEN SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) / COUNT(*) * 100 
                            ELSE 0 
                        END AS acceptance_rate,
                        CASE 
                            WHEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) * 100
                            ELSE 0 
                        END AS interview_to_hire_rate,
                        CASE 
                            WHEN COUNT(*) > 0 
                            THEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(*) * 100
                            ELSE 0 
                        END AS application_to_interview_rate
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN interviews i ON i.application_id = a.application_id
                '''
                cursor.execute(metrics_query)
            metrics = cursor.fetchone() or {}
        except Exception as metrics_exc:
            print(f'⚠️ Error fetching metrics: {metrics_exc}')
            import traceback
            print(traceback.format_exc())
            metrics = {}
        
        # Recruitment funnel data
        funnel_data = {}
        try:
            if base_where:
                funnel_query = f'''
                    SELECT 
                        COUNT(DISTINCT CASE WHEN a.status = 'applied' THEN a.application_id END) AS applied,
                        COUNT(DISTINCT CASE WHEN a.status = 'under_review' THEN a.application_id END) AS under_review,
                        COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) AS interviewed,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected
                    FROM applications a
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN interviews i ON i.application_id = a.application_id
                    {base_where}
                '''
                cursor.execute(funnel_query, tuple(params_reports))
            else:
                funnel_query = '''
                    SELECT 
                        COUNT(DISTINCT CASE WHEN a.status = 'applied' THEN a.application_id END) AS applied,
                        COUNT(DISTINCT CASE WHEN a.status = 'under_review' THEN a.application_id END) AS under_review,
                        COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) AS interviewed,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected
                    FROM applications a
                    LEFT JOIN interviews i ON i.application_id = a.application_id
                '''
                cursor.execute(funnel_query)
            funnel_data = cursor.fetchone() or {}
        except Exception as funnel_exc:
            print(f'⚠️ Error fetching funnel data: {funnel_exc}')
            funnel_data = {}
        
        # Source effectiveness (if source field exists in applications)
        source_effectiveness = []
        try:
            cursor.execute('SHOW COLUMNS FROM applications LIKE %s', ('source%',))
            has_source = cursor.fetchone() is not None
            if has_source and base_where:
                source_query = f'''
                    SELECT 
                        COALESCE(a.source, 'Unknown') AS source,
                        COUNT(*) AS total_applications,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hires,
                        CASE 
                            WHEN COUNT(*) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) / COUNT(*) * 100
                            ELSE 0 
                        END AS conversion_rate
                    FROM applications a
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {base_where}
                    GROUP BY a.source
                    ORDER BY total_applications DESC
                '''
                cursor.execute(source_query, tuple(params_reports))
                source_effectiveness = cursor.fetchall() or []
            elif has_source:
                source_query = '''
                    SELECT 
                        COALESCE(a.source, 'Unknown') AS source,
                        COUNT(*) AS total_applications,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hires,
                        CASE 
                            WHEN COUNT(*) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) / COUNT(*) * 100
                            ELSE 0 
                        END AS conversion_rate
                    FROM applications a
                    GROUP BY a.source
                    ORDER BY total_applications DESC
                '''
                cursor.execute(source_query)
                source_effectiveness = cursor.fetchall() or []
        except Exception:
            source_effectiveness = []
        
        # Trends (last 30 days)
        trends = []
        try:
            trends_where_parts = ['a.submitted_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)']
            trends_params = []
            if base_where:
                # Extract the WHERE condition from base_where and add it
                where_condition = base_where.replace('WHERE ', '')
                trends_where_parts.append(where_condition)
                trends_params.extend(params_reports)
            
            trends_where_clause = 'WHERE ' + ' AND '.join(trends_where_parts) if trends_where_parts else ''
            trends_query = f'''
                SELECT 
                    DATE(a.submitted_at) AS date,
                    COUNT(*) AS applications,
                    SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS accepted,
                    SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {trends_where_clause}
                GROUP BY DATE(a.submitted_at)
                ORDER BY date DESC
            '''
            cursor.execute(trends_query, tuple(trends_params) if trends_params else None)
            trends = cursor.fetchall() or []
        except Exception as trends_exc:
            print(f'⚠️ Error fetching trends: {trends_exc}')
            trends = []
        
        # Branch statistics (admin only)
        branch_stats = []
        if user.get('role') == 'admin':
            cursor.execute(
                '''
                SELECT 
                    b.branch_id,
                    b.branch_name,
                    COUNT(DISTINCT a.application_id) AS total_applications,
                    SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS accepted,
                    SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected,
                    COUNT(DISTINCT j.job_id) AS total_jobs,
                    COUNT(DISTINCT i.interview_id) AS total_interviews
                FROM branches b
                LEFT JOIN jobs j ON j.branch_id = b.branch_id
                LEFT JOIN applications a ON a.job_id = j.job_id
                LEFT JOIN interviews i ON i.application_id = a.application_id
                GROUP BY b.branch_id, b.branch_name
                ORDER BY total_applications DESC
                '''
            )
            branch_stats = cursor.fetchall()
        
        # Job performance
        job_performance = []
        try:
            cursor.execute(
                f'''
                SELECT 
                    j.job_id,
                    {job_title_expr} AS job_title,
                    COUNT(a.application_id) AS application_count,
                    SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS accepted_count,
                    SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected_count,
                    COUNT(DISTINCT i.interview_id) AS interview_count
                FROM jobs j
                LEFT JOIN applications a ON a.job_id = j.job_id
                LEFT JOIN interviews i ON i.application_id = a.application_id
                {base_where if base_where else ''}
                GROUP BY j.job_id, {job_title_expr}
                ORDER BY application_count DESC
                LIMIT 10
                ''',
                tuple(params_reports) if params_reports and base_where else None,
            )
            job_performance = cursor.fetchall() or []
        except Exception as job_perf_exc:
            print(f'⚠️ Error fetching job performance: {job_perf_exc}')
            job_performance = []
        
        # Status breakdown - compute percentages in Python for compatibility
        status_breakdown = []
        try:
            status_query = f'''
                SELECT 
                    a.status,
                    COUNT(*) AS count
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {base_where if base_where else ''}
                GROUP BY a.status
                ORDER BY count DESC
            '''
            cursor.execute(status_query, tuple(params_reports) if params_reports and base_where else None)
            status_breakdown = cursor.fetchall() or []
            total_status = sum((row.get('count') or 0) for row in status_breakdown)
            for row in status_breakdown:
                count_value = row.get('count') or 0
                row['percentage'] = (count_value / total_status * 100) if total_status else 0
        except Exception as status_exc:
            print(f'⚠️ Error fetching status breakdown: {status_exc}')
            status_breakdown = []
        
        # Cross-branch comparison (only for Super Admin)
        branch_comparison = []
        if user.get('role') == 'admin' and not branch_id:
            cursor.execute(
                '''
                SELECT 
                    b.branch_id,
                    b.branch_name,
                    COUNT(DISTINCT j.job_id) AS total_jobs,
                    COUNT(DISTINCT a.application_id) AS total_applications,
                    COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired_count,
                    COUNT(DISTINCT CASE WHEN a.status = 'pending' THEN a.application_id END) AS pending_count,
                    COUNT(DISTINCT i.interview_id) AS total_interviews,
                    COUNT(DISTINCT a.applicant_id) AS unique_applicants
                FROM branches b
                LEFT JOIN jobs j ON j.branch_id = b.branch_id
                LEFT JOIN applications a ON a.job_id = j.job_id
                LEFT JOIN interviews i ON i.application_id = a.application_id
                GROUP BY b.branch_id, b.branch_name
                ORDER BY total_applications DESC
                '''
            )
            branch_comparison = cursor.fetchall()
        
        # System usage statistics (only for Super Admin)
        system_stats = {}
        if user.get('role') == 'admin' and not branch_id:
            try:
                cursor.execute('SELECT COUNT(*) AS total_users FROM users')
                system_stats['total_users'] = cursor.fetchone()['total_users'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_admins FROM users WHERE user_type = "super_admin"')
                system_stats['total_admins'] = cursor.fetchone()['total_admins'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_hr FROM users WHERE user_type = "hr"')
                system_stats['total_hr'] = cursor.fetchone()['total_hr'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_applicants FROM users WHERE user_type = "applicant"')
                system_stats['total_applicants'] = cursor.fetchone()['total_applicants'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_jobs_posted FROM jobs')
                system_stats['total_jobs_posted'] = cursor.fetchone()['total_jobs_posted'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_applications_received FROM applications')
                system_stats['total_applications_received'] = cursor.fetchone()['total_applications_received'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_interviews_scheduled FROM interviews')
                system_stats['total_interviews_scheduled'] = cursor.fetchone()['total_interviews_scheduled'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_hires FROM applications WHERE status = "hired"')
                system_stats['total_hires'] = cursor.fetchone()['total_hires'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_sessions_24h FROM auth_sessions WHERE login_time >= DATE_SUB(NOW(), INTERVAL 24 HOUR)')
                system_stats['total_sessions_24h'] = cursor.fetchone()['total_sessions_24h'] or 0
            except Exception as e:
                print(f'⚠️ Error fetching system stats: {e}')
                system_stats = {
                    'total_users': 0,
                    'total_admins': 0,
                    'total_hr': 0,
                    'total_applicants': 0,
                    'total_jobs_posted': 0,
                    'total_applications_received': 0,
                    'total_interviews_scheduled': 0,
                    'total_hires': 0,
                    'total_sessions_24h': 0,
                }
        
        # Generate report sections data
        # 1. Applicant Summary Report
        applicant_summary = {}
        applicant_summary_details = []
        try:
            # Ensure date_filter is properly formatted
            if date_filter and not date_filter.strip().startswith('AND'):
                date_filter = f"AND {date_filter.strip()}"
            # Check if email_verified column exists in applicants or users table
            cursor.execute("SHOW COLUMNS FROM applicants LIKE 'email_verified'")
            has_applicant_email_verified = cursor.fetchone() is not None
            
            cursor.execute("SHOW COLUMNS FROM users LIKE 'email_verified'")
            has_user_email_verified = cursor.fetchone() is not None
            
            # Build verified_email expression
            if has_applicant_email_verified:
                verified_email_expr = "CASE WHEN ap.email_verified = 1 THEN ap.applicant_id END"
            elif has_user_email_verified:
                verified_email_expr = "CASE WHEN u.email_verified = 1 THEN ap.applicant_id END"
            else:
                verified_email_expr = "NULL"  # Return 0 if column doesn't exist
            
            # Build date filter clause for applicant query
            date_clause_applicant = ""
            if date_filter:
                # For applicant summary, filter by application submission date
                # Remove leading 'AND ' only, not all occurrences
                app_date_filter = date_filter.strip()
                if app_date_filter.startswith('AND '):
                    app_date_filter = app_date_filter[4:]  # Remove 'AND ' from start
                date_clause_applicant = f" AND {app_date_filter}"
            
            if branch_id:
                join_users = "LEFT JOIN users u ON ap.user_id = u.user_id" if has_user_email_verified and not has_applicant_email_verified else ""
                applicant_query = f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT CASE WHEN DATE(ap.created_at) >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) THEN ap.applicant_id END) AS new_applicants,
                        COUNT(DISTINCT CASE WHEN a.status IN ('pending', 'scheduled', 'interviewed') THEN a.application_id END) AS active_applications,
                        COUNT(DISTINCT CASE WHEN r.resume_id IS NOT NULL THEN ap.applicant_id END) AS with_resume,
                        COUNT(DISTINCT {verified_email_expr}) AS verified_email
                    FROM applicants ap
                    JOIN applications a ON ap.applicant_id = a.applicant_id
                    JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN resumes r ON ap.applicant_id = r.applicant_id
                    {join_users}
                    WHERE j.branch_id = %s{date_clause_applicant}
                '''
                applicant_params = [branch_id] + (date_params_all if date_params_all else [])
            else:
                join_users = "LEFT JOIN users u ON ap.user_id = u.user_id" if has_user_email_verified and not has_applicant_email_verified else ""
                where_clause = ""
                if date_filter:
                    # Remove leading 'AND ' only, not all occurrences
                    app_date_filter = date_filter.strip()
                    if app_date_filter.startswith('AND '):
                        app_date_filter = app_date_filter[4:]  # Remove 'AND ' from start
                    where_clause = f"WHERE {app_date_filter}"
                applicant_query = f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT CASE WHEN DATE(ap.created_at) >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) THEN ap.applicant_id END) AS new_applicants,
                        COUNT(DISTINCT CASE WHEN a.status IN ('pending', 'scheduled', 'interviewed') THEN a.application_id END) AS active_applications,
                        COUNT(DISTINCT CASE WHEN r.resume_id IS NOT NULL THEN ap.applicant_id END) AS with_resume,
                        COUNT(DISTINCT {verified_email_expr}) AS verified_email
                    FROM applicants ap
                    LEFT JOIN applications a ON ap.applicant_id = a.applicant_id
                    LEFT JOIN resumes r ON ap.applicant_id = r.applicant_id
                    {join_users}
                    {where_clause}
                '''
                applicant_params = date_params_all if date_params_all else []
            
            try:
                cursor.execute(applicant_query, tuple(applicant_params) if applicant_params else None)
                applicant_summary = cursor.fetchone() or {}
            except Exception as query_error:
                print(f'⚠️ Error in applicant query: {query_error}')
                applicant_summary = {}
            
            # Get detailed applicant list with job applications and status
            # Build WHERE clause using the same filters as other queries
            detail_where_parts = []
            detail_params = []
            
            # Apply branch filter (from request takes precedence, otherwise user's branch scope)
            if branch_filter_id:
                detail_where_parts.append("j.branch_id = %s")
                detail_params.append(branch_filter_id)
            elif branch_id:
                detail_where_parts.append("j.branch_id = %s")
                detail_params.append(branch_id)
            
            # Apply job filter
            if job_filter_id:
                detail_where_parts.append("j.job_id = %s")
                detail_params.append(job_filter_id)
            
            # Apply status filter
            if status_filter_condition:
                detail_where_parts.append("a.status = %s")
                detail_params.append(status_filter_condition)
            
            # Apply date filter
            if date_filter and date_params_all:
                app_date_filter = date_filter.strip()
                if app_date_filter.startswith('AND '):
                    app_date_filter = app_date_filter[4:]
                detail_where_parts.append(app_date_filter)
                detail_params.extend(date_params_all)
            
            # Build WHERE clause
            if detail_where_parts:
                detail_where_clause = "WHERE " + " AND ".join(detail_where_parts)
            else:
                detail_where_clause = ""
            
            detail_query = f'''
                SELECT 
                    ap.full_name AS applicant_name,
                    ap.email,
                    {job_title_expr} AS job_title,
                    b.branch_name,
                    a.status AS application_status,
                    a.submitted_at AS date_applied,
                    a.updated_at AS date_updated,
                    i.scheduled_date AS interview_scheduled_date,
                    CASE 
                        WHEN a.status = 'hired' AND a.submitted_at IS NOT NULL AND a.updated_at IS NOT NULL
                        THEN DATEDIFF(a.updated_at, a.submitted_at)
                        ELSE NULL
                    END AS time_to_hire_days,
                    CASE 
                        WHEN a.status = 'hired' THEN 'Hired'
                        WHEN a.status = 'rejected' THEN 'Not Hired'
                        WHEN a.status = 'pending' THEN 'Pending'
                        WHEN a.status = 'scheduled' THEN 'Interview Scheduled'
                        WHEN a.status = 'interviewed' THEN 'Interviewed'
                        ELSE 'Unknown'
                    END AS status_label,
                    CASE 
                        WHEN a.status = 'hired' THEN a.updated_at
                        WHEN a.status = 'rejected' THEN a.updated_at
                        WHEN a.status = 'scheduled' THEN i.scheduled_date
                        WHEN a.status = 'interviewed' THEN a.updated_at
                        WHEN a.status = 'pending' THEN a.submitted_at
                        ELSE a.updated_at
                    END AS status_date,
                    COALESCE(
                        (SELECT ad.full_name 
                         FROM activity_logs al_log 
                         LEFT JOIN admins ad ON al_log.admin_id = ad.admin_id
                         WHERE al_log.target_table = 'applications' 
                         AND al_log.target_id = a.application_id
                         AND al_log.action IN ('Updated application status', 'Hired applicant', 'Rejected applicant', 'Scheduled interview')
                         ORDER BY al_log.created_at DESC
                         LIMIT 1),
                        'Unassigned'
                    ) AS hr_assigned
                FROM applicants ap
                JOIN applications a ON ap.applicant_id = a.applicant_id
                LEFT JOIN jobs j ON a.job_id = j.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                LEFT JOIN (
                    SELECT i1.application_id, i1.scheduled_date
                    FROM interviews i1
                    INNER JOIN (
                        SELECT application_id, MAX(scheduled_date) AS max_date
                        FROM interviews
                        GROUP BY application_id
                    ) i2 ON i1.application_id = i2.application_id AND i1.scheduled_date = i2.max_date
                ) i ON i.application_id = a.application_id
                {detail_where_clause}
                ORDER BY a.submitted_at DESC
                LIMIT 200
            '''
            
            try:
                cursor.execute(detail_query, tuple(detail_params) if detail_params else None)
                applicant_summary_details = cursor.fetchall() or []
            except Exception as detail_error:
                print(f'⚠️ Error in applicant detail query: {detail_error}')
                print(f'Query: {detail_query}')
                print(f'Params: {detail_params}')
                applicant_summary_details = []
            
            # Format dates and enhance status information
            for app in applicant_summary_details:
                # Format date_applied
                if app.get('date_applied'):
                    try:
                        if isinstance(app['date_applied'], (datetime, date_class)):
                            app['date_applied'] = app['date_applied'].strftime('%b %d, %Y %I:%M %p')
                        elif isinstance(app['date_applied'], str):
                            dt = datetime.strptime(app['date_applied'], '%Y-%m-%d %H:%M:%S')
                            app['date_applied'] = dt.strftime('%b %d, %Y %I:%M %p')
                    except:
                        app['date_applied'] = app.get('date_applied', '—')
                
                # Format interview_scheduled_date
                if app.get('interview_scheduled_date'):
                    try:
                        if isinstance(app['interview_scheduled_date'], (datetime, date_class)):
                            app['interview_date'] = app['interview_scheduled_date'].strftime('%b %d, %Y %I:%M %p')
                        elif isinstance(app['interview_scheduled_date'], str):
                            dt = datetime.strptime(app['interview_scheduled_date'], '%Y-%m-%d %H:%M:%S')
                            app['interview_date'] = dt.strftime('%b %d, %Y %I:%M %p')
                    except:
                        app['interview_date'] = app.get('interview_scheduled_date', '—')
                else:
                    app['interview_date'] = '—'
                
                # Format time_to_hire
                time_to_hire_days = app.get('time_to_hire_days')
                if time_to_hire_days is not None:
                    if time_to_hire_days == 0:
                        app['time_to_hire'] = 'Same day'
                    elif time_to_hire_days == 1:
                        app['time_to_hire'] = '1 day'
                    else:
                        app['time_to_hire'] = f'{time_to_hire_days} days'
                else:
                    app['time_to_hire'] = '—'
                
                # Format date_updated (hired date if status is hired)
                if app.get('date_updated'):
                    try:
                        if isinstance(app['date_updated'], (datetime, date_class)):
                            app['date_updated'] = app['date_updated'].strftime('%b %d, %Y %I:%M %p')
                        elif isinstance(app['date_updated'], str):
                            dt = datetime.strptime(app['date_updated'], '%Y-%m-%d %H:%M:%S')
                            app['date_updated'] = dt.strftime('%b %d, %Y %I:%M %p')
                    except:
                        app['date_updated'] = app.get('date_updated', '—')
                
                # Set last_activity for backward compatibility
                app['last_activity'] = app.get('date_applied', '—')
                app['current_status'] = app.get('status_label', app.get('application_status', '—'))
            
        except Exception as e:
            print(f'⚠️ Error fetching applicant summary: {e}')
            applicant_summary = {}
            applicant_summary_details = []
        
        # Group applicants by status after fetching details
        applicants_by_status = {
            'hired': [],
            'rejected': [],
            'pending': [],
            'scheduled': [],
            'interviewed': []
        }
        for app in applicant_summary_details:
            status = app.get('application_status', '').lower()
            if status in applicants_by_status:
                applicants_by_status[status].append(app)
        
        # 2. Job Vacancy Report
        job_vacancy = {}
        job_vacancy_details = []
        try:
            job_where = base_where
            job_params = list(params_reports) if params_reports else []
            
            # Apply date filter for job postings
            if date_filter:
                job_date_filter = date_filter.replace('a.submitted_at', job_posted_expr)
                if job_where:
                    job_where += f" {job_date_filter}"
                else:
                    # Remove leading 'AND ' only, not all occurrences
                    job_date_clean = job_date_filter.strip()
                    if job_date_clean.startswith('AND '):
                        job_date_clean = job_date_clean[4:]  # Remove 'AND ' from start
                    job_where = f"WHERE {job_date_clean}"
                job_params.extend(date_params_all)
            
            try:
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT j.job_id) AS total_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'active' THEN j.job_id END) AS active_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'closed' THEN j.job_id END) AS closed_jobs,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        CASE 
                            WHEN COUNT(DISTINCT j.job_id) > 0 
                            THEN COUNT(DISTINCT a.application_id) / COUNT(DISTINCT j.job_id)
                            ELSE 0 
                        END AS avg_applications_per_job
                    FROM jobs j
                    LEFT JOIN applications a ON j.job_id = a.job_id
                    {job_where if job_where else ''}
                ''', tuple(job_params) if job_params else None)
                job_vacancy = cursor.fetchone() or {}
            except Exception as job_error:
                print(f'⚠️ Error in job vacancy query: {job_error}')
                job_vacancy = {}
            
            # Get detailed job list
            try:
                cursor.execute(f'''
                    SELECT 
                        {job_title_expr} AS job_title,
                        b.branch_name,
                        j.status,
                        COUNT(DISTINCT a.application_id) AS application_count,
                        DATE({job_posted_expr}) AS posted_date
                    FROM jobs j
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    LEFT JOIN applications a ON j.job_id = a.job_id
                    {job_where if job_where else ''}
                    GROUP BY j.job_id, {job_title_expr}, b.branch_name, j.status, posted_date
                    ORDER BY posted_date DESC
                    LIMIT 50
                ''', tuple(job_params) if job_params else None)
                job_vacancy_details = cursor.fetchall() or []
            except Exception as job_detail_error:
                print(f'⚠️ Error in job vacancy detail query: {job_detail_error}')
                job_vacancy_details = []
            
            # Format dates
            for job in job_vacancy_details:
                if job.get('posted_date'):
                    try:
                        if isinstance(job['posted_date'], (datetime, date_class)):
                            job['posted_date'] = job['posted_date'].strftime('%b %d, %Y')
                        elif isinstance(job['posted_date'], str):
                            dt = datetime.strptime(job['posted_date'], '%Y-%m-%d')
                            job['posted_date'] = dt.strftime('%b %d, %Y')
                    except:
                        pass
        except Exception as e:
            print(f'⚠️ Error fetching job vacancy: {e}')
            job_vacancy = {}
            job_vacancy_details = []
        
        # Enhanced Applicant Summary - Branch Breakdown
        applicants_by_branch = []
        try:
            # When "All Branches" is selected, start from branches table to show all branches
            # When a specific branch is selected, we can filter accordingly
            if branch_filter_id:
                # Specific branch selected - can use the original approach
                branch_applicant_where = base_where
                branch_applicant_params = list(params_reports) if params_reports else []
                
                if date_filter:
                    branch_applicant_date_filter = date_filter.replace('a.submitted_at', 'a.submitted_at')
                    if branch_applicant_where:
                        branch_applicant_where += f" {branch_applicant_date_filter}"
                    else:
                        branch_date_clean = branch_applicant_date_filter.strip()
                        if branch_date_clean.startswith('AND '):
                            branch_date_clean = branch_date_clean[4:]
                        branch_applicant_where = f"WHERE {branch_date_clean}"
                    branch_applicant_params.extend(date_params_all)
                
                cursor.execute(f'''
                    SELECT 
                        COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT CASE WHEN a.status = 'pending' THEN a.application_id END) AS pending_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'scheduled' THEN a.application_id END) AS scheduled_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'interviewed' THEN a.application_id END) AS interviewed_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    {branch_applicant_where if branch_applicant_where else ''}
                    GROUP BY b.branch_id, b.branch_name
                    ORDER BY total_applications DESC
                ''', tuple(branch_applicant_params) if branch_applicant_params else None)
                applicants_by_branch = cursor.fetchall() or []
            else:
                # All Branches selected - start from branches table to show ALL branches
                # But respect user's branch scope if they have one
                branch_params = []
                branch_scope_clause = ""
                
                # Apply user's branch scope if they have one (HR users with branch assignment)
                if branch_id:
                    branch_scope_clause = "WHERE b.branch_id = %s"
                    branch_params.append(branch_id)
                
                # Build date filter clause for JOIN condition
                date_filter_clause = ""
                if date_filter and date_params_all:
                    # Convert WHERE-style filter to JOIN condition
                    if "BETWEEN" in date_filter:
                        date_filter_clause = "AND DATE(a.submitted_at) BETWEEN %s AND %s"
                        branch_params.extend(date_params_all)
                    elif "YEAR(a.submitted_at)" in date_filter:
                        date_filter_clause = "AND YEAR(a.submitted_at) = %s"
                        branch_params.extend(date_params_all)
                    else:
                        date_filter_clause = date_filter.replace('AND ', 'AND ')
                        branch_params.extend(date_params_all)
                
                # Build job filter clause
                job_filter_clause = ""
                if job_filter_id:
                    job_filter_clause = "AND j.job_id = %s"
                    branch_params.append(job_filter_id)
                
                # Build the query starting from branches
                cursor.execute(f'''
                    SELECT 
                        COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                        b.branch_id,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT CASE WHEN a.status = 'pending' THEN a.application_id END) AS pending_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'scheduled' THEN a.application_id END) AS scheduled_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'interviewed' THEN a.application_id END) AS interviewed_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                    FROM branches b
                    {branch_scope_clause}
                    LEFT JOIN jobs j ON j.branch_id = b.branch_id {job_filter_clause}
                    LEFT JOIN applications a ON a.job_id = j.job_id {date_filter_clause}
                    LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    GROUP BY b.branch_id, b.branch_name
                    ORDER BY total_applications DESC, b.branch_name ASC
                ''', tuple(branch_params) if branch_params else None)
                applicants_by_branch = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching applicants by branch: {e}')
            import traceback
            print(traceback.format_exc())
            applicants_by_branch = []
        
        # Most/Least Applied Jobs
        most_applied_jobs = []
        least_applied_jobs = []
        try:
            job_applicant_where = base_where
            job_applicant_params = list(params_reports) if params_reports else []
            
            if date_filter:
                job_applicant_date_filter = date_filter.replace('a.submitted_at', 'a.submitted_at')
                if job_applicant_where:
                    job_applicant_where += f" {job_applicant_date_filter}"
                else:
                    job_date_clean = job_applicant_date_filter.strip()
                    if job_date_clean.startswith('AND '):
                        job_date_clean = job_date_clean[4:]
                    job_applicant_where = f"WHERE {job_date_clean}"
                job_applicant_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    {job_title_expr} AS job_title,
                    b.branch_name,
                    COUNT(DISTINCT a.application_id) AS application_count
                FROM jobs j
                LEFT JOIN applications a ON j.job_id = a.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                {job_applicant_where if job_applicant_where else ''}
                GROUP BY j.job_id, {job_title_expr}, b.branch_name
                HAVING application_count > 0
                ORDER BY application_count DESC
                LIMIT 10
            ''', tuple(job_applicant_params) if job_applicant_params else None)
            most_applied_jobs = cursor.fetchall() or []
            
            cursor.execute(f'''
                SELECT 
                    {job_title_expr} AS job_title,
                    b.branch_name,
                    COUNT(DISTINCT a.application_id) AS application_count
                FROM jobs j
                LEFT JOIN applications a ON j.job_id = a.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                {job_applicant_where if job_applicant_where else ''}
                GROUP BY j.job_id, {job_title_expr}, b.branch_name
                HAVING application_count > 0
                ORDER BY application_count ASC
                LIMIT 10
            ''', tuple(job_applicant_params) if job_applicant_params else None)
            least_applied_jobs = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching most/least applied jobs: {e}')
            most_applied_jobs = []
            least_applied_jobs = []
        
        # 3. Hiring Outcome Report
        hiring_outcome = {}
        try:
            hiring_where = base_where
            hiring_params = list(params_reports) if params_reports else []
            
            if date_filter:
                if hiring_where:
                    hiring_where += f" {date_filter}"
                else:
                    # Remove leading 'AND ' only, not all occurrences
                    date_filter_clean = date_filter.strip()
                    if date_filter_clean.startswith('AND '):
                        date_filter_clean = date_filter_clean[4:]  # Remove 'AND ' from start
                    hiring_where = f"WHERE {date_filter_clean}"
                hiring_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS total_hired,
                    CASE 
                        WHEN COUNT(*) > 0 
                        THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) / COUNT(*) * 100
                        ELSE 0 
                    END AS hire_rate,
                    AVG(CASE WHEN a.status = 'hired' AND a.submitted_at IS NOT NULL AND a.updated_at IS NOT NULL 
                        THEN DATEDIFF(a.updated_at, a.submitted_at) ELSE NULL END) AS avg_time_to_hire,
                    CASE 
                        WHEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) > 0 
                        THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) * 100
                        ELSE 0 
                    END AS interview_to_hire_rate,
                    CASE 
                        WHEN COUNT(*) > 0 
                        THEN COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) / COUNT(*) * 100
                        ELSE 0 
                    END AS rejection_rate
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                LEFT JOIN interviews i ON i.application_id = a.application_id
                {hiring_where if hiring_where else ''}
            ''', tuple(hiring_params) if hiring_params else None)
            hiring_outcome = cursor.fetchone() or {}
            
            # Round percentages
            if hiring_outcome.get('hire_rate'):
                hiring_outcome['hire_rate'] = round(hiring_outcome['hire_rate'], 1)
            if hiring_outcome.get('interview_to_hire_rate'):
                hiring_outcome['interview_to_hire_rate'] = round(hiring_outcome['interview_to_hire_rate'], 1)
            if hiring_outcome.get('rejection_rate'):
                hiring_outcome['rejection_rate'] = round(hiring_outcome['rejection_rate'], 1)
            if hiring_outcome.get('avg_time_to_hire'):
                hiring_outcome['avg_time_to_hire'] = round(hiring_outcome['avg_time_to_hire'] or 0, 1)
        except Exception as e:
            print(f'⚠️ Error fetching hiring outcome: {e}')
            hiring_outcome = {}
        
        # Interview Schedule Report (with no-shows and cancelled)
        interview_schedule = {}
        try:
            interview_where = base_where if base_where else ""
            interview_params = list(params_reports) if params_reports else []
            
            if date_filter:
                interview_date_filter = date_filter.replace('a.submitted_at', 'i.scheduled_date')
                if interview_where:
                    interview_where += f" {interview_date_filter}"
                else:
                    interview_date_clean = interview_date_filter.strip()
                    if interview_date_clean.startswith('AND '):
                        interview_date_clean = interview_date_clean[4:]
                    interview_where = f"WHERE {interview_date_clean}"
                interview_params.extend(date_params_all)
            
            # Check if interviews table has status column
            cursor.execute("SHOW COLUMNS FROM interviews LIKE 'status'")
            has_interview_status = cursor.fetchone() is not None
            
            if has_interview_status:
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT i.interview_id) AS interviews_scheduled,
                        COUNT(DISTINCT CASE WHEN i.status = 'completed' THEN i.interview_id END) AS interviews_completed,
                        COUNT(DISTINCT CASE WHEN i.status = 'cancelled' THEN i.interview_id END) AS interviews_cancelled,
                        COUNT(DISTINCT CASE WHEN i.status = 'no_show' THEN i.interview_id END) AS interviews_no_show,
                        COUNT(DISTINCT CASE WHEN i.status IN ('scheduled', 'confirmed', 'rescheduled') THEN i.interview_id END) AS interviews_upcoming
                    FROM interviews i
                    LEFT JOIN applications a ON i.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {interview_where if interview_where else ''}
                ''', tuple(interview_params) if interview_params else None)
            else:
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT i.interview_id) AS interviews_scheduled,
                        COUNT(DISTINCT CASE WHEN i.scheduled_date < NOW() THEN i.interview_id END) AS interviews_completed,
                        0 AS interviews_cancelled,
                        0 AS interviews_no_show,
                        COUNT(DISTINCT CASE WHEN i.scheduled_date >= NOW() THEN i.interview_id END) AS interviews_upcoming
                    FROM interviews i
                    LEFT JOIN applications a ON i.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {interview_where if interview_where else ''}
                ''', tuple(interview_params) if interview_params else None)
            
            interview_schedule = cursor.fetchone() or {}
            
            # Calculate completion rate
            total_scheduled = interview_schedule.get('interviews_scheduled', 0) or 0
            completed = interview_schedule.get('interviews_completed', 0) or 0
            if total_scheduled > 0:
                interview_schedule['completion_rate'] = round((completed / total_scheduled) * 100, 1)
            else:
                interview_schedule['completion_rate'] = 0
                
            # Calculate no-show rate
            no_shows = interview_schedule.get('interviews_no_show', 0) or 0
            if total_scheduled > 0:
                interview_schedule['no_show_rate'] = round((no_shows / total_scheduled) * 100, 1)
            else:
                interview_schedule['no_show_rate'] = 0
        except Exception as e:
            print(f'⚠️ Error fetching interview schedule: {e}')
            interview_schedule = {}
        
        # Enhanced Job Vacancy Activity (newly posted, updated, closed)
        job_activity = {}
        try:
            job_activity_where = base_where if base_where else ""
            job_activity_params = list(params_reports) if params_reports else []
            
            if date_filter:
                job_activity_date_filter = date_filter.replace('a.submitted_at', job_posted_expr)
                if job_activity_where:
                    job_activity_where += f" {job_activity_date_filter}"
                else:
                    job_date_clean = job_activity_date_filter.strip()
                    if job_date_clean.startswith('AND '):
                        job_date_clean = job_date_clean[4:]
                    job_activity_where = f"WHERE {job_date_clean}"
                job_activity_params.extend(date_params_all)
            
            # Check if updated_at column exists in jobs table
            cursor.execute("SHOW COLUMNS FROM jobs LIKE 'updated_at'")
            has_updated_at = cursor.fetchone() is not None
            
            # Use updated_at if exists, otherwise use created_at
            job_updated_expr = 'j.updated_at' if has_updated_at else 'j.created_at'
            
            if date_params_all and len(date_params_all) >= 2:
                # Use date range for newly posted, updated, and closed
                activity_params = list(job_activity_params) + list(date_params_all[:2]) * 3
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT CASE WHEN DATE({job_posted_expr}) BETWEEN %s AND %s THEN j.job_id END) AS newly_posted,
                        COUNT(DISTINCT CASE WHEN {job_updated_expr} IS NOT NULL AND DATE({job_updated_expr}) BETWEEN %s AND %s THEN j.job_id END) AS updated_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'closed' AND DATE({job_updated_expr}) BETWEEN %s AND %s THEN j.job_id END) AS closed_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'open' THEN j.job_id END) AS active_jobs
                    FROM jobs j
                    {job_activity_where if job_activity_where else ''}
                ''', tuple(activity_params))
            else:
                # No date filter, just count active jobs
                cursor.execute(f'''
                    SELECT 
                        0 AS newly_posted,
                        0 AS updated_jobs,
                        0 AS closed_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'open' THEN j.job_id END) AS active_jobs
                    FROM jobs j
                    {job_activity_where if job_activity_where else ''}
                ''', tuple(job_activity_params) if job_activity_params else None)
            job_activity = cursor.fetchone() or {}
        except Exception as e:
            print(f'⚠️ Error fetching job activity: {e}')
            job_activity = {}
        
        # 4. HR Performance Report (Enhanced with activity logs)
        hr_performance = {}
        hr_performance_details = []
        try:
            # Check if activity_logs table exists
            cursor.execute("SHOW TABLES LIKE 'activity_logs'")
            has_activity_logs = cursor.fetchone() is not None
            
            # Get interviews data
            interview_where = base_where if base_where else ""
            interview_params = list(params_reports) if params_reports else []
            
            if date_filter:
                interview_date_filter = date_filter.replace('a.submitted_at', 'i.scheduled_date')
                if interview_where:
                    interview_where += f" {interview_date_filter}"
                else:
                    interview_date_clean = interview_date_filter.strip()
                    if interview_date_clean.startswith('AND '):
                        interview_date_clean = interview_date_clean[4:]
                    interview_where = f"WHERE {interview_date_clean}"
                interview_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    COUNT(DISTINCT i.interview_id) AS interviews_scheduled,
                    COUNT(DISTINCT CASE WHEN i.status = 'completed' THEN i.interview_id END) AS interviews_completed
                FROM interviews i
                LEFT JOIN applications a ON i.application_id = a.application_id
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {interview_where if interview_where else ''}
            ''', tuple(interview_params) if interview_params else None)
            interview_data = cursor.fetchone() or {}
            
            # Get status updates and reviews from applications
            status_where = base_where if base_where else ""
            status_params = list(params_reports) if params_reports else []
            
            if date_filter:
                if status_where:
                    status_where += f" {date_filter}"
                else:
                    date_filter_clean = date_filter.strip()
                    if date_filter_clean.startswith('AND '):
                        date_filter_clean = date_filter_clean[4:]
                    status_where = f"WHERE {date_filter_clean}"
                status_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    COUNT(DISTINCT CASE WHEN a.updated_at IS NOT NULL AND a.updated_at != a.submitted_at THEN a.application_id END) AS status_updates,
                    COUNT(DISTINCT CASE WHEN a.submitted_at IS NOT NULL THEN a.application_id END) AS applications_reviewed,
                    AVG(CASE WHEN a.updated_at IS NOT NULL AND a.submitted_at IS NOT NULL 
                        THEN TIMESTAMPDIFF(HOUR, a.submitted_at, a.updated_at) ELSE NULL END) AS avg_response_time
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {status_where if status_where else ''}
            ''', tuple(status_params) if status_params else None)
            status_data = cursor.fetchone() or {}
            
            # Get HR performance from activity logs if available
            hr_activity_data = []
            if has_activity_logs:
                try:
                    activity_where = ""
                    activity_params = []
                    if date_filter:
                        activity_date_clean = date_filter.strip()
                        if activity_date_clean.startswith('AND '):
                            activity_date_clean = activity_date_clean[4:]
                        activity_where = f"WHERE {activity_date_clean.replace('a.submitted_at', 'al.created_at')}"
                        activity_params = date_params_all
                    
                    cursor.execute(f'''
                        SELECT 
                            al.admin_id,
                            COALESCE(ad.full_name, CONCAT('HR #', al.admin_id)) AS hr_name,
                            COUNT(*) AS total_actions,
                            COUNT(DISTINCT CASE WHEN al.target_table = 'applications' THEN al.target_id END) AS applications_processed,
                            COUNT(DISTINCT CASE WHEN al.action LIKE '%interview%' OR al.target_table = 'interviews' THEN al.target_id END) AS interviews_scheduled_count
                        FROM activity_logs al
                        LEFT JOIN admins ad ON al.admin_id = ad.admin_id
                        {activity_where}
                        GROUP BY al.admin_id, ad.full_name
                        ORDER BY total_actions DESC
                        LIMIT 10
                    ''', tuple(activity_params) if activity_params else None)
                    hr_activity_data = cursor.fetchall() or []
                except Exception as e:
                    print(f'⚠️ Error fetching HR activity: {e}')
                    hr_activity_data = []
            
            # Combine results
            hr_performance = {
                'interviews_scheduled': interview_data.get('interviews_scheduled', 0) or 0,
                'interviews_completed': interview_data.get('interviews_completed', 0) or 0,
                'status_updates': status_data.get('status_updates', 0) or 0,
                'applications_reviewed': status_data.get('applications_reviewed', 0) or 0,
                'avg_response_time': round(status_data.get('avg_response_time', 0) or 0, 1)
            }
            hr_performance_details = hr_activity_data
        except Exception as e:
            print(f'⚠️ Error fetching HR performance: {e}')
            hr_performance = {}
            hr_performance_details = []
        
        # Monthly/Yearly Trends (for monthly and yearly reports)
        monthly_trends = []
        yearly_trends = []
        branch_comparison_data = []
        fastest_hiring_branch = None
        monthly_trends = []
        
        # Fetch trend data based on period
        try:
            if period == 'week':
                # Weekly period: Get daily trends (Monday to Sunday)
                if date_filter and date_params_all:
                    trend_where = base_where if base_where else ''
                    if trend_where:
                        trend_where = trend_where + ' ' + date_filter
                    else:
                        # Remove 'AND ' prefix from date_filter if it exists
                        trend_where = date_filter.strip()
                        if trend_where.startswith('AND '):
                            trend_where = 'WHERE ' + trend_where[4:]
                        else:
                            trend_where = 'WHERE ' + trend_where
                    
                    trend_params = list(params_reports) if params_reports else []
                    trend_params.extend(date_params_all)
                    
                    # Get all applications grouped by submitted date
                    cursor.execute(f'''
                        SELECT 
                            DAYNAME(a.submitted_at) AS day_name,
                            DAYOFWEEK(a.submitted_at) AS day_of_week,
                            DATE(a.submitted_at) AS date_key,
                            DATE_FORMAT(a.submitted_at, '%a') AS day_key,
                            COUNT(DISTINCT a.application_id) AS total_applications,
                            COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                            0 AS hired_count,
                            COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {trend_where}
                        GROUP BY DAYNAME(a.submitted_at), DAYOFWEEK(a.submitted_at), DATE(a.submitted_at), DATE_FORMAT(a.submitted_at, '%a')
                    ''', tuple(trend_params))
                    daily_apps = cursor.fetchall() or []
                    
                    # Get hired applicants grouped by hire date (updated_at when status = 'hired')
                    # Build WHERE clause for hired applicants: filter by hire date (updated_at) within selected period
                    hired_where_parts = []
                    hired_params = []
                    
                    # Apply branch filter
                    if branch_filter_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_filter_id)
                    elif branch_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_id)
                    
                    # Apply job filter
                    if job_filter_id:
                        hired_where_parts.append('j.job_id = %s')
                        hired_params.append(job_filter_id)
                    
                    # Apply date filter based on hire date (updated_at) instead of submitted_at
                    if date_filter and date_params_all:
                        # Replace submitted_at with updated_at in date filter
                        hired_date_filter = date_filter.replace('a.submitted_at', 'a.updated_at').replace('submitted_at', 'updated_at')
                        hired_date_filter = hired_date_filter.strip()
                        if hired_date_filter.startswith('AND '):
                            hired_date_filter = hired_date_filter[4:]
                        hired_where_parts.append(hired_date_filter)
                        hired_params.extend(date_params_all)
                    
                    # Add status filter
                    hired_where_parts.append("a.status = 'hired'")
                    
                    # Build final WHERE clause
                    if hired_where_parts:
                        hired_trend_where = 'WHERE ' + ' AND '.join(hired_where_parts)
                    else:
                        hired_trend_where = "WHERE a.status = 'hired'"
                    
                    cursor.execute(f'''
                        SELECT 
                            DAYNAME(a.updated_at) AS day_name,
                            DAYOFWEEK(a.updated_at) AS day_of_week,
                            DATE(a.updated_at) AS date_key,
                            DATE_FORMAT(a.updated_at, '%a') AS day_key,
                            COUNT(DISTINCT a.application_id) AS hired_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {hired_trend_where}
                        GROUP BY DAYNAME(a.updated_at), DAYOFWEEK(a.updated_at), DATE(a.updated_at), DATE_FORMAT(a.updated_at, '%a')
                        ORDER BY DATE(a.updated_at)
                    ''', tuple(hired_params) if hired_params else None)
                    daily_hired = cursor.fetchall() or []
                    
                    # Debug: Print hired data if available
                    if daily_hired:
                        print(f'✅ Found {len(daily_hired)} hired date entries for weekly trends')
                        total_hired_count = sum(int(row.get('hired_count') or 0) for row in daily_hired)
                        print(f'   Total hired applicants: {total_hired_count}')
                        for h in daily_hired[:3]:
                            print(f'   - {h.get("date_key")}: {h.get("hired_count")} hired')
                    else:
                        print(f'⚠️ No hired applicants found for weekly period')
                        # Check if there are any hired applicants at all (for debugging)
                        cursor.execute('''
                            SELECT COUNT(*) as count 
                            FROM applications 
                            WHERE status = 'hired'
                        ''')
                        total_hired = cursor.fetchone()
                        if total_hired and total_hired.get('count', 0) > 0:
                            print(f'   But found {total_hired.get("count")} total hired applicants in database')
                    
                    # Merge the results - use date_key for matching
                    hired_dict = {}
                    for row in daily_hired:
                        date_key = row.get('date_key')
                        if date_key:
                            # Convert to string for consistent comparison
                            date_key_str = str(date_key)
                            hired_count = int(row.get('hired_count') or 0)
                            if hired_count > 0:  # Only add if there are actually hired applicants
                                hired_dict[date_key_str] = hired_dict.get(date_key_str, 0) + hired_count
                    
                    daily_trends = []
                    # Create a set to track which dates we've added
                    added_dates = set()
                    
                    # First, add all application dates with their hired_count
                    for app_row in daily_apps:
                        date_key = app_row.get('date_key')
                        if date_key:
                            date_key_str = str(date_key)
                            # Get hired count from hired_dict
                            hired_count_from_dict = hired_dict.get(date_key_str, 0)
                            app_row['hired_count'] = hired_count_from_dict
                            daily_trends.append(app_row)
                            added_dates.add(date_key_str)
                    
                    # Then, add any hired dates that don't have applications
                    for hired_row in daily_hired:
                        date_key = hired_row.get('date_key')
                        hired_count = int(hired_row.get('hired_count') or 0)
                        if date_key and hired_count > 0:
                            date_key_str = str(date_key)
                            if date_key_str not in added_dates:
                                daily_trends.append({
                                    'day_name': hired_row.get('day_name', ''),
                                    'day_of_week': hired_row.get('day_of_week', 0),
                                    'date_key': date_key,
                                    'day_key': hired_row.get('day_key', ''),
                                    'total_applications': 0,
                                    'total_applicants': 0,
                                    'hired_count': hired_count,
                                    'rejected_count': 0
                                })
                                added_dates.add(date_key_str)
                    
                    # Sort by date_key (handle both string and date objects)
                    daily_trends.sort(key=lambda x: str(x.get('date_key', '')) if x.get('date_key') else '')
                    # Convert to monthly_trends format for consistency
                    monthly_trends = daily_trends
                    
                    # Debug output
                    if monthly_trends:
                        total_hired = sum(int(row.get('hired_count', 0) or 0) for row in monthly_trends)
                        print(f'✅ Weekly trends: {len(monthly_trends)} days, {total_hired} total hired')
                else:
                    # No date filter - still try to get hired applicants for default period
                    # Query hired applicants for current week as fallback
                    today = datetime.now()
                    week_start = today - timedelta(days=today.weekday())
                    week_end = week_start + timedelta(days=6)
                    
                    hired_where_parts = []
                    hired_params = []
                    
                    # Apply branch filter
                    if branch_filter_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_filter_id)
                    elif branch_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_id)
                    
                    # Apply job filter
                    if job_filter_id:
                        hired_where_parts.append('j.job_id = %s')
                        hired_params.append(job_filter_id)
                    
                    # Add date filter for current week
                    hired_where_parts.append("DATE(a.updated_at) BETWEEN %s AND %s")
                    hired_params.extend([week_start.date(), week_end.date()])
                    hired_where_parts.append("a.status = 'hired'")
                    
                    hired_trend_where = 'WHERE ' + ' AND '.join(hired_where_parts)
                    
                    cursor.execute(f'''
                        SELECT 
                            DAYNAME(a.updated_at) AS day_name,
                            DAYOFWEEK(a.updated_at) AS day_of_week,
                            DATE(a.updated_at) AS date_key,
                            DATE_FORMAT(a.updated_at, '%a') AS day_key,
                            COUNT(DISTINCT a.application_id) AS hired_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {hired_trend_where}
                        GROUP BY DAYNAME(a.updated_at), DAYOFWEEK(a.updated_at), DATE(a.updated_at), DATE_FORMAT(a.updated_at, '%a')
                        ORDER BY DATE(a.updated_at)
                    ''', tuple(hired_params) if hired_params else None)
                    daily_hired = cursor.fetchall() or []
                    
                    # Build trends from hired data
                    daily_trends = []
                    for hired_row in daily_hired:
                        daily_trends.append({
                            'day_name': hired_row.get('day_name', ''),
                            'day_of_week': hired_row.get('day_of_week', 0),
                            'date_key': hired_row.get('date_key'),
                            'day_key': hired_row.get('day_key', ''),
                            'total_applications': 0,
                            'total_applicants': 0,
                            'hired_count': hired_row.get('hired_count', 0),
                            'rejected_count': 0
                        })
                    
                    daily_trends.sort(key=lambda x: str(x.get('date_key', '')) if x.get('date_key') else '')
                    monthly_trends = daily_trends
            
            elif period == 'month':
                # Monthly period: Get weekly trends (Week 1, 2, 3, 4)
                if date_filter and date_params_all:
                    trend_where = base_where if base_where else ''
                    if trend_where:
                        trend_where = trend_where + ' ' + date_filter
                    else:
                        # Remove 'AND ' prefix from date_filter if it exists
                        trend_where = date_filter.strip()
                        if trend_where.startswith('AND '):
                            trend_where = 'WHERE ' + trend_where[4:]
                        else:
                            trend_where = 'WHERE ' + trend_where
                    
                    trend_params = list(params_reports) if params_reports else []
                    trend_params.extend(date_params_all)
                    
                    # Get all applications grouped by submitted date (weekly)
                    cursor.execute(f'''
                        SELECT 
                            WEEK(a.submitted_at, 1) - WEEK(DATE_SUB(a.submitted_at, INTERVAL DAY(a.submitted_at)-1 DAY), 1) + 1 AS week_number,
                            CONCAT('Week ', WEEK(a.submitted_at, 1) - WEEK(DATE_SUB(a.submitted_at, INTERVAL DAY(a.submitted_at)-1 DAY), 1) + 1) AS week_key,
                            DATE_FORMAT(a.submitted_at, '%Y-%m') AS month_key,
                            COUNT(DISTINCT a.application_id) AS total_applications,
                            COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                            0 AS hired_count,
                            COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {trend_where}
                        GROUP BY WEEK(a.submitted_at, 1) - WEEK(DATE_SUB(a.submitted_at, INTERVAL DAY(a.submitted_at)-1 DAY), 1) + 1, 
                                 DATE_FORMAT(a.submitted_at, '%Y-%m')
                    ''', tuple(trend_params))
                    weekly_apps = cursor.fetchall() or []
                    
                    # Get hired applicants grouped by hire date (updated_at when status = 'hired')
                    hired_where_parts = []
                    hired_params = []
                    
                    # Apply branch filter
                    if branch_filter_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_filter_id)
                    elif branch_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_id)
                    
                    # Apply job filter
                    if job_filter_id:
                        hired_where_parts.append('j.job_id = %s')
                        hired_params.append(job_filter_id)
                    
                    # Apply date filter based on hire date (updated_at) instead of submitted_at
                    if date_filter and date_params_all:
                        hired_date_filter = date_filter.replace('a.submitted_at', 'a.updated_at').replace('submitted_at', 'updated_at')
                        hired_date_filter = hired_date_filter.strip()
                        if hired_date_filter.startswith('AND '):
                            hired_date_filter = hired_date_filter[4:]
                        hired_where_parts.append(hired_date_filter)
                        hired_params.extend(date_params_all)
                    
                    # Add status filter
                    hired_where_parts.append("a.status = 'hired'")
                    
                    # Build final WHERE clause
                    if hired_where_parts:
                        hired_trend_where = 'WHERE ' + ' AND '.join(hired_where_parts)
                    else:
                        hired_trend_where = "WHERE a.status = 'hired'"
                    
                    cursor.execute(f'''
                        SELECT 
                            WEEK(a.updated_at, 1) - WEEK(DATE_SUB(a.updated_at, INTERVAL DAY(a.updated_at)-1 DAY), 1) + 1 AS week_number,
                            CONCAT('Week ', WEEK(a.updated_at, 1) - WEEK(DATE_SUB(a.updated_at, INTERVAL DAY(a.updated_at)-1 DAY), 1) + 1) AS week_key,
                            DATE_FORMAT(a.updated_at, '%Y-%m') AS month_key,
                            COUNT(DISTINCT a.application_id) AS hired_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {hired_trend_where}
                        GROUP BY WEEK(a.updated_at, 1) - WEEK(DATE_SUB(a.updated_at, INTERVAL DAY(a.updated_at)-1 DAY), 1) + 1, 
                                 DATE_FORMAT(a.updated_at, '%Y-%m')
                    ''', tuple(hired_params) if hired_params else None)
                    weekly_hired = cursor.fetchall() or []
                    
                    # Merge the results
                    hired_dict = {(row['week_number'], row['month_key']): row['hired_count'] for row in weekly_hired}
                    weekly_trends = []
                    for app_row in weekly_apps:
                        key = (app_row['week_number'], app_row['month_key'])
                        app_row['hired_count'] = hired_dict.get(key, 0)
                        weekly_trends.append(app_row)
                    
                    # Add any hired weeks that don't have applications
                    for hired_row in weekly_hired:
                        key = (hired_row['week_number'], hired_row['month_key'])
                        if not any((row['week_number'], row['month_key']) == key for row in weekly_trends):
                            weekly_trends.append({
                                'week_number': hired_row['week_number'],
                                'week_key': hired_row['week_key'],
                                'month_key': hired_row['month_key'],
                                'total_applications': 0,
                                'total_applicants': 0,
                                'hired_count': hired_row['hired_count'],
                                'rejected_count': 0
                            })
                    
                    # Sort by week_number
                    weekly_trends.sort(key=lambda x: x['week_number'])
                    # Convert to monthly_trends format for consistency
                    monthly_trends = weekly_trends
                else:
                    monthly_trends = []
            
            elif period == 'year':
                # Yearly period: Get monthly trends
                # Get all applications grouped by submitted date (monthly)
                cursor.execute(f'''
                    SELECT 
                        YEAR(a.submitted_at) AS year,
                        MONTH(a.submitted_at) AS month,
                        DATE_FORMAT(a.submitted_at, '%Y-%m') AS month_key,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        0 AS hired_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {base_where if base_where else ''}
                    {date_filter if date_filter else ''}
                    GROUP BY YEAR(a.submitted_at), MONTH(a.submitted_at), DATE_FORMAT(a.submitted_at, '%Y-%m')
                ''', tuple(list(params_reports) + date_params_all) if params_reports and date_params_all else (tuple(params_reports) if params_reports else (tuple(date_params_all) if date_params_all else None)))
                monthly_apps = cursor.fetchall() or []
                
                # Get hired applicants grouped by hire date (updated_at when status = 'hired')
                hired_where_parts = []
                hired_params = []
                
                # Apply branch filter
                if branch_filter_id:
                    hired_where_parts.append('j.branch_id = %s')
                    hired_params.append(branch_filter_id)
                elif branch_id:
                    hired_where_parts.append('j.branch_id = %s')
                    hired_params.append(branch_id)
                
                # Apply job filter
                if job_filter_id:
                    hired_where_parts.append('j.job_id = %s')
                    hired_params.append(job_filter_id)
                
                # Apply date filter based on hire date (updated_at) instead of submitted_at
                if date_filter and date_params_all:
                    hired_date_filter = date_filter.replace('a.submitted_at', 'a.updated_at').replace('submitted_at', 'updated_at')
                    hired_date_filter = hired_date_filter.strip()
                    if hired_date_filter.startswith('AND '):
                        hired_date_filter = hired_date_filter[4:]
                    hired_where_parts.append(hired_date_filter)
                    hired_params.extend(date_params_all)
                
                # Add status filter
                hired_where_parts.append("a.status = 'hired'")
                
                # Build final WHERE clause
                if hired_where_parts:
                    hired_where_clause = 'WHERE ' + ' AND '.join(hired_where_parts)
                else:
                    hired_where_clause = "WHERE a.status = 'hired'"
                
                cursor.execute(f'''
                    SELECT 
                        YEAR(a.updated_at) AS year,
                        MONTH(a.updated_at) AS month,
                        DATE_FORMAT(a.updated_at, '%Y-%m') AS month_key,
                        COUNT(DISTINCT a.application_id) AS hired_count
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {hired_where_clause}
                    GROUP BY YEAR(a.updated_at), MONTH(a.updated_at), DATE_FORMAT(a.updated_at, '%Y-%m')
                ''', tuple(hired_params) if hired_params else None)
                monthly_hired = cursor.fetchall() or []
                
                # Merge the results
                hired_dict = {row['month_key']: row['hired_count'] for row in monthly_hired}
                monthly_trends = []
                for app_row in monthly_apps:
                    month_key = app_row['month_key']
                    app_row['hired_count'] = hired_dict.get(month_key, 0)
                    monthly_trends.append(app_row)
                
                # Add any hired months that don't have applications
                for hired_row in monthly_hired:
                    month_key = hired_row['month_key']
                    if not any(row['month_key'] == month_key for row in monthly_trends):
                        monthly_trends.append({
                            'year': hired_row['year'],
                            'month': hired_row['month'],
                            'month_key': month_key,
                            'total_applications': 0,
                            'total_applicants': 0,
                            'hired_count': hired_row['hired_count'],
                            'rejected_count': 0
                        })
                
                # Sort by year, month
                monthly_trends.sort(key=lambda x: (x['year'], x['month']))
            
            if period == 'month' or period == 'year':
                
                # Branch comparison with hiring metrics
                cursor.execute(f'''
                    SELECT 
                        COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired_count,
                        AVG(CASE WHEN a.status = 'hired' AND a.submitted_at IS NOT NULL AND a.updated_at IS NOT NULL 
                            THEN DATEDIFF(a.updated_at, a.submitted_at) ELSE NULL END) AS avg_time_to_hire,
                        CASE 
                            WHEN COUNT(DISTINCT a.application_id) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) / COUNT(DISTINCT a.application_id) * 100
                            ELSE 0 
                        END AS hire_rate
                    FROM applications a
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    {base_where if base_where else ''}
                    GROUP BY b.branch_id, b.branch_name
                    HAVING total_applications > 0
                    ORDER BY avg_time_to_hire ASC, hire_rate DESC
                ''', tuple(params_reports) if params_reports else None)
                branch_comparison_data = cursor.fetchall() or []
                
                # Find fastest hiring branch
                if branch_comparison_data:
                    fastest_hiring_branch = branch_comparison_data[0] if branch_comparison_data else None
                    if fastest_hiring_branch:
                        fastest_hiring_branch['avg_time_to_hire'] = round(fastest_hiring_branch.get('avg_time_to_hire') or 0, 1)
                        fastest_hiring_branch['hire_rate'] = round(fastest_hiring_branch.get('hire_rate') or 0, 1)
        except Exception as e:
            print(f'⚠️ Error fetching trends: {e}')
            import traceback
            print(traceback.format_exc())
            monthly_trends = []
            branch_comparison_data = []
        
        # Enhanced Applicant Summary - New applicants per job
        new_applicants_per_job = {}
        try:
            if date_filter:
                new_app_date_filter = date_filter.replace('a.submitted_at', 'a.submitted_at')
                new_app_where = base_where if base_where else ""
                new_app_params = list(params_reports) if params_reports else []
                
                if new_app_where:
                    new_app_where += f" {new_app_date_filter}"
                else:
                    new_date_clean = new_app_date_filter.strip()
                    if new_date_clean.startswith('AND '):
                        new_date_clean = new_date_clean[4:]
                    new_app_where = f"WHERE {new_date_clean}"
                new_app_params.extend(date_params_all)
                
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS new_applicants,
                        COUNT(DISTINCT j.job_id) AS active_jobs,
                        CASE 
                            WHEN COUNT(DISTINCT j.job_id) > 0 
                            THEN COUNT(DISTINCT ap.applicant_id) / COUNT(DISTINCT j.job_id)
                            ELSE 0 
                        END AS new_applicants_per_job
                    FROM applicants ap
                    JOIN applications a ON ap.applicant_id = a.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {new_app_where if new_app_where else ''}
                ''', tuple(new_app_params) if new_app_params else None)
                new_applicants_per_job = cursor.fetchone() or {}
            else:
                new_applicants_per_job = {'new_applicants': 0, 'active_jobs': 0, 'new_applicants_per_job': 0}
        except Exception as e:
            print(f'⚠️ Error fetching new applicants per job: {e}')
            new_applicants_per_job = {}
        
        # 1. Time-to-Hire (Average Days) - Enhanced KPI
        avg_time_to_hire = 0
        try:
            tth_where = base_where if base_where else ""
            tth_params = list(params_reports) if params_reports else []
            
            if date_filter:
                tth_date_filter = date_filter.replace('a.submitted_at', 'a.updated_at')
                if tth_where:
                    tth_where += f" {tth_date_filter}"
                else:
                    tth_date_clean = tth_date_filter.strip()
                    if tth_date_clean.startswith('AND '):
                        tth_date_clean = tth_date_clean[4:]
                    tth_where = f"WHERE {tth_date_clean}"
                tth_params.extend(date_params_all)
            
            if tth_where:
                tth_where += " AND a.status = 'hired'"
            else:
                tth_where = "WHERE a.status = 'hired'"
            
            cursor.execute(f'''
                SELECT 
                    AVG(DATEDIFF(a.updated_at, a.submitted_at)) AS avg_time_to_hire
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {tth_where}
            ''', tuple(tth_params) if tth_params else None)
            tth_result = cursor.fetchone()
            avg_time_to_hire = round(tth_result.get('avg_time_to_hire', 0) or 0, 1) if tth_result else 0
        except Exception as e:
            print(f'⚠️ Error fetching average time-to-hire: {e}')
            avg_time_to_hire = 0
        
        # 2. Most Applied Job - Enhanced KPI
        most_applied_job_info = {}
        try:
            maj_where = base_where if base_where else ""
            maj_params = list(params_reports) if params_reports else []
            
            if date_filter:
                if maj_where:
                    maj_where += f" {date_filter}"
                else:
                    maj_date_clean = date_filter.strip()
                    if maj_date_clean.startswith('AND '):
                        maj_date_clean = maj_date_clean[4:]
                    maj_where = f"WHERE {maj_date_clean}"
                maj_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    j.job_id,
                    {job_title_expr} AS job_title,
                    COUNT(DISTINCT a.application_id) AS total_applications,
                    COUNT(DISTINCT ap.applicant_id) AS total_applicants
                FROM applications a
                JOIN applicants ap ON a.applicant_id = ap.applicant_id
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {maj_where if maj_where else ''}
                GROUP BY j.job_id, {job_title_expr}
                ORDER BY total_applications DESC
                LIMIT 1
            ''', tuple(maj_params) if maj_params else None)
            most_applied_job_info = cursor.fetchone() or {}
        except Exception as e:
            print(f'⚠️ Error fetching most applied job: {e}')
            most_applied_job_info = {}
        
        # 3. Application Status Distribution (for Pie/Donut Chart)
        status_distribution = []
        try:
            status_where_parts = []
            status_params = list(params_reports) if params_reports else []
            
            # Apply branch filter
            if branch_filter_id:
                status_where_parts.append('j.branch_id = %s')
                status_params.append(branch_filter_id)
            elif branch_id:
                status_where_parts.append('j.branch_id = %s')
                status_params.append(branch_id)
            
            # Apply job filter
            if job_filter_id:
                status_where_parts.append('j.job_id = %s')
                status_params.append(job_filter_id)
            
            # Apply date filter
            if date_filter and date_params_all:
                date_clean = date_filter.strip()
                if date_clean.startswith('AND '):
                    date_clean = date_clean[4:]
                status_where_parts.append(date_clean)
                status_params.extend(date_params_all)
            
            # Build WHERE clause
            if status_where_parts:
                status_where = 'WHERE ' + ' AND '.join(status_where_parts)
            else:
                status_where = ''
            
            cursor.execute(f'''
                SELECT 
                    a.status,
                    COUNT(DISTINCT a.application_id) AS count
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {status_where}
                GROUP BY a.status
                ORDER BY count DESC
            ''', tuple(status_params) if status_params else None)
            status_distribution = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching status distribution: {e}')
            status_distribution = []
        
        # 4. Interview Schedule Report Details (Full List)
        interview_schedule_details = []
        try:
            interview_details_where = base_where if base_where else ""
            interview_details_params = list(params_reports) if params_reports else []
            
            if date_filter:
                interview_details_date_filter = date_filter.replace('a.submitted_at', 'i.scheduled_date')
                if interview_details_where:
                    interview_details_where += f" {interview_details_date_filter}"
                else:
                    interview_details_date_clean = interview_details_date_filter.strip()
                    if interview_details_date_clean.startswith('AND '):
                        interview_details_date_clean = interview_details_date_clean[4:]
                    interview_details_where = f"WHERE {interview_details_date_clean}"
                interview_details_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    ap.full_name AS applicant_name,
                    {job_title_expr} AS job_title,
                    COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                    i.scheduled_date AS interview_date,
                    COALESCE(i.interview_mode, 'In-person') AS interview_mode,
                    COALESCE(i.location, 'TBD') AS interview_location,
                    COALESCE(i.status, 'scheduled') AS interview_status,
                    COALESCE(ad.full_name, 'Unassigned') AS hr_assigned
                FROM interviews i
                LEFT JOIN applications a ON i.application_id = a.application_id
                LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
                LEFT JOIN jobs j ON a.job_id = j.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                LEFT JOIN activity_logs al ON al.target_table = 'interviews' AND al.target_id = i.interview_id
                LEFT JOIN admins ad ON al.admin_id = ad.admin_id
                {interview_details_where if interview_details_where else ''}
                ORDER BY i.scheduled_date DESC
                LIMIT 500
            ''', tuple(interview_details_params) if interview_details_params else None)
            interview_schedule_details = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching interview schedule details: {e}')
            interview_schedule_details = []
        
        # 5. HR Activity Log Report Details
        hr_activity_log_details = []
        try:
            cursor.execute("SHOW TABLES LIKE 'activity_logs'")
            has_activity_logs = cursor.fetchone() is not None
            
            if has_activity_logs:
                activity_where = "WHERE u.user_type = 'hr'"
                activity_params = []
                
                if date_filter:
                    activity_date_filter = date_filter.replace('a.submitted_at', 'al.created_at')
                    activity_where += f" {activity_date_filter}"
                    activity_params.extend(date_params_all)
                
                cursor.execute(f'''
                    SELECT 
                        COALESCE(ad.full_name, CONCAT('HR #', al.admin_id)) AS hr_name,
                        al.action,
                        al.target_table,
                        al.details,
                        al.created_at,
                        CASE 
                            WHEN al.target_table = 'applications' THEN (SELECT ap.full_name FROM applicants ap JOIN applications a ON ap.applicant_id = a.applicant_id WHERE a.application_id = al.target_id)
                            WHEN al.target_table = 'jobs' THEN (SELECT {job_title_expr} FROM jobs j WHERE j.job_id = al.target_id)
                            WHEN al.target_table = 'interviews' THEN (SELECT ap.full_name FROM applicants ap JOIN applications a ON ap.applicant_id = a.applicant_id JOIN interviews i ON i.application_id = a.application_id WHERE i.interview_id = al.target_id)
                            ELSE 'N/A'
                        END AS affected_item
                    FROM activity_logs al
                    LEFT JOIN admins ad ON al.admin_id = ad.admin_id
                    LEFT JOIN users u ON ad.user_id = u.user_id
                    {activity_where}
                    ORDER BY al.created_at DESC
                    LIMIT 500
                ''', tuple(activity_params) if activity_params else None)
                hr_activity_log_details = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching HR activity log details: {e}')
            import traceback
            traceback.print_exc()
            hr_activity_log_details = []
        
        # 6. Get HR Users list for filter
        hr_users_list = []
        try:
            cursor.execute('''
                SELECT DISTINCT ad.admin_id, ad.full_name
                FROM admins ad
                JOIN users u ON ad.user_id = u.user_id
                WHERE u.user_type = 'hr' AND u.is_active = 1 AND ad.is_active = 1
                ORDER BY ad.full_name
            ''')
            hr_users_list = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching HR users list: {e}')
            hr_users_list = []
        
        # Handle export requests - must be after all data is generated
        if export_format:
            try:
                return handle_report_export(cursor, export_format, export_section, export_type, period_summary, 
                                          applicant_summary, applicant_summary_details, job_vacancy, job_vacancy_details,
                                          hiring_outcome, hr_performance, job_title_expr, date_filter, date_params_all)
            except Exception as export_error:
                print(f'⚠️ Error in export: {export_error}')
                import traceback
                print(traceback.format_exc())
                flash(f'Error generating export: {str(export_error)}', 'error')
                # Redirect back to reports page without export parameter
                redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
                params = {'period': period}
                if selected_week_month:
                    params['week_month'] = selected_week_month
                if selected_week_number:
                    params['week_number'] = selected_week_number
                if selected_week:  # Legacy support
                    params['week'] = selected_week
                if selected_month:
                    params['month'] = selected_month
                if selected_year:
                    params['year'] = selected_year
                # Include branch and job filters in redirect
                if branch_filter_id:
                    params['branch_id'] = branch_filter_id
                if job_filter_id:
                    params['job_id'] = job_filter_id
                return redirect(url_for('admin_reports_analytics' if user.get('role') == 'admin' else 'hr_reports_analytics', **params))
        
        # Render HR template if user is HR, otherwise admin template
        template = 'hr/reports_analytics.html' if user.get('role') == 'hr' else 'admin/reports.html'
        branch_info = None
        if user.get('role') == 'hr':
            branch_id_session = session.get('branch_id')
            if branch_id_session:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id_session,))
                if branch_rows:
                    branch_info = branch_rows[0]
        # Group applicants by status for display
        applicants_by_status = {
            'hired': [],
            'rejected': [],
            'pending': [],
            'scheduled': [],
            'interviewed': []
        }
        for app in applicant_summary_details:
            status = app.get('application_status', '').lower()
            if status in applicants_by_status:
                applicants_by_status[status].append(app)
        
        # Calculate status counts
        status_counts = {
            'hired': len(applicants_by_status['hired']),
            'rejected': len(applicants_by_status['rejected']),
            'pending': len(applicants_by_status['pending']),
            'scheduled': len(applicants_by_status['scheduled']),
            'interviewed': len(applicants_by_status['interviewed'])
        }
        
        # Ensure all data is passed correctly
        return render_template(
            template,
            stats={**stats, **(metrics or {})},
            metrics=metrics or {},
            trends=trends or [],
            branch_stats=branch_stats or [],
            branch_comparison=branch_comparison or [],
            system_stats=system_stats or {},
            job_performance=job_performance or [],
            status_breakdown=status_breakdown or [],
            funnel_data=funnel_data or {},
            source_effectiveness=source_effectiveness or [],
            branch_info=branch_info,
            period=period,
            period_summary=period_summary,
            available_years=available_years,
            selected_week=selected_week,
            selected_week_month=selected_week_month,
            selected_week_number=selected_week_number,
            selected_month=selected_month,
            selected_year=selected_year,
            applicant_summary=applicant_summary,
            job_vacancy=job_vacancy,
            hiring_outcome=hiring_outcome,
            hr_performance=hr_performance,
            applicant_summary_details=applicant_summary_details,
            job_vacancy_details=job_vacancy_details,
            applicants_by_status=applicants_by_status,
            status_counts=status_counts,
            generate_section=generate_section,
            applicants_by_branch=applicants_by_branch,
            most_applied_jobs=most_applied_jobs,
            least_applied_jobs=least_applied_jobs,
            interview_schedule=interview_schedule,
            job_activity=job_activity,
            hr_performance_details=hr_performance_details,
            monthly_trends=monthly_trends,
            branch_comparison_data=branch_comparison_data,
            fastest_hiring_branch=fastest_hiring_branch,
            new_applicants_per_job=new_applicants_per_job,
            branches=branches,
            jobs=jobs,
            # New comprehensive report data
            avg_time_to_hire=avg_time_to_hire,
            most_applied_job_info=most_applied_job_info,
            status_distribution=status_distribution,
            interview_schedule_details=interview_schedule_details,
            hr_activity_log_details=hr_activity_log_details,
            hr_users_list=hr_users_list,
            hr_filter_condition=hr_filter_condition,
            status_filter_condition=status_filter_condition,
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Reports analytics error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'An error occurred while loading reports: {str(exc)}', 'error')
        template = 'hr/reports_analytics.html' if user.get('role') == 'hr' else 'admin/reports.html'
        # Get period from request to maintain state
        period = request.args.get('period', 'all')
        selected_week = request.args.get('week', '')
        selected_week_month = request.args.get('week_month', '')
        selected_week_number = request.args.get('week_number', '')
        selected_month = request.args.get('month', '')
        selected_year = request.args.get('year', '')
        # Get available years for dropdown
        try:
            cursor.execute("SELECT DISTINCT YEAR(submitted_at) as year FROM applications WHERE submitted_at IS NOT NULL ORDER BY year DESC")
            available_years = [str(row['year']) for row in cursor.fetchall() if row.get('year')]
        except:
            available_years = []
        return render_template(
            template,
            stats={},
            metrics={},
            trends=[],
            branch_stats=[],
            branch_comparison=[],
            system_stats={},
            job_performance=[],
            status_breakdown=[],
            funnel_data={},
            source_effectiveness=[],
            branch_info=None,
            period=period,
            period_summary={},
            available_years=available_years,
            selected_week=selected_week,
            selected_week_month=selected_week_month,
            selected_week_number=selected_week_number,
            selected_month=selected_month,
            selected_year=selected_year,
            applicant_summary={},
            job_vacancy={},
            hiring_outcome={},
            hr_performance={},
            applicant_summary_details=[],
            job_vacancy_details=[],
            applicants_by_status={'hired': [], 'rejected': [], 'pending': [], 'scheduled': [], 'interviewed': []},
            status_counts={'hired': 0, 'rejected': 0, 'pending': 0, 'scheduled': 0, 'interviewed': 0},
            generate_section='',
            applicants_by_branch=[],
            interview_schedule={'interviews_scheduled': 0, 'interviews_completed': 0},
            monthly_trends=[],
            branches=[],
            jobs=[],
        )
    finally:
        if cursor:
            cursor.close()


@app.route('/admin/applicants/<int:applicant_id>')
@login_required('admin', 'hr')
def view_applicant(applicant_id):
    """View detailed applicant profile."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicants'))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Ensure schema compatibility before querying
        ensure_schema_compatibility()
        # Admin can view all applicants, HR can only view applicants from their branch
        branch_id = get_branch_scope(user)
        if branch_id:  # Only restrict HR users, not admin
            cursor.execute(
                '''
                SELECT DISTINCT ap.applicant_id
                FROM applicants ap
                JOIN applications a ON ap.applicant_id = a.applicant_id
                JOIN jobs j ON a.job_id = j.job_id
                WHERE ap.applicant_id = %s AND j.branch_id = %s
                LIMIT 1
                ''',
                (applicant_id, branch_id),
            )
            if not cursor.fetchone():
                flash('You can only view applicants from your branch.', 'error')
                return redirect(url_for('applicants'))
        
        # Get applicant details
        cursor.execute(
            '''
            SELECT ap.applicant_id, ap.full_name, ap.email, ap.phone_number, 
                   ap.created_at, ap.last_login
            FROM applicants ap
            WHERE ap.applicant_id = %s
            LIMIT 1
            ''',
            (applicant_id,),
        )
        applicant = cursor.fetchone()
        
        if not applicant:
            flash('Applicant not found.', 'error')
            return redirect(url_for('applicants'))
        
        # Get applicant's applications
        where_clause = 'a.applicant_id = %s'
        params = [applicant_id]
        if branch_id:
            where_clause += ' AND j.branch_id = %s'
            params.append(branch_id)
        
        cursor.execute(
            f'''
            SELECT a.application_id, a.status, a.submitted_at,
                   j.job_id, j.title AS job_title,
                   COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                   COALESCE((
                       SELECT COUNT(*)
                       FROM interviews i
                       WHERE i.application_id = a.application_id
                   ), 0) AS interview_count
            FROM applications a
            JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_clause}
            ORDER BY a.submitted_at DESC
            ''',
            tuple(params),
        )
        applications = cursor.fetchall()
        if applications:
            for app in applications:
                try:
                    interview_count = app.get('interview_count', 0) or 0
                    status_value = (app.get('status') or 'pending').strip().lower()
                    # Note: Don't auto-change status based on interview_count
                    # Status should be: pending -> scheduled (when interview scheduled) -> interviewed (when interview completed)
                    # If status is 'scheduled', it should remain 'scheduled' until interview is marked as completed
                except Exception:
                    continue
        
        # Get applicant's resumes
        cursor.execute(
            '''
            SELECT resume_id, file_name, file_path, uploaded_at, file_size_bytes
            FROM resumes
            WHERE applicant_id = %s
            ORDER BY uploaded_at DESC
            ''',
            (applicant_id,),
        )
        resumes = cursor.fetchall() or []
        
        # Format resumes for template
        formatted_resumes = []
        for resume in resumes:
            file_path = (resume.get('file_path') or '').replace('\\', '/')
            file_name = resume.get('file_name') or os.path.basename(file_path) or 'Resume'
            file_size = format_file_size(resume.get('file_size_bytes') or 0) if resume.get('file_size_bytes') else 'Unknown'
            
            formatted_resumes.append({
                'resume_id': resume.get('resume_id'),
                'file_name': file_name,
                'file_size': file_size,
                'uploaded_at': format_human_datetime(resume.get('uploaded_at')) if resume.get('uploaded_at') else 'N/A',
                'view_url': url_for('admin_view_resume', resume_id=resume.get('resume_id')),
                'download_url': url_for('admin_download_resume', resume_id=resume.get('resume_id')),
                'is_pdf': file_name.lower().endswith('.pdf'),
                'file_path': file_path,
            })
        
        # Mark all applications as viewed when HR/Admin views the applicant profile
        if applications:
            application_ids = [app['application_id'] for app in applications]
            # Check if viewed_at column exists, if not, add it dynamically
            try:
                cursor.execute('SHOW COLUMNS FROM applications LIKE "viewed_at"')
                has_viewed_at = cursor.fetchone() is not None
                
                if not has_viewed_at:
                    # Add viewed_at column if it doesn't exist
                    cursor.execute('ALTER TABLE applications ADD COLUMN viewed_at DATETIME NULL')
                    db.commit()
                
                # Update viewed_at for all applications (only if not already viewed)
                placeholders = ','.join(['%s'] * len(application_ids))
                cursor.execute(
                    f'''
                    UPDATE applications 
                    SET viewed_at = NOW() 
                    WHERE application_id IN ({placeholders}) 
                    AND viewed_at IS NULL
                    ''',
                    tuple(application_ids)
                )
                db.commit()
            except Exception as e:
                print(f'⚠️ Error updating viewed_at: {e}')
                # Continue even if viewed_at column doesn't exist
        
        template = 'hr/applicants.html' if user.get('role') == 'hr' else 'admin/applicants.html'
        branch_info = None
        if user.get('role') == 'hr':
            if branch_id:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id,))
                if branch_rows:
                    branch_info = branch_rows[0]
        
        # Use HR template for HR users, admin template for admin users
        view_template = 'hr/view_applicant.html' if user.get('role') == 'hr' else 'admin/view_applicant.html'
        
        return render_template(view_template, 
                             applicant=applicant, 
                             applications=applications,
                             resumes=formatted_resumes,
                             branch_info=branch_info)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ View applicant error: {exc}')
        print(f'Full traceback: {error_details}')
        flash('An error occurred while loading applicant details.', 'error')
        return redirect(url_for('applicants'))
    finally:
        if cursor:
            cursor.close()


@app.route('/admin/applicants/<int:applicant_id>/verify', methods=['POST'])
@login_required('admin', 'hr')
def verify_applicant(applicant_id):
    """Verify an applicant's account. HR can only verify applicants from their branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        return jsonify({'success': False, 'error': 'Database error'}), 500
    
    cursor = db.cursor(dictionary=True)
    try:
        # If HR, verify applicant applied to their branch jobs
        branch_id = get_branch_scope(user)
        if branch_id:
            cursor.execute(
                '''
                SELECT a.applicant_id
                FROM applicants a
                JOIN applications ap ON a.applicant_id = ap.applicant_id
                JOIN jobs j ON ap.job_id = j.job_id
                WHERE a.applicant_id = %s AND j.branch_id = %s
                LIMIT 1
                ''',
                (applicant_id, branch_id),
            )
            if not cursor.fetchone():
                return jsonify({'success': False, 'error': 'You can only verify applicants from your branch.'}), 403
        
        cursor.execute(
            'UPDATE applicants SET is_verified = TRUE, email_verified_at = NOW() WHERE applicant_id = %s',
            (applicant_id,),
        )
        db.commit()
        return jsonify({'success': True})
    except Exception as exc:
        db.rollback()
        print(f'❌ Verify applicant error: {exc}')
        return jsonify({'success': False, 'error': str(exc)}), 500
    finally:
        cursor.close()


@app.route('/admin/add-job-posting')
@login_required('admin', 'hr')
def add_job_posting():
    return redirect(url_for('job_postings'))


@app.route('/admin/add-hr-account')
@login_required('admin')
def add_hr_account():
    return redirect(url_for('hr_accounts'))


@app.route('/admin/reset-all-data', methods=['GET', 'POST'])
@csrf.exempt
@login_required('admin')
def reset_all_data():
    """Reset all system data to zero while keeping admin/HR accounts and branches."""
    user = get_current_user()
    # Only allow admin users (not HR)
    if user.get('role') != 'admin':
        flash('Access denied. Only administrators can reset system data.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'GET':
        # Show confirmation page
        return render_template('admin/reset_data_confirmation.html')
    
    # POST request - validate confirmation text
    confirm_text = request.form.get('confirm_text', '').strip().upper()
    if confirm_text != 'RESET':
        flash('Invalid confirmation. Please type "RESET" to confirm the reset.', 'error')
        return redirect(url_for('reset_all_data'))
    
    # Debug: Check if CSRF token is present
    csrf_token_received = request.form.get('csrf_token')
    if not csrf_token_received:
        flash('CSRF token is missing. Please refresh the page and try again.', 'error')
        return redirect(url_for('reset_all_data'))
    
    # POST request - perform the reset
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    cursor = db.cursor(dictionary=True)
    
    try:
        # Disable foreign key checks temporarily
        cursor.execute('SET FOREIGN_KEY_CHECKS = 0')
        
        # Delete all data from tables (in order to respect foreign keys)
        # Keep: branches, admin table, users table (but only admin/hr users)
        
        tables_to_clear = [
            'activity_logs',
            'results',
            'interviews',
            'notifications',
            'saved_jobs',
            'applications',
            'jobs',
            'resumes',
            'profile_changes',
            'password_resets',
            'auth_sessions',
        ]
        
        deleted_counts = {}
        
        for table in tables_to_clear:
            try:
                cursor.execute(f'DELETE FROM {table}')
                deleted_counts[table] = cursor.rowcount
                print(f'✅ Cleared {table}: {deleted_counts[table]} rows')
            except Exception as e:
                print(f'⚠️ Error clearing {table}: {e}')
                deleted_counts[table] = 0
        
        # Delete all applicants (but keep admin/hr accounts in users table)
        cursor.execute('DELETE FROM applicants')
        applicants_deleted = cursor.rowcount
        deleted_counts['applicants'] = applicants_deleted
        print(f'✅ Cleared applicants: {applicants_deleted} rows')
        
        # Delete only applicant users from users table (keep admin and hr users)
        cursor.execute("DELETE FROM users WHERE user_type = 'applicant'")
        applicant_users_deleted = cursor.rowcount
        deleted_counts['users (applicants only)'] = applicant_users_deleted
        print(f'✅ Cleared applicant users: {applicant_users_deleted} rows')
        
        # Reset AUTO_INCREMENT for all cleared tables
        auto_increment_tables = [
            'activity_logs', 'results', 'interviews', 'notifications', 
            'saved_jobs', 'applications', 'jobs', 'resumes', 
            'profile_changes', 'password_resets', 'auth_sessions', 'applicants'
        ]
        
        for table in auto_increment_tables:
            try:
                cursor.execute(f'ALTER TABLE {table} AUTO_INCREMENT = 1')
                print(f'✅ Reset AUTO_INCREMENT for {table}')
            except Exception as e:
                print(f'⚠️ Could not reset AUTO_INCREMENT for {table}: {e}')
        
        # Re-enable foreign key checks
        cursor.execute('SET FOREIGN_KEY_CHECKS = 1')
        
        # Commit the transaction
        db.commit()
        
        total_deleted = sum(deleted_counts.values())
        flash(f'✅ System data reset successfully! Deleted {total_deleted} records. Admin/HR accounts and branches have been preserved.', 'success')
        print(f'✅ System reset complete. Total records deleted: {total_deleted}')
        
        return redirect(url_for('admin_dashboard'))
        
    except Exception as e:
        db.rollback()
        print(f'❌ Error resetting system data: {e}')
        import traceback
        traceback.print_exc()
        flash(f'Error resetting system data: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))
    finally:
        cursor.close()


@app.route('/admin/profile', methods=['GET', 'POST'])
@login_required('admin', 'hr')
def admin_profile():
    """Admin profile management."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_dashboard'))

    cursor = db.cursor(dictionary=True)
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'update_profile':
                full_name = request.form.get('full_name', '').strip()
                email = request.form.get('email', '').strip().lower()
                
                if not full_name or not email:
                    flash('Full name and email are required.', 'error')
                else:
                    # Check if email is already in use in users table
                    cursor.execute(
                        '''
                        SELECT u.user_id 
                        FROM users u
                        JOIN admins a ON a.user_id = u.user_id
                        WHERE u.email = %s AND a.admin_id <> %s
                        LIMIT 1
                        ''',
                        (email, user.get('id')),
                    )
                    if cursor.fetchone():
                        flash('Email address is already in use.', 'error')
                    else:
                        # Get user_id first
                        cursor.execute(
                            'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                            (user.get('id'),),
                        )
                        admin_record = cursor.fetchone()
                        if admin_record:
                            user_id = admin_record['user_id']
                            # Update email in users table
                            cursor.execute(
                                'UPDATE users SET email = %s WHERE user_id = %s',
                                (email, user_id),
                            )
                            # Update full_name and email in admins table
                            cursor.execute(
                                'UPDATE admins SET full_name = %s, email = %s WHERE admin_id = %s',
                                (full_name, email, user.get('id')),
                            )
                            # Determine if user is admin or HR
                            user_role = user.get('role', '').lower()
                            role_label = 'Admin' if user_role == 'admin' else 'HR'
                            branch_label = user.get('branch_name') or 'Unassigned Branch'
                            
                            # Only show branch label for HR, not for admin
                            if user_role == 'admin':
                                notification_msg = f'Admin {full_name} updated their profile information.'
                            else:
                                notification_msg = f'HR {full_name} ({branch_label}) updated their profile information.'
                            
                            create_admin_notification(cursor, notification_msg)
                            db.commit()
                            session['user_name'] = full_name
                            session['user_email'] = email
                            flash('Profile updated successfully.', 'success')
                        else:
                            flash('Account not found.', 'error')
            
            elif action == 'change_password':
                current_password = request.form.get('current_password', '').strip()
                new_password = request.form.get('new_password', '').strip()
                confirm_password = request.form.get('confirm_password', '').strip()
                
                # Get user_id from admins table, then get password_hash from users table
                cursor.execute(
                    'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                    (user.get('id'),),
                )
                admin_record = cursor.fetchone()
                
                if not admin_record:
                    flash('Account not found.', 'error')
                else:
                    user_id = admin_record['user_id']
                    cursor.execute(
                        'SELECT password_hash FROM users WHERE user_id = %s LIMIT 1',
                        (user_id,),
                    )
                    record = cursor.fetchone()
                    
                    if not record or not check_password(record.get('password_hash'), current_password):
                        flash('Current password is incorrect.', 'error')
                    elif len(new_password) < 6:
                        flash('New password must be at least 6 characters.', 'error')
                    elif new_password != confirm_password:
                        flash('New passwords do not match.', 'error')
                    else:
                        cursor.execute(
                            'UPDATE users SET password_hash = %s WHERE user_id = %s',
                            (hash_password(new_password), user_id),
                        )
                        # Determine if user is admin or HR
                        user_role = user.get('role', '').lower()
                        user_name = session.get('user_name') or user.get('user_name') or user.get('name') or user.get('full_name') or 'User'
                        branch_label = user.get('branch_name') or 'Unassigned Branch'
                        
                        # Only show branch label for HR, not for admin
                        if user_role == 'admin':
                            notification_msg = f'Admin {user_name} changed their account password.'
                        else:
                            notification_msg = f'HR {user_name} ({branch_label}) changed their account password.'
                        
                        create_admin_notification(cursor, notification_msg)
                        db.commit()
                        flash('Password updated successfully.', 'success')
            
            return redirect(url_for('admin_profile'))
        
        # Get profile data from admins and users tables
        cursor.execute(
            '''
            SELECT 
                a.admin_id,
                a.full_name,
                a.email,
                u.user_id,
                u.user_type,
                u.is_active,
                a.last_login,
                a.created_at
            FROM admins a
            JOIN users u ON u.user_id = a.user_id
            WHERE a.admin_id = %s
            LIMIT 1
            ''',
            (user.get('id'),),
        )
        profile = cursor.fetchone()
        if profile:
            # Map user_type to role for backward compatibility
            profile['role'] = 'admin' if profile.get('user_type') == 'super_admin' else 'hr'
        
        # Get login history using user_id from users table
        login_history = []
        if profile and profile.get('user_id'):
            try:
                cursor.execute('SHOW COLUMNS FROM auth_sessions')
                session_columns_raw = cursor.fetchall()
                session_columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in session_columns_raw}
                
                if 'last_activity' in session_columns and 'logout_time' in session_columns:
                    logout_expr = 'COALESCE(last_activity, logout_time)'
                elif 'logout_time' in session_columns:
                    logout_expr = 'logout_time'
                elif 'last_activity' in session_columns:
                    logout_expr = 'last_activity'
                else:
                    logout_expr = 'NULL'
                
                cursor.execute(
                    f'''
                    SELECT login_time, {logout_expr} AS logout_time, COALESCE(is_active, 1) AS is_active
                    FROM auth_sessions
                    WHERE user_id = %s
                    ORDER BY login_time DESC
                    LIMIT 10
                    ''',
                    (profile['user_id'],),
                )
                login_history = []
                for row in cursor.fetchall() or []:
                    is_active = bool(row.get('is_active', 1))
                    logout_value = format_human_datetime(row.get('logout_time')) if row.get('logout_time') else None
                    login_history.append({
                        'login_time': format_human_datetime(row.get('login_time')),
                        'logout_time': logout_value,
                        'is_active': is_active,
                    })
            except Exception as hist_error:
                print(f'⚠️ Login history error: {hist_error}')
                login_history = []
        
        return render_template('admin/profile.html', profile=profile, login_history=login_history, branches=fetch_branches())
    except Exception as exc:
        db.rollback()
        print(f'❌ Admin profile error: {exc}')
        flash('Unable to load profile.', 'error')
        return redirect(url_for('admin_dashboard'))
    finally:
        cursor.close()


@app.route('/hr/profile', methods=['GET', 'POST'])
@login_required('hr')
def hr_profile():
    """HR profile management."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_dashboard'))

    cursor = db.cursor(dictionary=True)
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'update_profile':
                full_name = request.form.get('full_name', '').strip()
                email = request.form.get('email', '').strip().lower()
                
                if not full_name or not email:
                    flash('Full name and email are required.', 'error')
                else:
                    # Check if email is already in use in users table
                    cursor.execute(
                        '''
                        SELECT u.user_id 
                        FROM users u
                        JOIN admins a ON a.user_id = u.user_id
                        WHERE u.email = %s AND a.admin_id <> %s
                        LIMIT 1
                        ''',
                        (email, user.get('id')),
                    )
                    if cursor.fetchone():
                        flash('Email address is already in use.', 'error')
                    else:
                        # Get user_id first
                        cursor.execute(
                            'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                            (user.get('id'),),
                        )
                        admin_record = cursor.fetchone()
                        if admin_record:
                            user_id = admin_record['user_id']
                            # Update email in users table
                            cursor.execute(
                                'UPDATE users SET email = %s WHERE user_id = %s',
                                (email, user_id),
                            )
                            # Update full_name and email in admins table
                            cursor.execute(
                                'UPDATE admins SET full_name = %s, email = %s WHERE admin_id = %s',
                                (full_name, email, user.get('id')),
                            )
                            db.commit()
                            session['user_name'] = full_name
                            session['user_email'] = email
                            flash('Profile updated successfully.', 'success')
                        else:
                            flash('Account not found.', 'error')
            
            elif action == 'change_password':
                current_password = request.form.get('current_password', '').strip()
                new_password = request.form.get('new_password', '').strip()
                confirm_password = request.form.get('confirm_password', '').strip()
                
                # Get user_id from admins table, then get password_hash from users table
                cursor.execute(
                    'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                    (user.get('id'),),
                )
                admin_record = cursor.fetchone()
                
                if not admin_record:
                    flash('Account not found.', 'error')
                else:
                    user_id = admin_record['user_id']
                    cursor.execute(
                        'SELECT password_hash FROM users WHERE user_id = %s LIMIT 1',
                        (user_id,),
                    )
                    record = cursor.fetchone()
                    
                    if not record or not check_password(record.get('password_hash'), current_password):
                        flash('Current password is incorrect.', 'error')
                    elif len(new_password) < 6:
                        flash('New password must be at least 6 characters.', 'error')
                    elif new_password != confirm_password:
                        flash('New passwords do not match.', 'error')
                    else:
                        cursor.execute(
                            'UPDATE users SET password_hash = %s WHERE user_id = %s',
                            (hash_password(new_password), user_id),
                        )
                        db.commit()
                        flash('Password updated successfully.', 'success')
            
            return redirect(url_for('hr_profile'))
        
        # Get profile data from admins and users tables
        cursor.execute(
            '''
            SELECT 
                a.admin_id,
                a.full_name,
                a.email,
                u.user_id,
                u.user_type,
                u.is_active,
                a.last_login,
                a.created_at
            FROM admins a
            JOIN users u ON u.user_id = a.user_id
            WHERE a.admin_id = %s
            LIMIT 1
            ''',
            (user.get('id'),),
        )
        profile = cursor.fetchone()
        if profile:
            profile['role'] = 'hr'
        
        # Get branch info from session (HR users manage branches via session, not direct branch_id)
        branch_info = None
        branch_id = session.get('branch_id')
        if branch_id:
            cursor.execute(
                'SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s LIMIT 1',
                (branch_id,),
            )
            branch_info = cursor.fetchone()
        
        return render_template('hr/profile.html', profile=profile, login_history=[], branch_info=branch_info)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ HR profile error: {exc}')
        print(f'Full traceback: {error_details}')
        flash('Unable to load profile.', 'error')
        return redirect(url_for('hr_dashboard'))
    finally:
        cursor.close()


@app.route('/admin/security', methods=['GET', 'POST'])
@login_required('admin')
def admin_security():
    """Security administration and audit logs."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/security.html', audit_logs=[], activity_logs=[], security_events=[])
    
    cursor = db.cursor(dictionary=True)
    try:
        # Get activity logs from activity_logs table - show only HR users
        activity_logs = []
        try:
            cursor.execute(
                '''
                SELECT 
                    al.log_id,
                    al.action,
                    al.target_table,
                    al.target_id,
                    al.details,
                    al.created_at,
                    COALESCE(a.full_name, 'System') AS admin_name,
                    a.email AS admin_email,
                    'HR' AS user_role
                FROM activity_logs al
                LEFT JOIN admins a ON a.admin_id = al.admin_id
                LEFT JOIN users u ON u.user_id = a.user_id
                WHERE u.user_type = 'hr'
                ORDER BY al.created_at DESC
                LIMIT 500
                '''
            )
            activity_logs = cursor.fetchall()
            # Format timestamps
            for log in activity_logs:
                if log.get('created_at'):
                    log['created_at'] = format_human_datetime(log['created_at'])
        except Exception:
            # Table might not exist, that's okay
            pass
        
        # Get security events (failed logins, suspicious activity)
        # For now, we'll identify suspicious patterns from auth_sessions
        cursor.execute(
            '''
            SELECT 
                u.email,
                COUNT(*) AS attempt_count,
                MAX(s.login_time) AS last_attempt
            FROM auth_sessions s
            JOIN users u ON u.user_id = s.user_id
            WHERE s.login_time >= DATE_SUB(NOW(), INTERVAL 24 HOUR)
            GROUP BY u.user_id, u.email
            HAVING attempt_count > 10
            ORDER BY attempt_count DESC
            LIMIT 50
            '''
        )
        security_events = cursor.fetchall()
        
        return render_template('admin/security.html', 
                             activity_logs=activity_logs,
                             security_events=security_events)
    except Exception as exc:
        db.rollback()
        print(f'❌ Security admin error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to load security logs.', 'error')
        return render_template('admin/security.html', audit_logs=[], activity_logs=[], security_events=[])
    finally:
        cursor.close()


@app.route('/jobs')
def jobs():
    """Public job listings with filters and smart matching.
    
    Applicants can see ALL jobs from ALL branches by default.
    Branch filtering only applies when explicitly selected via branch_id filter.
    """
    filters = {
        'keyword': request.args.get('keyword', '').strip(),
        'branch_id': request.args.get('branch_id', type=int),
        'position_id': request.args.get('position_id', type=int),
        'saved_only': request.args.get('saved_only', '').strip(),
    }
    # Keep saved_only if it's checked (value is '1'), otherwise remove it
    if filters.get('saved_only') != '1':
        filters.pop('saved_only', None)
    # Only keep filters with actual values (remove empty/None)
    filters = {k: v for k, v in filters.items() if v}
    
    applicant_id = session.get('user_id') if is_logged_in() and session.get('user_role') == 'applicant' else None
    
    # IMPORTANT: Pass filters dict (even if empty) or None - applicants see ALL branches by default
    # Only filter by branch if branch_id is explicitly provided in request
    job_listings = fetch_open_jobs(filters if filters else None, applicant_id)
    
    branches = fetch_branches()
    positions = fetch_positions()
    
    return render_template(
        'applicant/jobs.html',
        jobs=job_listings,
        branches=branches,
        positions=positions,
        current_filters=filters,
    )


@app.route('/about')
def about():
    return render_template('about.html')


@app.route('/logout')
def logout():
    """Logout user and redirect to login page."""
    # Logout user first (updates database)
    if is_logged_in():
        logout_user()
    
    # Force clear ALL session data to prevent any redirect loops
    session.clear()
    
    # Ensure session is completely empty
    for key in list(session.keys()):
        session.pop(key, None)
    
    # Flash logout message
    flash('You have been logged out successfully.', 'success')
    
    # Always redirect to login page using immediate redirect
    return immediate_redirect(url_for('login', _external=True))

# Handle favicon requests to prevent 404 errors
@app.route('/favicon.ico')
def favicon():
    """Handle favicon requests."""
    try:
        return send_from_directory('static/images', 'whitehat_logo.jpg', mimetype='image/jpeg')
    except Exception:
        # Return empty response if favicon not found
        return ('', 204)

# Handle Chrome DevTools requests to prevent 404 errors in logs
@app.route('/.well-known/appspecific/com.chrome.devtools.json')
def chrome_devtools():
    """Handle Chrome DevTools requests."""
    return ('', 204)


@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    """Handle CSRF token errors."""
    import traceback
    print(f'❌ CSRF Error: {e}')
    print(f'Traceback: {traceback.format_exc()}')
    flash('Security error: Your session has expired. Please refresh the page and try again.', 'error')
    # If it's a login route, return login page
    if request.endpoint == 'login' or '/login' in request.path:
        return render_template('login.html'), 400
    referrer = request.referrer or url_for('index', _external=True)
    return immediate_redirect(referrer), 400


# Security headers
@app.after_request
def set_security_headers(response):
    """Add security headers to all responses."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    
    # Use Content-Security-Policy instead of X-Frame-Options
    response.headers['Content-Security-Policy'] = "frame-ancestors 'none'"
    
    # Only add HSTS in production with HTTPS
    if not app.debug and request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains; preload'
    
    # Set cache control for static resources
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    
    # Set UTF-8 charset for HTML responses
    if response.mimetype and 'text/html' in response.mimetype:
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
    elif response.mimetype == 'application/json':
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
    
    # Remove redirect pages - ensure all redirects are immediate HTTP 302 redirects
    if response.status_code in (301, 302, 303, 307, 308):
        if 'Location' in response.headers:
            # Add cache control headers to prevent redirect page caching
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            
            # Remove redirect page HTML if present
            if hasattr(response, 'data') and response.data:
                try:
                    body_text = response.data.decode('utf-8', errors='ignore')
                    # Check if it's Flask's default redirect page
                    if 'Redirecting' in body_text or 'should be redirected' in body_text.lower() or 'redirect automatically' in body_text.lower():
                        # Replace redirect page with empty body for immediate redirect
                        response.data = b''
                except:
                    pass
    
    return response


if __name__ == '__main__':
    import os
    print('[*] Starting J&T Express Recruitment System...')
    with app.app_context():
        ensure_default_accounts()

    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    port = int(os.environ.get('PORT', 5000))
    host = os.environ.get('HOST', '0.0.0.0')
    
    app.run(debug=debug_mode, host=host, port=port)